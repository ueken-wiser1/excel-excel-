import numpy as np
from scipy.optimize import minimize
import pandas as pd
import glob
from datetime import timedelta
import datetime
import os
from openpyxl import load_workbook

t = datetime.datetime.now().time()
d = datetime.datetime.now()
d1 = d.strftime('%Y%m%d')
wb_record_list = load_workbook('C:/Users/touko/OneDrive/株価分析/excel/株式データ/test/20230516/20230517_重み付け記録.xlsx')
ws_record_list = wb_record_list.active
watch_list_folder = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/test/20230516/OSCI/'
file_list = glob.glob(watch_list_folder+'*.xlsx')

# 多項式
def f(coefficients, x):
    return np.dot(x, coefficients)

# 最小二乗誤差
def objective_function(coefficients):
    return np.sum((pseudo_scores - f(coefficients, rankings))**2)

for l in file_list:
    lastrow_rec = ws_record_list.max_row
# Excelファイルからデータを読み込む
    df = pd.read_excel(l)

    # 3変数の値を含む列を抽出してnumpy配列に変換
    rankings = df[['RSIスコア', 'ボリンジャーバンドスコア', 'MACDスコア']].values

    # 疑似スコアを含む列を抽出してnumpy配列に変換
    pseudo_scores = df['項.1'].values





    # 初期値（a,b,c）
    initial_guess = [1, 1, 1]

    # 最適化
    result = minimize(objective_function, initial_guess)

    a_optimized, b_optimized, c_optimized = result.x

    # 結果
    print(f'Optimized coefficients are {result.x}')
    file_name = os.path.basename(l)
    first_8_chars = file_name[:8]
    #print(str(a_optimized) + str(b_optimized) + str(c_optimized))
    ws_record_list.cell(lastrow_rec+1,1).value=first_8_chars
    ws_record_list.cell(lastrow_rec+1,2).value=a_optimized
    ws_record_list.cell(lastrow_rec+1,3).value=b_optimized
    ws_record_list.cell(lastrow_rec+1,4).value=c_optimized

wb_record_list.save('C:/Users/touko/OneDrive/株価分析/excel/株式データ/test/20230516/20230517_重み付け記録.xlsx')