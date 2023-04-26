import os
from re import L
import openpyxl
import pandas
import requests
import bs4
import time
import datetime
import glob
import re
import sys
import shutil
import winsound

#どんな動きをさせるのか
#1.     株式データフォルダ内のallkabu1について、銘柄スプレッドする
#2.     銘柄スプレッド：その日の市場終値を各銘柄ファイルに書き込む
#3.     銘柄スプレッドが完了したら、対象ファイルは名前を日付付きに変更する
#4.     対象ファイルを完了フォルダに移動する
#5.     
#6.     
#7.     
#8.     
#9.     
#10.    
#11.    
#12.    

#要確認事項
#1. file_
#プログラム

dirdaily = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/"
dirmerge = "D:/株取引/20221031_銘柄データ/"
dirstorage = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/完了/"
file_list = glob.glob(dirdaily + '*.xlsx')
stock_list = glob.glob(dirmerge + '*.xlsx')

t = datetime.datetime.now().time()
d = datetime.datetime.now()
d1 = d.strftime('%Y%m%d')

for l in file_list:

    wb_market = openpyxl.load_workbook(l)
    print(l)
    sheetmarket = wb_market.worksheets[0]
    for j in range(2, sheetmarket.max_row-1):
        stock_code = sheetmarket.cell(row=j, column=2).value
        print(stock_code)
        book_search_list = glob.glob(dirmerge + str(stock_code) + '*.xlsx')
        
        if len(book_search_list) == 0:
                stock_name = sheetmarket.cell(row=j, column=3).value
                wb_new = openpyxl.Workbook()
                sheet_new = wb_new.active
                wb_new.save(dirmerge+stock_code+'_'+stock_name+'.xlsx')
#                print('book_search_listが空の場合')
                for h in range(1, sheetmarket.max_column+1):
                        itemname = sheetmarket.cell(row=1, column=h).value
                        row_data = sheetmarket.cell(row=j, column=h).value
                        sheet_new.cell(row=1, column=h, value=itemname)
                        sheet_new.cell(row=2, column=h, value=row_data)
                        h += 1
                wb_new.save(dirmerge+stock_code+'_'+stock_name+'.xlsx')

        else:
                company_book = book_search_list[0]
#                print('book_search_listがある場合')
                print(company_book)

                wb_company = openpyxl.load_workbook(company_book) #フォルダ-銘柄のexcelを開く
                sheetcompany = wb_company.worksheets[0]
                last_row = sheetcompany.max_row
                last_column = sheetcompany.max_column

                for k in range(1, sheetcompany.max_column+1):
                        row_copy = sheetmarket.cell(row=j, column=k).value

                        sheetcompany.cell(row = last_row+1, column=k, value=row_copy)

                        k += 1
                wb_company.save(company_book)
                wb_company.close()
        j += 1

#8.     excel-銘柄を閉じる

#10.    excel-市場の全てのデータをコピーしたら、excel-市場を閉じる
    wb_market.close()
#    os.rename(dirdaily+'allkabu1.xlsx', dirdaily+d1+'_allkabu1.xlsx')
#    new_path = shutil.move(dirdaily+d1+'_allkabu1.xlsx', dirstorage)
#12.    フォルダ-Aの全てのファイルを走査したら、プログラムを終了する
print(t)
t = datetime.datetime.now().time()
print(t)

#稼働終了アナウンス
import winsound
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾