import os
import openpyxl
import requests
import bs4
import time
import datetime

#どんな動きをさせるのか
#excelを開く
#参照excelの各シートに記載された証券コードを読み込む
#シート名は"株式", "マーケット", "為替", "投信"
#読み込んだ証券コードにより株探URLを作成し、銘柄ページへ移動
#必要な情報を取得する

#excelを開く

wb = openpyxl.load_workbook('screeningcodelist.xlsx')
name = wb.get_sheet_names

#"株式"
sheet = wb.get_sheet_by_name('株式')
print(sheet.title)
print(name)

#読み込んだ証券コードにより株探URLを作成し、銘柄ページへ移動
#必要な情報を取得する

#日付を入れる
d_today = datetime.date.today()
sheet.cell(row=1, column=1).value = d_today

#参照excelの各シートに記載された証券コードを読み込む

sheet = wb.get_sheet_by_name('株式')
print(sheet)
for j in range(3, sheet.max_row + 1):
    time.sleep(0.2)
    kabutan_URL_base = 'http://kabutan.jp/stock/?code='
    stock_code = sheet.cell(row=j, column=2).value
    kabutan_URL = kabutan_URL_base + str(stock_code)
    res = requests.get(kabutan_URL)
    print(kabutan_URL)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')

#読み込んだ情報をexcelファイルに書き込む
#書き出す初めのセルの列を指定
    write_column = 4

#           現在株価
    try:
        stock_price = str(soup.select('td')[32])
    except IndexError as e:
        print('現在株価存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
    write_column += 1

#           前日比
    try:
        DoD = str(soup.select('span')[15])
    except IndexError as e:
        print('前日比存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = DoD
    write_column += 1
 
#           出来高
    try:
        dekidaka_yield = str(soup.select('td')[35])
    except IndexError as e:
        print('出来高存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = dekidaka_yield
    write_column += 1

#           決算日
    try:
        industry_type = str(soup.select('time')[5])
    except IndexError as e:
        print('決算日存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = industry_type
    write_column += 1

wb.save('screeningkabu1.xlsx')

