import imp
import os
from re import L
from turtle import color
import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd
import numpy as np
import bottleneck as bn
import requests
import bs4
import time
import datetime
import glob
import re
import sys
import winsound


#どんな動きをさせるのか
#1. 銘柄データ格納フォルダを取得
#2. 銘柄データを順番に開く
#3. 指定の列に日次データを計算
#4.     
#5.     
#6.     
#7.     
#8.     
#9.     
#10.    
#11.    
#12.    

#要確認事項
#1. file_

#プログラム

#testのため、仮フォルダ設定
dirmerge = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/test/20230421/"


stock_list = glob.glob(dirmerge + '*.xlsx')

#開始時間取得
t = datetime.datetime.now()
d = datetime.datetime.now()
d1 = d.strftime('%Y%m%d')
#開始時間取得

#testのため、exceptリストは使わなくておｋ
#except_list = ['C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/0000', 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/0010', '0011', '0012', '0090', '0091', '0092', '0093', '0094', '0095', '0101', '0102', '0108', '0800', '0802']

for l in stock_list:
        if "None.xlsx" in l:
                continue #ファイル名に"None.xlsx"が含まれていた場合はスキップ

        wb = openpyxl.load_workbook(l)
        sheetstock = wb.worksheets[0]
        lastrow_stockbook = sheetstock.max_row+1
        lastcolumn_stockbook = sheetstock.max_column

#配列宣言
        stockprice = [] #株価
        turnover = [] #出来高
        upprice = [] #上昇幅
        downprice = [] #下落幅
        stockprice_5dmovemean = [] #5日移動平均
        stockprice_25dmovemean = [] #25日移動平均
        stockprice_75dmovemean = [] #75日移動平均
        stockprice_10dmovemean = [] #10日移動平均
        stockprice_15dmovemean = [] #15日移動平均
        stockprice_20dmovemean = [] #20日移動平均
        stockprice_30dmovemean = [] #30日移動平均
        stockprice_65dmovemean = [] #13週移動平均
        stockprice_130dmovemean = [] #26週移動平均
        turnover_5movemean = [] #5日移動平均出来高
        turnover_10movemean = [] #10日移動平均出来高
        stockprice_25dmovestd = [] #25日移動標準偏差
        stockprice_65dmovestd = [] #13週移動標準偏差
#20230409追加
        stockprice_20dmovestd = [] #20日移動標準偏差ーボリンジャーバンド
        stockprice_20dmovedtl = [] #20日移動分散ーボリンジャーバンド
        stockprice_14dupmean = [] #14日上昇平均ーRSI
        stockprice_14ddownmean = [] #14日下落平均ーRSI
        stockprice_12dEmovemean = [] #12日指数平滑移動平均ーMACD
        stockprice_26dEmovemean = [] #26日指数平滑移動平均ーMACD
        stockprice_9dEmovemean = [] #9日指数平滑移動平均ーMACD
        macdscore=[] #macdスコア用配列
        diffarray=[] #macdスコア計算用
        diffarray_max=[]
        diffarray_min=[]



#シートの2行目~最終行をループ
        for i in range(2,lastrow_stockbook):

#A列を配列へ格納
#                print(i)
                stockprice = np.append(stockprice, sheetstock.cell(row=i, column=4).value) #株価の配列格納
                turnover = np.append(turnover, sheetstock.cell(row=i, column=11).value) #出来高の配列格納
                #前日比のセルから、upprice, downpriceの配列格納をしたい。
                #前日比マイナスの時はuppriceにゼロ、プラスの時はdownpriceにゼロを入れる形にする。
                if int(sheetstock.cell(i, 5).value) >= 0:
                        upprice = np.append(upprice, sheetstock.cell(i, 5).value)
                        downprice = np.append(downprice, 0)
                        #print(sheetstock.cell(i,5).value)
                else:
                        upprice = np.append(upprice, 0)
                        downprice = np.append(downprice, abs(sheetstock.cell(i,5).value))
#移動平均数値の取込
        if lastrow_stockbook > 7:
                stockprice_5dmovemean = bn.move_mean(stockprice, window=5)
                turnover_5dmovemean = bn.move_mean(turnover, window=5)

                if lastrow_stockbook > 12:
                        stockprice_10dmovemean = bn.move_mean(stockprice, window=10)
                        turnover_10dmovemean = bn.move_mean(turnover, window=10)
                        if lastrow_stockbook > 14:
                                stockprice_12dEmovemean = bn.move_mean(stockprice, window=12)

        for j in range(2, lastrow_stockbook):
                if lastrow_stockbook > 7:
                        sheetstock.cell(row=j,column=199).value = stockprice_5dmovemean[j-2]
                        sheetstock.cell(row=j,column=230).value = turnover_5dmovemean[j-2]

                        if lastrow_stockbook > 12:
                                sheetstock.cell(row=j,column=205).value = stockprice_10dmovemean[j-2]
                                sheetstock.cell(row=j,column=211).value = turnover_10dmovemean[j-2]
                                if lastrow_stockbook > 14:
                                        sheetstock.cell(row=j,column=317).value = stockprice_12dEmovemean[j-2] + (2/13)*(sheetstock.cell(j,4).value - stockprice_12dEmovemean[j-2])
                                                                                        

                if lastrow_stockbook > 7:
                        if sheetstock.cell(j,15).value-sheetstock.cell(j,199).value>0 and sheetstock.cell(j-1,15).value-sheetstock.cell(j-1,199).value<=0:
                                sheetstock.cell(j,229).value=1
                        elif sheetstock.cell(j,15).value-sheetstock.cell(j,199).value<=0 and sheetstock.cell(j-1,15).value-sheetstock.cell(j-1,199).value>0:
                                sheetstock.cell(j,229).value=4
                        else:
                                pass



        for k in range(2, lastrow_stockbook):
                if sheetstock.cell(k,319).value is not None:
                    macdscore = np.append(macdscore, sheetstock.cell(row=k, column=319).value) #macdの配列格納
                if k > 11:
                        stockprice_9dEmovemean = bn.move_mean(macdscore, window=9)
                        sheetstock.cell(row=k,column=320).value = stockprice_9dEmovemean[k-2]
                        diffarray = np.append(macdscore, sheetstock.cell(row=k, column=320).value) #macdの配列格納


print(macdscore)
