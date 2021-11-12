import os
import openpyxl
import requests
import bs4
import time
import datetime
import re
import urllib.request
import urllib
from urllib.parse import urljoin
import urllib3

#------------お約束開始---冒頭
#稼働時間計測開始
import datetime
t = datetime.datetime.now().time()
#------------お約束終了---冒頭

#------------プログラム本文---ここから

today = datetime.date.today()
d = today.strftime('%Y%m%d')

t = datetime.datetime.now().time()
##########################################################################
#信用規制情報を取得
#今日の日付を取得
#対象：信用規制
dir_data01 = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/ダウンロードデータ/01.信用規制中銘柄/"
dir_data02 = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/ダウンロードデータ/02.信用規制解除銘柄/"

regulationbook01 = openpyxl.Workbook()
sheet01 = regulationbook01.worksheets[0]
regulationbook01.save(dir_data01+str(d)+"_信用規制.xlsx")

regulationbook02 = openpyxl.Workbook()
sheet02 = regulationbook02.worksheets[0]
regulationbook02.save(dir_data02+str(d)+"_信用規制解除.xlsx")



res = requests.get("https://www.jpx.co.jp/markets/equities/margin-reg/index.html")
res.raise_for_status()
soup = bs4.BeautifulSoup(res.text, 'html.parser')
td_datetags = soup.find_all("td",class_="a-left")
j = 2
# print(len(td_datetags))
for i in range(5,len(td_datetags)+1):
    result = soup.select(".a-center")[i]
    # print(i,type(result),result)
    strresult = str(result)
    sliceresult = strresult[21:25]
    # print(sliceresult)
    if sliceresult == "202*":
        pass
    elif sliceresult == "b-co":
        break
    else:
        sheet01.cell(row=j,column=2).value = sliceresult
        j += 1
        
regulationbook01.save(dir_data01+str(d)+"_信用規制.xlsx")
m = 2
# print(len(td_datetags))
for k in reversed(range(len(td_datetags)-1)):
        # print(k)
        # print(soup.select(".a-center")[30])
        result = soup.select(".a-center")[k]
    # print(i,type(result),result)
        strresult = str(result)
        sliceresult = strresult[21:25]
        # print(sliceresult)
        if sliceresult == "202*":
            pass
        elif sliceresult == "b-co":
            break
        else:
            sheet02.cell(row=m,column=2).value = sliceresult
            m += 1
regulationbook02.save(dir_data02+str(d)+"_信用規制解除.xlsx")
#信用規制情報取得を完了
##########################################################################

##########################################################################
#空売規制情報DL
karauri_URL = 'https://www.jpx.co.jp/markets/equities/ss-reg/index.html'
res01 = requests.get(karauri_URL)
print(karauri_URL)
res01.raise_for_status()
soup01 = bs4.BeautifulSoup(res01.text, 'html.parser')
dtags = soup01.find_all("td",class_="a-center")
dtags_con = soup01.select("td a")[1]
url = dtags_con.get("href")
karauri_file_base = 'https://www.jpx.co.jp/'
karauri_file_url = karauri_file_base + url
print(karauri_file_url)

savepoint01 = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/ダウンロードデータ/03.空売り価格規制トリガー抵触銘柄一覧/"+str(d)+"_karauri.csv"
csv = urllib.request.urlopen(karauri_file_url).read()

with open(savepoint01, mode="wb") as f:
    f.write(csv)
#空売規制情報DL
##########################################################################

##########################################################################
#増担措置情報DL
mashitan_URL = 'https://www.taisyaku.jp/brand/'
res02 = requests.get(mashitan_URL)
print(mashitan_URL)
res02.raise_for_status()
soup02 = bs4.BeautifulSoup(res02.text, 'html.parser')
ultags = soup02.find_all("ul",class_="download")
dtags_con = soup02.select("ul a")[12]
url = dtags_con.get("href")
print(url)


savepoint02 = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/ダウンロードデータ/04.貸借取引銘柄別増担保金徴収措置一覧/"+str(d)+"_mashitan.xlsx"
xlsx = urllib.request.urlopen(url).read()

with open(savepoint02, mode="wb") as f:
    f.write(xlsx)
#増担措置情報DL
##########################################################################

##########################################################################
#品残料率情報DL
shinazan_URL = 'https://www.taisyaku.jp/search/result/index/1/'
res03 = requests.get(shinazan_URL)
print(shinazan_URL)
res03.raise_for_status()
soup03 = bs4.BeautifulSoup(res03.text, 'html.parser')
atags = soup03.find_all("a")
dtags_con_shina = soup03.select("a")[22]
dtags_con_kashi = soup03.select("a")[28]
url_shina = dtags_con_shina.get("href")
url_kashi = dtags_con_kashi.get("href")

savepoint03 = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/ダウンロードデータ/05.品貸料率/"+str(d)+"_shinakashi.csv"
savepoint04 = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/ダウンロードデータ/06.銘柄別融資・貸株残高一覧表/"+str(d)+"_kashikabu.csv"

csv_shina = urllib.request.urlopen(url_shina).read()
print(url_shina)
with open(savepoint03, mode="wb") as f:
    f.write(csv_shina)
csv_kashi = urllib.request.urlopen(url_kashi).read()
print(url_kashi)
with open(savepoint04, mode="wb") as f:
    f.write(csv_kashi)
#品貸料率情報DL
##########################################################################

#------------プログラム本文---ここまで

#------------お約束開始---末尾
#稼働時間表示
print(t)
t = datetime.datetime.now().time()
print(t)

#稼働終了アナウンス
import winsound
# winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
# winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
# winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
# winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
# winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾