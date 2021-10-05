
#
#20210926_createStationary_CompanyData
#市場データから銘柄データにコピペするときの、銘柄データのひな形を作成する。
#参照元はC:\Users/touko\program/20210802_銘柄集計allkabu0.xlsx←名前は変更予定：stockcodelist00
#参照元の証券コードと会社名を読み取って、指定したフォルダにexcelを開いて、読み取ったコードと名称をファイル名として保存する。
#一行目に、各列の名称を記載する。


#import
import openpyxl
import datetime
import winsound

#使用ディレクトリ、ファイル指定
dir_code = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/銘柄データ集計/"
dir_company = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/"

#稼働時間計測開始
t = datetime.datetime.now().time()
print(t)

#本文
codelist = openpyxl.load_workbook(dir_code + 'stockcodelist00.xlsx')
sheet01 = codelist.worksheets[0]
lastrow = sheet01.max_row + 1
print(lastrow)

for i in range(2, lastrow):
    code_company = str(sheet01.cell(row=i,column=2).value)
    name_company = str(sheet01.cell(row=i,column=3).value)
    company_book = openpyxl.Workbook()
    sheet02 = company_book.worksheets[0]
#   列名称記載
    sheet02.cell(row=1,column=1).value = "日付"
    sheet02.cell(row=1,column=2).value = "コード"
    sheet02.cell(row=1,column=3).value = "名称"
    sheet02.cell(row=1,column=4).value = "株価"
    sheet02.cell(row=1,column=5).value = "前日比"
    sheet02.cell(row=1,column=6).value = "値幅"
    sheet02.cell(row=1,column=7).value = "売買代金"
    sheet02.cell(row=1,column=8).value = "約定回数"
    sheet02.cell(row=1,column=9).value = "決算日"
    sheet02.cell(row=1,column=10).value = "前日終値"
    sheet02.cell(row=1,column=11).value = "出来高"
    sheet02.cell(row=1,column=12).value = "始値"
    sheet02.cell(row=1,column=13).value = "高値"
    sheet02.cell(row=1,column=14).value = "安値"
    sheet02.cell(row=1,column=15).value = "終値"
    sheet02.cell(row=1,column=16).value = "前日比%"
    sheet02.cell(row=1,column=17).value = "PER"
    sheet02.cell(row=1,column=18).value = "PBR"
    sheet02.cell(row=1,column=19).value = "上場市場"
    sheet02.cell(row=1,column=20).value = "VWAP"
    sheet02.cell(row=1,column=21).value = "発行済み株式数"
    sheet02.cell(row=1,column=22).value = "最新信用売残"
    sheet02.cell(row=1,column=23).value = "最新信用買残"
    sheet02.cell(row=1,column=24).value = "信用倍率"
    sheet02.cell(row=1,column=25).value = "売買代金2"
    sheet02.cell(row=1,column=26).value = "信用売残前週比"
    sheet02.cell(row=1,column=27).value = "信用買残前週比"
    sheet02.cell(row=1,column=28).value = "出来高前日比"
    sheet02.cell(row=1,column=29).value = "約定回数前日比"
    sheet02.cell(row=1,column=30).value = "時価総額"
    sheet02.cell(row=1,column=31).value = "浮動株総額"
    sheet02.cell(row=1,column=32).value = "取引規模"
    sheet02.cell(row=1,column=33).value = "平均約定金額"
    sheet02.cell(row=1,column=51).value = "当日IR等有無"
    sheet02.cell(row=1,column=101).value = "信用取引規制中"
    sheet02.cell(row=1,column=102).value = "増担保措置有無"
    sheet02.cell(row=1,column=103).value = "増担保措置内容"
    sheet02.cell(row=1,column=106).value = "空売規制対象"
    sheet02.cell(row=1,column=111).value = "融資新規"
    sheet02.cell(row=1,column=112).value = "融資返済"
    sheet02.cell(row=1,column=113).value = "融資残高"
    sheet02.cell(row=1,column=114).value = "貸株新規"
    sheet02.cell(row=1,column=115).value = "貸株返済"
    sheet02.cell(row=1,column=116).value = "貸株残高"
    sheet02.cell(row=1,column=117).value = "差引残高"
    sheet02.cell(row=1,column=118).value = "回転日数"
    sheet02.cell(row=1,column=121).value = "貸株超過"
    sheet02.cell(row=1,column=122).value = "最高料率"
    sheet02.cell(row=1,column=123).value = "当日料率"
    sheet02.cell(row=1,column=124).value = "前日料率"
    sheet02.cell(row=1,column=151).value = "みんかぶ目標株価"
    sheet02.cell(row=1,column=152).value = "目標との差分"
    sheet02.cell(row=1,column=201).value = "移動平均5日"
    sheet02.cell(row=1,column=202).value = "移動平均25日"
    sheet02.cell(row=1,column=203).value = "移動平均75日"
    sheet02.cell(row=1,column=204).value = "移動平均乖離率5日"
    sheet02.cell(row=1,column=205).value = "移動平均乖離率25日"
    sheet02.cell(row=1,column=206).value = "移動平均乖離率75日"
    sheet02.cell(row=1,column=251).value = "平均出来高"
    sheet02.cell(row=1,column=252).value = "平均約定回数"
    sheet02.cell(row=1,column=253).value = "平均回転日数"
    sheet02.cell(row=1,column=254).value = "平均移動平均乖離率5日"
    sheet02.cell(row=1,column=255).value = "平均移動平均乖離率25日"
    sheet02.cell(row=1,column=256).value = "平均移動平均乖離率75日"
    sheet02.cell(row=1,column=257).value = "平均信用買残"
    sheet02.cell(row=1,column=258).value = "平均融資新規"
    sheet02.cell(row=1,column=259).value = "平均融資返済"
    sheet02.cell(row=1,column=260).value = "平均信用売残"
    sheet02.cell(row=1,column=261).value = "平均貸株新規"
    sheet02.cell(row=1,column=262).value = "平均貸株返済"
    sheet02.cell(row=1,column=263).value = "平均貸株超過"
    sheet02.cell(row=1,column=264).value = "平均出来高変化率"
    sheet02.cell(row=1,column=265).value = "平均約定回数変化率"
    sheet02.cell(row=1,column=266).value = "平均信用買残変化率"
    sheet02.cell(row=1,column=267).value = "平均信用売残変化率"
    sheet02.cell(row=1,column=268).value = "平均平均約定金額"
    sheet02.cell(row=1,column=269).value = "標準偏差出来高"
    sheet02.cell(row=1,column=270).value = "標準偏差約定回数"
    sheet02.cell(row=1,column=271).value = "標準偏差回転日数"
    sheet02.cell(row=1,column=272).value = "標準偏差移動平均乖離率5日"
    sheet02.cell(row=1,column=273).value = "標準偏差移動平均乖離率25日"
    sheet02.cell(row=1,column=274).value = "標準偏差移動平均乖離率75日"
    sheet02.cell(row=1,column=275).value = "標準偏差信用買残"
    sheet02.cell(row=1,column=276).value = "標準偏差融資新規"
    sheet02.cell(row=1,column=277).value = "標準偏差融資返済"
    sheet02.cell(row=1,column=278).value = "標準偏差信用売残"
    sheet02.cell(row=1,column=279).value = "標準偏差貸株新規"
    sheet02.cell(row=1,column=280).value = "標準偏差貸株返済"
    sheet02.cell(row=1,column=281).value = "標準偏差貸株超過"
    sheet02.cell(row=1,column=282).value = "標準偏差出来高変化率"
    sheet02.cell(row=1,column=283).value = "標準偏差約定回数変化率"
    sheet02.cell(row=1,column=284).value = "標準偏差信用買残変化率"
    sheet02.cell(row=1,column=285).value = "標準偏差信用売残変化率"
    sheet02.cell(row=1,column=286).value = "標準偏差平均約定金額"
    sheet02.cell(row=1,column=501).value = "標準化出来高"
    sheet02.cell(row=1,column=502).value = "標準化約定回数"
    sheet02.cell(row=1,column=503).value = "標準化回転日数"
    sheet02.cell(row=1,column=504).value = "標準化移動平均乖離率5日"
    sheet02.cell(row=1,column=505).value = "標準化移動平均乖離率25日"
    sheet02.cell(row=1,column=506).value = "標準化移動平均乖離率75日"
    sheet02.cell(row=1,column=507).value = "標準化信用買残"
    sheet02.cell(row=1,column=508).value = "標準化融資新規"
    sheet02.cell(row=1,column=509).value = "標準化融資返済"
    sheet02.cell(row=1,column=510).value = "標準化信用売残"
    sheet02.cell(row=1,column=511).value = "標準化貸株新規"
    sheet02.cell(row=1,column=512).value = "標準化貸株返済"
    sheet02.cell(row=1,column=513).value = "標準化貸株超過"
    sheet02.cell(row=1,column=514).value = "標準化出来高変化率"
    sheet02.cell(row=1,column=515).value = "標準化出来高約定回数変化率"
    sheet02.cell(row=1,column=516).value = "標準化信用買残変化率"
    sheet02.cell(row=1,column=517).value = "標準化信用売残変化率"
    sheet02.cell(row=1,column=518).value = "標準化平均約定金額"
    sheet02.cell(row=1,column=1000).value = "売/買時スコア"

#   保存
    company_book.save(dir_company + str(code_company) + '_' +str(name_company) +'.xlsx')

print(t)
t = datetime.datetime.now().time()
print(t)
winsound.Beep(500,50)  #ビープ音（500Hzの音を50msec流す）