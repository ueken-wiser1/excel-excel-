import os
import openpyxl


wb = openpyxl.load_workbook('kabu.xlsx')
print(type(wb))
#name = wb.get_sheet_names
#print(name)
print(wb.get_sheet_names())

for i in range(1,13):
    print(i)
    a = '月配当'
    b = i
    c =str(b) + a
    sheet = wb.get_sheet_by_name(c)
    print(sheet)

    for j in range(1, sheet.max_row):
        print(sheet.cell(row=j, column=1).value)