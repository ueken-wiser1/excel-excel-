import os
import openpyxl
import requests
import bs4
import time
import datetime
import win32com.client
import re
import winsound

#どんな機能か
#株探サイト各銘柄のニュースタブにおいて、過去2ページ分含む全てのニュース日付を記録
#ニュースは分類分けしておく
#1. 開示情報、材料、決算
#2. テク
#3. 5%
#注目、市況、特集は除外
#nglist2のワードについても除外
#

#どんな動きをさせるのか



#------------お約束開始---冒頭
#稼働時間計測開始
import datetime
t = datetime.datetime.now().time()
#------------お約束終了---冒頭

#------------プログラム本文---ここから
#不要なニュースワードリスト
#タグ
ng_list01 = ["ctg1", "ctg5", "ctg9"]
#文字列
ng_list02 = ["Notification", "Report", "Notice", "Financial", "Delayed", "Announcement", "保有状況報告書", "筆頭株主の異動","取得状況", "招集", "コーポレート・ガバナンス", "Governance", "人事異動", "に出展", "前日に動いた銘柄", "本日の", "個人投資家の予想", "法定事前開示書類", "売上高のお知らせ", "月次情報のお知らせ", "説明会", "月次売上", "独立役員届出書", "株主総会招集通知", "Meeting", "動意株・", "社会報告書"]

#関数定義
def is_include_listed_word(text, word_list):
    for listed_word in word_list:
        if listed_word in text:
            return True
    return False

#処理
#stockcodelistを開く
#書込用excelfileを開く
wb_ref_lowpricestock = openpyxl.load_workbook('C:/Users/touko/OneDrive/株価分析/excel/株式データ/スクレイピング生データ/stockcodelist.xlsx')
wb_write_lowpricestock = openpyxl.Workbook()

#stockcodelistのシートを指定
#書込用excelfileのシートを指定
sheet01 = wb_ref_lowpricestock.worksheets[0]
sheet02 = wb_write_lowpricestock.active
k = 2

#パラメータ名の記載
sheet02.cell(row=1, column=1).value = "日付"
sheet02.cell(row=1, column=2).value = "時間"
sheet02.cell(row=1, column=3).value = "証券コード"
sheet02.cell(row=1, column=4).value = "名称"
sheet02.cell(row=1, column=5).value = "株価"
sheet02.cell(row=1, column=6).value = "株価"
sheet02.cell(row=1, column=7).value = "ニュース"
sheet02.cell(row=1, column=8).value = "URL"

for j in range(2, sheet01.max_row + 1):
#    time.sleep(0.1)
    kabutan_newsURL_base = 'https://kabutan.jp/stock/news?code='
    stock_code = sheet01.cell(row=j, column=2).value
#読み込んだ証券コードにより株探URLを作成し、銘柄ページへ移動
    kabutan_newsURL = kabutan_newsURL_base + str(stock_code)
    res = requests.get(kabutan_newsURL)
#    print(kabutan_newsURL)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')
    stock_name = str(soup.select('title')[0].get_text())
    write_column = 1
    stock_price01 = str(soup.select('span')[15].get_text())
    stock_price02 = str(soup.select('span')[16].get_text())

#当日の日付を取得する
    d_today = datetime.date.today()
    str_d_today = str(d_today)
    tdtags = soup.find_all('td')       #全aタグ取得
#    print('tdタグ数：', len(tdtags))  #aタグ数取得
    num_tdtags = int(len(tdtags))
    for i in range(23, 45):
        t_news = str(soup.select('td')[i-2].get_text())
        td_elem = str(soup.select('td')[i-2])
        news_tag = str(soup.select('td')[i-1])   
        td_elem2 = str(soup.select('td')[i].get_text())
        td_elem3 = str(soup.select('a')[i+28].get('href'))
        tf01 = is_include_listed_word(news_tag, ng_list01)
        tf02 = is_include_listed_word(td_elem2, ng_list02)
#        cond02 = [td_elem2 for td_elem2 in l if not in td_elem
#当日の日付と銘柄ニュースページの一致する日付を探す
        if str_d_today in str(td_elem):
#不要なタグがnews_tagに含まれる場合、それは外す
            if tf01 == False :
                if tf02 == False:
#必要な情報を取得する
                    print(stock_code)
#            print(d_today)
                    tdtags=soup.find_all('td')       #全aタグ取得
#            print('tdタグ数：', len(tdtags))  #aタグ数取得
                    sheet02.cell(row=k, column=write_column).value = d_today
                    sheet02.cell(row=k, column=write_column+1).value = t_news
                    sheet02.cell(row=k, column=write_column+2).value = stock_code
                    sheet02.cell(row=k, column=write_column+3).value = stock_name
                    sheet02.cell(row=k, column=write_column+4).value = stock_price01
                    sheet02.cell(row=k, column=write_column+5).value = stock_price02
#                    sheet02.cell(row=k, column=write_column+5).value = news_tag
                    sheet02.cell(row=k, column=write_column+6).value = td_elem2
                    sheet02.cell(row=k, column=write_column+7).value = td_elem3
#            wb_lowpricestock.save('stockcodelist02.xlsx')
                    k += 1

print(t)
t = datetime.datetime.now().time()
print(t)
path = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/ダウンロードデータ/07.IR情報入力処理/"
wb_write_lowpricestock.save(path+str(d_today)+"newsallkabu.xlsx")
#------------プログラム本文---ここまで

#------------お約束開始---末尾
#稼働時間表示
print(t)
t = datetime.datetime.now().time()
print(t)

#稼働終了アナウンス
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾