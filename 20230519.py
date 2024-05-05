
import openpyxl
import datetime
import glob
import jpholiday
import winsound

#開始時間取得
t = datetime.datetime.now()
d = datetime.datetime.now()
d1 = d.strftime('%Y%m%d')
#開始時間取得

#対象フォルダ指定
dirdaily = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/"
dirstorage = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/追跡調査/現在/1日目/"
dividend_data ="C:/Users/touko/OneDrive/株価分析/excel/株式データ/test/20230508/20230508_配当カレンダー.xlsx"

file_list = glob.glob(dirdaily + '*allkabu1.xlsx')

wb_dividend = openpyxl.load_workbook(dividend_data)
sheetdividend = wb_dividend.worksheets[0]
lastrow_dividend = sheetdividend.max_row+1

for l in file_list:
    #日付データを順番に開く
    wb_daily = openpyxl.load_workbook(l)
    wb_sim = openpyxl.Workbook()
    sheetsim = wb_sim.worksheets[0]
    print(l)
    sheetdaily = wb_daily.worksheets[0]
    lastrow = sheetdaily.max_row+1
    lastcolumn = sheetdaily.max_column
    daycode = sheetdaily.cell(2,1).value
    daycode_format = daycode.strftime('%Y%m%d')
    basis_date = daycode
    basis_date2 = basis_date + datetime.timedelta(days=1)

    while True:
        # 指定日が土曜日または日曜日の場合
        if basis_date2.weekday() == 5 or basis_date2.weekday() == 6:
            #print(f"{basis_date}は土曜日または日曜日です。")
            # 指定日を翌日に上書き
            basis_date2 = basis_date2 + datetime.timedelta(days=1)
        # 指定日が祝日の場合
        elif jpholiday.is_holiday(basis_date2):
            #print(f"{basis_date}は{str(jpholiday.is_holiday_name(basis_date))}です。")
            # 指定日を翌日に上書き
            basis_date2 = basis_date2 + datetime.timedelta(days=1)
        # 指定日が平日の場合
        else:
            #print(f"{basis_date}は平日です。")
            # ループを終了
            break

    

#パラメータ名の設定
    sheetsim.cell(1,1).value = "注目完了フラグ"     #コピペ対象(ただし値は無し)
    sheetsim.cell(1,2).value = "日付"               #コピペ対象
    sheetsim.cell(1,3).value = "項"                 #コピペ対象(ただし値は無し)
    sheetsim.cell(1,4).value = "証券コード"         #コピペ対象
    sheetsim.cell(1,5).value = "会社名"             #コピペ対象
    sheetsim.cell(1,6).value = "株価"               #コピペ対象
    sheetsim.cell(1,7).value = "前日比"             #コピペ対象
    sheetsim.cell(1,8).value = "業界"               #コピペ対象
    sheetsim.cell(1,9).value = "PER"                #コピペ対象
    sheetsim.cell(1,10).value = "PBR"               #コピペ対象
    sheetsim.cell(1,11).value = "利回り"            #コピペ対象
    sheetsim.cell(1,12).value = "配当金"
    sheetsim.cell(1,13).value = "配当月"
    sheetsim.cell(1,14).value = "買残"
    sheetsim.cell(1,15).value = "売残"
    sheetsim.cell(1,16).value = "信用倍率"
    sheetsim.cell(1,17).value = "IR有無"
    sheetsim.cell(1,18).value = "リンク"
    sheetsim.cell(1,19).value = "5日線比率"
    sheetsim.cell(1,20).value = "25日線比率"
    sheetsim.cell(1,21).value = "75日線比率"
    sheetsim.cell(1,22).value = "RSIスコア"
    sheetsim.cell(1,23).value = "ボリンジャーバンドスコア"
    sheetsim.cell(1,24).value = "MACDスコア"
    sheetsim.cell(1,25).value = "テクニカルスコア"
    sheetsim.cell(1,26).value = "項"
    sheetsim.cell(1,27).value = "証券コード"
    sheetsim.cell(1,28).value = "会社名"
    sheetsim.cell(1,29).value = "追跡開始価格"
    sheetsim.cell(1,30).value = "利益"
    sheetsim.cell(1,31).value = "利益%"
    k=2
    for i in range(2, lastrow):
        if sheetdaily.cell(i,4).value is None:
            pass
        else:
            if sheetdaily.cell(i,199).value is not None and sheetdaily.cell(i,200).value is not None and sheetdaily.cell(i,201).value is not None:
                if sheetdaily.cell(i,229).value == 1 and sheetdaily.cell(i,231).value == 1:
                    sheetsim.cell(k,2).value = sheetdaily.cell(i,1).value #日付
                    sheetsim.cell(k,4).value = sheetdaily.cell(i,2).value #証券コード
                    sheetsim.cell(k,5).value = sheetdaily.cell(i,3).value #会社名
                    sheetsim.cell(k,6).value = sheetdaily.cell(i,4).value #株価
                    sheetsim.cell(k,7).value = sheetdaily.cell(i,5).value #前日比
                    sheetsim.cell(k,8).value = sheetdaily.cell(i,34).value #業界
                    sheetsim.cell(k,9).value = sheetdaily.cell(i,17).value #PER
                    sheetsim.cell(k,10).value = sheetdaily.cell(i,18).value #PBR
                    sheetsim.cell(k,11).value = sheetdaily.cell(i,61).value #利回り

                    for j in range(2,lastrow_dividend):
                        if str(sheetsim.cell(k,4).value) == str(sheetdividend.cell(j,1).value) and sheetdividend.cell(j,4).value is True:
                            if sheetdaily.cell(i,32).value is not None and isinstance(sheetdaily.cell(i,32).value, str):
                                continue
                            elif sheetdaily.cell(i,32).value is None:
                                sheetsim.cell(k,12).value = 0 #配当金
                                continue
                            else:
                                sheetsim.cell(k,12).value = sheetdaily.cell(i,32).value/2 #配当金
                                sheetsim.cell(k,13).value = sheetdividend.cell(j,3).value #配当月
                        elif str(sheetsim.cell(k,4).value) == str(sheetdividend.cell(j,1).value) and sheetdividend.cell(j,4).value is False:
                            sheetsim.cell(k,12).value = sheetdaily.cell(i,32).value #配当金
                            sheetsim.cell(k,13).value = sheetdividend.cell(j,3).value #配当月

                    sheetsim.cell(k,14).value = sheetdaily.cell(i,23).value #買残
                    sheetsim.cell(k,15).value = sheetdaily.cell(i,22).value #売残
                    sheetsim.cell(k,16).value = sheetdaily.cell(i,24).value #信用倍率
                    sheetsim.cell(k,17).value = sheetdaily.cell(i,51).value #IR有無
                    url = 'http://kabutan.jp/stock/?code=' + str(sheetdaily.cell(i,2).value) #リンク
                    sheetsim.cell(k,18).hyperlink = url #リンク
                    sheetsim.cell(k,19).value = 100-100*(sheetdaily.cell(i,4).value/sheetdaily.cell(i,199).value)
                    sheetsim.cell(k,20).value = 100-100*(sheetdaily.cell(i,4).value/sheetdaily.cell(i,200).value)
                    sheetsim.cell(k,21).value = 100-100*(sheetdaily.cell(i,4).value/sheetdaily.cell(i,201).value)
                    sheetsim.cell(k,22).value = sheetdaily.cell(i,324).value #RSIスコア
                    sheetsim.cell(k,23).value = sheetdaily.cell(i,325).value #ボリンジャーバンドスコア
                    sheetsim.cell(k,24).value = sheetdaily.cell(i,326).value #MACDスコア
                    sheetsim.cell(k,25).value = sheetdaily.cell(i,327).value #テクニカルスコア
                    sheetsim.cell(k,27).value = sheetdaily.cell(i,2).value #証券コード
                    sheetsim.cell(k,28).value = sheetdaily.cell(i,3).value #会社名
                    print(sheetdaily.cell(i,2).value + '_' + sheetdaily.cell(i,3).value)
                    k += 1
            else:
                pass

    sheetsim.cell(row=1,column=32).value = basis_date2
    wb_sim.save(dirstorage+daycode_format+'_'+'OSCI.xlsx')
#終了時間取得-経過時間
print(t)
t1 = datetime.datetime.now()
print(t1)
dt = t1-t
print(dt)
#終了時間取得-経過時間

#稼働終了アナウンス
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾