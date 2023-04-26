#5日線をプラス越えした銘柄リスト
#2022/10/28
#日付データフォルダに適用
#日付データnを開いたら、最終行取得
#2行目から最終行まで繰り返し
#プラス越えフラグあるデータを取得して、新規作成ファイルに記録

import os
from re import L
import openpyxl
import pandas
import requests
import bs4
import time
import datetime
import glob
import re
import sys
import shutil
import winsound

#開始時間取得
t = datetime.datetime.now()
d = datetime.datetime.now()
d1 = d.strftime('%Y%m%d')
#開始時間取得

#対象フォルダ指定
dirdaily = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/"
dirmerge = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/test/20221027/銘柄/"
dirsim = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/test/20221027/sim/"
dirstorage = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/完了/"

file_list = glob.glob(dirdaily + '*.xlsx')
stock_list = glob.glob(dirmerge + '*.xlsx')

#日付データのリストをglob関数で作成
for l in file_list:
    #日付データを順番に開く
    wb_daily = openpyxl.load_workbook(l)
    wb_sim = openpyxl.Workbook()
    sheetsim = wb_sim.worksheets[0]
    print(l)
    sheetdaily = wb_daily.worksheets[0]
    lastrow = sheetdaily.max_row+1
    lastcolumn = sheetdaily.max_column
    target=0
    daycode = sheetdaily.cell(2,1).value
    daycode_format = daycode.strftime('%Y%m%d')
    k=2
    for i in range(2, lastrow):
        if sheetdaily.cell(i,229).value==1:
            stockcode = sheetdaily.cell(i,2).value
            stockname = sheetdaily.cell(i,3).value
            print(stockcode + '＿' + stockname)
            sheetsim.cell(1,k).value = 1
            sheetsim.cell(2,k).value = stockcode
            sheetsim.cell(3,k).value = stockname
            sheetsim.cell(4,1).value = daycode
            k+=1
        else:
            pass
    
        wb_sim.save(dirsim+daycode_format+'_'+'buyselsim.xlsx')
        wb_daily.close()
        os.rename(dirdaily+'allkabu1.xlsx', dirdaily+d1+'_allkabu1.xlsx')
        new_path = shutil.move(dirdaily+d1+'_allkabu1.xlsx', dirstorage)



#終了時間取得-経過時間
print(t)
t1 = datetime.datetime.now()
print(t1)
dt = t1-t
print(dt)
#終了時間取得-経過時間

#稼働終了アナウンス
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾

#    for j in range(2, lastrow):
#        if sheetdaily.cell(j,229).value=='プラス超え':
#            target+=1
#        else:
#            pass
#    print(target)


#        print(target)