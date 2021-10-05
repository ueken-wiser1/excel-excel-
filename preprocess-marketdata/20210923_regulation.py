
# coding: utf-8

import os
import openpyxl
import requests
import bs4
import time
import datetime
import sys
import sys
import codecs
import requests
from selenium import webdriver
from selenium.webdriver.common.keys import Keys as keys
import winsound
import glob
import numpy as np

t = datetime.datetime.now().time()


#対象：信用規制
dir_data01 = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/ダウンロードデータ/01.信用規制中銘柄/"
dir_data02 = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/ダウンロードデータ/02.信用規制解除銘柄/"
dir_market = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/"
#ダウンロードしたexcel(以下databook)を開く
data01_list = glob.glob(dir_data01 + '*.xlsx')
data02_list = glob.glob(dir_data02 + '*.xlsx')
market_list = glob.glob(dir_market + '*.xlsx')

#print(market_list)

#print(market_list[0])

databook01 = openpyxl.load_workbook(data01_list[0])
databook02 = openpyxl.load_workbook(data02_list[0])
marketbook = openpyxl.load_workbook(market_list[0])

print(str(marketbook))

sheet01 = databook01.worksheets[0]
sheet02 = databook02.worksheets[0]
sheet03 = marketbook.worksheets[0]
lastrow_databook01 = sheet01.max_row + 1
lastrow_databook02 = sheet02.max_row + 1
lastrow_marketbook = sheet03.max_row

#databook内の対象cell(code_databook)を指定する
for i in range(2, lastrow_databook01+1):
    code_databook = str(sheet01.cell(row=i,column=2).value)
#    treatment = str(sheet01.cell(row=i,column=5).value)

#marketbook内をcode_databookで検索
    for j in range(2, lastrow_marketbook):

#code_databookでヒットした行の指定列にコピペ
        if code_databook in str(sheet03.cell(row=j, column=2).value):
            sheet03.cell(row=j,column=101).value = 1
#            sheet02.cell(row=j,column=103).value = treatment
#            print(code_databook)

#databook内の対象cell(code_databook)を指定する
for i in range(2, lastrow_databook02+1):
    code_databook = str(sheet02.cell(row=i,column=2).value)
#    treatment = str(sheet01.cell(row=i,column=5).value)

#marketbook内をcode_databookで検索
    for j in range(2, lastrow_marketbook):

#code_databookでヒットした行の指定列にコピペ
        if code_databook in str(sheet03.cell(row=j, column=2).value):
            sheet03.cell(row=j,column=101).value = 0
#            sheet02.cell(row=j,column=103).value = treatment
#            print(code_databook)

marketbook.save(market_list[0])

print(t)
t = datetime.datetime.now().time()
print(t)

winsound.Beep(500,50)  #ビープ音（500Hzの音を50msec流す）