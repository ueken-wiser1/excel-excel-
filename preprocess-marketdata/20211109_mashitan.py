'''
プログラミング
python
・プログラミング：ダウンロードしたプログラムからallkabuアウトプットプログラムへの実装-これは別プログラムにした方が良いか
1. 銘柄別融資・貸株残高、品貸料率データを日証金からダウンロードする
2. ダウンロードしたデータを開く
3. ダウンロードしたデータ内で証券コードを指定する
4. 指定した証券コードをキーに市場データで一致する証券コードを検索する
5. 一致した行に対して、指定した証券コードの対応データを指定した列にコピペする

対応データ
銘柄別融資・貸株残高、品貸料率データURL
https://www.taisyaku.jp/search/result/index/1/
品貸料率：aタグ22
銘柄別融資・貸株残高：aタグ28

ポイント
・ファイルをダウンロード
    →2021/09/21-実装凍結-データ処理プログラムの作成を優先
    →selenium/webdriverで実装いけるかな
・book1-cell1の数値をキーに、book2-column1の数値内検索。対応する行を特定。
    →対応する行の指定した列にbook1のデータをコピペする
'''
import os
import glob
import openpyxl
import pandas as pa
import datetime
import winsound
import shutil

t = datetime.datetime.now().time()
'''
#ダウンロード機能
taisho_URL = 'https://www.taisyaku.jp/search/result/index/1/'
res = requests.get(taisho_URL)
res.raise_for_status()
soup = bs4.BeautifulSoup(res.content,'lxml')
'''
#対象：貸借取引銘柄別増担保金徴収措置一覧
dir_data = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/ダウンロードデータ/04.貸借取引銘柄別増担保金徴収措置一覧/"
dir_market = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/"
#ダウンロードしたexcel(以下databook)を開く
data_list = glob.glob(dir_data + '*.xlsx')
market_list = glob.glob(dir_market + '*.xlsx')

print(market_list)

print(market_list[0])

databook = openpyxl.load_workbook(data_list[0])
marketbook = openpyxl.load_workbook(market_list[0])

print(str(marketbook))

sheet01 = databook.worksheets[0]
sheet02 = marketbook.worksheets[0]
lastrow_databook = sheet01.max_row + 1
lastrow_marketbook = sheet02.max_row + 1

#databook内の対象cell(code_databook)を指定する
for i in range(7, lastrow_databook):
    code_databook = str(sheet01.cell(row=i,column=2).value)
    treatment = str(sheet01.cell(row=i,column=5).value)

#marketbook内をcode_databookで検索
    for j in range(2, lastrow_marketbook):

#code_databookでヒットした行の指定列にコピペ
        if code_databook in str(sheet02.cell(row=j, column=2).value):
            sheet02.cell(row=j,column=102).value = 1
            sheet02.cell(row=j,column=103).value = treatment
            print(code_databook)


marketbook.save(market_list[0])
databook.close()
new_path = shutil.move(data_list[0],"C:/Users/touko/OneDrive/株価分析/excel/株式データ/ダウンロードデータ/04.貸借取引銘柄別増担保金徴収措置一覧/完了/")

print(t)
t = datetime.datetime.now().time()
print(t)

winsound.Beep(500,50)  #ビープ音（500Hzの音を50msec流す）