
# coding: utf-8

import collections
import os
import openpyxl
import requests
import bs4
import time
import datetime
import sys
import sys
import codecs
import requests
from selenium import webdriver
from selenium.webdriver.common.keys import Keys as keys
import winsound
import glob
import numpy as np
import shutil

#------------お約束開始---冒頭
#稼働時間計測開始
import datetime
t = datetime.datetime.now().time()
#------------お約束終了---冒頭

#------------プログラム本文---ここから
'''
信用規制のページを開いて、各タグを取得する
タイトルタグをワード検索して、ヒットしたところから下は規制解除のシートに記録
それまでは規制中のシートに記録
欲しいのは証券コードと銘柄名
基本的に毎日自動で取得しに行くようにする
'''
########################################################################################################
#信用規制情報を取得
#今日の日付を取得
today = datetime.date.today()
#対象：信用規制
dir_data01 = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/ダウンロードデータ/01.信用規制中銘柄/"
dir_data02 = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/ダウンロードデータ/02.信用規制解除銘柄/"

data_list01 = glob.glob(dir_data01 + '*.xlsx')
regulationbook01 = openpyxl.load_workbook(data_list01[0])
# print(data_list01[0])
n1 = os.path.basename(data_list01[0])
# print(n)
sheet01 = regulationbook01.worksheets[0]

data_list02 = glob.glob(dir_data02 + '*.xlsx')
regulationbook02 = openpyxl.load_workbook(data_list02[0])
# print(data_list02[0])
n2 = os.path.basename(data_list02[0])
# print(n)
sheet02 = regulationbook02.worksheets[0]

dir_market = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/"

market_list = glob.glob(dir_market + '*.xlsx')

marketbook = openpyxl.load_workbook(market_list[0])
sheet03 = marketbook.worksheets[0]

lastrow_databook01 = sheet01.max_row
lastrow_databook02 = sheet02.max_row
lastrow_marketbook = sheet03.max_row

i=2
j=2
#databook内の対象cell(code_databook)を指定する
for i in range(2, lastrow_databook01):
    # print(i)
    code_databook = str(sheet01.cell(row=i,column=2).value)
    # print(code_databook)

#marketbook内をcode_databookで検索
    for j in range(2, lastrow_marketbook):
        # print(j)

#code_databookでヒットした行の指定列にコピペ
        if code_databook in str(sheet03.cell(row=j, column=2).value):
            # print(sheet03.cell(row=j,column=3).value)
            sheet03.cell(row=j,column=101).value = 1
            # print(j, sheet03.cell(row=j,column=101).value)
            sheet01.cell(row=i,column=3).value = sheet03.cell(row=j,column=3).value
marketbook.save(market_list[0])

#databook内の対象cell(code_databook)を指定する
for i in range(2, lastrow_databook02):
    # print(i)
    code_databook = str(sheet02.cell(row=i,column=2).value)

#marketbook内をcode_databookで検索
    for j in range(2, lastrow_marketbook):
        # print(j)

#code_databookでヒットした行の指定列にコピペ
        if code_databook in str(sheet03.cell(row=j, column=2).value):
            # print(sheet03.cell(row=j,column=3).value)
            sheet03.cell(row=j,column=101).value = 0
            # print(j, sheet03.cell(row=j,column=101).value)
            sheet02.cell(row=i,column=3).value = sheet03.cell(row=j,column=3).value

print(market_list[0])
regulationbook01.close()
regulationbook02.close()
new_path = shutil.move(data_list01[0],"C:/Users/touko/OneDrive/株価分析/excel/株式データ/ダウンロードデータ/01.信用規制中銘柄/完了/")
new_path = shutil.move(data_list02[0],"C:/Users/touko/OneDrive/株価分析/excel/株式データ/ダウンロードデータ/02.信用規制解除銘柄/完了/")

marketbook.save(market_list[0])

#------------プログラム本文---ここまで

#------------お約束開始---末尾
#稼働時間表示
print(t)
t = datetime.datetime.now().time()
print(t)

#稼働終了アナウンス
import winsound
# winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
# winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
# winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
# winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾