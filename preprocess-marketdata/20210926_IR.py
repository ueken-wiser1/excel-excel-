#
#20210926_IR情報入力処理
#市場データに対して、当日にIR等ニュースの情報があったかなかったかを入力する。
#対応としては基本的に信用規制情報と同じ。＝IR等ニュースがあれば、当該列に"1"を記載。
#

import os
import glob
import openpyxl
import pandas as pa
import datetime
import winsound

t = datetime.datetime.now().time()

#対象：IR情報入力処理
dir_data = "C:/Users/touko/OneDrive/株価分析/ダウンロードデータ/07.IR情報入力処理/"
dir_market = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/"
#ダウンロードしたexcel(以下databook)を開く
data_list = glob.glob(dir_data + '*.xlsx')
market_list = glob.glob(dir_market + '*.xlsx')

#print(market_list)

print(market_list[0])

databook = openpyxl.load_workbook(data_list[0])
marketbook = openpyxl.load_workbook(market_list[0])

print(str(marketbook))

sheet01 = databook.worksheets[0]
sheet02 = marketbook.worksheets[0]
lastrow_databook = sheet01.max_row + 1
lastrow_marketbook = sheet02.max_row + 1

#databook内の対象cell(code_databook)を指定する
for i in range(2, lastrow_databook):
    code_databook = str(sheet01.cell(row=i,column=3).value)
#    treatment = str(sheet01.cell(row=i,column=5).value)

#marketbook内をcode_databookで検索
    for j in range(2, lastrow_marketbook):

#code_databookでヒットした行の指定列にコピペ
        if code_databook in str(sheet02.cell(row=j, column=2).value):
            sheet02.cell(row=j,column=51).value = 1
            print(code_databook)

marketbook.save(market_list[0])

print(t)
t = datetime.datetime.now().time()
print(t)

winsound.Beep(500,50)