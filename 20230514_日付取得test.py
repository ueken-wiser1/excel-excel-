#何を見るのか
#日付のフォーマットとそれらを照合して、一致と見なすかどうかの確認

import os
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import datetime
import jpholiday
from datetime import timedelta

t = datetime.datetime.now().time()
d = datetime.datetime.now()
d1 = d.strftime('%Y%m%d')

watch_list_file = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/test/20230514/20230104_OSCI.xlsx'
completed_folder = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/完了/完/'


wb_watch_list = load_workbook(watch_list_file)
ws_watch_list = wb_watch_list.active

last_row = ws_watch_list.max_row
#print('最終行は'+str(last_row))
last_col = ws_watch_list.max_column
#print('最終列は'+str(last_col))
file_date = ws_watch_list.cell(row=1, column=last_col).value
basis_date = ws_watch_list.cell(row=2, column=2).value
#print('対象日付は'+str(file_date))
print(file_date)
print(basis_date)
diff = file_date-basis_date
print(diff)

if file_date>basis_date:
    print("差分は1以上")
else:
    print("差分は0以下")

    
print(t)
t = datetime.datetime.now().time()
print(t)

#稼働終了アナウンス
import winsound
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾