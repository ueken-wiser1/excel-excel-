import os
import openpyxl
import requests
import bs4
import time
import datetime
import sys
import winsound
import glob

#どんな動きをさせるのか
#excelを開く
#参照excelの各シートに記載された証券コードを読み込む
#シート名は"株式", "マーケット", "為替", "投信"
#読み込んだ証券コードにより株探URLを作成し、銘柄ページへ移動
#必要な情報を取得する
t = datetime.datetime.now().time()
#excelを開く
#対象：信用規制
dir_market = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/"
#ダウンロードしたexcel(以下databook)を開く
market_list = glob.glob(dir_market + '*.xlsx')

marketbook = openpyxl.load_workbook(market_list[0])

print(str(marketbook))


sheet01 = marketbook.worksheets[0]
lastrow_marketbook = sheet01.max_row

for j in range(2, lastrow_marketbook):

#売買代金2=VWAP(T)*出来高(K)
    sheet01.cell(row=j,column=25).value = sheet01.cell(row=j,column=20).value * sheet01.cell(row=j,column=11).value
#平均約定金額平均約定金額=出来高(K)/約定回数(H)*VWAP(T)
    sheet01.cell(row=j,column=33).value = sheet01.cell(row=j,column=11).value / sheet01.cell(row=j,column=8).value / sheet01.cell(row=j,column=20).value
#回転日数=((融資残(DI)+貸株残(DL))*2)/(融資新規(DG)+融資返済(DH)+貸株新規(DJ)+貸株返済(DK)
    sheet01.cell(row=j,column=118).value = (sheet01.cell(row=j,column=113).value + sheet01.cell(row=j,column=116).value)*2/(sheet01.cell(row=j,column=111).value + sheet01.cell(row=j,column=112).value + sheet01.cell(row=j,column=114).value + sheet01.cell(row=j,column=115).value)
#現在株価との差=株価(D)-みんかぶ目標株価(EU)
    sheet01.cell(row=j,column=152).value = sheet01.cell(row=j,column=4).value - sheet01.cell(row=j,column=151).value

print(t)
t = datetime.datetime.now().time()
print(t)

marketbook.save(market_list[0])
winsound.Beep(500,50)  #ビープ音（500Hzの音を50msec流す）
