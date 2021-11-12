#使ってない証券コードのデータを削除するプログラム
#株式会社の証券コードでないコードのリストはできた
#上記リストにあるコードを持つファイル/データがあれば、それを削除する
#結果：市場データ、銘柄データから、使われていない証券コードデータが消えている
#1. 不使用証券コードを取得する
#2. 市場データを読み込む
#3. 不使用証券コードと市場データの銘柄コードを比較して、同じであればその行を削除する
#4. 不使用証券コードと銘柄データのあるフォルダ内のファイル名を比較し、同じモノがあればそのファイルを削除する
import os
import openpyxl
import xlrd
import time
import sys
import winsound
import glob

#------------お約束開始---冒頭
#稼働時間計測開始
import datetime
t = datetime.datetime.now().time()
#------------お約束終了---冒頭

#------------プログラム本文---ここから

#使用フォルダ指定
folder01 = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/銘柄データ集計/" #不使用証券コードリスト
folder02 = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/完了/" #市場データフォルダ
folder03 = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/完了/" #銘柄データフォルダ

#証券コードリストを開く
uselist = openpyxl.load_workbook(folder01 + "stockcodelist_use.xlsx")
sheet01 = uselist.worksheets[0]
lastrow_uselist = sheet01.max_row

#使用証券コードを配列に格納する
array_uselist =[]
for i in range(2, lastrow_uselist+1):
    array_uselist.append(sheet01.cell(row=i,column=1).value)

tagscrapedata = openpyxl.load_workbook(folder01 + "stockalltagresult01.xlsx")
sheet02 = tagscrapedata.worksheets[0]
lastrow_tagscrape = sheet02.max_row

#市場データを順に開き、配列にないコードがある行を削除する
for j in reversed(range(2,lastrow_tagscrape+1)):
    if sheet02.cell(row=j,column=2).value in array_uselist:
        pass
    else:
        print(sheet02.cell(row=j,column=2).value,sheet02.cell(row=j,column=3).value)
        sheet02.delete_rows(j)

    tagscrapedata.save()

#------------プログラム本文---ここまで

#------------お約束開始---末尾
#稼働時間表示
print(t)
t = datetime.datetime.now().time()
print(t)

#稼働終了アナウンス
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾