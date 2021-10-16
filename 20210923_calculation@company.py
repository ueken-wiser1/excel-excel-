import os
from re import L
import openpyxl
import requests
import bs4
import time
import datetime
import sys
import winsound
import glob
import xlrd
import numpy as np

#どんな動きをさせるのか
#excelを開く
#参照excelの各シートに記載された証券コードを読み込む
#シート名は"株式", "マーケット", "為替", "投信"
#読み込んだ証券コードにより株探URLを作成し、銘柄ページへ移動
#必要な情報を取得する

#excel数式入力を全てpythonでの計算に置き換える
#→標準化数値の導出に対して、配列への数値入力がexcel数式では出来なかったため

t = datetime.datetime.now().time()
#excelを開く
#対象：信用規制
dir_stock = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/"
#ダウンロードしたexcel(以下databook)を開く
stock_list = glob.glob(dir_stock + '*.xlsx')

for l in stock_list:
        stockbook = openpyxl.load_workbook(l)
#    print(str(stockbook))
        sheet01 = stockbook.worksheets[0]
        lastrow_stockbook = sheet01.max_row
        lastcolumn_stockbook = sheet01.max_column
#全部を配列で取込
        array_entire = []
        for g in range(2,lastrow_stockbook+1):
                for f in range(1,lastcolumn_stockbook+1):
                        if sheet01.cell(row=g,column=f).value == '－':
                                sheet01.cell(row=g,column=f).value = 0
                        else:
                                pass
                        array_entire.append(sheet01.cell(row=g,column=f).value)

        for j in range(2,lastrow_stockbook+1):
                print(l)

#値幅の計算式入れ直し
                sheet01.cell(row=j, column=6).value = str('=M')+str(j)+str('-N')+str(j)
#参照元のデータの無いセルを0で埋める

#DG融資新規　DH融資返済　DI融資残高　DJ貸株新規　DK貸株返済　DL貸株残高　DM差引残高
                if sheet01.cell(row=j,column=111).value is None:
                        sheet01.cell(row=j,column=111).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=112).value is None:
                        sheet01.cell(row=j,column=112).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=113).value is None:
                        sheet01.cell(row=j,column=113).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=114).value is None:
                        sheet01.cell(row=j,column=114).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=115).value is None:
                        sheet01.cell(row=j,column=115).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=116).value is None:
                        sheet01.cell(row=j,column=116).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=117).value is None:
                        sheet01.cell(row=j,column=117).value = 0
                else:
                        pass
                
#売買代金2=VWAP(T)*出来高(K)
#                sheet01.cell(row=j,column=25).value = str('=T')+str(j)+str('*K')+str(j)
                sheet01.cell(row=j,column=25).value = sheet01.cell(row=j,column=20).value * sheet01.cell(row=j-1,column=11).value
#信用売残前週比：(V)
#もしj=2なら、sheet01...は0
#もしsheet01...の結果が0なら、その前の0でない値を入力
                if j == 2:
                        sheet01.cell(row=j,column=26).value = 0
                else:
#                        sheet01.cell(row=j,column=26).value = str('=V')+str(j)+str('-V')+str(j-1)
                        sheet01.cell(row=j,column=26).value = sheet01.cell(row=j,column=23).value - sheet01.cell(row=j-1,column=23).value
                        i = 1
                        while sheet01.cell(row=j,column=26).value == 0:
                                sheet01.cell(row=j,column=26).value == sheet01.cell(row=j-i, column=26).value
                                i -= 1
#信用買残前週比：(W)
#もしj=2なら、sheet01...は0
#もしsheet01...の結果が0なら、その前の0でない値を入力
#                sheet01.cell(row=j,column=27).value = str('=W')+str(j)+str('-W')+str(j-1)
                if j == 2:
                        sheet01.cell(row=j,column=27).value = 0
                else:
#                        sheet01.cell(row=j,column=27).value = str('=V')+str(j)+str('-V')+str(j-1)
                        sheet01.cell(row=j,column=27).value = sheet01.cell(row=j,column=22).value - sheet01.cell(row=j-1,column=22).value
                        i = 1
                        while sheet01.cell(row=j,column=27).value == 0:
                                sheet01.cell(row=j,column=27).value == sheet01.cell(row=j-i, column=27).value
                                i -= 1
#出来高前日比：(K)
#もしj=2なら、sheet01...は0
#                sheet01.cell(row=j,column=28).value = str('=K')+str(j)+str('-K')+str(j-1)
                if j == 2:
                        sheet01.cell(row=j,column=28).value = 0
                else:
                        sheet01.cell(row=j,column=28).value = sheet01.cell(row=j,column=11).value -sheet01.cell(row=j-1,column=11).value

#約定回数前日比：(H)
#もしj=2なら、sheet01...は0
#                sheet01.cell(row=j,column=29).value = str('=H')+str(j)+str('-H')+str(j-1)
                if j == 2:
                        sheet01.cell(row=j,column=29).value = 0
                else:
                        sheet01.cell(row=j,column=29).value = sheet01.cell(row=j,column=8).value - sheet01.cell(row=j-1,column=8).value

#平均約定金額平均約定金額=出来高(K)/約定回数(H)*VWAP(T)
                sheet01.cell(row=j,column=33).value = sheet01.cell(row=j,column=11).value / sheet01.cell(row=j,column=8).value * sheet01.cell(row=j,column=20).value
#回転日数=((融資残(DI)+貸株残(DL))*2)/(融資新規(DG)+融資返済(DH)+貸株新規(DJ)+貸株返済(DK)
#もし参照セルの値が一つでも0なら、回転日数は0を返す
                if sheet01.cell(row=j,column=111).value == 0 and sheet01.cell(row=j,column=112).value == 0 and sheet01.cell(row=j,column=114).value == 0 and sheet01.cell(row=j,column=115).value == 0:
                        sheet01.cell(row=j,column=118).value = 0
                elif sheet01.cell(row=j,column=111).value == str(0) and sheet01.cell(row=j,column=112).value == str(0) and sheet01.cell(row=j,column=114).value == str(0) and sheet01.cell(row=j,column=115).value == str(0):
                        sheet01.cell(row=j,column=118).value = 0
                else:
                        sheet01.cell(row=j,column=118).value = (sheet01.cell(row=j,column=113).value + sheet01.cell(row=j,column=116).value)*2/(sheet01.cell(row=j,column=111).value +sheet01.cell(row=j,column=112).value + sheet01.cell(row=j,column=114).value + sheet01.cell(row=j,column=115).value)
#現在株価との差=株価(D)-みんかぶ目標株価(EU)
                sheet01.cell(row=j,column=152).value = sheet01.cell(row=j,column=4).value - sheet01.cell(row=j,column=151).value
#株価のリストを作って、その中で一部の要素を抽出するやり方を取るか
#株価の配列を取得する
                array_kabuka = []
                for h in range(2, lastrow_stockbook):
                        array_kabuka.append(float(sheet01.cell_value(i,0)))
                        print(array_kabuka)

#移動平均線数値　5日＝当日〜4日前の株価総和/5
                if j < 6 :
                        pass
                else:
                        for i in range(5):
                                sheet01.cell(row=j,column=201).value = sum(array_kabuka[i-5:i])
#移動平均線数値　25日＝当日〜24日前の株価総和/25
                if j < 26 :
                        pass
                else:
                        sheet01.cell(row=j,column=202).value = str('=SUM(D')+str(j-25)+str(':D')+str(j)+str(')/25')
#移動平均線数値　75日＝当日〜74日前の株価総和/75
                if j < 76 :
                        pass
                else:
                        sheet01.cell(row=j,column=203).value = str('=SUM(D')+str(j-75)+str(':D')+str(j)+str(')/75')
#移動平均乖離率　5日＝（株価ー移動平均5日）/移動平均5日
                if j < 6 :
                        pass
                else:
                        sheet01.cell(row=j,column=204).value = str('=(D')+str(j)+str('-GS')+str(j)+str(')/GS')+str(j)
#移動平均乖離率　25日＝（株価ー移動平均25日）/移動平均25日
                if j < 26 :
                        pass
                else:
                        sheet01.cell(row=j,column=205).value = str('=(D')+str(j)+str('-GT')+str(j)+str(')/GT')+str(j)
#移動平均乖離率　75日＝（株価ー移動平均75日）/移動平均75日
                if j < 76 :
                        pass
                else:
                        sheet01.cell(row=j,column=206).value = str('=(D')+str(j)+str('-GU')+str(j)+str(')/GU')+str(j)


#平均出来高＝平均値（average(対象列)）
                sheet01.cell(row=j,column=251).value = str('=AVERAGE(K:K)')
#平均約定回数＝平均値（average(対象列)）
                sheet01.cell(row=j,column=252).value = str('=AVERAGE(H:H)')
#平均回転日数＝平均値（average(対象列)）
                sheet01.cell(row=j,column=253).value = str('=AVERAGE(DN:DN)')
#平均移動平均乖離率5日＝平均値（average(対象列)）
                sheet01.cell(row=j,column=254).value = str('=AVERAGE(GV:GV)')
#平均移動平均乖離率25日＝平均値（average(対象列)）
                sheet01.cell(row=j,column=255).value = str('=AVERAGE(GW:GW)')
#平均移動平均乖離率75日＝平均値（average(対象列)）
                sheet01.cell(row=j,column=256).value = str('=AVERAGE(GX:GX)')
#平均信用買残＝平均値（average(対象列)）
                sheet01.cell(row=j,column=257).value = str('=AVERAGE(W:W)')
#平均融資新規＝平均値（average(対象列)）
#参照元に対して、何も入力されていなければ、0を入力しておく
                if sheet01.cell(row=j,column=111).value is None:
                        sheet01.cell(row=j,column=258).value = 0
                else:
                        sheet01.cell(row=j,column=258).value = str('=AVERAGE(DG:DG)')
#平均融資返済＝平均値（average(対象列)）
#参照元に対して、何も入力されていなければ、0を入力しておく
                if sheet01.cell(row=j,column=112).value is None:
                        sheet01.cell(row=j,column=259).value = 0
                else:
                        sheet01.cell(row=j,column=259).value = str('=AVERAGE(DH:DH)')
#平均信用売残＝平均値（average(対象列)）
                sheet01.cell(row=j,column=260).value = str('=AVERAGE(V:V)')
#平均貸株新規＝平均値（average(対象列)）
#参照元に対して、何も入力されていなければ、0を入力しておく
                if sheet01.cell(row=j,column=114).value is None:
                        sheet01.cell(row=j,column=261).value = 0
                else:
                        sheet01.cell(row=j,column=261).value = str('=AVERAGE(DJ:DJ)')
#平均貸株返済＝平均値（average(対象列)）
#参照元に対して、何も入力されていなければ、0を入力しておく
                if sheet01.cell(row=j,column=115).value is None:
                        sheet01.cell(row=j,column=262).value = 0
                else:
                        sheet01.cell(row=j,column=262).value = str('=AVERAGE(DK:DK)')
#平均貸株超過＝平均値（average(対象列)）
#参照元に対して、何も入力されていなければ、0を入力しておく
                if sheet01.cell(row=j,column=116).value is None:
                        sheet01.cell(row=j,column=263).value = 0
                else:
                        sheet01.cell(row=j,column=263).value = str('=AVERAGE(DL:DL)')
                        
#平均出来高変化率＝平均値（average(対象列)）
                sheet01.cell(row=j,column=264).value = str('=AVERAGE(AB:AB)')
#平均約定回数変化率＝平均値（average(対象列)）
                sheet01.cell(row=j,column=265).value = str('=AVERAGE(AC:AC)')
#平均信用買残変化率＝平均値（average(対象列)）
                sheet01.cell(row=j,column=266).value = str('=AVERAGE(AA:AA)')
#平均信用売残変化率＝平均値（average(対象列)）
                sheet01.cell(row=j,column=267).value = str('=AVERAGE(Z:Z)')
#平均平均約定金額＝平均値（average(対象列)）
                sheet01.cell(row=j,column=268).value = str('=AVERAGE(AG:AG)')


#以下標準偏差計算について、@が入力されている
#出来高　約定回数　移動平均乖離率5日　移動平均乖離率25日　移動平均乖離率75日
                if sheet01.cell(row=j,column=11).value is None:
                        sheet01.cell(row=j,column=11).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=8).value is None:
                        sheet01.cell(row=j,column=8).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=204).value is None:
                        sheet01.cell(row=j,column=204).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=205).value is None:
                        sheet01.cell(row=j,column=205).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=206).value is None:
                        sheet01.cell(row=j,column=206).value = 0
                else:
                        pass
#信用買残　融資新規　融資返済　信用売残　貸株新規　貸株返済　貸株超過
                if sheet01.cell(row=j,column=23).value is None:
                        sheet01.cell(row=j,column=23).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=111).value is None:
                        sheet01.cell(row=j,column=111).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=112).value is None:
                        sheet01.cell(row=j,column=112).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=22).value is None:
                        sheet01.cell(row=j,column=22).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=114).value is None:
                        sheet01.cell(row=j,column=114).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=115).value is None:
                        sheet01.cell(row=j,column=115).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=117).value is None:
                        sheet01.cell(row=j,column=117).value = 0
                else:
                        pass

#出来高変化率　約定回数変化率　信用買残変化率　信用売残変化率　平均約定金額
                if sheet01.cell(row=j,column=28).value is None:
                        sheet01.cell(row=j,column=28).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=29).value is None:
                        sheet01.cell(row=j,column=29).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=27).value is None:
                        sheet01.cell(row=j,column=27).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=26).value is None:
                        sheet01.cell(row=j,column=26).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=33).value is None:
                        sheet01.cell(row=j,column=33).value = 0
                else:
                        pass


#以下標準偏差計算について、@が入力されている
#標準偏差出来高＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=269).value = str('=STDEV.P(K:K)')
                array_dekidaka =[]
                for k in range(2, lastrow_stockbook):
                        array_dekidaka.append(int(sheet01.cell(row=k,column=11).value))
#                print(array_dekidaka)
                data = np.array(array_dekidaka)
#                print(data)
                std = np.std(data)
                sheet01.cell(row=j, column=269).value = std
#標準偏差約定回数＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=270).value = str('=STDEV.P(H:H)')
                array_yakujo =[]
                for k in range(2, lastrow_stockbook):
#                        print(array_yakujo)
                        if sheet01.cell(row=k,column=8).value == '－' or sheet01.cell(row=k,column=8).value is None:
                                sheet01.cell(row=k,column=8).value = 0
                        else:
                                pass
                        array_yakujo.append(sheet01.cell(row=k,column=8).value)
                data = np.array(array_yakujo)
                std = np.std(data)
                sheet01.cell(row=j, column=270).value = std
#標準偏差回転日数＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=271).value = str('=STDEV.P(DN:DN)')
                array_kaiten =[]
                for k in range(2, lastrow_stockbook):
                        array_kaiten.append(float(sheet01.cell(row=k,column=118).value))
                data = np.array(array_kaiten)
                std = np.std(data)
                sheet01.cell(row=j, column=271).value = std
#標準偏差移動平均乖離率5日＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=272).value = str('=STDEV.P(GV:GV)')
                array_idoheikin5 =[]
                for k in range(2, lastrow_stockbook):
                        array_idoheikin5.append(sheet01.cell(row=k,column=204).value)
                data = np.array(array_idoheikin5)
                print(array_idoheikin5)
                std = np.std(data)
                sheet01.cell(row=j, column=272).value = std
#標準偏差移動平均乖離率25日＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=273).value = str('=STDEV.P(GW:GW)')
                array_idoheikin25 =[]
                for k in range(2, lastrow_stockbook):
                        array_idoheikin25.append(float(sheet01.cell(row=k,column=205).value))
                data = np.array(array_idoheikin25)
                std = np.std(data)
                sheet01.cell(row=j, column=273).value = std
#標準偏差移動平均乖離率75日＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=274).value = str('=STDEV.P(GX:GX)')
                array_idoheikin75 =[]
                for k in range(2, lastrow_stockbook):
                        array_idoheikin75.append(float(sheet01.cell(row=k,columne=206).value))
                data = np.array(array_idoheikin75)
                std = np.std(data)
                sheet01.cell(row=j, column=274).value = std
#標準偏差信用買残＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=275).value = str('=STDEV.P(W:W)')
                array_kaizan =[]
                for k in range(2, lastrow_stockbook):
                        array_kaizan.append(float(sheet01.cell(row=k,column=23).value))
                data = np.array(array_kaizan)
                std = np.std(data)
                sheet01.cell(row=j, column=275).value = std
#標準偏差融資新規＝標準偏差(stdevp(対象列)）入力無し
#                sheet01.cell(row=j,column=276).value = str('=STDEV.P(DG:DG))')
                array_yushishinki =[]
                for k in range(2, lastrow_stockbook):
                        array_yushishinki.append(float(sheet01.cell(row=k,column=111).value))
                data = np.array(array_yushishinki)
                std = np.std(data)
                sheet01.cell(row=j, column=276).value = std
#標準偏差融資返済＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=277).value = str('=STDEV.P(DH:DH)')
                array_yusihensai =[]
                for k in range(2, lastrow_stockbook):
                        array_yusihensai.append(float(sheet01.cell(row=k,column=112).value))
                data = np.array(array_yusihensai)
                std = np.std(data)
                sheet01.cell(row=j, column=277).value = std
#標準偏差信用売残＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=278).value = str('=STDEV.P(V:V)')
                array_urizan =[]
                for k in range(2, lastrow_stockbook):
                        array_urizan.append(float(sheet01.cell(row=k,column=22).value))
                data = np.array(array_urizan)
                std = np.std(data)
                sheet01.cell(row=j, column=278).value = std
#標準偏差貸株新規＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=279).value = str('=STDEV.P(DJ:DJ)')
                array_kashishinki =[]
                for k in range(2, lastrow_stockbook):
                        array_kashishinki.append(float(sheet01.cell(row=k,column=114).value))
                data = np.array(array_kashishinki)
                std = np.std(data)
                sheet01.cell(row=j, column=279).value = std
#標準偏差貸株返済＝標準偏差(stdevp(対象列)）入力無し
#                sheet01.cell(row=j,column=280).value = str('=STDEV.P(DK:DK))')
                array_kashihensai =[]
                for k in range(2, lastrow_stockbook):
                        array_kashihensai.append(float(sheet01.cell(row=k,column=115).value))
                data = np.array(array_kashihensai)
                std = np.std(data)
                sheet01.cell(row=j, column=280).value = std
#標準偏差貸株超過＝標準偏差(stdevp(対象列)）入力無し
#                sheet01.cell(row=j,column=281).value = str('=STDEV.P(DL:DL))')
                array_kashichoka =[]
                for k in range(2, lastrow_stockbook):
                        array_kashichoka.append(float(sheet01.cell(row=k,column=116).value))
                data = np.array(array_kashichoka)
                std = np.std(data)
                sheet01.cell(row=j, column=281).value = std
#標準偏差出来高変化率＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=282).value = str('=STDEV.P(AB:AB)')
                array_dekidakahenka =[]
                for k in range(2, lastrow_stockbook):
                        array_dekidakahenka.append(int(sheet01.cell(row=k,column=28).value))
                data = np.array(array_dekidakahenka)
                std = np.std(data)
                sheet01.cell(row=j, column=282).value = std
#標準偏差約定回数変化率＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=283).value = str('=STDEV.P(AC:AC)')
                array_yakujohenka =[]
                for k in range(2, lastrow_stockbook):
                        array_yakujohenka.append(int(sheet01.cell(row=k,column=29).value))
                data = np.array(array_yakujohenka)
                std = np.std(data)
                sheet01.cell(row=j, column=283).value = std
#標準偏差信用買残変化率＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=284).value = str('=STDEV.P(AA:AA)')
                array_kaizanhenka =[]
                for k in range(2, lastrow_stockbook):
                        array_kaizanhenka.append(float(sheet01.cell(row=k,column=27).value))
                data = np.array(array_kaizanhenka)
                std = np.std(data)
                sheet01.cell(row=j, column=284).value = std
#標準偏差信用売残変化率＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=285).value = str('=STDEV.P(Z:Z))')
                array_urizanhenka =[]
                for k in range(2, lastrow_stockbook):
                        array_urizanhenka.append(float(sheet01.cell(row=k,column=26).value))
                data = np.array(array_urizanhenka)
                std = np.std(data)
                sheet01.cell(row=j, column=285).value = std
#標準偏差平均約定金額＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=286).value = str('=STDEV.P(AG:AG))')
                array_heikinyakujo =[]
                for k in range(2, lastrow_stockbook):
                        array_heikinyakujo.append(float(sheet01.cell(row=k,column=33).value))
                data = np.array(array_heikinyakujo)
                std = np.std(data)
                sheet01.cell(row=j, column=286).value = std


#標準化出来高＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                sheet01.cell(row=j,column=501).value = str('=STANDARDIZE(K')+str(j)+str(',IQ')+str(j)+str(',JI')+str(j)+str(')')

#標準化約定回数＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                sheet01.cell(row=j,column=502).value = str('=STANDARDIZE(H')+str(j)+str(',IR')+str(j)+str(',JJ')+str(j)+str(')')

#標準化回転日数＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                sheet01.cell(row=j,column=503).value = str('=STANDARDIZE(DN')+str(j)+str(',IS')+str(j)+str(',JK')+str(j)+str(')')

#標準化移動平均乖離率5日＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                sheet01.cell(row=j,column=504).value = str('=STANDARDIZE(GV')+str(j)+str(',IT')+str(j)+str(',JL')+str(j)+str(')')

#標準化移動平均乖離率25日＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                sheet01.cell(row=j,column=505).value = str('=STANDARDIZE(GW')+str(j)+str(',IU')+str(j)+str(',JM')+str(j)+str(')')

#標準化移動平均乖離率75日＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                sheet01.cell(row=j,column=506).value = str('=STANDARDIZE(GX')+str(j)+str(',IV')+str(j)+str(',JN')+str(j)+str(')')

#標準化信用買残＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                sheet01.cell(row=j,column=507).value = str('=STANDARDIZE(W')+str(j)+str(',IW')+str(j)+str(',JO')+str(j)+str(')')

#標準化融資新規＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                sheet01.cell(row=j,column=508).value = str('=STANDARDIZE(DG')+str(j)+str(',IX')+str(j)+str(',JP')+str(j)+str(')')

#標準化融資返済＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                sheet01.cell(row=j,column=509).value = str('=STANDARDIZE(DH')+str(j)+str(',IY')+str(j)+str(',JQ')+str(j)+str(')')

#標準化信用売残＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                sheet01.cell(row=j,column=510).value = str('=STANDARDIZE(V')+str(j)+str(',IZ')+str(j)+str(',JR')+str(j)+str(')')

#標準化貸株新規＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                sheet01.cell(row=j,column=511).value = str('=STANDARDIZE(DJ')+str(j)+str(',JA')+str(j)+str(',JS')+str(j)+str(')')

#標準化貸株返済＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                sheet01.cell(row=j,column=512).value = str('=STANDARDIZE(DK')+str(j)+str(',JB')+str(j)+str(',JT')+str(j)+str(')')

#標準化貸株超過＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                sheet01.cell(row=j,column=513).value = str('=STANDARDIZE(DL')+str(j)+str(',JC')+str(j)+str(',JU')+str(j)+str(')')

#標準化出来高変化率＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                sheet01.cell(row=j,column=514).value = str('=STANDARDIZE(AB')+str(j)+str(',JD')+str(j)+str(',JV')+str(j)+str(')')

#標準化約定回数変化率＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                sheet01.cell(row=j,column=515).value = str('=STANDARDIZE(AC')+str(j)+str(',JE')+str(j)+str(',JW')+str(j)+str(')')

#標準化信用買残変化率＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                sheet01.cell(row=j,column=516).value = str('=STANDARDIZE(AA')+str(j)+str(',JF')+str(j)+str(',JX')+str(j)+str(')')

#標準化信用売残変化率＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                sheet01.cell(row=j,column=517).value = str('=STANDARDIZE(Z')+str(j)+str(',JG')+str(j)+str(',JY')+str(j)+str(')')

#標準化平均約定金額＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                sheet01.cell(row=j,column=518).value = str('=STANDARDIZE(AG')+str(j)+str(',JH')+str(j)+str(',JZ')+str(j)+str(')')

#売り時/買い時スコア＝12*標準化出来高(SG)+15*標準化約定回数(SH)+6*標準化平均約定金額(SX)+3/標準化回転日数(SI)+9*標準化移動平均乖離率5日(SJ)+6*標準化信用買残(SM)+10*信用規制(CW)+8*増担規制(CX)+4*標準化融資新規(SN)+2/標準化融資返済(SO)+4/標準化信用売残(SP)+5*空売り規制(DB)+3/標準化貸株新規(SQ)+2*標準化貸株返済(SR)+1/標準化貸株超過(SS)+(20*IR)
                sheet01.cell(row=j,column=1000).value = str('=12*SG')+str(j)+str('+15*SH')+str(j)+str('+6*SX')+str(j)+str('+3/(1+SI')+str(j)+str(')+9*SJ')+str(j)+str('+6*SM')+str(j)+str('+10*CW')+str(j)+str('+8*CX')+str(j)+str('+4*SN')+str(j)+str('+2/(1+SO')+str(j)+str(')+4/(1+SP')+str(j)+str(')+5*DB')+str(j)+str('+3/(1+SQ')+str(j)+str(')+2*SR')+str(j)+str('+1/(1+SS')+str(j)+str(')+20*AY')+str(j)

        stockbook.save(l)
        stockbook.close()
print(t)
t = datetime.datetime.now().time()
print(t)


winsound.Beep(500,50)  #ビープ音（500Hzの音を50msec流す）
'''
#以下標準偏差計算について、@が入力されている
#標準偏差出来高＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=269).value = str('=STDEV.P(K:K)')
                array_dekidaka =[]
                for k in range(2, lastrow_stockbook):
                        array_dekidaka.append(int(sheet01.cell(row=k,column=11).value))
#                print(array_dekidaka)
                data = np.array(array_dekidaka)
#                print(data)
                std = np.std(data)
                sheet01.cell(row=j, column=269).value = std
#標準偏差約定回数＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=270).value = str('=STDEV.P(H:H)')
                array_yakujo =[]
                for k in range(2, lastrow_stockbook):
                        array_yakujo.append(sheet01.cell(row=k,column=8).value)
                data = np.array(array_yakujo)
                std = np.std(data)
                sheet01.cell(row=j, column=270).value = std
#標準偏差回転日数＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=271).value = str('=STDEV.P(DN:DN)')
                array_kaiten =[]
                for k in range(2, lastrow_stockbook):
                        array_kaiten.append(sheet01.cell(row=k,column=118).value)
                data = np.array(array_kaiten)
                std = np.std(data)
                sheet01.cell(row=j, column=271).value = std
#標準偏差移動平均乖離率5日＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=272).value = str('=STDEV.P(GV:GV)')
                array_idoheikin5 =[]
                for k in range(2, lastrow_stockbook):
                        array_idoheikin5.append(sheet01.cell(row=k,column=204).value)
                data = np.array(array_idoheikin5)
                std = np.std(data)
                sheet01.cell(row=j, column=272).value = std
#標準偏差移動平均乖離率25日＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=273).value = str('=STDEV.P(GW:GW)')
                array_idoheikin25 =[]
                for k in range(2, lastrow_stockbook):
                        array_idoheikin25.append(sheet01.cell(row=k,column=205).value)
                data = np.array(array_idoheikin25)
                std = np.std(data)
                sheet01.cell(row=j, column=273).value = std
#標準偏差移動平均乖離率75日＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=274).value = str('=STDEV.P(GX:GX)')
                array_idoheikin75 =[]
                for k in range(2, lastrow_stockbook):
                        array_idoheikin75.append(sheet01.cell(row=k,columne=206).value)
                data = np.array(array_idoheikin75)
                std = np.std(data)
                sheet01.cell(row=j, column=274).value = std
#標準偏差信用買残＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=275).value = str('=STDEV.P(W:W)')
                array_kaizan =[]
                for k in range(2, lastrow_stockbook):
                        array_kaizan.append(sheet01.cell(row=k,column=23).value)
                data = np.array(array_kaizan)
                std = np.std(data)
                sheet01.cell(row=j, column=275).value = std
#標準偏差融資新規＝標準偏差(stdevp(対象列)）入力無し
#                sheet01.cell(row=j,column=276).value = str('=STDEV.P(DG:DG))')
                array_yushishinki =[]
                for k in range(2, lastrow_stockbook):
                        array_yushishinki.append(sheet01.cell(row=k,column=111).value)
                data = np.array(array_yushishinki)
                std = np.std(data)
                sheet01.cell(row=j, column=276).value = std
#標準偏差融資返済＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=277).value = str('=STDEV.P(DH:DH)')
                array_yusihensai =[]
                for k in range(2, lastrow_stockbook):
                        array_yusihensai.append(sheet01.cell(row=k,column=112).value)
                data = np.array(array_yusihensai)
                std = np.std(data)
                sheet01.cell(row=j, column=277).value = std
#標準偏差信用売残＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=278).value = str('=STDEV.P(V:V)')
                array_urizan =[]
                for k in range(2, lastrow_stockbook):
                        array_urizan.append(sheet01.cell(row=k,column=22).value)
                data = np.array(array_urizan)
                std = np.std(data)
                sheet01.cell(row=j, column=278).value = std
#標準偏差貸株新規＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=279).value = str('=STDEV.P(DJ:DJ)')
                array_kashishinki =[]
                for k in range(2, lastrow_stockbook):
                        array_kashishinki.append(sheet01.cell(row=k,column=114).value)
                data = np.array(array_kashishinki)
                std = np.std(data)
                sheet01.cell(row=j, column=279).value = std
#標準偏差貸株返済＝標準偏差(stdevp(対象列)）入力無し
#                sheet01.cell(row=j,column=280).value = str('=STDEV.P(DK:DK))')
                array_kashihensai =[]
                for k in range(2, lastrow_stockbook):
                        array_kashihensai.append(sheet01.cell(row=k,column=115).value)
                data = np.array(array_kashihensai)
                std = np.std(data)
                sheet01.cell(row=j, column=280).value = std
#標準偏差貸株超過＝標準偏差(stdevp(対象列)）入力無し
#                sheet01.cell(row=j,column=281).value = str('=STDEV.P(DL:DL))')
                array_kashichoka =[]
                for k in range(2, lastrow_stockbook):
                        array_kashichoka.append(sheet01.cell(row=k,column=116).value)
                data = np.array(array_kashichoka)
                std = np.std(data)
                sheet01.cell(row=j, column=281).value = std
#標準偏差出来高変化率＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=282).value = str('=STDEV.P(AB:AB)')
                array_dekidakahenka =[]
                for k in range(2, lastrow_stockbook):
                        array_dekidakahenka.append(sheet01.cell(row=k,column=28).value)
                data = np.array(array_dekidakahenka)
                std = np.std(data)
                sheet01.cell(row=j, column=282).value = std
#標準偏差約定回数変化率＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=283).value = str('=STDEV.P(AC:AC)')
                array_yakujohenka =[]
                for k in range(2, lastrow_stockbook):
                        array_yakujohenka.append(sheet01.cell(row=k,column=29).value)
                data = np.array(array_yakujohenka)
                std = np.std(data)
                sheet01.cell(row=j, column=283).value = std
#標準偏差信用買残変化率＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=284).value = str('=STDEV.P(AA:AA)')
                array_kaizanhenka =[]
                for k in range(2, lastrow_stockbook):
                        array_kaizanhenka.append(sheet01.cell(row=k,column=27).value)
                data = np.array(array_kaizanhenka)
                std = np.std(data)
                sheet01.cell(row=j, column=284).value = std
#標準偏差信用売残変化率＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=285).value = str('=STDEV.P(Z:Z))')
                array_urizanhenka =[]
                for k in range(2, lastrow_stockbook):
                        array_urizanhenka.append(sheet01.cell(row=k,column=26).value)
                data = np.array(array_urizanhenka)
                std = np.std(data)
                sheet01.cell(row=j, column=285).value = std
#標準偏差平均約定金額＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=286).value = str('=STDEV.P(AG:AG))')
                array_heikinyakujo =[]
                for k in range(2, lastrow_stockbook):
                        array_heikinyakujo.append(sheet01.cell(row=k,column=33).value)
                data = np.array(array_heikinyakujo)
                std = np.std(data)
                sheet01.cell(row=j, column=286).value = std
'''