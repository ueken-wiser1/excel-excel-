#移動平均データを計算した銘柄データから、日付データに回帰させる
#日付データフォルダを走査
##重複日付の削除
#2022/10/26
#銘柄データフォルダに適用
#銘柄データnを開いたら、最終行取得
#2行目から最終行まで繰り返し
#i行目の日付とi+j行目の日付が同じなら、i+j行目を削除

import os
from re import L
import openpyxl
import pandas
import requests
import bs4
import time
import datetime
import glob
import re
import sys
import winsound

#開始時間取得
t = datetime.datetime.now()
d = datetime.datetime.now()
d1 = d.strftime('%Y%m%d')
#開始時間取得

#対象フォルダ指定
dirdaily = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/"
dirmerge = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/"

file_list = glob.glob(dirdaily + '*.xlsx')
stock_list = glob.glob(dirmerge + '*.xlsx')

#銘柄データのリストをglob関数で作成
for l in file_list:
    wb_daily = openpyxl.load_workbook(l)
    print(l)
    sheetdaily = wb_daily.worksheets[0]
    lastrow = sheetdaily.max_row+1
    lastcolumn = 330

    for i in range(2, lastrow):
        daycode = sheetdaily.cell(i,1).value
        stockcode = sheetdaily.cell(i,2).value
#        print(str(daycode) + ' ' + str(stockcode))
        if daycode is None:
            pass
        else:
            daycode_format = daycode.strftime('%Y%m%d')
            print(daycode)
#        print(daycode_format)
        book_search_list = glob.glob(dirmerge+str(stockcode)+'*.xlsx')
#        print(len(book_search_list))
#        print(book_search_list)
        if len(book_search_list) == 0:
            print('銘柄データなし')
            pass
        else:
            stockbook = book_search_list[0]
            #print(str(daycode) +'_'+stockbook)
            wb_stock = openpyxl.load_workbook(stockbook)
            sheetstock = wb_stock.worksheets[0]
            lastrow2=sheetstock.max_row+1
            #print(stockbook)

            for j in range(2, lastrow2):
                daycode_same = sheetstock.cell(j,1).value
                stockcode_same = sheetstock.cell(j,2).value
                #print(daycode)
                #print(daycode_same)
                #print(stockcode)
                #print(stockcode_same)
                if daycode == daycode_same and stockcode == stockcode_same:
                    #print(daycode)
                    #print(str(daycode_same) + ' , ' + str(stockcode) + '=' + str(stockcode_same))
    #                print(daycode_same)
                    for k in range(1,lastcolumn):
                        row_copy = sheetstock.cell(j,k).value
                        sheetdaily.cell(i,k,value=row_copy)
#                        print(row_copy)
                else:
                        #print('何もしない')
                    pass
        #wb_stock.save(stockbook)

    wb_daily.save(l)

#終了時間取得-経過時間
print(t)
t1 = datetime.datetime.now()
print(t1)
dt = t1-t
print(dt)
#終了時間取得-経過時間

#稼働終了アナウンス
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾