import imp
import os
from re import L
from turtle import color
import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd
import numpy as np
import bottleneck as bn
import requests
import bs4
import time
import datetime
import glob
import re
import sys
import winsound

#開始時間取得
t = datetime.datetime.now()
d = datetime.datetime.now()
d1 = d.strftime('%Y%m%d')
#開始時間取得

#対象フォルダ指定
dirdaily = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/"
dirreferrence = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/test/20230417/"

file_list = glob.glob(dirdaily + '*.xlsx')
file_ref =dirreferrence+'保持銘柄リスト.xlsx'

for l in file_list:
    #日付データを順番に開く
    wb_daily = openpyxl.load_workbook(l)
    wb_sim = openpyxl.load_workbook(filename=file_ref)
    sheetsim = wb_sim.worksheets[0]
    sheetdaily = wb_daily.worksheets[0]
    lastrow_daily = sheetdaily.max_row+1
    lastrow_sim = sheetsim.max_row+1
    lastcolumn = sheetdaily.max_column

    for k in range(2, lastrow_sim):
        print(str(sheetsim.cell(k,2).value) +'_'+ sheetsim.cell(k,3).value)
        for i in range(2, lastrow_daily):
            stockcode = sheetdaily.cell(i,2).value
            stockcode_hold = sheetsim.cell(k,2).value
            
            if sheetdaily.cell(i,4).value is None:
                #print(str(sheetsim.cell(k,2).value)+'_'+sheetsim.cell(k,3).value)
                pass
            else:
                if sheetdaily.cell(i,4).value - sheetdaily.cell(i,199).value <= 0 and sheetdaily.cell(i,2).value==sheetsim.cell(k,2).value:
                    print(str(sheetsim.cell(k,2).value)+'_'+sheetsim.cell(k,3).value+'が、本日終値'+str(sheetdaily.cell(i,4).value)+'で売りアラートになっています。')
                else:
                    
                    pass
                

#終了時間取得-経過時間
print(t)
t1 = datetime.datetime.now()
print(t1)
dt = t1-t
print(dt)
#終了時間取得-経過時間

#稼働終了アナウンス
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾