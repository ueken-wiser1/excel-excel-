import os
from re import L
import openpyxl
from openpyxl.worksheet.dimensions import SheetFormatProperties
from openpyxl.worksheet.pagebreak import RowBreak
import requests
import bs4
import time
import datetime
import sys
import winsound
import glob
import xlrd
import numpy as np
import pandas as pd
from sklearn import preprocessing
import statistics
import pprint

#どんな動きをさせるのか
#excelを開く
#参照excelの各シートに記載された証券コードを読み込む
#シート名は"株式", "マーケット", "為替", "投信"
#読み込んだ証券コードにより株探URLを作成し、銘柄ページへ移動
#必要な情報を取得する

#excel数式入力を全てpythonでの計算に置き換える
#→標準化数値の導出に対して、配列への数値入力がexcel数式では出来なかったため

t = datetime.datetime.now().time()
#excelを開く
#対象：信用規制
dir_stock = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/"
#ダウンロードしたexcel(以下databook)を開く
stock_list = glob.glob(dir_stock + '*.xlsx')

for l in stock_list:
        stockbook = openpyxl.load_workbook(l)
#    print(str(stockbook))
        sheet01 = stockbook.worksheets[0]
        lastrow_stockbook = sheet01.max_row
        lastcolumn_stockbook = sheet01.max_column

        for g in range(2, lastrow_stockbook):
                sheet01.cell(row=g,column=25).value = 0
                sheet01.cell(row=g,column=30).value = 0
                sheet01.cell(row=g,column=31).value = 0
                sheet01.cell(row=g,column=32).value = 0
                sheet01.cell(row=g,column=33).value = 0
                sheet01.cell(row=g,column=118).value = 0
                sheet01.cell(row=g,column=152).value = 0
#全部を配列で取込
        array_entire = []
        for g in range(2,lastrow_stockbook+1):
                for f in range(1,lastcolumn_stockbook+1):
                        if sheet01.cell(row=g,column=f).value == '－':
                                sheet01.cell(row=g,column=f).value = 0
                        else:
                                pass
                        array_entire.append(sheet01.cell(row=g,column=f).value)
#        print(array_entire)
        count_plus=0

        for j in range(2,lastrow_stockbook):
                print(l)

#何連騰しているかを表示するコード
#AH-34列目に記載
#騰がったかどうかはE-前日比のプラスマイナスで判断する。
#0は騰がったとしてカウントしない。
#カウンタを設定して、前日比がプラスならカウンタをプラス1、0以下ならカウンタを0にする。
#前日比がNoneであれば、0を入れる。
                zenjituhi = sheet01.cell(row=j,column=5).value
                if zenjituhi is None:
                        zenjituhi = 0
                else:
                        pass

                if zenjituhi > 0:
                        count_plus += 1
                else:
                        count_plus = 0
                sheet01.cell(row=j,column=34).value = count_plus



#参照元のデータの無いセルを0で埋める
#M高値
                if sheet01.cell(row=j,column=13).value is None:
                        sheet01.cell(row=j,column=13).value = 0
                else:
                        pass
#N安値
                if sheet01.cell(row=j,column=14).value is None:
                        sheet01.cell(row=j,column=14).value = 0
                else:
                        pass


#T-VWAP
                if sheet01.cell(row=j,column=20).value is None:
                        sheet01.cell(row=j,column=20).value = 0
                else:
                        pass
#DG融資新規　DH融資返済　DI融資残高　DJ貸株新規　DK貸株返済　DL貸株残高　DM差引残高
#none_to_resultで関数化
                if sheet01.cell(row=j,column=111).value is None:
                        sheet01.cell(row=j,column=111).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=112).value is None:
                        sheet01.cell(row=j,column=112).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=113).value is None:
                        sheet01.cell(row=j,column=113).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=114).value is None:
                        sheet01.cell(row=j,column=114).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=115).value is None:
                        sheet01.cell(row=j,column=115).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=116).value is None:
                        sheet01.cell(row=j,column=116).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=117).value is None:
                        sheet01.cell(row=j,column=117).value = 0
                else:
                        pass
#データのないセルに対して、0を入れる
#出来高　約定回数　移動平均乖離率5日　移動平均乖離率25日　移動平均乖離率75日
                if sheet01.cell(row=j,column=11).value is None:
                        sheet01.cell(row=j,column=11).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=8).value is None:
                        sheet01.cell(row=j,column=8).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=204).value is None:
                        sheet01.cell(row=j,column=204).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=205).value is None:
                        sheet01.cell(row=j,column=205).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=206).value is None:
                        sheet01.cell(row=j,column=206).value = 0
                else:
                        pass
#信用買残　融資新規　融資返済　信用売残　貸株新規　貸株返済　貸株超過
                if sheet01.cell(row=j,column=23).value is None:
                        sheet01.cell(row=j,column=23).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=22).value is None:
                        sheet01.cell(row=j,column=22).value = 0
                else:
                        pass

#出来高変化率　約定回数変化率　信用買残変化率　信用売残変化率　平均約定金額
                if sheet01.cell(row=j,column=28).value is None:
                        sheet01.cell(row=j,column=28).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=29).value is None:
                        sheet01.cell(row=j,column=29).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=27).value is None:
                        sheet01.cell(row=j,column=27).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=26).value is None:
                        sheet01.cell(row=j,column=26).value = 0
                else:
                        pass
                if sheet01.cell(row=j,column=33).value is None:
                        sheet01.cell(row=j,column=33).value = 0
                else:
                        pass

#値幅の計算式入れ直し
                sheet01.cell(row=j, column=6).value = sheet01.cell(row=j,column=13).value - sheet01.cell(row=j,column=14).value

#売買代金2=VWAP(T)*出来高(K)
#                sheet01.cell(row=j,column=25).value = str('=T')+str(j)+str('*K')+str(j)
                if j == 1:
                        pass
                else:
#                        print(sheet01.cell(row=j,column=20).value)
#                        print(sheet01.cell(row=j-1,column=11).value)
                        sheet01.cell(row=j,column=25).value = sheet01.cell(row=j,column=20).value * sheet01.cell(row=j-1,column=11).value
#信用売残前週比：(V)
#もしj=2なら、sheet01...は0
#もしsheet01...の結果が0なら、その前の0でない値を入力
                if j > 1:
                        sheet01.cell(row=j,column=26).value = 0
                else:
#                        sheet01.cell(row=j,column=26).value = str('=V')+str(j)+str('-V')+str(j-1)
                        sheet01.cell(row=j,column=26).value = sheet01.cell(row=j,column=22).value - sheet01.cell(row=j-1,column=22).value

#信用買残前週比：(W)
#もしj=2なら、sheet01...は0
#もしsheet01...の結果が0なら、その前の0でない値を入力
#                sheet01.cell(row=j,column=27).value = str('=W')+str(j)+str('-W')+str(j-1)
                if j > 1:
                        sheet01.cell(row=j,column=27).value = 0
                else:
#                        sheet01.cell(row=j,column=27).value = str('=V')+str(j)+str('-V')+str(j-1)
                        sheet01.cell(row=j,column=27).value = sheet01.cell(row=j,column=23).value - sheet01.cell(row=j-1,column=23).value

#出来高前日比：(K)
#もしj=2なら、sheet01...は0
#                sheet01.cell(row=j,column=28).value = str('=K')+str(j)+str('-K')+str(j-1)
                if j > 1:
                        sheet01.cell(row=j,column=28).value = 0
                else:
                        sheet01.cell(row=j,column=28).value = sheet01.cell(row=j,column=11).value -sheet01.cell(row=j-1,column=11).value

#約定回数前日比：(H)
#もしj=2なら、sheet01...は0
#                sheet01.cell(row=j,column=29).value = str('=H')+str(j)+str('-H')+str(j-1)
                if j > 1:
                        sheet01.cell(row=j,column=29).value = 0
                else:
                        sheet01.cell(row=j,column=29).value = sheet01.cell(row=j,column=8).value - sheet01.cell(row=j-1,column=8).value

#平均約定金額=出来高(K)/約定回数(H)*VWAP(T)
#もし約定回数の値が0なら、平均約定金額は0を返す
                if sheet01.cell(row=j,column=8).value == 0:
                        sheet01.cell(row=j,column=33).value = 0
                else:
                        sheet01.cell(row=j,column=33).value = sheet01.cell(row=j,column=11).value / sheet01.cell(row=j,column=8).value * sheet01.cell(row=j,column=20).value
#回転日数=((融資残(DI)+貸株残(DL))*2)/(融資新規(DG)+融資返済(DH)+貸株新規(DJ)+貸株返済(DK)
#もし参照セルの値が一つでも0なら、回転日数は0を返す

                if sheet01.cell(row=j,column=111).value == 0 and sheet01.cell(row=j,column=112).value == 0 and sheet01.cell(row=j,column=114).value == 0 and sheet01.cell(row=j,column=115).value == 0:
                        sheet01.cell(row=j,column=118).value = 0
                elif sheet01.cell(row=j,column=111).value == str(0) and sheet01.cell(row=j,column=112).value == str(0) and sheet01.cell(row=j,column=114).value == str(0) and sheet01.cell(row=j,column=115).value == str(0):
                        sheet01.cell(row=j,column=118).value = 0
                else:
                        sheet01.cell(row=j,column=118).value = (int(sheet01.cell(row=j,column=113).value) + int(sheet01.cell(row=j,column=116).value))*2/(int(sheet01.cell(row=j,column=111).value) + int(sheet01.cell(row=j,column=112).value) + int(sheet01.cell(row=j,column=114).value) + int(sheet01.cell(row=j,column=115).value))
#現在株価との差=株価(D)-みんかぶ目標株価(EU)
#みんかぶスクレイピングは現状不要
#                sheet01.cell(row=j,column=152).value = sheet01.cell(row=j,column=4).value - sheet01.cell(row=j,column=151).value
#株価のリストを作って、その中で一部の要素を抽出するやり方を取るか
#株価の配列を取得する
                array_kabuka = []
                for h in range(1, lastrow_stockbook):
                        if sheet01.cell(row=h+1,column=4).value is None:
                                sheet01.cell(row=h+1,column=4).value = 0
                        else:
                                array_kabuka.append(float(sheet01.cell(row=h+1,column=4).value))
#                        print(array_kabuka)
#                print(array_kabuka)
#移動平均線数値　5日＝当日〜4日前の株価総和/5
                s = pd.Series(array_kabuka)
                if j < 6 :
                        sheet01.cell(row=j,column=201).value = 0
                else:
#                        print(s)
                        idoheikin5 = s.rolling(5).mean()
#                        print(idoheikin5[j])
#                        print(idoheikin5)
#                        print(idoheikin5[0])
                        sheet01.cell(row=j,column=201).value = idoheikin5[j-1]
#                        print(sheet01.cell(row=j,column=201).value)

#移動平均線数値　25日＝当日〜24日前の株価総和/25
                if j < 26 :
                        sheet01.cell(row=j,column=202).value = 0
                else:
                        idoheikin25 = s.rolling(25).mean()
                        sheet01.cell(row=j,column=202).value = idoheikin25[j-1]
#移動平均線数値　75日＝当日〜74日前の株価総和/75
                if j < 76 :
                        sheet01.cell(row=j,column=203).value = 0
                else:
#                        sheet01.cell(row=j,column=203).value = str('=SUM(D')+str(j-75)+str(':D')+str(j)+str(')/75')
                        idoheikin75 = s.rolling(75).mean()
                        sheet01.cell(row=j,column=203).value = idoheikin75[j-1]
#移動平均乖離率　5日＝（株価ー移動平均5日）/移動平均5日
                if j < 6 :
                        sheet01.cell(row=j,column=204).value = 0
                else:
                        idoheikinkairi5 = (array_kabuka[j-1]-s.rolling(5).mean())/s.rolling(5).mean()
#                        print(idoheikinkairi5[0])
                        sheet01.cell(row=j,column=204).value = idoheikinkairi5[j-1]
#                        print(j)
#                        print(idoheikinkairi5[j-1])
#                        print(i)
#                        print(array_kabuka)
#                        print(sum(array_kabuka[i-5:i]))
#                        print(sheet01.cell(row=j,column=4).value)
#                        print(sheet01.cell(row=j,column=204).value)

#移動平均乖離率　25日＝（株価ー移動平均25日）/移動平均25日
                if j < 26 :
                        sheet01.cell(row=j,column=205).value = 0
                else:
                        idoheikinkairi25 = (array_kabuka[j-1]-s.rolling(25).mean())/s.rolling(25).mean()
                        sheet01.cell(row=j,column=205).value = idoheikinkairi25[j-1]
#移動平均乖離率　75日＝（株価ー移動平均75日）/移動平均75日
                if j < 76 :
                        sheet01.cell(row=j,column=206).value = 0
                else:
                        idoheikinkairi75 = (array_kabuka[j-1]-s.rolling(75).mean())/s.rolling(75).mean()
                        sheet01.cell(row=j,column=206).value = idoheikinkairi75[j-1]
                        
#以下標準偏差計算について、@が入力されている
        def standardization(x):
                x_mean = statistics.mean(x)
                x_stdev = statistics.stdev(x)
                if x_stdev == 0:
                        return 0
                else:
                        return [(i - x_mean) / x_stdev for i in x]

        for i in range(1,lastrow_stockbook):
#平均出来高＝平均値（average(対象列)）
                array_dekidaka =[]
                for k in range(1, lastrow_stockbook):
                        array_dekidaka.append(int(sheet01.cell(row=k+1,column=11).value))
#                sheet01.cell(row=j,column=251).value = str('=AVERAGE(K:K)')
                sheet01.cell(row=i+1,column=251).value = sum(array_dekidaka)/len(array_dekidaka)
#平均約定回数＝平均値（average(対象列)）
                array_yakujo =[]
                for k in range(1, lastrow_stockbook):
                        if sheet01.cell(row=k+1,column=8).value is None:
                                sheet01.cell(row=k+1,column=8).value = 0
                        else:
                                pass
                        array_yakujo.append(sheet01.cell(row=k+1,column=8).value)
#                sheet01.cell(row=j,column=252).value = str('=AVERAGE(H:H)')
#                print(array_yakujo)
                sheet01.cell(row=i+1,column=252).value = sum(array_yakujo)/len(array_yakujo)
#平均回転日数＝平均値（average(対象列)）
                array_kaiten =[]
                for k in range(1, lastrow_stockbook):
                        array_kaiten.append(float(sheet01.cell(row=k+1,column=118).value))
                        print(array_kaiten)
#                sheet01.cell(row=j,column=253).value = str('=AVERAGE(DN:DN)')
                sheet01.cell(row=i+1,column=253).value = sum(array_kaiten)/len(array_kaiten)
#平均移動平均乖離率5日＝平均値（average(対象列)）
                array_idoheikin5 =[]
                
                for k in range(1, lastrow_stockbook):
#                        print(sheet01.cell(row=k+1,column=204).value)
                        array_idoheikin5.append(sheet01.cell(row=k+1,column=204).value)
#                print(array_idoheikin5)
#                sheet01.cell(row=j,column=254).value = str('=AVERAGE(GV:GV)')
#                print(array_idoheikin5)
                sheet01.cell(row=i+1,column=254).value = sum(array_idoheikin5)/len(array_idoheikin5)
#平均移動平均乖離率25日＝平均値（average(対象列)）
                array_idoheikin25 =[]
                for k in range(1, lastrow_stockbook):
                        array_idoheikin25.append(float(sheet01.cell(row=k+1,column=205).value))
#                sheet01.cell(row=j,column=255).value = str('=AVERAGE(GW:GW)')
                sheet01.cell(row=i+1,column=255).value = sum(array_idoheikin25)/len(array_idoheikin25)
#平均移動平均乖離率75日＝平均値（average(対象列)）
                array_idoheikin75 =[]
                for k in range(1, lastrow_stockbook):
                        array_idoheikin75.append(float(sheet01.cell(row=k+1,column=206).value))
#                sheet01.cell(row=j,column=256).value = str('=AVERAGE(GX:GX)')
                sheet01.cell(row=i+1,column=256).value = sum(array_idoheikin75)/len(array_idoheikin75)
#平均信用買残＝平均値（average(対象列)）
                array_kaizan =[]
                for k in range(1, lastrow_stockbook):
                        array_kaizan.append(float(sheet01.cell(row=k+1,column=23).value))
#                sheet01.cell(row=j,column=257).value = str('=AVERAGE(W:W)')
                sheet01.cell(row=i+1,column=257).value = sum(array_kaizan)/len(array_kaizan)
#平均融資新規＝平均値（average(対象列)）
                array_yushishinki =[]
                for k in range(1, lastrow_stockbook):
                        array_yushishinki.append(float(sheet01.cell(row=k+1,column=111).value))
#                sheet01.cell(row=j,column=258).value = str('=AVERAGE(DG:DG)')
                sheet01.cell(row=i+1,column=258).value = sum(array_yushishinki)/len(array_yushishinki)
#平均融資返済＝平均値（average(対象列)）
#参照元に対して、何も入力されていなければ、0を入力しておく
                array_yusihensai =[]
                for k in range(1, lastrow_stockbook):
                        array_yusihensai.append(float(sheet01.cell(row=k+1,column=112).value))
#                sheet01.cell(row=j,column=259).value = str('=AVERAGE(DH:DH)')
                sheet01.cell(row=i+1,column=259).value = sum(array_yusihensai)/len(array_yusihensai)
#平均信用売残＝平均値（average(対象列)）
                array_urizan =[]
                for k in range(1, lastrow_stockbook):
                        array_urizan.append(float(sheet01.cell(row=k+1,column=22).value))
#                sheet01.cell(row=j,column=260).value = str('=AVERAGE(V:V)')
                sheet01.cell(row=i+1,column=260).value = sum(array_urizan)/len(array_urizan)
#平均貸株新規＝平均値（average(対象列)）
                array_kashishinki =[]
                for k in range(1, lastrow_stockbook):
                        array_kashishinki.append(float(sheet01.cell(row=k+1,column=114).value))
#                sheet01.cell(row=j,column=261).value = str('=AVERAGE(DJ:DJ)')
                sheet01.cell(row=i+1,column=261).value = sum(array_kashishinki)/len(array_kashishinki)
#平均貸株返済＝平均値（average(対象列)）
                array_kashihensai =[]
                for k in range(1, lastrow_stockbook):
                        array_kashihensai.append(float(sheet01.cell(row=k+1,column=115).value))
#                sheet01.cell(row=j,column=262).value = str('=AVERAGE(DK:DK)')
                sheet01.cell(row=i+1,column=262).value = sum(array_kashihensai)/len(array_kashihensai)
#平均貸株超過＝平均値（average(対象列)）
#参照元に対して、何も入力されていなければ、0を入力しておく
                array_kashichoka =[]
                for k in range(1, lastrow_stockbook):
                        array_kashichoka.append(float(sheet01.cell(row=k+1,column=116).value))
#                sheet01.cell(row=j,column=263).value = str('=AVERAGE(DL:DL)')
                sheet01.cell(row=i+1,column=263).value = sum(array_kashichoka)/len(array_kashichoka)
                        
#平均出来高変化率＝平均値（average(対象列)）
                array_dekidakahenka =[]
                for k in range(1, lastrow_stockbook):
                        array_dekidakahenka.append(int(sheet01.cell(row=k+1,column=28).value))
#                sheet01.cell(row=j,column=264).value = str('=AVERAGE(AB:AB)')
                sheet01.cell(row=i+1,column=264).value = sum(array_dekidakahenka)/len(array_dekidakahenka)
#平均約定回数変化率＝平均値（average(対象列)）
                array_yakujohenka =[]
                for k in range(1, lastrow_stockbook):
                        array_yakujohenka.append(int(sheet01.cell(row=k+1,column=29).value))
#                sheet01.cell(row=j,column=265).value = str('=AVERAGE(AC:AC)')
                sheet01.cell(row=i+1,column=265).value = sum(array_yakujohenka)/len(array_yakujohenka)
#平均信用買残変化率＝平均値（average(対象列)）
                array_kaizanhenka =[]
                for k in range(1, lastrow_stockbook):
                        array_kaizanhenka.append(float(sheet01.cell(row=k+1,column=27).value))
#                sheet01.cell(row=j,column=266).value = str('=AVERAGE(AA:AA)')
                sheet01.cell(row=i+1,column=266).value = sum(array_kaizanhenka)/len(array_kaizanhenka)
#平均信用売残変化率＝平均値（average(対象列)）
                array_urizanhenka =[]
                for k in range(1, lastrow_stockbook):
                        array_urizanhenka.append(float(sheet01.cell(row=k+1,column=27).value))
#                sheet01.cell(row=j,column=267).value = str('=AVERAGE(Z:Z)')
                sheet01.cell(row=i+1,column=267).value = sum(array_urizanhenka)/len(array_urizanhenka)
#平均平均約定金額＝平均値（average(対象列)）
                array_heikinyakujo =[]
                for k in range(1, lastrow_stockbook):
                        array_heikinyakujo.append(float(sheet01.cell(row=k+1,column=33).value))
#                sheet01.cell(row=j,column=268).value = str('=AVERAGE(AG:AG)')
                sheet01.cell(row=i+1,column=268).value = sum(array_heikinyakujo)/len(array_heikinyakujo)
#以下標準偏差計算について、@が入力されている
#標準偏差出来高＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=269).value = str('=STDEV.P(K:K)')
                array_dekidaka =[]
                for k in range(1, lastrow_stockbook):
                        array_dekidaka.append(int(sheet01.cell(row=k+1,column=11).value))
#                print(array_dekidaka)
                data = np.array(array_dekidaka)
#                print(data)
                std = np.std(data)
                sheet01.cell(row=i+1, column=269).value = std
#標準偏差約定回数＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=270).value = str('=STDEV.P(H:H)')
                array_yakujo =[]
                for k in range(1, lastrow_stockbook):
#                        print(array_yakujo)
                        if sheet01.cell(row=k+1,column=8).value == '－' or sheet01.cell(row=k+1,column=8).value is None:
                                sheet01.cell(row=k+1,column=8).value = 0
                        else:
                                pass
                        array_yakujo.append(sheet01.cell(row=k+1,column=8).value)
                data = np.array(array_yakujo)
                std = np.std(data)
                sheet01.cell(row=i+1, column=270).value = std
#標準偏差回転日数＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=271).value = str('=STDEV.P(DN:DN)')
                array_kaiten =[]
                for k in range(1, lastrow_stockbook):
                        array_kaiten.append(float(sheet01.cell(row=k+1,column=118).value))
                data = np.array(array_kaiten)
                std = np.std(data)
                sheet01.cell(row=i+1, column=271).value = std
#標準偏差移動平均乖離率5日＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=272).value = str('=STDEV.P(GV:GV)')
                array_idoheikin5 =[]
                for k in range(1, lastrow_stockbook):
                        array_idoheikin5.append(sheet01.cell(row=k+1,column=204).value)
#                print(array_idoheikin5)
                data = np.array(array_idoheikin5)
#                print(array_idoheikin5)
                std = np.std(data)
#                print(i)
#                print(data)
#                print(std)
                sheet01.cell(row=i+1, column=272).value = std
#標準偏差移動平均乖離率25日＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=273).value = str('=STDEV.P(GW:GW)')
                array_idoheikin25 =[]
                for k in range(1, lastrow_stockbook):
                        array_idoheikin25.append(float(sheet01.cell(row=k+1,column=205).value))
                data = np.array(array_idoheikin25)
                std = np.std(data)
                sheet01.cell(row=i+1, column=273).value = std
#標準偏差移動平均乖離率75日＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=274).value = str('=STDEV.P(GX:GX)')
                array_idoheikin75 =[]
                for k in range(1, lastrow_stockbook):
                        array_idoheikin75.append(float(sheet01.cell(row=k+1,column=206).value))
                data = np.array(array_idoheikin75)
                std = np.std(data)
                sheet01.cell(row=i+1, column=274).value = std
#標準偏差信用買残＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=275).value = str('=STDEV.P(W:W)')
                array_kaizan =[]
                for k in range(1, lastrow_stockbook):
                        array_kaizan.append(float(sheet01.cell(row=k+1,column=23).value))
                data = np.array(array_kaizan)
                std = np.std(data)
                sheet01.cell(row=i+1, column=275).value = std
#標準偏差融資新規＝標準偏差(stdevp(対象列)）入力無し
#                sheet01.cell(row=j,column=276).value = str('=STDEV.P(DG:DG))')
                array_yushishinki =[]
                for k in range(1, lastrow_stockbook):
                        array_yushishinki.append(float(sheet01.cell(row=k+1,column=111).value))
                data = np.array(array_yushishinki)
                std = np.std(data)
                sheet01.cell(row=i+1, column=276).value = std
#標準偏差融資返済＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=277).value = str('=STDEV.P(DH:DH)')
                array_yusihensai =[]
                for k in range(1, lastrow_stockbook):
                        array_yusihensai.append(float(sheet01.cell(row=k+1,column=112).value))
                data = np.array(array_yusihensai)
                std = np.std(data)
                sheet01.cell(row=i+1, column=277).value = std
#標準偏差信用売残＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=278).value = str('=STDEV.P(V:V)')
                array_urizan =[]
                for k in range(1, lastrow_stockbook):
                        array_urizan.append(float(sheet01.cell(row=k+1,column=22).value))
                data = np.array(array_urizan)
                std = np.std(data)
                sheet01.cell(row=i+1, column=278).value = std
#標準偏差貸株新規＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=279).value = str('=STDEV.P(DJ:DJ)')
                array_kashishinki =[]
                for k in range(1, lastrow_stockbook):
                        array_kashishinki.append(float(sheet01.cell(row=k+1,column=114).value))
                data = np.array(array_kashishinki)
                std = np.std(data)
                sheet01.cell(row=i+1, column=279).value = std
#標準偏差貸株返済＝標準偏差(stdevp(対象列)）入力無し
#                sheet01.cell(row=j,column=280).value = str('=STDEV.P(DK:DK))')
                array_kashihensai =[]
                for k in range(1, lastrow_stockbook):
                        array_kashihensai.append(float(sheet01.cell(row=k+1,column=115).value))
                data = np.array(array_kashihensai)
                std = np.std(data)
                sheet01.cell(row=i+1, column=280).value = std
#標準偏差貸株超過＝標準偏差(stdevp(対象列)）入力無し
#                sheet01.cell(row=j,column=281).value = str('=STDEV.P(DL:DL))')
                array_kashichoka =[]
                for k in range(1, lastrow_stockbook):
                        array_kashichoka.append(float(sheet01.cell(row=k+1,column=116).value))
                data = np.array(array_kashichoka)
                std = np.std(data)
                sheet01.cell(row=i+1, column=281).value = std
#標準偏差出来高変化率＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=282).value = str('=STDEV.P(AB:AB)')
                array_dekidakahenka =[]
                for k in range(1, lastrow_stockbook):
                        array_dekidakahenka.append(int(sheet01.cell(row=k+1,column=28).value))
                data = np.array(array_dekidakahenka)
                std = np.std(data)
                sheet01.cell(row=i+1, column=282).value = std
#標準偏差約定回数変化率＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=283).value = str('=STDEV.P(AC:AC)')
                array_yakujohenka =[]
                for k in range(1, lastrow_stockbook):
                        array_yakujohenka.append(int(sheet01.cell(row=k+1,column=29).value))
                data = np.array(array_yakujohenka)
                std = np.std(data)
                sheet01.cell(row=i+1, column=283).value = std
#標準偏差信用買残変化率＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=284).value = str('=STDEV.P(AA:AA)')
                array_kaizanhenka =[]
                for k in range(1, lastrow_stockbook):
                        array_kaizanhenka.append(float(sheet01.cell(row=k+1,column=27).value))
                data = np.array(array_kaizanhenka)
                std = np.std(data)
                sheet01.cell(row=i+1, column=284).value = std
#標準偏差信用売残変化率＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=285).value = str('=STDEV.P(Z:Z))')
                array_urizanhenka =[]
                for k in range(1, lastrow_stockbook):
                        array_urizanhenka.append(float(sheet01.cell(row=k+1,column=26).value))
                data = np.array(array_urizanhenka)
                std = np.std(data)
                sheet01.cell(row=i+1, column=285).value = std
#標準偏差平均約定金額＝標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=286).value = str('=STDEV.P(AG:AG))')
                array_heikinyakujo =[]
                for k in range(1, lastrow_stockbook):
                        array_heikinyakujo.append(float(sheet01.cell(row=k+1,column=33).value))
                data = np.array(array_heikinyakujo)
                std = np.std(data)
                sheet01.cell(row=i+1, column=286).value = std


#標準化出来高＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=501).value = str('=STANDARDIZE(K')+str(j)+str(',IQ')+str(j)+str(',JI')+str(j)+str(')')
#                print(standardization(array_dekidaka))
#                print(array_dekidaka[i])
                std_dekidaka = standardization(array_dekidaka)
#                print(array_dekidaka[1])
#                print(i)
#                print(lastrow_stockbook)
#                print(std_dekidaka[i-1])
#                print(std_dekidaka[-1])
#                print(std_dekidaka[len(std_dekidaka)])
#                print(std_dekidaka[i+2])
#                print(len(std_dekidaka))


                sheet01.cell(row=i+1,column=501).value = std_dekidaka[i-1]
#標準化約定回数＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
#                sheet01.cell(row=j,column=502).value = str('=STANDARDIZE(H')+str(j)+str(',IR')+str(j)+str(',JJ')+str(j)+str(')')
                std_yakujo = standardization(array_yakujo)
                sheet01.cell(row=i+1,column=502).value = std_yakujo[i-1]
#標準化回転日数＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                std_kaiten = standardization(array_kaiten)
                sheet01.cell(row=i+1,column=503).value = std_kaiten[i-1]

#標準化移動平均乖離率5日＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                if i < 6:
                        pass
                else:
                        std_idoheikin5 = standardization(array_idoheikin5)
#                        print(array_idoheikin5)
#                        print(std_idoheikin5)
                        sheet01.cell(row=i+1,column=504).value = std_idoheikin5[i-1]
#                print(array_idoheikin5)
#標準化移動平均乖離率25日＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                if i < 26:
                        pass
                else:
                        std_idoheikin25 = standardization(array_idoheikin25)
                        sheet01.cell(row=i+1,column=505).value = std_idoheikin25[i-1]
#標準化移動平均乖離率75日＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                if i < 76:
                        pass
                else:
                        std_idoheikin75 = standardization(array_idoheikin75)
                        sheet01.cell(row=i+1,column=506).value = std_idoheikin75[i-1]

#標準化信用買残＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                std_kaizan = standardization(array_kaizan)
                sheet01.cell(row=i+1,column=507).value = std_kaizan[i-1]

#標準化融資新規＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                std_yushishinki = standardization(array_yushishinki)
                sheet01.cell(row=i+1,column=508).value = std_yushishinki[i-1]

#標準化融資返済＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                std_yusihensai = standardization(array_yusihensai)
                sheet01.cell(row=i+1,column=509).value = std_yusihensai[i-1]

#標準化信用売残＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                std_urizan = standardization(array_urizan)
                sheet01.cell(row=i+1,column=510).value = std_urizan[i-1]

#標準化貸株新規＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                std_kashishinki = standardization(array_kashishinki)
                sheet01.cell(row=i+1,column=511).value = std_kashishinki[i-1]

#標準化貸株返済＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                std_kashihensai = standardization(array_kashihensai)
                sheet01.cell(row=i+1,column=512).value = std_kashihensai[i-1]

#標準化貸株超過＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                std_kashichoka = standardization(array_kashichoka)
                sheet01.cell(row=i+1,column=513).value = std_kashichoka[i-1]

#標準化出来高変化率＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                std_dekidakahenka = standardization(array_dekidakahenka)
                sheet01.cell(row=i+1,column=514).value = std_dekidakahenka[i-1]

#標準化約定回数変化率＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                std_yakujohenka = standardization(array_yakujohenka)
                sheet01.cell(row=i+1,column=515).value = std_yakujohenka[i-1]

#標準化信用買残変化率＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                std_kaizanhenka = standardization(array_kaizanhenka)
                sheet01.cell(row=i+1,column=516).value = std_kaizanhenka[i-1]

#標準化信用売残変化率＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                std_urizanhenka = standardization(array_urizanhenka)
                sheet01.cell(row=i+1,column=517).value = std_urizanhenka[i-1]

#標準化平均約定金額＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
                std_heikinyakujo = standardization(array_heikinyakujo)
                sheet01.cell(row=i+1,column=518).value = std_heikinyakujo[i-1]

#売り時/買い時スコア＝12*標準化出来高(SG)+15*標準化約定回数(SH)+6*標準化平均約定金額(SX)+3/標準化回転日数(SI)+9*標準化移動平均乖離率5日(SJ)+6*標準化信用買残(SM)+10*信用規制(CW)+8*増担規制(CX)+4*標準化融資新規(SN)+2/標準化融資返済(SO)+4/標準化信用売残(SP)+5*空売り規制(DB)+3/標準化貸株新規(SQ)+2*標準化貸株返済(SR)+1/標準化貸株超過(SS)+(20*IR)
#                print(sheet01.cell(row=i,column=501).value)
#                print(sheet01.cell(row=i,column=502).value)
#                print(sheet01.cell(row=i,column=518).value)
#                print(sheet01.cell(row=i,column=503).value)
#                print(sheet01.cell(row=i,column=504).value)#
                if sheet01.cell(row=i+1,column=504).value is None:
                        sheet01.cell(row=i+1,column=504).value = 0
                else:
                        pass
#                print(sheet01.cell(row=i,column=507).value)
#                print(sheet01.cell(row=i,column=101).value)#
                if sheet01.cell(row=i+1,column=101).value is None:
                        sheet01.cell(row=i+1,column=101).value = 0
                else:
                        pass
#                print(sheet01.cell(row=i,column=102).value)#
                if sheet01.cell(row=i+1,column=102).value is None:
                        sheet01.cell(row=i+1,column=102).value = 0
                else:
                        pass
#                print(sheet01.cell(row=i,column=508).value)
#                print(sheet01.cell(row=i,column=509).value)
#                print(sheet01.cell(row=i,column=510).value)
#                print(sheet01.cell(row=i,column=106).value)#
                if sheet01.cell(row=i+1,column=106).value is None:
                        sheet01.cell(row=i+1,column=106).value = 0
                else:
                        pass
#                print(sheet01.cell(row=i,column=511).value)
#                print(sheet01.cell(row=i,column=512).value)
#                print(sheet01.cell(row=i,column=513).value)
#                print(sheet01.cell(row=i,column=51).value)#
                if sheet01.cell(row=i+1,column=51).value is None:
                        sheet01.cell(row=i+1,column=51).value = 0
                else:
                        pass

                sheet01.cell(row=i+1,column=1000).value = sheet01.cell(row=i+1,column=501).value*12 + sheet01.cell(row=i+1,column=502).value*15 + sheet01.cell(row=i+1,column=518).value*6 + 3/(1+sheet01.cell(row=i+1,column=503).value) + sheet01.cell(row=i+1,column=504).value*9 + sheet01.cell(row=i+1,column=507).value*6 + sheet01.cell(row=i+1,column=101).value*10 + sheet01.cell(row=i+1,column=102).value*8 + sheet01.cell(row=i+1,column=508).value*4 + 2/(1+sheet01.cell(row=i+1,column=509).value) + 4/(1+sheet01.cell(row=i+1,column=510).value) + sheet01.cell(row=i+1,column=106).value*5 + 3/(1+sheet01.cell(row=i+1,column=511).value) + sheet01.cell(row=i+1,column=512).value*2 + 1/(1+sheet01.cell(row=i+1,column=513).value) + sheet01.cell(row=i+1,column=51).value*20

        stockbook.save(l)
        stockbook.close()
print(t)
t = datetime.datetime.now().time()
print(t)


winsound.Beep(500,50)  #ビープ音（500Hzの音を50msec流す）
