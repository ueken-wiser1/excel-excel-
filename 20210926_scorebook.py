#
#スコア集計プログラム
#銘柄データから出来高、約定回数の変化等を取得し、当日の全銘柄を並べたシートを作成する。
#上記シートについて、各項目を昇順/降順ソートし、上から順番に別のシートに並べる。
#並べるのは通常版と、株価500円以下で篩い分けした2パターン。
#証券コードと会社名、できれば当日株価も記載。
#通常版、低位株版と昇順、降順の4パターンを一覧したい。
#→フォーマット側で、1シート目:昇順=買い時スコア、2シート目:降順=売り時スコアとするか。
#1.銘柄データから各日のデータを引っ張ってきて、その日のスコアを取得していく
#　→これは最終日のみでいいか：頭はバラバラだが、お尻は決まっているため


#import
from os import close
import openpyxl
import datetime
import winsound
import pandas as pd
import glob
import os

#使用ディレクトリ、ファイル指定
dir_stock ="C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式"
dir_outcome = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/スコアブック/"
scorebook = openpyxl.load_workbook("C:/Users/touko/OneDrive/株価分析/excel/株式データ/スコアブック/20210926_売買時スコア一覧表-フォーマット.xlsx")
stock_list = glob.glob(dir_stock + '/*.xlsx')
sheet01 = scorebook.worksheets[0]
sheet02 = scorebook.worksheets[1]
sheet03 = scorebook.worksheets[2]

#稼働時間計測開始
t = datetime.datetime.now().time()
day = datetime.date.today()
print(day)

#os._exit()

#本文
#sheet03に当日データを入力
for l in stock_list:
    stockbook = openpyxl.load_workbook(l)
    print(l)
#    print(stockbook[l])
    sheet_stock = stockbook.worksheets[0]
    lastrow_stockbook = sheet_stock.max_row
    lastcolumn_stockbook = sheet_stock.max_column
#    print(lastrow_stockbook)
#    print(lastcolumn_stockbook)
#    print(stock_list.index(l))
    for j in (2,stock_list.index(l)+1):

        for i in range(1,lastcolumn_stockbook+1):
            cell_value = sheet_stock.cell(row=lastrow_stockbook,column=i).value
            sheet03.cell(row=j,column=i).value = cell_value
#            print(j)
#            print(i)
#            print(cell_value)


        

    
    stockbook.close()

scorebook.save('C:/Users/touko/OneDrive/株価分析/excel/株式データ/スコアブック/'+str(day)+'score.xlsx')

#sheet01に降順ソートしたデータを入力
#ワークシートを読み込んで、読み込んだデータフレームを以て弄るというのが常道か
#pandasでワークシートを読み込み、データフレーム化
#読み込んだデータフレームに対して、指定した列のソート
#上から順に読み込んだもの、上から読み込みかつ500円以下のものを出力

print(t)
t = datetime.datetime.now().time()
print(t)

winsound.Beep(500,50)  #ビープ音（500Hzの音を50msec流す）