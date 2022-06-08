import os
import openpyxl
from openpyxl.worksheet.dimensions import SheetDimension
import requests
import bs4
import time
import datetime
import re
import glob
#------------お約束開始---冒頭
#稼働時間計測開始
import datetime
t = datetime.datetime.now().time()
#------------お約束終了---冒頭

'''
仕様
1. 銘柄データ1列目の日付と株探各銘柄ニュースで使っている日付を比較
2. 一致したら、IR有無の列にチェックする
3. タグの種類を分けること
    td_kaiji:開示情報
    ctg1:市況-ng
    ctg2:材料
    ctg3_kk:決算
    ctg3_ks:修正
    ctg4:テク
    ctg5:特集-ng
    ctg9:注目-ng
    ctg12:5%
4. 走査した日付には上記IRチェックとは別の、走査済みフラグを付与して、次回の走査時にはその日付はスキップするフローにする

'''

#------------プログラム本文---ここから
folder01 = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/"
kabutan_newsURL_base = 'https://kabutan.jp/stock/news?code='


#関数定義
def is_include_listed_word(text, word_list):
    for listed_word in word_list:
        if listed_word in text:
            return True
    return False
    
#各銘柄データを開く

t1 = datetime.datetime.now().time()
print(t1)
today_file = glob.glob(folder01+"*.xlsx")
stockbook = openpyxl.load_workbook(today_file[0])
sheet01 = stockbook.worksheets[0]
lastrow_stockbook =sheet01.max_row
for i in range(2,lastrow_stockbook+1):
    if sheet01.cell(row=i,column=999).value == 1:
        print(lastrow_stockbook,sheet01.cell(row=i,column=1).value,"passします")
        continue
    else:
        date_stockbook = sheet01.cell(row=i,column=1).value
        date_stockbookstr = str(date_stockbook)
        slicedate_stockbook = date_stockbookstr[0:10]
        stockcode = sheet01.cell(row=i,column=2).value
    #対象の銘柄データの証券コードを開く
    #対象銘柄の株探ニュースページの各ページ送り

        time.sleep(0.3)
        
        res = requests.get(kabutan_newsURL_base+stockcode)
        res.raise_for_status()
        soup = bs4.BeautifulSoup(res.text, 'html.parser')
        datetags = soup.find_all("td",class_="news_time")
        num_datetags = len(datetags)
    #対象ニュースページの記載日付を取り込む
        for k in range(num_datetags):
            date_site = soup.select(".news_time")[k]
            date_sitestr = str(date_site)
            slicedate_site = date_sitestr[38:48]
            parent_sitedate = date_site.parent
    #銘柄データで取り込んだ日付とニュースページで取り込んだ日付を突き合わせる
            if slicedate_stockbook == slicedate_site:
                # print(num_datetags)
                print(sheet01.cell(row=i,column=2).value,sheet01.cell(row=i,column=3).value)
    #ニュースカテゴリーで分類分け
    #ニュース掲載時間で分類分け－当日or翌日
                if "ctg_kaiji" in str(parent_sitedate) \
                    or "ctg2" in str(parent_sitedate):
                    if int(date_sitestr[49:51]) >=15:
                        sheet01.cell(row=i,column=56).value = 1
                        print(parent_sitedate)
                    else:
                        sheet01.cell(row=i,column=51).value = 1
                        print(parent_sitedate)
                elif "ctg3_kk" in str(parent_sitedate):
                    if int(date_sitestr[49:51]) >=15:
                        sheet01.cell(row=i,column=57).value = 1
                        print(parent_sitedate)
                    else:
                        sheet01.cell(row=i,column=52).value = 1
                        print(parent_sitedate)
                elif "ctg3_ks" in str(parent_sitedate):
                    if int(date_sitestr[49:51]) >=15:
                        sheet01.cell(row=i,column=58).value = 1
                        print(parent_sitedate)
                    else:
                        sheet01.cell(row=i,column=53).value = 1
                        print(parent_sitedate)
                elif "ctg4" in str(parent_sitedate):
                    if int(date_sitestr[49:51]) >=15:
                        sheet01.cell(row=i,column=59).value = 1
                        print(parent_sitedate)
                    else:
                        sheet01.cell(row=i,column=54).value = 1
                        print(parent_sitedate)
                elif "ctg12" in str(parent_sitedate):
                    if int(date_sitestr[49:51]) >=15:
                        sheet01.cell(row=i,column=60).value = 1
                        print(parent_sitedate)
                    else:
                        sheet01.cell(row=i,column=55).value = 1
                        print(parent_sitedate)
                else:
                    pass
                    print(sheet01.cell(row=i,column=2).value,sheet01.cell(row=i,column=3).value,"check完了")
    sheet01.cell(row=i,column=999).value = 1
stockbook.save(today_file[0])

#------------プログラム本文---ここまで

#------------お約束開始---末尾
#稼働時間表示
print(t)
t = datetime.datetime.now().time()
print(t)

#稼働終了アナウンス
import winsound
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾