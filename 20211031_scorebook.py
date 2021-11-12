import os
from re import L
import openpyxl
import time
import datetime
import sys
import winsound
import glob
import xlrd
import numpy as np
import pandas as pd

#スコアブックプログラム
#1. シートを3つ用意する
#2. シート3に"株式"フォルダ内の全銘柄の最終行をコピーペーストする
#3. 貼り付けたデータを配列にする
#4. 配列の指定した要素でソートする
#5. 昇順ソートの上位10位をシート1に、降順ソートの上位10位をシート2に転記する
#6. 昇順ソートで価格で条件分岐：500円以下の上位10位をシート1に転記する
#7. 降順ソートで価格で条件分岐：500円以下の上位10位をシート2に転記する
#8. 終わったブックを保存して閉じる

t = datetime.datetime.now().time()
d = datetime.date.today()
#print(d)

dir_stock = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/完了/"
stock_list = glob.glob(dir_stock + '*.xlsx')

scorebook = openpyxl.Workbook()
scoresheet_allscore = scorebook.create_sheet(title='allscore')
#allscoreシートへの列名の記載
#これ、コードとして長すぎ
#別プログラムにして、呼び出す感じにしたい

scoresheet_allscore.cell(row=1,column=1).value = '日付'
scoresheet_allscore.cell(row=1,column=2).value = 'コード'
scoresheet_allscore.cell(row=1,column=3).value = '会社名'
scoresheet_allscore.cell(row=1,column=4).value = '株価'
scoresheet_allscore.cell(row=1,column=5).value = '前日比'
scoresheet_allscore.cell(row=1,column=6).value = '値幅'
scoresheet_allscore.cell(row=1,column=7).value = '売買代金'
scoresheet_allscore.cell(row=1,column=8).value = '約定回数'
scoresheet_allscore.cell(row=1,column=9).value = '決算日'
scoresheet_allscore.cell(row=1,column=10).value = '前日終値'
scoresheet_allscore.cell(row=1,column=11).value = '出来高'
scoresheet_allscore.cell(row=1,column=12).value = '始値'
scoresheet_allscore.cell(row=1,column=13).value = '高値'
scoresheet_allscore.cell(row=1,column=14).value = '安値'
scoresheet_allscore.cell(row=1,column=15).value = '終値'
scoresheet_allscore.cell(row=1,column=16).value = '前日比%'#不使用
scoresheet_allscore.cell(row=1,column=17).value = 'PER'
scoresheet_allscore.cell(row=1,column=18).value = 'PBR'
scoresheet_allscore.cell(row=1,column=19).value = '上場市場'#不使用
scoresheet_allscore.cell(row=1,column=20).value = 'VWAP'
scoresheet_allscore.cell(row=1,column=21).value = '発行済株式数'
scoresheet_allscore.cell(row=1,column=22).value = '最新信用売残'
scoresheet_allscore.cell(row=1,column=23).value = '最新信用買残'
scoresheet_allscore.cell(row=1,column=24).value = '信用倍率'
scoresheet_allscore.cell(row=1,column=25).value = '売買代金2'#計算値
scoresheet_allscore.cell(row=1,column=26).value = '信用売残前週比'
scoresheet_allscore.cell(row=1,column=27).value = '信用買残前週比'
scoresheet_allscore.cell(row=1,column=28).value = '出来高前日比'
scoresheet_allscore.cell(row=1,column=29).value = '約定回数前日比'
scoresheet_allscore.cell(row=1,column=30).value = '時価総額'
scoresheet_allscore.cell(row=1,column=31).value = '浮動株総額'#不使用
scoresheet_allscore.cell(row=1,column=32).value = '取引規模'#不使用
scoresheet_allscore.cell(row=1,column=33).value = '平均約定金額'#不使用
scoresheet_allscore.cell(row=1,column=34).value = '連騰回数'
scoresheet_allscore.cell(row=1,column=35).value = '時価総額2'
scoresheet_allscore.cell(row=1,column=51).value = '当日IR有無'
scoresheet_allscore.cell(row=1,column=101).value = '信用取引規制中'
scoresheet_allscore.cell(row=1,column=102).value = '貸借取引銘柄別増担保金徴収措置'
scoresheet_allscore.cell(row=1,column=103).value = '貸借取引銘柄別増担保金徴収措置内容'
scoresheet_allscore.cell(row=1,column=106).value = '空売規制対象'
scoresheet_allscore.cell(row=1,column=111).value = '融資新規'
scoresheet_allscore.cell(row=1,column=112).value = '融資返済'
scoresheet_allscore.cell(row=1,column=113).value = '融資残高'
scoresheet_allscore.cell(row=1,column=114).value = '貸株新規'
scoresheet_allscore.cell(row=1,column=115).value = '貸株返済'
scoresheet_allscore.cell(row=1,column=116).value = '貸株残高'
scoresheet_allscore.cell(row=1,column=117).value = '差引残高'
scoresheet_allscore.cell(row=1,column=118).value = '回転日数'
scoresheet_allscore.cell(row=1,column=121).value = '貸株超過株数'
scoresheet_allscore.cell(row=1,column=122).value = '最高料率'
scoresheet_allscore.cell(row=1,column=123).value = '当日品貸料率'
scoresheet_allscore.cell(row=1,column=124).value = '前日品貸料率'
scoresheet_allscore.cell(row=1,column=151).value = 'みんかぶ目標株価'
scoresheet_allscore.cell(row=1,column=152).value = '現在株価との差'
scoresheet_allscore.cell(row=1,column=201).value = '移動平均線数値5日'
scoresheet_allscore.cell(row=1,column=202).value = '移動平均線数値25日'
scoresheet_allscore.cell(row=1,column=203).value = '移動平均線数値75日'
scoresheet_allscore.cell(row=1,column=204).value = '移動平均乖離率5日'
scoresheet_allscore.cell(row=1,column=205).value = '移動平均乖離率25日'
scoresheet_allscore.cell(row=1,column=206).value = '移動平均乖離率75日'
scoresheet_allscore.cell(row=1,column=251).value = '平均出来高'
scoresheet_allscore.cell(row=1,column=252).value = '平均約定回数'
scoresheet_allscore.cell(row=1,column=253).value = '平均回転日数'
scoresheet_allscore.cell(row=1,column=254).value = '平均移動平均乖離率5日'
scoresheet_allscore.cell(row=1,column=255).value = '平均移動平均乖離率25日'
scoresheet_allscore.cell(row=1,column=256).value = '平均移動平均乖離率75日'
scoresheet_allscore.cell(row=1,column=257).value = '平均信用買残'
scoresheet_allscore.cell(row=1,column=258).value = '平均融資新規'
scoresheet_allscore.cell(row=1,column=259).value = '平均融資返済'
scoresheet_allscore.cell(row=1,column=260).value = '平均信用売残'
scoresheet_allscore.cell(row=1,column=261).value = '平均貸株新規'
scoresheet_allscore.cell(row=1,column=262).value = '平均貸株返済'
scoresheet_allscore.cell(row=1,column=263).value = '平均貸株超過'
scoresheet_allscore.cell(row=1,column=264).value = '平均出来高変化率'
scoresheet_allscore.cell(row=1,column=265).value = '平均約定回数変化率'
scoresheet_allscore.cell(row=1,column=266).value = '平均買残変化率'
scoresheet_allscore.cell(row=1,column=267).value = '平均売残変化率'
scoresheet_allscore.cell(row=1,column=268).value = '平均平均約定金額'
scoresheet_allscore.cell(row=1,column=269).value = '標準偏差出来高'
scoresheet_allscore.cell(row=1,column=270).value = '標準偏差約定回数'
scoresheet_allscore.cell(row=1,column=271).value = '標準偏差回転日数'
scoresheet_allscore.cell(row=1,column=272).value = '標準偏差移動平均乖離率5日'
scoresheet_allscore.cell(row=1,column=273).value = '標準偏差移動平均乖離率25日'
scoresheet_allscore.cell(row=1,column=274).value = '標準偏差移動平均乖離率75日'
scoresheet_allscore.cell(row=1,column=275).value = '標準偏差信用買残'
scoresheet_allscore.cell(row=1,column=276).value = '標準偏差融資新規'
scoresheet_allscore.cell(row=1,column=277).value = '標準偏差融資返済'
scoresheet_allscore.cell(row=1,column=278).value = '標準偏差信用売残'
scoresheet_allscore.cell(row=1,column=279).value = '標準偏差貸株新規'
scoresheet_allscore.cell(row=1,column=280).value = '標準偏差貸株返済'
scoresheet_allscore.cell(row=1,column=281).value = '標準偏差貸株超過'
scoresheet_allscore.cell(row=1,column=282).value = '標準偏差出来高変化率'
scoresheet_allscore.cell(row=1,column=283).value = '標準偏差約定回数変化率'
scoresheet_allscore.cell(row=1,column=284).value = '標準偏差信用買残変化率'
scoresheet_allscore.cell(row=1,column=285).value = '標準偏差信用売残変化率'
scoresheet_allscore.cell(row=1,column=286).value = '標準偏差平均約定金額'
scoresheet_allscore.cell(row=1,column=501).value = '標準化出来高'
scoresheet_allscore.cell(row=1,column=502).value = '標準化約定回数'
scoresheet_allscore.cell(row=1,column=503).value = '標準化回転日数'
scoresheet_allscore.cell(row=1,column=504).value = '標準化移動平均乖離率5日'
scoresheet_allscore.cell(row=1,column=505).value = '標準化移動平均乖離率25日'
scoresheet_allscore.cell(row=1,column=506).value = '標準化移動平均乖離率75日'
scoresheet_allscore.cell(row=1,column=507).value = '標準化信用買残'
scoresheet_allscore.cell(row=1,column=508).value = '標準化融資新規'
scoresheet_allscore.cell(row=1,column=509).value = '標準化融資返済'
scoresheet_allscore.cell(row=1,column=510).value = '標準化信用売残'
scoresheet_allscore.cell(row=1,column=511).value = '標準化貸株新規'
scoresheet_allscore.cell(row=1,column=512).value = '標準化貸株返済'
scoresheet_allscore.cell(row=1,column=513).value = '標準化貸株超過'
scoresheet_allscore.cell(row=1,column=514).value = '標準化出来高変化率'
scoresheet_allscore.cell(row=1,column=515).value = '標準化約定回数変化率'
scoresheet_allscore.cell(row=1,column=516).value = '標準化信用買残変化率'
scoresheet_allscore.cell(row=1,column=517).value = '標準化信用売残変化率'
scoresheet_allscore.cell(row=1,column=518).value = '標準化平均約定金額'
scoresheet_allscore.cell(row=1,column=1000).value = '売り時/買い時スコア'

##買い時ランキング　昇順ソート
scoresheet_downsort = scorebook.create_sheet(title='downsort')

scoresheet_downsort.cell(row=1,column=1).value = '売買スコア'
scoresheet_downsort.cell(row=1,column=2).value = 'コード'
scoresheet_downsort.cell(row=1,column=3).value = '会社名'
scoresheet_downsort.cell(row=1,column=4).value = '株価'
scoresheet_downsort.cell(row=1,column=5).value = '連騰回数'

scoresheet_downsort.cell(row=1,column=6).value = '出来高変化'
scoresheet_downsort.cell(row=1,column=7).value = 'コード'
scoresheet_downsort.cell(row=1,column=8).value = '会社名'
scoresheet_downsort.cell(row=1,column=9).value = '株価'
scoresheet_downsort.cell(row=1,column=10).value = '連騰回数'

scoresheet_downsort.cell(row=1,column=11).value = '約定変化'
scoresheet_downsort.cell(row=1,column=12).value = 'コード'
scoresheet_downsort.cell(row=1,column=13).value = '会社名'
scoresheet_downsort.cell(row=1,column=14).value = '株価'
scoresheet_downsort.cell(row=1,column=15).value = '連騰回数'

scoresheet_downsort.cell(row=1,column=16).value = '時価総額'
scoresheet_downsort.cell(row=1,column=17).value = 'コード'
scoresheet_downsort.cell(row=1,column=18).value = '会社名'
scoresheet_downsort.cell(row=1,column=19).value = '株価'
scoresheet_downsort.cell(row=1,column=20).value = '連騰回数'

scoresheet_downsort.cell(row=1,column=21).value = '買残変化'
scoresheet_downsort.cell(row=1,column=22).value = 'コード'
scoresheet_downsort.cell(row=1,column=23).value = '会社名'
scoresheet_downsort.cell(row=1,column=24).value = '株価'
scoresheet_downsort.cell(row=1,column=25).value = '連騰回数'

##売り時ランキング　降順ソート
scoresheet_upsort = scorebook.create_sheet(title='upsort')

scoresheet_upsort.cell(row=1,column=1).value = '売買スコア'
scoresheet_upsort.cell(row=1,column=2).value = 'コード'
scoresheet_upsort.cell(row=1,column=3).value = '会社名'
scoresheet_upsort.cell(row=1,column=4).value = '株価'
scoresheet_upsort.cell(row=1,column=5).value = '連騰回数'

scoresheet_upsort.cell(row=1,column=6).value = '出来高変化'
scoresheet_upsort.cell(row=1,column=7).value = 'コード'
scoresheet_upsort.cell(row=1,column=8).value = '会社名'
scoresheet_upsort.cell(row=1,column=9).value = '株価'
scoresheet_upsort.cell(row=1,column=10).value = '連騰回数'

scoresheet_upsort.cell(row=1,column=11).value = '約定変化'
scoresheet_upsort.cell(row=1,column=12).value = 'コード'
scoresheet_upsort.cell(row=1,column=13).value = '会社名'
scoresheet_upsort.cell(row=1,column=14).value = '株価'
scoresheet_upsort.cell(row=1,column=15).value = '連騰回数'

scoresheet_upsort.cell(row=1,column=16).value = '時価総額'
scoresheet_upsort.cell(row=1,column=17).value = 'コード'
scoresheet_upsort.cell(row=1,column=18).value = '会社名'
scoresheet_upsort.cell(row=1,column=19).value = '株価'
scoresheet_upsort.cell(row=1,column=20).value = '連騰回数'

scoresheet_upsort.cell(row=1,column=21).value = '買残変化'
scoresheet_upsort.cell(row=1,column=22).value = 'コード'
scoresheet_upsort.cell(row=1,column=23).value = '会社名'
scoresheet_upsort.cell(row=1,column=24).value = '株価'
scoresheet_upsort.cell(row=1,column=25).value = '連騰回数'


#買い時ランキング　低位株
scoresheet_downsort.cell(row=13,column=1).value = '売買スコア'
scoresheet_downsort.cell(row=13,column=2).value = 'コード'
scoresheet_downsort.cell(row=13,column=3).value = '会社名'
scoresheet_downsort.cell(row=13,column=4).value = '株価'
scoresheet_downsort.cell(row=13,column=5).value = '連騰回数'

scoresheet_downsort.cell(row=13,column=6).value = '出来高変化'
scoresheet_downsort.cell(row=13,column=7).value = 'コード'
scoresheet_downsort.cell(row=13,column=8).value = '会社名'
scoresheet_downsort.cell(row=13,column=9).value = '株価'
scoresheet_downsort.cell(row=13,column=10).value = '連騰回数'

scoresheet_downsort.cell(row=13,column=11).value = '約定変化'
scoresheet_downsort.cell(row=13,column=12).value = 'コード'
scoresheet_downsort.cell(row=13,column=13).value = '会社名'
scoresheet_downsort.cell(row=13,column=14).value = '株価'
scoresheet_downsort.cell(row=13,column=15).value = '連騰回数'

scoresheet_downsort.cell(row=13,column=16).value = '時価総額'
scoresheet_downsort.cell(row=13,column=17).value = 'コード'
scoresheet_downsort.cell(row=13,column=18).value = '会社名'
scoresheet_downsort.cell(row=13,column=19).value = '株価'
scoresheet_downsort.cell(row=13,column=20).value = '連騰回数'

scoresheet_downsort.cell(row=13,column=21).value = '買残変化'
scoresheet_downsort.cell(row=13,column=22).value = 'コード'
scoresheet_downsort.cell(row=13,column=23).value = '会社名'
scoresheet_downsort.cell(row=13,column=24).value = '株価'
scoresheet_downsort.cell(row=13,column=25).value = '連騰回数'

#売り時ランキング　低位株
scoresheet_upsort.cell(row=13,column=1).value = '売買スコア'
scoresheet_upsort.cell(row=13,column=2).value = 'コード'
scoresheet_upsort.cell(row=13,column=3).value = '会社名'
scoresheet_upsort.cell(row=13,column=4).value = '株価'
scoresheet_upsort.cell(row=13,column=5).value = '連騰回数'

scoresheet_upsort.cell(row=13,column=6).value = '出来高変化'
scoresheet_upsort.cell(row=13,column=7).value = 'コード'
scoresheet_upsort.cell(row=13,column=8).value = '会社名'
scoresheet_upsort.cell(row=13,column=9).value = '株価'
scoresheet_upsort.cell(row=13,column=10).value = '連騰回数'

scoresheet_upsort.cell(row=13,column=11).value = '約定変化'
scoresheet_upsort.cell(row=13,column=12).value = 'コード'
scoresheet_upsort.cell(row=13,column=13).value = '会社名'
scoresheet_upsort.cell(row=13,column=14).value = '株価'
scoresheet_upsort.cell(row=13,column=15).value = '連騰回数'

scoresheet_upsort.cell(row=13,column=16).value = '時価総額'
scoresheet_upsort.cell(row=13,column=17).value = 'コード'
scoresheet_upsort.cell(row=13,column=18).value = '会社名'
scoresheet_upsort.cell(row=13,column=19).value = '株価'
scoresheet_upsort.cell(row=13,column=20).value = '連騰回数'

scoresheet_upsort.cell(row=13,column=21).value = '買残変化'
scoresheet_upsort.cell(row=13,column=22).value = 'コード'
scoresheet_upsort.cell(row=13,column=23).value = '会社名'
scoresheet_upsort.cell(row=13,column=24).value = '株価'
scoresheet_upsort.cell(row=13,column=25).value = '連騰回数'

#ソートキー：売買スコア　出来高変化　約定変化　時価総額　買残変化
j = 2

array_infocolum = [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,\
                    31,32,33,34,35,36,37,38,51,101,102,103,106,111,112,113,114,115,116,117,118,121,122,123,124,\
                    151,152,201,202,203,204,205,206,\
                    251,252,253,254,255,256,257,258,259,260,261,262,263,264,265,266,267,268,269,270,\
                    271,272,273,274,275,276,277,278,279,280,281,282,283,284,285,286,\
                    501,502,503,504,505,506,507,508,509,510,511,512,513,514,515,516,517,518,1000]

for l in stock_list:
    stockbook = openpyxl.load_workbook(l)
    sheet01 = stockbook.worksheets[0]
    lastrow_stockbook = sheet01.max_row
    lastcolumn_stockbook = sheet01.max_column
    print(sheet01.cell(row=2,column=2).value)

#0を入力する列は"allkabu1の列タイトル"にある列のみにする
#対象列は配列にする

    for i in range(2,lastcolumn_stockbook+1):
        scoresheet_allscore.cell(row=j,column=i).value = sheet01.cell(row=lastrow_stockbook,column=i).value
        if scoresheet_allscore.cell(row=j,column=i).value is None or scoresheet_allscore.cell(row=j,column=i).value == '－':
            if i in array_infocolum:
                scoresheet_allscore.cell(row=j,column=i).value = 0
            else:
                pass
        else:
                pass

    stockbook.close()
    j += 1

scorebook_dir = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/スコアブック/'
today_scorebook = scorebook_dir+str(d)+'score.xlsx'
scorebook.save(today_scorebook)

lastrow_scorebook = scoresheet_allscore.max_row
lastcolumn_scorebook = scoresheet_allscore.max_column
print(today_scorebook) 
#todayscore = []
#for i in range(2,lastrow_scorebook):
#    for k in range(2,lastcolumn_scorebook):
#        todayscore.append(scoresheet_allscore.cell(row=i,column=k).value)
df = pd.read_excel(today_scorebook, sheet_name='allscore', engine="openpyxl")

def sort_and_paste(sortkey, ascend, yokozure):
    df_s = df.sort_values(sortkey, ascending=ascend)
    n=df_s.columns.get_loc(sortkey)
    j=1
    k=1
    #sortkeyの変化でcolumnの書き込む列が変わる
    if ascend == True:
        for i in range(2,lastrow_scorebook+1):
            if j<10:
#ソートキー
#ソートキーは行列の何番目の列か？
                scoresheet_upsort.cell(row=j+1,column=yokozure*5+1).value = df_s.iat[i-2,n]
#証券コード
                scoresheet_upsort.cell(row=j+1,column=yokozure*5+2).value = df_s.iat[i-2,1]
#会社名
                scoresheet_upsort.cell(row=j+1,column=yokozure*5+3).value = df_s.iat[i-2,2]
#株価
                scoresheet_upsort.cell(row=j+1,column=yokozure*5+4).value = df_s.iat[i-2,3]
#連騰回数
                scoresheet_upsort.cell(row=j+1,column=yokozure*5+5).value = df_s.iat[i-2,33]
                j += 1
            else:
                pass

            if k<10:
                if df_s.iat[i-2,3] < 500:
                    scoresheet_upsort.cell(row=k+12,column=yokozure*5+1).value = df_s.iat[i-2,n]
                    scoresheet_upsort.cell(row=k+12,column=yokozure*5+2).value = df_s.iat[i-2,1]
                    scoresheet_upsort.cell(row=k+12,column=yokozure*5+3).value = df_s.iat[i-2,2]
                    scoresheet_upsort.cell(row=k+12,column=yokozure*5+4).value = df_s.iat[i-2,3]
                    scoresheet_upsort.cell(row=k+12,column=yokozure*5+5).value = df_s.iat[i-2,33]
                    k += 1
                else:
                    pass
            else:
                    pass
    else:            
        for i in range(2,lastrow_scorebook+1):
            if j<10:
                scoresheet_downsort.cell(row=j+1,column=yokozure*5+1).value = df_s.iat[i-2,n]
                scoresheet_downsort.cell(row=j+1,column=yokozure*5+2).value = df_s.iat[i-2,1]
                scoresheet_downsort.cell(row=j+1,column=yokozure*5+3).value = df_s.iat[i-2,2]
                scoresheet_downsort.cell(row=j+1,column=yokozure*5+4).value = df_s.iat[i-2,3]
                scoresheet_downsort.cell(row=j+1,column=yokozure*5+5).value = df_s.iat[i-2,33]
                j += 1
            else:
                pass

            if k<10:
                if df_s.iat[i-2,3] < 500:
                    scoresheet_downsort.cell(row=k+12,column=yokozure*5+1).value = df_s.iat[i-2,n]
                    scoresheet_downsort.cell(row=k+12,column=yokozure*5+2).value = df_s.iat[i-2,1]
                    scoresheet_downsort.cell(row=k+12,column=yokozure*5+3).value = df_s.iat[i-2,2]
                    scoresheet_downsort.cell(row=k+12,column=yokozure*5+4).value = df_s.iat[i-2,3]
                    scoresheet_downsort.cell(row=k+12,column=yokozure*5+5).value = df_s.iat[i-2,33]
                    k += 1
                else:
                    pass
            else:
                    pass

#print(df)
#ソートキー：売買スコア　出来高変化　約定変化　時価総額　買残変化

#df_s = df.sort_values('売り時/買い時スコア', ascending=True)
sort_and_paste('売り時/買い時スコア',True,0)
sort_and_paste('売り時/買い時スコア',False,0)
#df_s = df.sort_values("標準化出来高変化率", ascending=False)
sort_and_paste('出来高',True,1)
sort_and_paste('出来高',False,1)
#df_s = df.sort_values("標準化約定回数変化率", ascending=False)
sort_and_paste('PBR',True,2)
sort_and_paste('PBR',False,2)
#df_s = df.sort_values("時価総額", ascending=False)
sort_and_paste('時価総額2',True,3)
sort_and_paste('時価総額2',False,3)
#df_s = df.sort_values("標準化信用買残変化率", ascending=False)
sort_and_paste('最新信用買残',True,4)
sort_and_paste('最新信用買残',False,4)

scorebook.save(today_scorebook)

print(t)
t = datetime.datetime.now().time()
print(t)

winsound.Beep(500,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）