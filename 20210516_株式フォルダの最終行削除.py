#何を見るのか
#日付のフォーマットとそれらを照合して、一致と見なすかどうかの確認

import os
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import datetime
import jpholiday
from datetime import timedelta
import pandas as pd


t = datetime.datetime.now().time()
d = datetime.datetime.now()
d1 = d.strftime('%Y%m%d')

watch_list_file = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/'
completed_folder = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/完了/完/'

for watch_list_file in glob.glob(os.path.join(watch_list_file, '*.xlsx')):

    wb_watch_list = load_workbook(watch_list_file)
    ws_watch_list = wb_watch_list.active
    # Excelファイルを読み込みます
    
    last_row = ws_watch_list.max_row
    #print('最終行は'+str(last_row))
    last_col = ws_watch_list.max_column
    #print('最終列は'+str(last_col))
    ws_watch_list.delete_rows(last_row)
    wb_watch_list.save(watch_list_file)
    print(watch_list_file+"の最終行を削除しました。")

print(t)
t = datetime.datetime.now().time()
print(t)

#稼働終了アナウンス
import winsound
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾