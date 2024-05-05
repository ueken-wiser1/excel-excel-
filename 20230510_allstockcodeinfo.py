import requests
import bs4
import os
import openpyxl
import time
import datetime
import winsound


#どんな動きをさせるのか
#excelを開く
#0000-9999まで証券コードを走査して、該当した銘柄のデータを抽出する
#エラーが出た場合はスキップ
#抽出するデータはこれまでと同じ

t = datetime.datetime.now().time()

#excelを開く
wb = openpyxl.Workbook()
#print(type(wb))
name = wb.get_sheet_names
#print(name)
#print(wb.get_sheet_names())
sheet = wb.active
#print(sheet.title)
#print(name)
#0000-9999までの証券コードを分解する



#------------プログラム本文---ここから
today = datetime.date.today()
d = today.strftime('%Y%m%d')
dir = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/スクレイピング生データ/"
wb = openpyxl.load_workbook(dir + 'stockcodelist.xlsx')
name = wb.get_sheet_names

#"株式"
sheet = wb.get_sheet_by_name('株式')
t = datetime.datetime.now().time()



for j in range(2, sheet.max_row + 1):
    time.sleep(0.3)
    kabutan_URL_base = 'http://kabutan.jp/stock/?code='
    kabutankessan_URL_base = 'http://kabutan.jp/stock/finance/?code='
    stock_code = sheet.cell(row=j, column=2).value
    kabutan_URL = kabutan_URL_base + str(stock_code)
    kabutankessan_URL = kabutankessan_URL_base + str(stock_code)
    res = requests.get(kabutan_URL)
    res2 = requests.get(kabutankessan_URL)
    print(kabutan_URL)
    res.raise_for_status()
    res2.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')
    soup2 = bs4.BeautifulSoup(res2.text, 'html.parser')

    tags = soup.find_all('li')       #全aタグ取得
    #print('tdタグ数：', len(tdtags))  #aタグ数取得a53 td25
    num_tags = int(len(tags))
#読み込んだ情報をexcelファイルに書き込む
#書き出す初めのセルの列を指定
    write_column = 1

#日付を入れる
    d_today = datetime.date.today()
    sheet.cell(row=j, column=1).value = d_today
    write_column += 2
#    t = soup.select('time')[4]
#    print(t)
#           名称
    stock_name = soup.select('h3')[0]
    print(stock_name)
    sheet.cell(row=j, column=3).value = str(soup.select('h3')[0].get_text())
    write_column += 1

#           現在株価
    try:
        stock_price = str(soup.select('td')[32].get_text())
    except IndexError as e:
#        print('現在株価存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
        write_column += 1

#           前日比
    try:
        DoD = str(soup.select('dd')[0].get_text())
    except IndexError as e:
#        print('前日比存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = DoD
        write_column += 1


#           決算日
    try:
        stock_price = str(soup.select('dd')[7].get_text())
    except IndexError as e:
#        print('前日終値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
        write_column += 1


#           PER
    try:
        kabu_PER = str(soup.select('td')[18].get_text())
    except IndexError as e:
#        print('PER存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = kabu_PER
        write_column += 1

#           PBR
    try:
        kabu_PBR = str(soup.select('td')[19].get_text())
    except IndexError as e:
#        print('PBR存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = kabu_PBR
        write_column += 1

#           上場市場
    try:
        stock_price = str(soup.select('span')[13].get_text())
    except IndexError as e:
#        print('上場市場存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
        write_column += 1


#           発行済み株式数
    try:
        number_of_issued_shares = str(soup.select('td')[42].get_text())
    except IndexError as e:
#        print('発行済み株式数存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = number_of_issued_shares
        write_column += 1


#           単元株
    try:
        unit_share = str(soup.select('td')[40].get_text())
    except IndexError as e:
        print('単元株存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = unit_share
        write_column += 1

    write_column = 402
#           会社説明
    try:
        company_info = str(soup.select('td')[99].get_text())
    except IndexError as e:
        print('単元株存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = company_info
        write_column += 1

    write_column = 401
#           業界
    try:
        work = str(soup.select('td')[100].get_text())
    except IndexError as e:
        print('単元株存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = work
        write_column += 1
    
    write_column = 601

#           テーマ
    for i in range(num_tags):
        try:
            word = "theme"
            theme = str(soup.select('li')[i].get_text())
        except IndexError as e:
            print('単元株存在しない')
            write_column += 1
            pass
        else:
            if word in str(theme):
                sheet.cell(row=j, column=write_column).value = theme
                write_column += 1

wb.save('C:/Users/touko/OneDrive/株価分析/excel/株式データ/銘柄データ集計/stockcodelist_info.xlsx')

print(t)
t = datetime.datetime.now().time()
print(t)

winsound.Beep(500,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）