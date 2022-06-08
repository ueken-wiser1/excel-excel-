import os
from re import L
import openpyxl
import time
import datetime
import sys
import winsound
import glob
import xlrd
import numpy as np
import pandas as pd

#------------お約束開始---冒頭
#稼働時間計測開始
import datetime
t = datetime.datetime.now().time()
#------------お約束終了---冒頭

#------------プログラム本文---ここから
d = datetime.date.today()

scorebook_dir = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/スコアブック/'
scorebook = scorebook_dir+str(d)+'score.xlsx'
today_scorebook = openpyxl.load_workbook(scorebook)

scoresheet_allscore = today_scorebook['allscore']

lastrow_scorebook = scoresheet_allscore.max_row
lastcolumn_scorebook = scoresheet_allscore.max_column
print(scorebook) 

#------------列名指定---ここから
##買い時ランキング　昇順ソート
scoresheet_downsort = today_scorebook['downsort']

scoresheet_downsort.cell(row=1,column=1).value = '売買スコア'
scoresheet_downsort.cell(row=1,column=2).value = 'コード'
scoresheet_downsort.cell(row=1,column=3).value = '会社名'
scoresheet_downsort.cell(row=1,column=4).value = '株価'
scoresheet_downsort.cell(row=1,column=5).value = '連騰回数'

scoresheet_downsort.cell(row=1,column=6).value = '出来高'
scoresheet_downsort.cell(row=1,column=7).value = 'コード'
scoresheet_downsort.cell(row=1,column=8).value = '会社名'
scoresheet_downsort.cell(row=1,column=9).value = '株価'
scoresheet_downsort.cell(row=1,column=10).value = '連騰回数'

scoresheet_downsort.cell(row=1,column=11).value = 'PBR'
scoresheet_downsort.cell(row=1,column=12).value = 'コード'
scoresheet_downsort.cell(row=1,column=13).value = '会社名'
scoresheet_downsort.cell(row=1,column=14).value = '株価'
scoresheet_downsort.cell(row=1,column=15).value = '連騰回数'

scoresheet_downsort.cell(row=1,column=16).value = '時価総額'
scoresheet_downsort.cell(row=1,column=17).value = 'コード'
scoresheet_downsort.cell(row=1,column=18).value = '会社名'
scoresheet_downsort.cell(row=1,column=19).value = '株価'
scoresheet_downsort.cell(row=1,column=20).value = '連騰回数'

scoresheet_downsort.cell(row=1,column=21).value = '最新信用買残'
scoresheet_downsort.cell(row=1,column=22).value = 'コード'
scoresheet_downsort.cell(row=1,column=23).value = '会社名'
scoresheet_downsort.cell(row=1,column=24).value = '株価'
scoresheet_downsort.cell(row=1,column=25).value = '連騰回数'

##売り時ランキング　降順ソート
scoresheet_upsort = today_scorebook['upsort']

scoresheet_upsort.cell(row=1,column=1).value = '売買スコア'
scoresheet_upsort.cell(row=1,column=2).value = 'コード'
scoresheet_upsort.cell(row=1,column=3).value = '会社名'
scoresheet_upsort.cell(row=1,column=4).value = '株価'
scoresheet_upsort.cell(row=1,column=5).value = '連騰回数'

scoresheet_upsort.cell(row=1,column=6).value = '出来高'
scoresheet_upsort.cell(row=1,column=7).value = 'コード'
scoresheet_upsort.cell(row=1,column=8).value = '会社名'
scoresheet_upsort.cell(row=1,column=9).value = '株価'
scoresheet_upsort.cell(row=1,column=10).value = '連騰回数'

scoresheet_upsort.cell(row=1,column=11).value = 'PBR'
scoresheet_upsort.cell(row=1,column=12).value = 'コード'
scoresheet_upsort.cell(row=1,column=13).value = '会社名'
scoresheet_upsort.cell(row=1,column=14).value = '株価'
scoresheet_upsort.cell(row=1,column=15).value = '連騰回数'

scoresheet_upsort.cell(row=1,column=16).value = '時価総額'
scoresheet_upsort.cell(row=1,column=17).value = 'コード'
scoresheet_upsort.cell(row=1,column=18).value = '会社名'
scoresheet_upsort.cell(row=1,column=19).value = '株価'
scoresheet_upsort.cell(row=1,column=20).value = '連騰回数'

scoresheet_upsort.cell(row=1,column=21).value = '最新信用買残'
scoresheet_upsort.cell(row=1,column=22).value = 'コード'
scoresheet_upsort.cell(row=1,column=23).value = '会社名'
scoresheet_upsort.cell(row=1,column=24).value = '株価'
scoresheet_upsort.cell(row=1,column=25).value = '連騰回数'


#買い時ランキング　低位株
scoresheet_downsort.cell(row=12,column=1).value = '売買スコア'
scoresheet_downsort.cell(row=12,column=2).value = 'コード'
scoresheet_downsort.cell(row=12,column=3).value = '会社名'
scoresheet_downsort.cell(row=12,column=4).value = '株価'
scoresheet_downsort.cell(row=12,column=5).value = '連騰回数'

scoresheet_downsort.cell(row=12,column=6).value = '出来高'
scoresheet_downsort.cell(row=12,column=7).value = 'コード'
scoresheet_downsort.cell(row=12,column=8).value = '会社名'
scoresheet_downsort.cell(row=12,column=9).value = '株価'
scoresheet_downsort.cell(row=12,column=10).value = '連騰回数'

scoresheet_downsort.cell(row=12,column=11).value = 'PBR'
scoresheet_downsort.cell(row=12,column=12).value = 'コード'
scoresheet_downsort.cell(row=12,column=13).value = '会社名'
scoresheet_downsort.cell(row=12,column=14).value = '株価'
scoresheet_downsort.cell(row=12,column=15).value = '連騰回数'

scoresheet_downsort.cell(row=12,column=16).value = '時価総額'
scoresheet_downsort.cell(row=12,column=17).value = 'コード'
scoresheet_downsort.cell(row=12,column=18).value = '会社名'
scoresheet_downsort.cell(row=12,column=19).value = '株価'
scoresheet_downsort.cell(row=12,column=20).value = '連騰回数'

scoresheet_downsort.cell(row=12,column=21).value = '最新信用買残'
scoresheet_downsort.cell(row=12,column=22).value = 'コード'
scoresheet_downsort.cell(row=12,column=23).value = '会社名'
scoresheet_downsort.cell(row=12,column=24).value = '株価'
scoresheet_downsort.cell(row=12,column=25).value = '連騰回数'

#売り時ランキング　低位株
scoresheet_upsort.cell(row=12,column=1).value = '売買スコア'
scoresheet_upsort.cell(row=12,column=2).value = 'コード'
scoresheet_upsort.cell(row=12,column=3).value = '会社名'
scoresheet_upsort.cell(row=12,column=4).value = '株価'
scoresheet_upsort.cell(row=12,column=5).value = '連騰回数'

scoresheet_upsort.cell(row=12,column=6).value = '出来高'
scoresheet_upsort.cell(row=12,column=7).value = 'コード'
scoresheet_upsort.cell(row=12,column=8).value = '会社名'
scoresheet_upsort.cell(row=12,column=9).value = '株価'
scoresheet_upsort.cell(row=12,column=10).value = '連騰回数'

scoresheet_upsort.cell(row=12,column=11).value = 'PBR'
scoresheet_upsort.cell(row=12,column=12).value = 'コード'
scoresheet_upsort.cell(row=12,column=13).value = '会社名'
scoresheet_upsort.cell(row=12,column=14).value = '株価'
scoresheet_upsort.cell(row=12,column=15).value = '連騰回数'

scoresheet_upsort.cell(row=12,column=16).value = '時価総額'
scoresheet_upsort.cell(row=12,column=17).value = 'コード'
scoresheet_upsort.cell(row=12,column=18).value = '会社名'
scoresheet_upsort.cell(row=12,column=19).value = '株価'
scoresheet_upsort.cell(row=12,column=20).value = '連騰回数'

scoresheet_upsort.cell(row=12,column=21).value = '最新信用買残'
scoresheet_upsort.cell(row=12,column=22).value = 'コード'
scoresheet_upsort.cell(row=12,column=23).value = '会社名'
scoresheet_upsort.cell(row=12,column=24).value = '株価'
scoresheet_upsort.cell(row=12,column=25).value = '連騰回数'
#------------列名指定---ここまで

#ソートキー：売買スコア　出来高　PBR　時価総額　最新信用買残
j = 2

df = pd.read_excel(today_scorebook, sheet_name='allscore', engine="openpyxl")

def sort_and_paste(sortkey, ascend, yokozure):
    df_s = df.sort_values(sortkey, ascending=ascend)
    n=df_s.columns.get_loc(sortkey)
    j=1
    k=1
    kabutan_URL_base = 'http://kabutan.jp/stock/?code='
    #sortkeyの変化でcolumnの書き込む列が変わる
    if ascend == True:
        for i in range(2,lastrow_scorebook+1):
            if j<10:
#ソートキー
#ソートキーは行列の何番目の列か？
                scoresheet_upsort.cell(row=j+1,column=yokozure*5+1).value = df_s.iat[i-2,n]
#証券コード ハイパーリンクも設定
                scoresheet_upsort.cell(row=j+1,column=yokozure*5+2).value = df_s.iat[i-2,1]
                scoresheet_upsort.cell(row=j+1,column=yokozure*5+2).hyperlink = kabutan_URL_base + str(df_s.iat[i-2,1])
#会社名
                scoresheet_upsort.cell(row=j+1,column=yokozure*5+3).value = df_s.iat[i-2,2]
#株価
                scoresheet_upsort.cell(row=j+1,column=yokozure*5+4).value = df_s.iat[i-2,3]
#連騰回数
                scoresheet_upsort.cell(row=j+1,column=yokozure*5+5).value = df_s.iat[i-2,33]
                j += 1
            else:
                pass

            if k<10:
                if df_s.iat[i-2,3] < 500 and df_s.iat[i-2,n] != 0:
                    scoresheet_upsort.cell(row=k+12,column=yokozure*5+1).value = df_s.iat[i-2,n]
                    scoresheet_upsort.cell(row=k+12,column=yokozure*5+2).value = df_s.iat[i-2,1]
                    scoresheet_upsort.cell(row=k+12,column=yokozure*5+2).hyperlink = kabutan_URL_base + str(df_s.iat[i-2,1])
                    scoresheet_upsort.cell(row=k+12,column=yokozure*5+3).value = df_s.iat[i-2,2]
                    scoresheet_upsort.cell(row=k+12,column=yokozure*5+4).value = df_s.iat[i-2,3]
                    scoresheet_upsort.cell(row=k+12,column=yokozure*5+5).value = df_s.iat[i-2,33]
                    k += 1
                else:
                    pass
            else:
                    pass
    else:            
        for i in range(2,lastrow_scorebook+1):
            if j<10:
                scoresheet_downsort.cell(row=j+1,column=yokozure*5+1).value = df_s.iat[i-2,n]
                scoresheet_downsort.cell(row=j+1,column=yokozure*5+2).value = df_s.iat[i-2,1]
                scoresheet_downsort.cell(row=j+1,column=yokozure*5+2).hyperlink = kabutan_URL_base + str(df_s.iat[i-2,1])
                scoresheet_downsort.cell(row=j+1,column=yokozure*5+3).value = df_s.iat[i-2,2]
                scoresheet_downsort.cell(row=j+1,column=yokozure*5+4).value = df_s.iat[i-2,3]
                scoresheet_downsort.cell(row=j+1,column=yokozure*5+5).value = df_s.iat[i-2,33]
                j += 1
            else:
                pass

            if k<10:
                if df_s.iat[i-2,3] < 500 and df_s.iat[i-2,n] != 0:
                    scoresheet_downsort.cell(row=k+12,column=yokozure*5+1).value = df_s.iat[i-2,n]
                    scoresheet_downsort.cell(row=k+12,column=yokozure*5+2).value = df_s.iat[i-2,1]
                    scoresheet_downsort.cell(row=k+12,column=yokozure*5+2).hyperlink = kabutan_URL_base + str(df_s.iat[i-2,1])
                    scoresheet_downsort.cell(row=k+12,column=yokozure*5+3).value = df_s.iat[i-2,2]
                    scoresheet_downsort.cell(row=k+12,column=yokozure*5+4).value = df_s.iat[i-2,3]
                    scoresheet_downsort.cell(row=k+12,column=yokozure*5+5).value = df_s.iat[i-2,33]
                    k += 1
                else:
                    pass
            else:
                    pass

#print(df)
#ソートキー：売買スコア　出来高変化　約定変化　時価総額　買残変化

#df_s = df.sort_values('売り時/買い時スコア', ascending=True)
sort_and_paste('売り時/買い時スコア',True,0)
sort_and_paste('売り時/買い時スコア',False,0)
#df_s = df.sort_values("標準化出来高変化率", ascending=False)
sort_and_paste('出来高',True,1)
sort_and_paste('出来高',False,1)
#df_s = df.sort_values("標準化約定回数変化率", ascending=False)
sort_and_paste('PBR',True,2)
sort_and_paste('PBR',False,2)
#df_s = df.sort_values("時価総額", ascending=False)
sort_and_paste('時価総額2',True,3)
sort_and_paste('時価総額2',False,3)
#df_s = df.sort_values("標準化信用買残変化率", ascending=False)
sort_and_paste('最新信用買残',True,4)
sort_and_paste('最新信用買残',False,4)

today_scorebook.save(scorebook)

#------------プログラム本文---ここまで

#------------お約束開始---末尾
#稼働時間表示
print(t)
t = datetime.datetime.now().time()
print(t)

#稼働終了アナウンス
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾