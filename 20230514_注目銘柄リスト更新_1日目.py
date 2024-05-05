
import os
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import datetime
import jpholiday
from datetime import timedelta
import shutil
from openpyxl.utils import get_column_letter

t = datetime.datetime.now().time()
d = datetime.datetime.now()
d1 = d.strftime('%Y%m%d')

def find_date_row(ws, date):
    for row2 in range(2, ws.max_row + 1):
        #print(code)
        if ws.cell(row=row2, column=1).value == date:
            return row2
    return None

watch_list_folder = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/追跡調査/1日目/'
stock_folder = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/'
dirstorage = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/追跡調査/2日目以降/'
n_columns = 41

plus_fill = PatternFill(patternType='solid', fgColor='ee82ee') #前日より値上がり
minus_fill = PatternFill(patternType='solid', fgColor='00bfff') #前日より値下がり
attained_fill = PatternFill(patternType='solid', fgColor='adff2f') #+2%目標を達成したら
unattained_fill = PatternFill(patternType='solid', fgColor='696969') #+2%目標未達で、5日目を迎えたら
for watch_list_file in glob.glob(os.path.join(watch_list_folder, '*.xlsx')):
    wb_watch_list = load_workbook(watch_list_file)
    ws_watch_list = wb_watch_list.active

    last_row = ws_watch_list.max_row
    #print('最終行は'+str(last_row))
    last_col = ws_watch_list.max_column
    last_col_first = ws_watch_list.max_column
    #print('最終列は'+str(last_col))
    last_date = ws_watch_list.cell(row=1, column=last_col).value

    basis_date = ws_watch_list.cell(row=1, column=last_col).value
    #print('対象日付は'+str(file_date))
    #初日データ処理が終わっていれば、見るべきはlast_colの日付だけ∴ここのbasis_dateは意味のないパラメータか
    last_col_fifth = ws_watch_list.max_column
    for row in range(2, last_row + 1):
        last_col_sixth = ws_watch_list.max_column
        completed_flag = ws_watch_list.cell(row=row, column=1).value
        stock_code = ws_watch_list.cell(row=row, column=4).value
        company_name= ws_watch_list.cell(row=row, column=5).value
#        if completed_flag == True:
#            print(str(security_code)+'は2%目標を達成しました。')
#            continue
#        elif completed_flag == False:
#            print(str(security_code)+'は-2%目標を達成しました。あるいは5日期限を満了しました。')
#            continue


        print("証券コード："+str(stock_code)+"を検索中")
        stock_data_file = os.path.join(stock_folder, f'{stock_code}_{company_name}.xlsx')
        print(stock_data_file)
        #print(watch_list_file)
        last_col_fourth = ws_watch_list.max_column

        if os.path.exists(stock_data_file):
            wb_stock_data = load_workbook(stock_data_file)
            ws_stock_data = wb_stock_data.active

            date_row = find_date_row(ws_stock_data, last_date)
            #print(security_code_row)
            last_col_third=ws_watch_list.max_column

            if date_row:
                closing_price = ws_stock_data.cell(row=date_row, column=4).value
                start_price = ws_stock_data.cell(row=date_row, column=12).value
                
                ws_watch_list.cell(row=row, column=last_col).value = closing_price
                ws_watch_list.cell(row=row, column=29).value = start_price
                print("証券コード"+str(stock_code) +"は、終値"+str(ws_watch_list.cell(row=row, column=last_col).value)+"を示しました")
                #print(date_row)               
                #print(start_price)

                diff=closing_price - start_price
                if diff > 0:
                    ws_watch_list.cell(row=row, column=last_col).fill = plus_fill
                elif diff < 0:
                    ws_watch_list.cell(row=row,column=last_col).fill = minus_fill

                if closing_price is not None and start_price is not None:
                    ratio = ((closing_price - start_price) / start_price)*100
                    ws_watch_list.cell(row=row, column=31).value = ratio
                    ws_watch_list.cell(row=row, column=30).value = closing_price - start_price
                    if ratio > 2:
                        ws_watch_list.cell(row=row, column=1).value = True
                        for i in range(1, n_columns):
                            ws_watch_list.cell(row=row, column=i).fill =attained_fill

                        profit = ws_watch_list.cell(row=row,column=last_col).value - ws_watch_list.cell(row=row,column=29).value
                        print(str(ws_watch_list.cell(row=row,column=4).value)+"_"+str(ws_watch_list.cell(row=row,column=5).value) + "は2%目標を達成しました。利益は"+str(profit)+"円です。")

                    elif ratio < -2:
                        ws_watch_list.cell(row=row, column=1).value = False
                        for i in range(1, n_columns):
                            ws_watch_list.cell(row=row, column=i).fill =unattained_fill
                        loss = ws_watch_list.cell(row=row,column=last_col).value - ws_watch_list.cell(row=row,column=29).value
                        print(str(ws_watch_list.cell(row=row,column=4).value)+"_"+str(ws_watch_list.cell(row=row,column=5).value) + "は-2%目標に達してしまいました。損失は"+str(loss)+"円です。")

#                    elif last_col == n_columns:
#                        ws_watch_list.cell(row=row, column=1).value = False
#                        for i in range(1, n_columns):
#                            ws_watch_list.cell(row=row, column=i).fill =unattained_fill
#                        loss = ws_watch_list.cell(row=row,column=last_col).value - ws_watch_list.cell(row=row,column=29).value
#                        print(str(ws_watch_list.cell(row=row,column=4).value)+"_"+str(ws_watch_list.cell(row=row,column=5).value) + "は目標を達成できず、5日期限を迎えました。損益は"+str(loss)+"円です。")
            last_col_second=ws_watch_list.max_column

            basis_date2 = basis_date + datetime.timedelta(days=1)
                # 指定日が平日になるまでループ
            while True:
                # 指定日が土曜日または日曜日の場合
                if basis_date2.weekday() == 5 or basis_date2.weekday() == 6:
                    #print(f"{basis_date2}は土曜日または日曜日です。")
                    # 指定日を翌日に上書き
                    basis_date2 = basis_date2 + datetime.timedelta(days=1)
                # 指定日が祝日の場合
                elif jpholiday.is_holiday(basis_date2):
                    #print(basis_date2)
                    #print(f"{basis_date2}は{str(jpholiday.is_holiday_name(basis_date2))}です。")
                    # 指定日を翌日に上書き
                    basis_date2 = basis_date2 + datetime.timedelta(days=1)
                # 指定日が平日の場合
                else:
                    #print(f"{basis_date2}は平日です。")
                    # ループを終了
                    break

            ws_watch_list.cell(row=1,column=last_col+1).value = basis_date2

            wb_stock_data.close()
    last_col_last = ws_watch_list.max_column
    for i in range(last_col_last+1, 1, -1):
        #print(ws_watch_list.cell(row=1, column=i).value)
        #last_col_last = ws_watch_list.max_column
        #print(last_col_last)
        if ws_watch_list.cell(row=1, column=i).value is None:
            ws_watch_list.delete_cols(i)
            print(str(i)+"列目は1行目に何もないため削除しました。")
    last_col_last = ws_watch_list.max_column
    wb_watch_list.save(watch_list_file)

    last_col_last = ws_watch_list.max_column


    print(watch_list_file+'を保存しました')

    wb_watch_list.close()
    shutil.move(watch_list_file, dirstorage)
    
print(t)
t = datetime.datetime.now().time()
print(t)

#稼働終了アナウンス
import winsound
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾