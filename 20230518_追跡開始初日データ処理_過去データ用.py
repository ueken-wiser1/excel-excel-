
import os
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import datetime
import jpholiday
import shutil

t = datetime.datetime.now().time()
d = datetime.datetime.now()
d1 = d.strftime('%Y%m%d')

def find_security_code_row(ws, code):
    for row2 in range(2, ws.max_row + 1):
        if int(ws.cell(row=row2, column=2).value) == int(code):
            #print(code)
            return row2
    raise ValueError(f"security code {code} not found")

watch_list_folder = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/追跡調査/0日目/'
completed_folder = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/完了/'
dividend = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/test/20230508/20230508_配当カレンダー.xlsx'
dirstorage = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/追跡調査/1日目/'
n_columns = 29
plus_fill = PatternFill(patternType='solid', fgColor='ee82ee') #前日より値上がり
minus_fill = PatternFill(patternType='solid', fgColor='00bfff') #前日より値下がり
attained_fill = PatternFill(patternType='solid', fgColor='adff2f') #+2%目標を達成したら
unattained_fill = PatternFill(patternType='solid', fgColor='696969') #-2%達成したら

for watch_list_file in glob.glob(os.path.join(watch_list_folder, '*.xlsx')):
    wb_watch_list = load_workbook(watch_list_file)
    ws_watch_list = wb_watch_list.active

    last_row = ws_watch_list.max_row
    #print('最終行は'+str(last_row))
    last_col = ws_watch_list.max_column
    #print('最終列は'+str(last_col))
    if ws_watch_list.cell(row=2, column=2).value is None:
        print("この日は対象無し")
        continue
    basis_date = ws_watch_list.cell(row=2, column=2).value + datetime.timedelta(days=1)
    print(basis_date)

    #print('対象日付は'+str(file_date))

    # 指定日が平日になるまでループ
    while True:
        # 指定日が土曜日または日曜日の場合
        if basis_date.weekday() == 5 or basis_date.weekday() == 6:
            print(f"{basis_date}は土曜日または日曜日です。")
            # 指定日を翌日に上書き
            basis_date = basis_date + datetime.timedelta(days=1)
        # 指定日が祝日の場合
        elif jpholiday.is_holiday(basis_date):
            print(f"{basis_date}は{str(jpholiday.is_holiday_name(basis_date))}です。")
            # 指定日を翌日に上書き
            basis_date = basis_date + datetime.timedelta(days=1)
        # 指定日が平日の場合
        else:
            print(f"{basis_date}は平日です。")
            # ループを終了
            break
    file_date = basis_date.strftime('%Y%m%d')
    ws_watch_list.cell(row=1,column=last_col+1).value = basis_date

    for row in range(2, last_row + 1):

        security_code = ws_watch_list.cell(row=row, column=4).value
        #print(security_code)
        daily_data_file = os.path.join(completed_folder, f'{file_date}_allkabu1.xlsx')
        print(daily_data_file)
        #print(watch_list_file)

        if os.path.exists(daily_data_file):
            wb_daily_data = load_workbook(daily_data_file)
            ws_daily_data = wb_daily_data.active

            security_code_row = find_security_code_row(ws_daily_data, security_code)
            #print(security_code_row)

            if security_code_row is not None:
                opening_price = ws_daily_data.cell(row=security_code_row, column=12).value
                closing_price = ws_daily_data.cell(row=security_code_row, column=4).value
                #print(opening_price)
                ws_watch_list.cell(row=row, column=29).value = opening_price
                ws_watch_list.cell(row=row, column=last_col+1).value = closing_price
                #print(ws_watch_list.cell(row=row, column=last_col+1).value)
                #print(ws_watch_list.cell(row=row, column=22).value)
                diff=ws_watch_list.cell(row=row, column=last_col+1).value - ws_watch_list.cell(row=row, column=29).value
                if diff > 0:
                    ws_watch_list.cell(row=row, column=last_col+1).fill = plus_fill
                elif diff < 0:
                    ws_watch_list.cell(row=row,column=last_col+1).fill = minus_fill

                if closing_price is not None and opening_price is not None:
                    ratio = ((closing_price - opening_price) / opening_price)*100
                    ws_watch_list.cell(row=row, column=31).value = ratio
                    ws_watch_list.cell(row=row, column=30).value = closing_price - opening_price
                    if ratio > 2:
                        ws_watch_list.cell(row=row, column=1).value = True
                        for i in range(1, n_columns):
                            ws_watch_list.cell(row=row, column=i).fill =attained_fill
                        profit = closing_price - opening_price
                        print(str(ws_watch_list.cell(row=row,column=4).value)+"_"+str(ws_watch_list.cell(row=row,column=5).value) + "は2%目標を達成しました。利益は"+str(profit)+"円です。")

                    elif ratio < -2:
                        ws_watch_list.cell(row=row, column=1).value = False
                        for i in range(1, n_columns):
                            ws_watch_list.cell(row=row, column=i).fill =unattained_fill
                        loss = ws_watch_list.cell(row=row,column=last_col).value - ws_watch_list.cell(row=row,column=22).value
                        print(str(ws_watch_list.cell(row=row,column=4).value)+"_"+str(ws_watch_list.cell(row=row,column=5).value) + "は目標を達成できず、終了しました。損失は"+str(loss)+"円です。")
    
            basis_date2 = basis_date + datetime.timedelta(days=1)
                # 指定日が平日になるまでループ
            while True:
                # 指定日が土曜日または日曜日の場合
                if basis_date2.weekday() == 5 or basis_date2.weekday() == 6:
                    print(f"{basis_date2}は土曜日または日曜日です。")
                    # 指定日を翌日に上書き
                    basis_date2 = basis_date2 + datetime.timedelta(days=1)
                # 指定日が祝日の場合
                elif jpholiday.is_holiday(basis_date2):
                    print(f"{basis_date2}は{str(jpholiday.holiday_name(basis_date2))}です。")
                    # 指定日を翌日に上書き
                    basis_date2 = basis_date2 + datetime.timedelta(days=1)
                # 指定日が平日の場合
                else:
                    print(f"{basis_date2}は平日です。")
                    # ループを終了
                    break
            print(basis_date2)
            ws_watch_list.cell(row=1,column=last_col+2).value = basis_date2

            wb_daily_data.close()

    
        


    wb_watch_list.save(watch_list_file)
    print(watch_list_file+'を保存しました')
    wb_watch_list.close()
    shutil.move(watch_list_file, dirstorage)

print(t)
t = datetime.datetime.now().time()
print(t)

#稼働終了アナウンス
import winsound
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾