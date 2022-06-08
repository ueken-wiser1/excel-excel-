import os
import openpyxl
import requests
import bs4
import time
import datetime
import sys
import sys
from selenium import webdriver
from selenium.webdriver.common.keys import Keys as keys
import winsound
import pathlib
import glob

#------------お約束開始---冒頭
#稼働時間計測開始
import datetime
t = datetime.datetime.now().time()
#------------お約束終了---冒頭

'''
仕様


'''

#------------プログラム本文---ここから
today = datetime.date.today()
d = today.strftime('%Y%m%d')
dir01 = "C:/Users/touko/OneDrive/自動化用/05.FGO画像/FGO画像/起動tips/"#フォルダパス入力
file_list = glob.glob(dir01 + '*.PNG')
wb = openpyxl.load_workbook('C:/Users/touko/OneDrive/自動化用/05.FGO画像/FGO画像/起動tipsリスト.xlsx')
sheet = wb.active
i = 1
sheet.cell(row=i,column=1).value = "No"#項番
sheet.cell(row=i,column=2).value = "ゲーム内ID"#項番
sheet.cell(row=i,column=3).value = "ファイル名"
sheet.cell(row=i,column=4).value = "サーヴァント名"#項番
sheet.cell(row=i,column=5).value = "クラス"#項番
sheet.cell(row=i,column=6).value = "フルパス"
sheet.cell(row=i,column=7).value = "備考"
sheet.cell(row=i,column=8).value = "最終行"
#print(file_list)
#print(i)
for l in file_list:
    print(i)
    sheet.cell(row=i+1, column=1).value = i
#    print(l)
    filename = os.path.basename(l)
#    print(filename)
    filename_ext = os.path.splitext(filename)[0]
#    print(filename_ext)
    sheet.cell(row=i+1, column=2).value = filename_ext
    file_fullpath = dir01 + filename
#    print(file_fullpath)
    sheet.cell(row=i+1, column=3).value = file_fullpath
    sheet.cell(row=i+1, column=5).value = "=COUNTA(A:A)-1"
    i += 1
wb.save('C:/Users/touko/OneDrive/自動化用/05.FGO画像/FGO画像/起動tipsリスト.xlsx')#リスト保存先
#------------プログラム本文---ここまで

#------------お約束開始---末尾
#稼働時間表示
print(t)
t = datetime.datetime.now().time()
print(t)

#稼働終了アナウンス
import winsound
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾