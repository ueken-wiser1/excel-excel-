import imp
import os
from re import L
from turtle import color
import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd
import numpy as np
import bottleneck as bn
import requests
import bs4
import time
import datetime
import glob
import re
import sys
import winsound


#どんな動きをさせるのか
#1. 銘柄データ格納フォルダを取得
#2. 銘柄データを順番に開く
#3. 指定の列に日次データを計算
#4.     
#5.     
#6.     
#7.     
#8.     
#9.     
#10.    
#11.    
#12.    

#要確認事項
#1. file_

#プログラム

dirmerge = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/"

stock_list = glob.glob(dirmerge + '*.xlsx')

#開始時間取得
t = datetime.datetime.now()
d = datetime.datetime.now()
d1 = d.strftime('%Y%m%d')
#開始時間取得


for l in stock_list:
    wb = openpyxl.load_workbook(l)
    ws = wb.worksheets[0]
    sheetstock = wb.worksheets[0]
    lastrow_stockbook = sheetstock.max_row+1
    if sheetstock.cell(lastrow_stockbook, 4).value == '－':
        sheetstock.cell(lastrow_stockbook,4).value = sheetstock.cell(lastrow_stockbook-1,4).value
        sheetstock.cell(lastrow_stockbook,11).value = 0
        sheetstock.cell(lastrow_stockbook,12).value = sheetstock.cell(lastrow_stockbook-1,4).value
        sheetstock.cell(lastrow_stockbook,13).value = sheetstock.cell(lastrow_stockbook-1,4).value
        sheetstock.cell(lastrow_stockbook,14).value = sheetstock.cell(lastrow_stockbook-1,4).value
        sheetstock.cell(lastrow_stockbook,15).value = sheetstock.cell(lastrow_stockbook-1,4).value
        if sheetstock.cell(lastrow_stockbook,3).value == '株探からのお知らせ':
            ws.delete_rows(lastrow_stockbook)
            print(l+'削除')           

    wb.save(l)
    print(l)

#終了時間取得-経過時間
print(t)
t1 = datetime.datetime.now()
print(t1)
dt = t1-t
print(dt)
#終了時間取得-経過時間

#稼働終了アナウンス
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾
#株価列はD=4列目
#13週(65日)移動平均     HA=209列目
#26週(130日)移動平均    HB=210列目
#移動平均5日線          GQ=199列目
#移動平均10日線         GW=205列目
#移動平均15日線         GX=206列目
#移動平均20日線         GY=207列目
#移動平均25日線         GR=200列目
#移動平均30日線         GZ=208列目
#ボリンジャーバンド 13週、25日
#標準偏差
#+1σ 25日               HD=212列目
#-1σ 25日               HE=213列目
#+2σ 25日               HF=214列目
#-2σ 25日               HG=215列目
#+3σ 25日               HH=216列目
#-3σ 25日               HI=217列目
#+1σ 13週               HJ=218列目
#-1σ 13週               HK=219列目
#+2σ 13週               HL=220列目
#-2σ 13週               HM=221列目
#+3σ 13週               HN=222列目
#-3σ 13週               HO=223列目
#25日ボリンジャーバンドでの株価位置   HP=224列目
#13週ボリンジャーバンドでの株価位置   HQ=225列目
#ローソク足分類         HQ=225列目




