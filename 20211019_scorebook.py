import os
from re import L
import openpyxl
import time
import datetime
import sys
import winsound
import glob
import xlrd
import numpy as np
import pandas as pd

#スコアブックプログラム
#1. シートを3つ用意する
#2. シート3に"株式"フォルダ内の全銘柄の最終行をコピーペーストする
#3. 貼り付けたデータを配列にする
#4. 配列の指定した要素でソートする
#5. 昇順ソートの上位10位をシート1に、降順ソートの上位10位をシート2に転記する
#6. 昇順ソートで価格で条件分岐：500円以下の上位10位をシート1に転記する
#7. 降順ソートで価格で条件分岐：500円以下の上位10位をシート2に転記する
#8. 終わったブックを保存して閉じる

t = datetime.datetime.now().time()
d = datetime.date.today()
#print(d)

dir_stock = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/test/"
stock_list = glob.glob(dir_stock + '*.xlsx')

scorebook = openpyxl.Workbook()
scoresheet_allscore = scorebook.create_sheet(title='allscore')
#allscoreシートへの列名の記載
scoresheet_allscore.cell(row=1,column=1).value = '日付'
scoresheet_allscore.cell(row=1,column=2).value = 'コード'
scoresheet_allscore.cell(row=1,column=3).value = '会社名'
scoresheet_allscore.cell(row=1,column=4).value = '株価'
scoresheet_allscore.cell(row=1,column=5).value = '前日比'
scoresheet_allscore.cell(row=1,column=6).value = '値幅'
scoresheet_allscore.cell(row=1,column=7).value = '売買代金'
scoresheet_allscore.cell(row=1,column=8).value = '約定回数'
scoresheet_allscore.cell(row=1,column=9).value = '決算日'
scoresheet_allscore.cell(row=1,column=10).value = '前日終値'
scoresheet_allscore.cell(row=1,column=11).value = '出来高'
scoresheet_allscore.cell(row=1,column=12).value = '始値'
scoresheet_allscore.cell(row=1,column=13).value = '高値'
scoresheet_allscore.cell(row=1,column=14).value = '安値'
scoresheet_allscore.cell(row=1,column=15).value = '終値'
scoresheet_allscore.cell(row=1,column=16).value = '前日比%'#不使用
scoresheet_allscore.cell(row=1,column=17).value = 'PER'
scoresheet_allscore.cell(row=1,column=18).value = 'PBR'
scoresheet_allscore.cell(row=1,column=19).value = '上場市場'#不使用
scoresheet_allscore.cell(row=1,column=20).value = 'VWAP'
scoresheet_allscore.cell(row=1,column=21).value = '発行済株式数'
scoresheet_allscore.cell(row=1,column=22).value = '最新信用売残'
scoresheet_allscore.cell(row=1,column=23).value = '最新信用買残'
scoresheet_allscore.cell(row=1,column=24).value = '信用倍率'
scoresheet_allscore.cell(row=1,column=25).value = '売買代金2'#計算値
scoresheet_allscore.cell(row=1,column=26).value = '信用売残前週比'
scoresheet_allscore.cell(row=1,column=27).value = '信用買残前週比'
scoresheet_allscore.cell(row=1,column=28).value = '出来高前日比'
scoresheet_allscore.cell(row=1,column=29).value = '約定回数前日比'
scoresheet_allscore.cell(row=1,column=30).value = '時価総額'
scoresheet_allscore.cell(row=1,column=31).value = '浮動株総額'#不使用
scoresheet_allscore.cell(row=1,column=32).value = '取引規模'#不使用
scoresheet_allscore.cell(row=1,column=33).value = '平均約定金額'#不使用
scoresheet_allscore.cell(row=1,column=34).value = '連騰回数'
scoresheet_allscore.cell(row=1,column=51).value = '当日IR有無'
scoresheet_allscore.cell(row=1,column=101).value = '信用取引規制中'
scoresheet_allscore.cell(row=1,column=102).value = '貸借取引銘柄別増担保金徴収措置'
scoresheet_allscore.cell(row=1,column=103).value = '貸借取引銘柄別増担保金徴収措置内容'
scoresheet_allscore.cell(row=1,column=106).value = '空売規制対象'
scoresheet_allscore.cell(row=1,column=111).value = '融資新規'
scoresheet_allscore.cell(row=1,column=112).value = '融資返済'
scoresheet_allscore.cell(row=1,column=113).value = '融資残高'
scoresheet_allscore.cell(row=1,column=114).value = '貸株新規'
scoresheet_allscore.cell(row=1,column=115).value = '貸株返済'
scoresheet_allscore.cell(row=1,column=116).value = '貸株残高'
scoresheet_allscore.cell(row=1,column=117).value = '差引残高'
scoresheet_allscore.cell(row=1,column=118).value = '回転日数'
scoresheet_allscore.cell(row=1,column=121).value = '貸株超過株数'
scoresheet_allscore.cell(row=1,column=122).value = '最高料率'
scoresheet_allscore.cell(row=1,column=123).value = '当日品貸料率'
scoresheet_allscore.cell(row=1,column=124).value = '前日品貸料率'
scoresheet_allscore.cell(row=1,column=151).value = 'みんかぶ目標株価'
scoresheet_allscore.cell(row=1,column=152).value = '現在株価との差'
scoresheet_allscore.cell(row=1,column=201).value = '移動平均線数値5日'
scoresheet_allscore.cell(row=1,column=202).value = '移動平均線数値25日'
scoresheet_allscore.cell(row=1,column=203).value = '移動平均線数値75日'
scoresheet_allscore.cell(row=1,column=204).value = '移動平均乖離率5日'
scoresheet_allscore.cell(row=1,column=205).value = '移動平均乖離率25日'
scoresheet_allscore.cell(row=1,column=206).value = '移動平均乖離率75日'
scoresheet_allscore.cell(row=1,column=251).value = '平均出来高'
scoresheet_allscore.cell(row=1,column=252).value = '平均約定回数'
scoresheet_allscore.cell(row=1,column=253).value = '平均回転日数'
scoresheet_allscore.cell(row=1,column=254).value = '平均移動平均乖離率5日'
scoresheet_allscore.cell(row=1,column=255).value = '平均移動平均乖離率25日'
scoresheet_allscore.cell(row=1,column=256).value = '平均移動平均乖離率75日'
scoresheet_allscore.cell(row=1,column=257).value = '平均信用買残'
scoresheet_allscore.cell(row=1,column=258).value = '平均融資新規'
scoresheet_allscore.cell(row=1,column=259).value = '平均融資返済'
scoresheet_allscore.cell(row=1,column=260).value = '平均信用売残'
scoresheet_allscore.cell(row=1,column=261).value = '平均貸株新規'
scoresheet_allscore.cell(row=1,column=262).value = '平均貸株返済'
scoresheet_allscore.cell(row=1,column=263).value = '平均貸株超過'
scoresheet_allscore.cell(row=1,column=264).value = '平均出来高変化率'
scoresheet_allscore.cell(row=1,column=265).value = '平均約定回数変化率'
scoresheet_allscore.cell(row=1,column=266).value = '平均買残変化率'
scoresheet_allscore.cell(row=1,column=267).value = '平均売残変化率'
scoresheet_allscore.cell(row=1,column=268).value = '平均平均約定金額'
scoresheet_allscore.cell(row=1,column=269).value = '標準偏差出来高'
scoresheet_allscore.cell(row=1,column=270).value = '標準偏差約定回数'
scoresheet_allscore.cell(row=1,column=271).value = '標準偏差回転日数'
scoresheet_allscore.cell(row=1,column=272).value = '標準偏差移動平均乖離率5日'
scoresheet_allscore.cell(row=1,column=273).value = '標準偏差移動平均乖離率25日'
scoresheet_allscore.cell(row=1,column=274).value = '標準偏差移動平均乖離率75日'
scoresheet_allscore.cell(row=1,column=275).value = '標準偏差信用買残'
scoresheet_allscore.cell(row=1,column=276).value = '標準偏差融資新規'
scoresheet_allscore.cell(row=1,column=277).value = '標準偏差融資返済'
scoresheet_allscore.cell(row=1,column=278).value = '標準偏差信用売残'
scoresheet_allscore.cell(row=1,column=279).value = '標準偏差貸株新規'
scoresheet_allscore.cell(row=1,column=280).value = '標準偏差貸株返済'
scoresheet_allscore.cell(row=1,column=281).value = '標準偏差貸株超過'
scoresheet_allscore.cell(row=1,column=282).value = '標準偏差出来高変化率'
scoresheet_allscore.cell(row=1,column=283).value = '標準偏差約定回数変化率'
scoresheet_allscore.cell(row=1,column=284).value = '標準偏差信用買残変化率'
scoresheet_allscore.cell(row=1,column=285).value = '標準偏差信用売残変化率'
scoresheet_allscore.cell(row=1,column=286).value = '標準偏差平均約定金額'
scoresheet_allscore.cell(row=1,column=501).value = '標準化出来高'
scoresheet_allscore.cell(row=1,column=502).value = '標準化約定回数'
scoresheet_allscore.cell(row=1,column=503).value = '標準化回転日数'
scoresheet_allscore.cell(row=1,column=504).value = '標準化移動平均乖離率5日'
scoresheet_allscore.cell(row=1,column=505).value = '標準化移動平均乖離率25日'
scoresheet_allscore.cell(row=1,column=506).value = '標準化移動平均乖離率75日'
scoresheet_allscore.cell(row=1,column=507).value = '標準化信用買残'
scoresheet_allscore.cell(row=1,column=508).value = '標準化融資新規'
scoresheet_allscore.cell(row=1,column=509).value = '標準化融資返済'
scoresheet_allscore.cell(row=1,column=510).value = '標準化信用売残'
scoresheet_allscore.cell(row=1,column=511).value = '標準化貸株新規'
scoresheet_allscore.cell(row=1,column=512).value = '標準化貸株返済'
scoresheet_allscore.cell(row=1,column=513).value = '標準化貸株超過'
scoresheet_allscore.cell(row=1,column=514).value = '標準化出来高変化率'
scoresheet_allscore.cell(row=1,column=515).value = '標準化約定回数変化率'
scoresheet_allscore.cell(row=1,column=516).value = '標準化信用買残変化率'
scoresheet_allscore.cell(row=1,column=517).value = '標準化信用売残変化率'
scoresheet_allscore.cell(row=1,column=518).value = '標準化平均約定金額'
scoresheet_allscore.cell(row=1,column=1000).value = '売り時/買い時スコア'

scoresheet_downsort = scorebook.create_sheet(title='downsort')
scoresheet_upsort = scorebook.create_sheet(title='upsort')

j = 2

for l in stock_list:
    stockbook = openpyxl.load_workbook(l)
    sheet01 = stockbook.worksheets[0]
    lastrow_stockbook = sheet01.max_row
    lastcolumn_stockbook = sheet01.max_column
    
    for i in range(1,lastcolumn_stockbook):
        scoresheet_allscore.cell(row=j,column=i).value = sheet01.cell(row=lastrow_stockbook,column=i).value
        if scoresheet_allscore.cell(row=j,column=i).value is None:
            scoresheet_allscore.cell(row=j,column=i).value = 0
        else:
            if scoresheet_allscore.cell(row=j,column=i).value == '－':
                scoresheet_allscore.cell(row=j,column=i).value = 0
            else:
                pass

    stockbook.close()
    j += 1

today_scorebook = dir_stock+'scorebook/'+str(d)+'score.xlsx'
scorebook.save(today_scorebook)

lastrow_scorebook = scoresheet_allscore.max_row
lastcolumn_scorebook = scoresheet_allscore.max_column
print(today_scorebook)
#todayscore = []
#for i in range(2,lastrow_scorebook):
#    for k in range(2,lastcolumn_scorebook):
#        todayscore.append(scoresheet_allscore.cell(row=i,column=k).value)
df = pd.read_excel(today_scorebook, sheet_name='allscore', index_col=1, engine="openpyxl")

#print(df)

df_s = df.sort_values('値幅', ascending=True)

#print(df_s)
for i in range(1, 3):
#    print(df_s.iat[i,1]) , print(df_s.iat[i,2])
    scoresheet_upsort.cell(row=i+3,column=i+3).value = scoresheet_allscore.cell(row=i,column=1).value
    scoresheet_upsort.cell(row=i+4,column=i+4).value = scoresheet_allscore.cell(row=i,column=2).value

df_s = df.sort_values("値幅", ascending=False)
for i in range(1, 3):
#    print(df_s.iat[i,1]) , print(df_s.iat[i,2])
    scoresheet_downsort.cell(row=i+10,column=i+3).value = scoresheet_allscore.cell(row=i,column=1).value
    scoresheet_downsort.cell(row=i+11,column=i+4).value = scoresheet_allscore.cell(row=i,column=2).value

#df_s = df.sort_values('コード')
#print(df)

scorebook.save(today_scorebook)

print(t)
t = datetime.datetime.now().time()
print(t)

winsound.Beep(500,50)  #ビープ音（500Hzの音を50msec流す）