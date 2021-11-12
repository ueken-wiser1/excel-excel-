import os
import openpyxl
import requests
import bs4
import time
import datetime
import re
import glob
#------------お約束開始---冒頭
#稼働時間計測開始
import datetime
t = datetime.datetime.now().time()
#------------お約束終了---冒頭

'''
仕様
1. 銘柄データ1列目の日付と株探各銘柄ニュースで使っている日付を比較
2. 
'''

#------------プログラム本文---ここから
folder01 = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/完了/"
#銘柄データを順に開く
# filelist01 = glob.glob(folder01+"*.xlsx")
# for l in filelist01:
#     stockbook = openpyxl.load_workbook(l)
#     print(l[47:51])
    
file01 = "1301_極洋.xlsx"
kabutan_newsURL_base = 'https://kabutan.jp/stock/news?code=1301&nmode=0&page=2'
res = requests.get(kabutan_newsURL_base)
res.raise_for_status()
soup = bs4.BeautifulSoup(res.text, 'html.parser')
tdtags = soup.find_all('td')
atags = soup.find_all('a')
num_tdtags = len(tdtags)
num_atags = len(atags)


samplestock = openpyxl.load_workbook(folder01 + file01)
sheet01 = samplestock.worksheets[0]
sampledate = sheet01.cell(row=8,column=1).value
sample = str(sampledate)
slicesample = sample[0:10]

datetags_news = soup.find_all("td",class_="news_time")

# print(datetags_news)

num_datetags_news = len(datetags_news)


for i in range(num_datetags_news):
    sitedate = soup.select(".news_time")[i]
    parent_sitedate = sitedate.parent
    sitedatestr = str(sitedate)
    print(i,num_datetags_news,sitedate)
    # print(sitedatestr)
    # print(sitedatestr[49:51])
    # print(len(sitedatestr))
    # print(parent_sitedate)
    # if "ctg5" in str(parent_sitedate) or "ctg4" in str(parent_sitedate):
    #     print(True)
    # else:
    #     print(False)
    # if slicesample in sitedatestr:
    #     print(True)
    # else:
    #     print(False)
#     sitedatestr =str(sitedate)
# # sitedate = soup.find_all(".news_time",id=datetime)
# # print(sitedate)
#     print(sitedatestr[38:48])
# # for date in sitedate:
# #     print(date.get()[0:10])
#     datestr = str(sampledate)
#     print(datestr[0:10])
#     if sampledate == sitedatestr:
#         print(True)
#     else:
#         print(False)
#------------プログラム本文---ここまで

#------------お約束開始---末尾
#稼働時間表示
print(t)
t = datetime.datetime.now().time()
print(t)

#稼働終了アナウンス
import winsound
# winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
# winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
# winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
# winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
# winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾