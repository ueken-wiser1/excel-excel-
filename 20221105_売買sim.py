#売買sim
#2022/10/28
#銘柄データフォルダに適用
#銘柄データnを開いたら、最終行取得
#2行目から最終行まで繰り返し
#i行目の日付とi+j行目の日付が同じなら、i+j行目を削除

import os
from re import L
import openpyxl
import pandas
import requests
import bs4
import time
import datetime
import glob
import re
import sys
import winsound

#開始時間取得
t = datetime.datetime.now()
d = datetime.datetime.now()
d1 = d.strftime('%Y%m%d')
#開始時間取得


#対象フォルダ指定
dirdaily = "D:/株取引/20221031_日付データ/完了/"
dirmerge = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/test/20221027/銘柄/"
dirsim = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/test/20221027/sim/"

file_list = glob.glob(dirdaily + '*.xlsx')
stock_list = glob.glob(dirmerge + '*.xlsx')
sim_list = glob.glob(dirsim + '*.xlsx')

#simデータのリストをglob関数で作成
for l in sim_list:
    #simデータを順番に開く
    wb_sim = openpyxl.load_workbook(l)
    sheetsim = wb_sim.worksheets[0]
    k=5
    print(l)
    lastrow = sheetsim.max_row+1
    lastcolumn = sheetsim.max_column
    basename = os.path.basename(l)
    daycode_format = basename[:8]
    daycode_format_int = int(basename[:8])
#    print(daycode_format)
    for i in range(2,lastcolumn):
        if sheetsim.cell(1,i).value == '売り' or sheetsim.cell(1,i).value is None:
            print('パスしました')
            #売りフラグが立っているか、cellが空であればパスする
            pass
        elif sheetsim.cell(1,i).value == 'プラス超え':
            #プラス超えフラグであれば、証券コードと最終行日付を取得
            codestock_sim = sheetsim.cell(2,i).value
            namestock_sim = sheetsim.cell(3,i).value
            daycode_sim = sheetsim.cell(lastrow-1,1).value
            print(namestock_sim)
            print(daycode_sim)
#            print(daycode_sim)
#            print(lastrow)
            #日付は'yyyymmdd'フォーマットする-1
            daycode_sim_format = daycode_sim.strftime('%Y%m%d')
#            print(daycode_sim_format)
            #日付データフォルダ内のファイルを順番に開き、ファイル名の最初の8字を取得する-2
            for m in file_list:
                basename_daily = os.path.basename(m)
                daycode_daily_format = basename_daily[:8]
                print(daycode_daily_format)
                #print(m)
            #1の日付を超える2の日付を持つファイルを開く
                if daycode_daily_format > daycode_sim_format:
                    wb_daily = openpyxl.load_workbook(m)
#                    print(m)
                    sheetdaily = wb_daily.worksheets[0]
                    lastrow_daily = sheetdaily.max_row
            #証券コードを参照し、始値を記録する
                    for j in range(2,lastrow_daily):
                        if sheetdaily.cell(j,2).value == codestock_sim:
                            price_begin = sheetdaily.cell(j,12).value
                            sheetsim.cell(lastrow,i).value = price_begin                
            #フラグを買いに変更
                            sheetsim.cell(1,i).value = '買い'
                            print(codestock_sim + '＿' + namestock_sim + 'を' + str(price_begin) + 'で買いました')
                            

            #continue
            pass
        elif sheetsim.cell(1,i).value == '買い' or '保有':
            #買いフラグであれば、証券コードと最終行日付を取得
            codestock_sim = sheetsim.cell(2,i).value
            namestock_sim = sheetsim.cell(3,i).value
            daycode_sim = sheetsim.cell(lastrow,1).value
            #日付は'yyyymmdd'フォーマットする-1
            daycode_sim_format = daycode_sim.strftime('%Y%m%d')
            #日付データフォルダ内のファイルを順番に開き、ファイル名の最初の8字を取得する-2
            for m in file_list:
                basename_daily = os.path.basename(m)
                daycode_daily_format = basename[:8]
            #1の日付を超える2の日付を持つファイルを開く
                if daycode_daily_format > daycode_sim_format:
                    wb_daily = openpyxl.load_workbook(m)
                    sheetdaily = wb_daily.worksheets[0]
                    lastrow_daily = sheetdaily.max_row+1
            #証券コードを参照し、終値を記録する
                    for j in range(2,lastrow_daily):
                        if sheetdaily.cell(j,2).value == codestock_sim:
                            price_end = sheetdaily.cell(j,15).value
                            sheetsim.cell(lastrow,i).value = price_end
            #マイナス超えフラグが無いか、確認する
                        elif sheetdaily.cell(j,229).value == 'マイナス超え':
                            sheetsim.cell(1,i).value = 'マイナス超え'
                        else:
                            sheetsim.cell(1,i).value = '保有'
                            print(codestock_sim + '＿' + namestock_sim + 'を保有中です')
                        break
            #フラグを保有に変更
            #マイナス超えフラグがあれば、フラグをマイナス超えに変更
            #continue
                sheetsim.cell(4,i).value = sheetsim.cell(lastrow,i).value - sheetsim.cell(4,i).value
            pass
        elif sheetsim.cell(1,i).value == 'マイナス超え':
            #マイナス超えフラグであれば、証券コードと最終行日付を取得
            codestock_sim = sheetsim.cell(2,i).value
            namestock_sim = sheetsim.cell(3,i).value
            daycode_sim = sheetsim.cell(lastrow,1).value
            #日付は'yyyymmdd'フォーマットする-1
            daycode_sim_format = daycode_sim.strftime('%Y%m%d')
            #日付データフォルダ内のファイルを順番に開き、ファイル名の最初の8字を取得する-2
            for m in file_list:
                basename_daily = os.path.basename(m)
                daycode_daily_format = basename[:8]
            #1の日付を超える2の日付を持つファイルを開く
                if daycode_daily_format > daycode_sim_format:
                    wb_daily = openpyxl.load_workbook(m)
                    sheetdaily = wb_daily.worksheets[0]
                    lastrow_daily = sheetdaily.max_row+1
            #証券コードを参照し、始値を記録する
                    for j in range(2,lastrow_daily):
                        if sheetdaily.cell(j,2).value == codestock_sim:
                            price_begin = sheetdaily.cell(j,15).value
                            sheetsim.cell(lastrow,i).value = price_begin    
            #フラグを売りに変更
                            sheetsim.cell(1,i).value = '売り'
                            print(codestock_sim + '＿' + namestock_sim + 'を売りました')
                        break
            #continue
                sheetsim.cell(4,i).value = sheetsim.cell(lastrow,i).value - sheetsim.cell(4,i).value
            pass
    rieki = 0
#    for k in range(2,lastcolumn):
#        rieki = int(rieki) + int(sheetsim.cell(4,k).value)
#        sheetsim.cell(1,1).value = rieki
#    print(str(daycode_format) + 'からの利益は' + str(rieki) + '円です')
    wb_sim.save(dirsim+daycode_format+'_'+'buyselsim.xlsx')










#終了時間取得-経過時間
print(t)
t1 = datetime.datetime.now()
print(t1)
dt = t1-t
print(dt)
#終了時間取得-経過時間

#稼働終了アナウンス
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾