import requests
import bs4
import os
import openpyxl
import time
import datetime
import winsound


#どんな動きをさせるのか
#excelを開く
#0000-9999まで証券コードを走査して、該当した銘柄のデータを抽出する
#エラーが出た場合はスキップ
#抽出するデータは全てのタグにある情報

t = datetime.datetime.now().time()

#excelを開く
wb = openpyxl.Workbook()
name = wb.get_sheet_names
sheet = wb.active
kabutan_URL_base = 'http://kabutan.jp/stock/?code='
#0000-9999までの証券コードを分解する

#0-9→'000'+0-9
for i in range(10):
    time.sleep(0.1)
    stock_code = '000' + str(i)


    kabutan_URL = kabutan_URL_base + str(stock_code)
    print(kabutan_URL)

#銘柄ページから情報を読み込む

    res = requests.get(kabutan_URL)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')

#読み込んだ情報をexcelファイルに書き込む
#           コード
    row_number = i + 1

    d_today = datetime.date.today()

    sheet.cell(row=row_number,column=1).value = d_today
    sheet.cell(row=row_number,column=2).value = stock_code
    sheet.cell(row=row_number,column=3).value = str(soup.select('h3')[0].get_text())
    sheet.cell(row=row_number,column=4).value = len(soup.find_all("h2"))
    sheet.cell(row=row_number,column=5).value = len(soup.find_all("span"))
    sheet.cell(row=row_number,column=6).value = len(soup.find_all("dd"))
    sheet.cell(row=row_number,column=7).value = len(soup.find_all("a"))
    sheet.cell(row=row_number,column=8).value = len(soup.find_all("td"))
    sheet.cell(row=row_number,column=9).value = len(soup.find_all("time"))
    sheet.cell(row=row_number,column=10).value = len(soup.find_all("h3"))
    sheet.cell(row=row_number,column=11).value = len(soup.find_all("li"))

for i in range(10, 100):
    time.sleep(0.2)
    stock_code = '00' + str(i)

    kabutan_URL = kabutan_URL_base + str(stock_code)
    print(kabutan_URL)

#銘柄ページから情報を読み込む

    res = requests.get(kabutan_URL)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')


#読み込んだ情報をexcelファイルに書き込む
#           コード
    row_number = i + 1

    d_today = datetime.date.today()

    sheet.cell(row=row_number,column=1).value = d_today
    sheet.cell(row=row_number,column=2).value = stock_code
    sheet.cell(row=row_number,column=3).value = str(soup.select('h3')[0].get_text())
    sheet.cell(row=row_number,column=4).value = len(soup.find_all("h2"))
    sheet.cell(row=row_number,column=5).value = len(soup.find_all("span"))
    sheet.cell(row=row_number,column=6).value = len(soup.find_all("dd"))
    sheet.cell(row=row_number,column=7).value = len(soup.find_all("a"))
    sheet.cell(row=row_number,column=8).value = len(soup.find_all("td"))
    sheet.cell(row=row_number,column=9).value = len(soup.find_all("time"))
    sheet.cell(row=row_number,column=10).value = len(soup.find_all("h3"))
    sheet.cell(row=row_number,column=11).value = len(soup.find_all("li"))

#100-999→'00'+100-999
for i in range(100, 1000):
    time.sleep(0.2)
    stock_code = '0' + str(i)

    kabutan_URL = kabutan_URL_base + str(stock_code)
    print(kabutan_URL)

#銘柄ページから情報を読み込む

    res = requests.get(kabutan_URL)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')

#読み込んだ情報をexcelファイルに書き込む
#           コード
    row_number = i + 1

    d_today = datetime.date.today()

    sheet.cell(row=row_number,column=1).value = d_today
    sheet.cell(row=row_number,column=2).value = stock_code
    sheet.cell(row=row_number,column=3).value = str(soup.select('h3')[0].get_text())
    sheet.cell(row=row_number,column=4).value = len(soup.find_all("h2"))
    sheet.cell(row=row_number,column=5).value = len(soup.find_all("span"))
    sheet.cell(row=row_number,column=6).value = len(soup.find_all("dd"))
    sheet.cell(row=row_number,column=7).value = len(soup.find_all("a"))
    sheet.cell(row=row_number,column=8).value = len(soup.find_all("td"))
    sheet.cell(row=row_number,column=9).value = len(soup.find_all("time"))
    sheet.cell(row=row_number,column=10).value = len(soup.find_all("h3"))
    sheet.cell(row=row_number,column=11).value = len(soup.find_all("li"))

#1000-9999→+1000-9999
for i in range(1000, 10000):
    time.sleep(0.2)
    stock_code = str(i)

    kabutan_URL = kabutan_URL_base + str(stock_code)
    print(kabutan_URL)

#銘柄ページから情報を読み込む

    res = requests.get(kabutan_URL)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')

#読み込んだ情報をexcelファイルに書き込む
#           コード
    row_number = i + 1

    d_today = datetime.date.today()

    sheet.cell(row=row_number,column=1).value = d_today
    sheet.cell(row=row_number,column=2).value = stock_code
    sheet.cell(row=row_number,column=3).value = str(soup.select('h3')[0].get_text())
    sheet.cell(row=row_number,column=4).value = len(soup.find_all("h2"))
    sheet.cell(row=row_number,column=5).value = len(soup.find_all("span"))
    sheet.cell(row=row_number,column=6).value = len(soup.find_all("dd"))
    sheet.cell(row=row_number,column=7).value = len(soup.find_all("a"))
    sheet.cell(row=row_number,column=8).value = len(soup.find_all("td"))
    sheet.cell(row=row_number,column=9).value = len(soup.find_all("time"))
    sheet.cell(row=row_number,column=10).value = len(soup.find_all("h3"))
    sheet.cell(row=row_number,column=11).value = len(soup.find_all("li"))

    wb.save('C:/Users/touko/OneDrive/株価分析/excel/株式データ/スコアブック/allkabutaginfo.xlsx')