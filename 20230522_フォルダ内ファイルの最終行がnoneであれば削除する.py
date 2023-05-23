import os
from openpyxl import load_workbook
import glob

# フォルダ内のExcelファイルを取得
folder_path = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/'
excel_files = glob.glob(os.path.join(folder_path, '*.xlsx'))

# 各Excelファイルの処理
for file_path in excel_files:
    # Excelファイルを開く
    wb = load_workbook(file_path)
    ws = wb.active
    
    # 最終行から順番に処理
    max_row = ws.max_row
    row_index = max_row
    
    while row_index >= 1:
        cell_value = ws.cell(row=row_index, column=1).value
        print(file_path)
        if cell_value is None:
            # 行を削除
            ws.delete_rows(row_index)
            print('最終行1列目は記載ないため、削除しました。')
        else:
            # Noneでなくなったら次のファイルに移る
            print('このファイルは問題なし。次に移ります。')
            break
        
        row_index -= 1
    
    # Excelファイルを保存
    wb.save(file_path)
    wb.close()
