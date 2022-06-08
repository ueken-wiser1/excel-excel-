import requests
import bs4
import os
import openpyxl
import time
import datetime
import winsound


#どんな動きをさせるのか
#excelを開く
#0000-9999まで証券コードを走査して、該当した銘柄のデータを抽出する
#エラーが出た場合はスキップ
#抽出するデータはこれまでと同じ

t = datetime.datetime.now().time()

#excelを開く
wb = openpyxl.Workbook()
#print(type(wb))
name = wb.get_sheet_names
#print(name)
#print(wb.get_sheet_names())
sheet = wb.active
#print(sheet.title)
#print(name)
#0000-9999までの証券コードを分解する

#0-9→'000'+0-9
for code_number in range(10):
    time.sleep(0.1)
#    sheet = wb.get_sheet_by_name('Sheet1')
    stock_code = '000' + str(code_number)
#    print(stock_code)
#    print(sheet.title)

    kabutan_URL_base = 'http://kabutan.jp/stock/?code='
    kabutan_URL = kabutan_URL_base + str(stock_code)
    print(kabutan_URL)

#銘柄ページから情報を読み込む

    res = requests.get(kabutan_URL)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')
#    wb = openpyxl.load_workbook('allkabu1.xlsx')
#    name = wb.get_sheet_names

#読み込んだ情報をexcelファイルに書き込む
#           コード
    row_number = code_number + 2
    write_column = 1

#    print(row_number)
#    print(write_column)
#    print(sheet.cell(row=row_number, column=write_column).value)
#    print(sheet.title)
#    print(name)

    d_today = datetime.date.today()
    sheet.cell(row=row_number, column=write_column).value = d_today
    write_column += 1

    sheet.cell(row=row_number, column=write_column).value = str(stock_code)
    write_column += 1
#    wb.save('allkabu1.xlsx')
#           名称
    stock_name = soup.select('h3')[0]
    print(stock_name)
    sheet.cell(row=row_number, column=write_column).value = str(soup.select('h3')[0].get_text())
    write_column += 1

#           現在株価
    try:
        stock_price = str(soup.select('td')[32])
    except IndexError as e:
#        print('現在株価存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           前日比
    try:
        DoD = str(soup.select('dd')[0])
    except IndexError as e:
#        print('前日比存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = DoD
        write_column += 1

#           売買代金
    try:
        dekidaka_yield = str(soup.select('td')[36].get_text())
    except IndexError as e:
#        print('売買代金存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = dekidaka_yield
        write_column += 1

#           約定回数
    try:
        number_of_contracts = str(soup.select('td')[38].get_text())
    except IndexError as e:
#        print('約定回数存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = number_of_contracts
        write_column += 1

#           決算日
    try:
        stock_price = str(soup.select('dd')[7])
    except IndexError as e:
#        print('前日終値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           前日終値1
    try:
        stock_price = str(soup.select('dd')[7])
    except IndexError as e:
#        print('前日終値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           前日終値2
    try:
        stock_price = str(soup.select('dd')[8].get_text())
    except IndexError as e:
#        print('前日終値存在しない')
        write_column += 2
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 2

#           出来高
    try:
        dekidaka_yield = str(soup.select('td')[35].get_text())
    except IndexError as e:
#        print('出来高存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = dekidaka_yield
        write_column += 1

#           始値
    try:
        stock_price = str(soup.select('td')[23])
    except IndexError as e:
#        print('始値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           高値
    try:
        stock_price = str(soup.select('td')[26])
    except IndexError as e:
#        print('高値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           安値
    try:
        stock_price = str(soup.select('td')[29])
    except IndexError as e:
#        print('安値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           終値
    try:
        stock_price = str(soup.select('td')[32])
    except IndexError as e:
#        print('終値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           前日比%
    try:
        DoD = str(soup.select('span')[16].get_text())
    except IndexError as e:
#        print('前日比存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = DoD
        write_column += 1

#           PER
    try:
        kabu_PER = str(soup.select('td')[18].get_text())
    except IndexError as e:
#        print('PER存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = kabu_PER
        write_column += 1

#           PBR
    try:
        kabu_PBR = str(soup.select('td')[19].get_text())
    except IndexError as e:
#        print('PBR存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = kabu_PBR
        write_column += 1

#           上場市場
    try:
        stock_price = str(soup.select('span')[12].get_text())
    except IndexError as e:
#        print('上場市場存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           VWAP
    try:
        VWAP = str(soup.select('td')[37].get_text())
    except IndexError as e:
#        print('VWAP存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = VWAP
        write_column += 1

#           発行済み株式数
    try:
        number_of_issued_shares = str(soup.select('td')[42].get_text())
    except IndexError as e:
#        print('発行済み株式数存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = number_of_issued_shares
        write_column += 1

#           最新信用売残
    try:
        credit_unsold = str(soup.select('td')[46])
    except IndexError as e:
#        print('最新信用売残存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = credit_unsold
        write_column += 1

#           最新信用買残
    try:
        credit_unpurchased = str(soup.select('td')[47])
    except IndexError as e:
#        print('最新信用買残存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = credit_unpurchased
        write_column += 1

#           信用倍率
    try:
        credit_ratio = str(soup.select('td')[48])
    except IndexError as e:
#        print('信用倍率存在しない')
        write_column += 1
        pass
    else:
        
        sheet.cell(row=row_number, column=write_column).value = credit_ratio
        write_column += 1

#           単元株
    try:
        unit_share = str(soup.select('td')[40])
    except IndexError as e:
        print('単元株存在しない')
        print(unit_share)
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = unit_share
        write_column += 1

#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[82])
#    write_column = write_column + 1
#           決算タブに移動
#    kabutan_URL_base_finance = 'http://kabutan.jp/stock/finance?code='
#    kabutan_URL_finance = kabutan_URL_base_finance + str(stock_code)
#    res = requests.get(kabutan_URL_finance)
#    res.raise_for_status()
#    soup = bs4.BeautifulSoup(res.text, 'html.parser')
#    tag_title = soup.find_all('title')
#           総資産回転率
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[226])
#    pass
#    write_column = write_column + 1
#           1秒待機
wb.save('C:/Users/touko/OneDrive/株価分析/excel/株式データ/銘柄データ集計/stockcodelist.xlsx')
#time.sleep(1)
code_number += 1

#10-99→'00'+10-99
for code_number in range(10, 100):
    time.sleep(0.1)
    stock_code = '00' + str(code_number)

    kabutan_URL_base = 'http://kabutan.jp/stock/?code='
    kabutan_URL = kabutan_URL_base + str(stock_code)
    print(kabutan_URL)

#銘柄ページから情報を読み込む

    res = requests.get(kabutan_URL)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')
#    wb = openpyxl.load_workbook('allkabu1.xlsx')
#    name = wb.get_sheet_names

#読み込んだ情報をexcelファイルに書き込む
#           コード
    row_number = code_number + 1
    write_column = 1

#    print(row_number)
#    print(write_column)
#    print(sheet.cell(row=row_number, column=write_column).value)
#    print(sheet.title)
#    print(name)

    d_today = datetime.date.today()
    sheet.cell(row=row_number, column=write_column).value = d_today
    write_column += 1

    sheet.cell(row=row_number, column=write_column).value = str(stock_code)
    write_column += 1
#    wb.save('allkabu1.xlsx')
#           名称
    stock_name = soup.select('h3')[0]
    print(stock_name)
    sheet.cell(row=row_number, column=write_column).value = str(soup.select('h3')[0].get_text())
    write_column += 1

#           現在株価
    try:
        stock_price = str(soup.select('td')[32])
    except IndexError as e:
#        print('現在株価存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           前日比
    try:
        DoD = str(soup.select('dd')[0])
    except IndexError as e:
#        print('前日比存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = DoD
        write_column += 1

#           売買代金
    try:
        dekidaka_yield = str(soup.select('td')[36].get_text())
    except IndexError as e:
#        print('売買代金存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = dekidaka_yield
        write_column += 1

#           約定回数
    try:
        number_of_contracts = str(soup.select('td')[38].get_text())
    except IndexError as e:
#        print('約定回数存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = number_of_contracts
        write_column += 1

#           決算日
    try:
        stock_price = str(soup.select('dd')[7])
    except IndexError as e:
#        print('前日終値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           前日終値1
    try:
        stock_price = str(soup.select('dd')[7])
    except IndexError as e:
#        print('前日終値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           前日終値2
    try:
        stock_price = str(soup.select('dd')[8].get_text())
    except IndexError as e:
#        print('前日終値存在しない')
        write_column += 2
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 2

#           出来高
    try:
        dekidaka_yield = str(soup.select('td')[35].get_text())
    except IndexError as e:
#        print('出来高存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = dekidaka_yield
        write_column += 1

#           始値
    try:
        stock_price = str(soup.select('td')[23])
    except IndexError as e:
#        print('始値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           高値
    try:
        stock_price = str(soup.select('td')[26])
    except IndexError as e:
#        print('高値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           安値
    try:
        stock_price = str(soup.select('td')[29])
    except IndexError as e:
#        print('安値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           終値
    try:
        stock_price = str(soup.select('td')[32])
    except IndexError as e:
#        print('終値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           前日比%
    try:
        DoD = str(soup.select('span')[16].get_text())
    except IndexError as e:
#        print('前日比存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = DoD
        write_column += 1

#           PER
    try:
        kabu_PER = str(soup.select('td')[18].get_text())
    except IndexError as e:
#        print('PER存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = kabu_PER
        write_column += 1

#           PBR
    try:
        kabu_PBR = str(soup.select('td')[19].get_text())
    except IndexError as e:
#        print('PBR存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = kabu_PBR
        write_column += 1

#           上場市場
    try:
        stock_price = str(soup.select('span')[12].get_text())
    except IndexError as e:
#        print('上場市場存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           VWAP
    try:
        VWAP = str(soup.select('td')[37].get_text())
    except IndexError as e:
#        print('VWAP存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = VWAP
        write_column += 1

#           発行済み株式数
    try:
        number_of_issued_shares = str(soup.select('td')[42].get_text())
    except IndexError as e:
#        print('発行済み株式数存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = number_of_issued_shares
        write_column += 1

#           最新信用売残
    try:
        credit_unsold = str(soup.select('td')[46])
    except IndexError as e:
#        print('最新信用売残存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = credit_unsold
        write_column += 1

#           最新信用買残
    try:
        credit_unpurchased = str(soup.select('td')[47])
    except IndexError as e:
#        print('最新信用買残存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = credit_unpurchased
        write_column += 1

#           信用倍率
    try:
        credit_ratio = str(soup.select('td')[48])
    except IndexError as e:
#        print('信用倍率存在しない')
        write_column += 1
        pass
    else:
        
        sheet.cell(row=row_number, column=write_column).value = credit_ratio
        write_column += 1

#           単元株
    try:
        unit_share = str(soup.select('td')[40])
    except IndexError as e:
        print('単元株存在しない')
        print(unit_share)
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = unit_share
        write_column += 1
    
#           決算タブに移動
#    kabutan_URL_base_finance = 'http://kabutan.jp/stock/finance?code='
#    kabutan_URL_finance = kabutan_URL_base_finance + str(stock_code)
#    res = requests.get(kabutan_URL_finance)
#    res.raise_for_status()
#    soup = bs4.BeautifulSoup(res.text, 'html.parser')
#    tag_title = soup.find_all('title')
#           総資産回転率
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[226])
#    write_column += 1
#           1秒待機
#    time.sleep(1)
    code_number += 1



    wb.save('C:/Users/touko/OneDrive/株価分析/excel/株式データ/銘柄データ集計/stockcodelist.xlsx')

#100-999→'00'+100-999
for code_number in range(100, 1000):
    time.sleep(0.1)
    stock_code = '0' + str(code_number)

    kabutan_URL_base = 'http://kabutan.jp/stock/?code='
    kabutan_URL = kabutan_URL_base + str(stock_code)
    print(kabutan_URL)

#銘柄ページから情報を読み込む

    res = requests.get(kabutan_URL)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')
#    wb = openpyxl.load_workbook('allkabu1.xlsx')
#    name = wb.get_sheet_names

#読み込んだ情報をexcelファイルに書き込む
#           コード
    row_number = code_number + 1
    write_column = 1

#    print(row_number)
#    print(write_column)
#    print(sheet.cell(row=row_number, column=write_column).value)
#    print(sheet.title)
#    print(name)

    d_today = datetime.date.today()
    sheet.cell(row=row_number, column=write_column).value = d_today
    write_column += 1

    sheet.cell(row=row_number, column=write_column).value = str(stock_code)
    write_column += 1
#    wb.save('allkabu1.xlsx')
#           名称
    stock_name = soup.select('h3')[0]
    print(stock_name)
    sheet.cell(row=row_number, column=write_column).value = str(soup.select('h3')[0].get_text())
    write_column += 1

#           現在株価
    try:
        stock_price = str(soup.select('td')[32])
    except IndexError as e:
#        print('現在株価存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           前日比
    try:
        DoD = str(soup.select('dd')[0])
    except IndexError as e:
#        print('前日比存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = DoD
        write_column += 1

#           売買代金
    try:
        dekidaka_yield = str(soup.select('td')[36].get_text())
    except IndexError as e:
#        print('売買代金存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = dekidaka_yield
        write_column += 1

#           約定回数
    try:
        number_of_contracts = str(soup.select('td')[38].get_text())
    except IndexError as e:
#        print('約定回数存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = number_of_contracts
        write_column += 1

#           決算日
    try:
        stock_price = str(soup.select('dd')[7])
    except IndexError as e:
#        print('前日終値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           前日終値1
    try:
        stock_price = str(soup.select('dd')[7])
    except IndexError as e:
#        print('前日終値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           前日終値2
    try:
        stock_price = str(soup.select('dd')[8].get_text())
    except IndexError as e:
#        print('前日終値存在しない')
        write_column += 2
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 2

#           出来高
    try:
        dekidaka_yield = str(soup.select('td')[35].get_text())
    except IndexError as e:
#        print('出来高存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = dekidaka_yield
        write_column += 1

#           始値
    try:
        stock_price = str(soup.select('td')[23])
    except IndexError as e:
#        print('始値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           高値
    try:
        stock_price = str(soup.select('td')[26])
    except IndexError as e:
#        print('高値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           安値
    try:
        stock_price = str(soup.select('td')[29])
    except IndexError as e:
#        print('安値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           終値
    try:
        stock_price = str(soup.select('td')[32])
    except IndexError as e:
#        print('終値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           前日比%
    try:
        DoD = str(soup.select('span')[16].get_text())
    except IndexError as e:
#        print('前日比存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = DoD
        write_column += 1

#           PER
    try:
        kabu_PER = str(soup.select('td')[18].get_text())
    except IndexError as e:
#        print('PER存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = kabu_PER
        write_column += 1

#           PBR
    try:
        kabu_PBR = str(soup.select('td')[19].get_text())
    except IndexError as e:
#        print('PBR存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = kabu_PBR
        write_column += 1

#           上場市場
    try:
        stock_price = str(soup.select('span')[12].get_text())
    except IndexError as e:
#        print('上場市場存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           VWAP
    try:
        VWAP = str(soup.select('td')[37].get_text())
    except IndexError as e:
#        print('VWAP存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = VWAP
        write_column += 1

#           発行済み株式数
    try:
        number_of_issued_shares = str(soup.select('td')[42].get_text())
    except IndexError as e:
#        print('発行済み株式数存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = number_of_issued_shares
        write_column += 1

#           最新信用売残
    try:
        credit_unsold = str(soup.select('td')[46])
    except IndexError as e:
#        print('最新信用売残存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = credit_unsold
        write_column += 1

#           最新信用買残
    try:
        credit_unpurchased = str(soup.select('td')[47])
    except IndexError as e:
#        print('最新信用買残存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = credit_unpurchased
        write_column += 1

#           信用倍率
    try:
        credit_ratio = str(soup.select('td')[48])
    except IndexError as e:
#        print('信用倍率存在しない')
        write_column += 1
        pass
    else:
        
        sheet.cell(row=row_number, column=write_column).value = credit_ratio
        write_column += 1

#           単元株
    try:
        unit_share = str(soup.select('td')[40])
    except IndexError as e:
        print('単元株存在しない')
        print(unit_share)
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = unit_share
        write_column += 1

#    決算タブに移動
#    kabutan_URL_base_finance = 'http://kabutan.jp/stock/finance?code='
#    kabutan_URL_finance = kabutan_URL_base_finance + str(stock_code)
#    res = requests.get(kabutan_URL_finance)
#    res.raise_for_status()
#    soup = bs4.BeautifulSoup(res.text, 'html.parser')
#    tag_title = soup.find_all('title')
#           総資産回転率
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[226])
#    write_column += 1
#           1秒待機
#    time.sleep(1)
    code_number += 1
wb.save('C:/Users/touko/OneDrive/株価分析/excel/株式データ/銘柄データ集計/stockcodelist.xlsx')

#1000-9999→+1000-9999
for code_number in range(1000, 10000):
    time.sleep(0.2)
    stock_code = str(code_number)

    kabutan_URL_base = 'http://kabutan.jp/stock/?code='
    kabutan_URL = kabutan_URL_base + str(stock_code)
    print(kabutan_URL)

#銘柄ページから情報を読み込む

    res = requests.get(kabutan_URL)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')
#    wb = openpyxl.load_workbook('allkabu1.xlsx')
#    name = wb.get_sheet_names

#読み込んだ情報をexcelファイルに書き込む
#           コード
    row_number = code_number + 1
    write_column = 1

#    print(row_number)
#    print(write_column)
#    print(sheet.cell(row=row_number, column=write_column).value)
#    print(sheet.title)
#    print(name)

    d_today = datetime.date.today()
    sheet.cell(row=row_number, column=write_column).value = d_today
    write_column += 1

    sheet.cell(row=row_number, column=write_column).value = str(stock_code)
    write_column += 1
#    wb.save('allkabu1.xlsx')
#           名称
    stock_name = soup.select('h3')[0]
    print(stock_name)
    sheet.cell(row=row_number, column=write_column).value = str(soup.select('h3')[0].get_text())
    write_column += 1

#           現在株価
    try:
        stock_price = str(soup.select('td')[32])
    except IndexError as e:
#        print('現在株価存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           前日比
    try:
        DoD = str(soup.select('dd')[0])
    except IndexError as e:
#        print('前日比存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = DoD
        write_column += 1

#           売買代金
    try:
        dekidaka_yield = str(soup.select('td')[36].get_text())
    except IndexError as e:
#        print('売買代金存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = dekidaka_yield
        write_column += 1

#           約定回数
    try:
        number_of_contracts = str(soup.select('td')[38].get_text())
    except IndexError as e:
#        print('約定回数存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = number_of_contracts
        write_column += 1

#           決算日
    try:
        stock_price = str(soup.select('dd')[7])
    except IndexError as e:
#        print('前日終値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           前日終値1
    try:
        stock_price = str(soup.select('dd')[7])
    except IndexError as e:
#        print('前日終値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           前日終値2
    try:
        stock_price = str(soup.select('dd')[8].get_text())
    except IndexError as e:
#        print('前日終値存在しない')
        write_column += 2
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 2

#           出来高
    try:
        dekidaka_yield = str(soup.select('td')[35].get_text())
    except IndexError as e:
#        print('出来高存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = dekidaka_yield
        write_column += 1

#           始値
    try:
        stock_price = str(soup.select('td')[23])
    except IndexError as e:
#        print('始値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           高値
    try:
        stock_price = str(soup.select('td')[26])
    except IndexError as e:
#        print('高値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           安値
    try:
        stock_price = str(soup.select('td')[29])
    except IndexError as e:
#        print('安値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           終値
    try:
        stock_price = str(soup.select('td')[32])
    except IndexError as e:
#        print('終値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           前日比%
    try:
        DoD = str(soup.select('span')[16].get_text())
    except IndexError as e:
#        print('前日比存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = DoD
        write_column += 1

#           PER
    try:
        kabu_PER = str(soup.select('td')[18].get_text())
    except IndexError as e:
#        print('PER存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = kabu_PER
        write_column += 1

#           PBR
    try:
        kabu_PBR = str(soup.select('td')[19].get_text())
    except IndexError as e:
#        print('PBR存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = kabu_PBR
        write_column += 1

#           上場市場
    try:
        stock_price = str(soup.select('span')[12].get_text())
    except IndexError as e:
#        print('上場市場存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = stock_price
        write_column += 1

#           VWAP
    try:
        VWAP = str(soup.select('td')[37].get_text())
    except IndexError as e:
#        print('VWAP存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = VWAP
        write_column += 1

#           発行済み株式数
    try:
        number_of_issued_shares = str(soup.select('td')[42].get_text())
    except IndexError as e:
#        print('発行済み株式数存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = number_of_issued_shares
        write_column += 1

#           最新信用売残
    try:
        credit_unsold = str(soup.select('td')[46])
    except IndexError as e:
#        print('最新信用売残存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = credit_unsold
        write_column += 1

#           最新信用買残
    try:
        credit_unpurchased = str(soup.select('td')[47])
    except IndexError as e:
#        print('最新信用買残存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = credit_unpurchased
        write_column += 1

#           信用倍率
    try:
        credit_ratio = str(soup.select('td')[48])
    except IndexError as e:
#        print('信用倍率存在しない')
        write_column += 1
        pass
    else:
        
        sheet.cell(row=row_number, column=write_column).value = credit_ratio
        write_column += 1
        
        #    time.sleep(1)
    code_number += 1
#           単元株
    try:
        unit_share = str(soup.select('td')[40])
    except IndexError as e:
        print('単元株存在しない')
        print(unit_share)
        write_column += 1
        pass
    else:
        sheet.cell(row=row_number, column=write_column).value = unit_share
        write_column += 1

wb.save('C:/Users/touko/OneDrive/株価分析/excel/株式データ/銘柄データ集計/stockcodelist_raw.xlsx')

print(t)
t = datetime.datetime.now().time()
print(t)

winsound.Beep(500,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）