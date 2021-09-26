import os
import openpyxl
import requests
import bs4
import time
import datetime

#どんな動きをさせるのか
#excelを開く
#参照excelの各シートに記載された証券コードを読み込む
#シート名は"株式", "マーケット", "為替", "投信"
#読み込んだ証券コードにより株探URLを作成し、銘柄ページへ移動
#必要な情報を取得する

#excelを開く

wb = openpyxl.load_workbook('screeningcodelist02.xlsx')
name = wb.get_sheet_names

#"株式"
sheet = wb.get_sheet_by_name('株式')
print(sheet.title)
print(name)

#読み込んだ証券コードにより株探URLを作成し、銘柄ページへ移動
#必要な情報を取得する

#参照excelの各シートに記載された証券コードを読み込む

sheet = wb.get_sheet_by_name('株式')
for j in range(1, sheet.max_row + 1):
    print(j)
    stock_code = sheet.cell(row=j, column=1).value
    codearea = 10*j - 7
#読み込んだ情報をexcelファイルに書き込む
#書き出す初めのセルの列を指定
    write_column = 3
    print(codearea)
    write_row = codearea

#           証券コード
    print(stock_code)
    sheet.cell(row=codearea, column=3).value = stock_code

#           銘柄名
    stock_name = sheet.cell(row=j, column=2).value
    sheet.cell(row=codearea, column=4).value = stock_name

wb.save('screeningkabu3.xlsx')

