import openpyxl
from openpyxl import Workbook
import os
import requests
import bs4
import time
import datetime
import winsound

#------------お約束開始---冒頭
#稼働時間計測開始
import datetime
t = datetime.datetime.now().time()
#------------お約束終了---冒頭

'''
目的
前場開始直後の値下がり率ランキングの推移から、反発利益が得られそうかの感触を見る。

仕様
株探ページから値下がり率ランキングを取得。
9:20-10:20に実施←これはタスクスケジューラで指定。
値下がり率ランキングページの50件を取得し、コード、銘柄名、株価、前日比、前日比％、出来高を合わせて記録。
日経も併せて記録。株価と前日比、前日比％のみ。
5分後、値下がり率ランキングページの50件を取得し、同様の記録。
前のランキングリストのコードを、今のランキングリストでスキャン。
ヒットすれば、パス。ヒットしなければ、コードから株探銘柄ページを取得し、同様の数字を記録。
前のランキングリストのコードを全てスキャンした今のランキングリストを新たなランキングリストとして、次のランキング取得時の参照とする。
重要な数字は値下がり率だから、順位自体はそこまで重要でない。
全てのリストが出来たら、新しいシートに、9:00段階のランキングを値下がり率と併せて記録する。次の行には出来高を記録する。
そのランキングの上からコードを参照して、次のランキングリストをスキャン。
ヒットすれば、その時の値下がり率、出来高を記録し、次のランキングリストに移る。
9:00段階のランキングリストを全て終わったら、次のランキングリストに移り、コードを上から参照し、前のランキングリストでスキャン。
ヒットすればパス。ヒットしなければ、値下がり率、出来高を記録し、次のランキングリストをスキャン。
これを繰り返し、10:00までのランキングリストのソート完了。

1行目は日付と時間の記載。
使うフォルダは新規に設定。
ファイルは毎日新規に作成。
次のランキングの記録は右に続けていく。
'''

#------------プログラム本文---ここから

#ファイル準備-指定のフォルダにワークブックを作成。シートとして、データを入力する"data"とデータから結果を導く"summary"のシートを作る。
dir = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/値下がりランキング調査/前場/"
wb = Workbook()
sheet01 = wb.create_sheet(title="data")
sheet02 = wb.create_sheet(title="summary")
today = datetime.date.today()
d = today.strftime('%Y%m%d')
wb.save(dir+d+"_値下がりランキング.xlsx")
#記録開始行、列を定義。
i = 2
j = 1
k = 2

#株探ページから、ランキングデータをスクレイピングする。
kabutan_ranking_URL = 'https://kabutan.jp/warning/?mode=2_2'
res = requests.get(kabutan_ranking_URL)
res.raise_for_status()
soup = bs4.BeautifulSoup(res.text, 'html.parser')
rows = soup.find_all('tr')
time.sleep(1)
#ランキングデータの必要なデータ位置を定義。
row_indices = list(range(5,20))
cell_indices = [0, 1, 2, 5, 7, 8, 9]

#繰り返し　9:00ー10:00の13回
n = 1

not_in_now =[]

for n in range(1,13):

    kabutan_ranking_URL = 'https://kabutan.jp/warning/?mode=2_2'
    res = requests.get(kabutan_ranking_URL)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')
    rows = soup.find_all('tr')
    time.sleep(1)
#ランキングデータの必要なデータ位置を定義。

    i = 2
    n_now       = 8*(n-1)+2
    n_before    = 8*(n-2)+2

    #日時、列名の記載。
    t = datetime.datetime.now().time()
    sheet01.cell(1,n_now-1).value = t
    sheet01.cell(1,n_now+0).value = "コード"
    sheet01.cell(1,n_now+1).value = "会社名"
    sheet01.cell(1,n_now+2).value = "市場"
    sheet01.cell(1,n_now+3).value = "株価"
    sheet01.cell(1,n_now+4).value = "値下幅"
    sheet01.cell(1,n_now+5).value = "値下率"
    sheet01.cell(1,n_now+6).value = "出来高"

    #ランキングデータの取得
    for row_index in row_indices:
        if len(rows) > row_index:
            row = rows[row_index]
            cells = row.find_all(['td', 'th'])
            k = 0
            for cell_index in cell_indices:
                if len(cells) > cell_index:

                    cell = cells[cell_index]
                    sheet01.cell(i, n_now+k).value = cell.text
                    
                    #print(row0)
                    #print(column)
                    #print(cell.text)
                    j += 1
                    k += 1
                #print(sheet01.cell(row_index,2).value)
            i += 1
#nが2以上なら、その前のランキングデータがあるから、前のデータを参照して、今のデータになり銘柄についてデータを個別処理する。
#    for k in range(2,15):
#        print(sheet01.cell(k, n_now).value)
#    print(n)
    #print(sheet01.cell(2,10).value)
    if n >= 2:
        not_in_now =[]
        lastrow = sheet01.max_row
        #print(lastrow)
        for l in range(2, lastrow+1):

            code_before = sheet01.cell(l,n_before).value
            code_now    = sheet01.cell(l,n_now).value
            #print(code_before)
            #print(code_now)
            #print(n_now)
            values_before    = [sheet01.cell(row=i, column=n_before).value for i in range(2, lastrow + 1)]
            values_now       = [sheet01.cell(row=i, column=n_now).value for i in range(2, lastrow + 1)]
            #print(values_before)
            #print(values_now)
            #print(n)
            #print(sheet01.cell(2,10).value)

        for i, value_before in enumerate(values_before, start=2):
            #values_nowにvalues_beforeの要素がない場合の処理を開始。
            if value_before not in values_now:
                not_in_now.append(value_before)
            a = 1
            lastrow01 = sheet01.max_row
            print(not_in_now)
            for item in not_in_now:
                #株探ページをスクレイピングして欲しい情報を取り出す。
                kabutan_URL_base = 'http://kabutan.jp/stock/?code='
                kabutan_URL = kabutan_URL_base + str(item)
                res = requests.get(kabutan_URL)
                res.raise_for_status()
                soup = bs4.BeautifulSoup(res.text, 'html.parser')
                name_elem = soup.select('span')[8].text
                market_elem = soup.select('span')[11].text
                price_elem = soup.select('td')[32].text
                down_elem = soup.select('span')[14].text
                downper_elem = soup.select('span')[15].text
                amount = soup.select('td')[35].text
                #取得したデータをexcelに書き込む。
                sheet01.cell(lastrow01+a, n_now).value = item
                sheet01.cell(lastrow01+a, n_now+1).value = name_elem
                sheet01.cell(lastrow01+a, n_now+2).value =market_elem
                sheet01.cell(lastrow01+a, n_now+3).value =price_elem
                sheet01.cell(lastrow01+a, n_now+4).value =down_elem
                sheet01.cell(lastrow01+a, n_now+5).value =downper_elem
                sheet01.cell(lastrow01+a, n_now+6).value =amount
                #print(item)
                #リストの次の項目に移る場合、加算行を+1する。
                a+=1
                time.sleep(1)
        else:
            pass

    #print("nは1でしたよ")


                    
    print(n)
    n += 1
    time.sleep(60)
wb.save(dir+d+"_値下がりランキング.xlsx")

#データの集約プロセス
#まずは各ランキングをソートするところからか。
#最初のランキングリストからコードを取得し、次のランキングリストからそのコードを探す。
#該当があれば、その値下幅と値下率、出来高をコピペする。これを最後のランキングリストまで続ける。
#最初のランキングリストが終わったら、次のランキングリストで、今のランキングコードを前のランキングリストから探して、あればパス、なければ次のランキングリストから値下幅と値下率、出来高をコピペする。
#上を繰り返す。
#どう表示するか。
#銘柄横並びで、縦に値下幅、値下率、出来高を横並びで記録する。
# データが格納されている範囲（15行8列 × 12個）を指定
lastrow_aggre = sheet01.max_row
start_row = 2
end_row = lastrow_aggre
start_col = 1
end_col = 8 * 12

# パラメータに対応するリストを抽出
target_list = []
for col in range(start_col, end_col + 1, 8):
    column_values = [sheet01.cell(row=row, column=col).value for row in range(start_row, end_row + 1)]
    target_list.append(column_values)

# パラメータに対応するリストを昇順に並べる
sorted_list = sorted(target_list, key=lambda x: x[0])  # ここでは1行目の値で昇順に並べる例

# 結果を出力
for i, values in enumerate(sorted_list):
    for j, value in enumerate(values):
        sheet02.cell(row=start_row + j, column=start_col + i * 8).value = value


#------------プログラム本文---ここまで

#------------お約束開始---末尾
#稼働時間表示
print(t)
t = datetime.datetime.now().time()
print(t)

#稼働終了アナウンス
import winsound
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾