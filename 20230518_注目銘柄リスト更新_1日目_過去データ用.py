'''#コメントここから
これは追跡開始初日データ処理を実施するコードで、過去のデータ用になります。
対象フォルダ
作業ファイルフォルダ：'C:/Users/touko/OneDrive/株価分析/excel/株式データ/追跡調査/過去/0日目/'
参照ファイルフォルダ：'C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/'
格納ファイルフォルダ：'C:/Users/touko/OneDrive/株価分析/excel/株式データ/追跡調査/過去/1日目/'

実施する処理
OSCIファイルについて、最終列の日付を取得する。
OSCIファイルの2列目の行の証券コードをキーにして、銘柄データフォルダを検索。
ヒットした銘柄データファイルを開く。
OSCIファイルの最終列の日付をキーとして、銘柄データをスキャンし、ヒットした行から、それぞれデータを取得し、OSCIファイルに転記。
始値、終値、終値-始値の値により、網掛けの色を変える、利益%が±2%を越えたら、TRUE/FALSE記載。
10行分、OSCIファイルの最終列が43に達したら、ファイルを格納ファイルフォルダに移す。
次のOSCIファイルに移る。
'''#コメントここまで

import os
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import datetime
import jpholiday

import shutil


t = datetime.datetime.now().time()
d = datetime.datetime.now()
d1 = d.strftime('%Y%m%d')

def find_date_row(ws, date):
    for row2 in range(2, ws.max_row + 1):
        #print(code)
        if ws.cell(row=row2, column=1).value == date:
            return row2
    return None

watch_list_folder = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/追跡調査/過去/1日目/'
stock_folder = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/'
dirstorage = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/追跡調査/過去/2日目以降/'

n_columns = 43

plus_fill = PatternFill(patternType='solid', fgColor='ee82ee') #前日より値上がり
minus_fill = PatternFill(patternType='solid', fgColor='00bfff') #前日より値下がり
attained_fill = PatternFill(patternType='solid', fgColor='adff2f') #+2%目標を達成したら
unattained_fill = PatternFill(patternType='solid', fgColor='696969') #+2%目標未達で、5日目を迎えたら

file_list = glob.glob(watch_list_folder+'*.xlsx')

for l in file_list:
    wb_watch_list = load_workbook(l)
    ws_watch_list = wb_watch_list.active
    print(l+"に対して作業中。")
    last_row = ws_watch_list.max_row
    last_col = ws_watch_list.max_column

    if last_col > 42:
        print(l +"は調査期間を満了しました。")
        continue

    last_date = ws_watch_list.cell(row=1, column=last_col).value

    for i in range(2, last_row + 1):
        
        stock_code = ws_watch_list.cell(row=i, column=4).value
        company_name= ws_watch_list.cell(row=i, column=5).value

        
        stock_data_file = os.path.join(stock_folder, f'{stock_code}_{company_name}.xlsx')
        print("証券コード："+str(stock_code)+"を検索中")

        if os.path.exists(stock_data_file):
            wb_stock_data = load_workbook(stock_data_file)
            ws_stock_data = wb_stock_data.active

            date_row = find_date_row(ws_stock_data, last_date)


            if date_row:
                closing_price = ws_stock_data.cell(row=date_row, column=4).value
                start_price = ws_stock_data.cell(row=date_row, column=12).value
                
                ws_watch_list.cell(row=i, column=last_col).value = closing_price
                ws_watch_list.cell(row=i, column=29).value = start_price
                print("証券コード"+str(stock_code) +"は、終値"+str(ws_watch_list.cell(row=i, column=last_col).value)+"を示しました")


                diff=closing_price - start_price
                if diff > 0:
                    ws_watch_list.cell(row=i, column=last_col).fill = plus_fill
                elif diff < 0:
                    ws_watch_list.cell(row=i,column=last_col).fill = minus_fill

                if closing_price is not None and start_price is not None:
                    ratio = ((closing_price - start_price) / start_price)*100
                    ws_watch_list.cell(row=i, column=31).value = ratio
                    ws_watch_list.cell(row=i, column=30).value = diff
                    if ratio > 2:
                        ws_watch_list.cell(row=i, column=1).value = True
                        for j in range(1, last_col+1):
                            ws_watch_list.cell(row=i, column=j).fill =attained_fill

                        profit = ws_watch_list.cell(row=i,column=last_col).value - ws_watch_list.cell(row=i,column=29).value
                        print(str(ws_watch_list.cell(row=i,column=4).value)+"_"+str(ws_watch_list.cell(row=i,column=5).value) + "は2%目標を達成しました。利益は"+str(profit)+"円です。")

                    elif ratio < -2:
                        ws_watch_list.cell(row=i, column=1).value = False
                        for j in range(1, last_col+1):
                            ws_watch_list.cell(row=i, column=j).fill =unattained_fill
                        loss = ws_watch_list.cell(row=i,column=last_col).value - ws_watch_list.cell(row=i,column=29).value
                        print(str(ws_watch_list.cell(row=i,column=4).value)+"_"+str(ws_watch_list.cell(row=i,column=5).value) + "は-2%目標に達してしまいました。損失は"+str(loss)+"円です。")


            last_col_second=ws_watch_list.max_column

            last_date2 = last_date + datetime.timedelta(days=1)

            while True:
                if last_date2.weekday() == 5 or last_date2.weekday() == 6:

                    last_date2 = last_date2 + datetime.timedelta(days=1)

                elif jpholiday.is_holiday(last_date2):

                    last_date2 = last_date2 + datetime.timedelta(days=1)

                else:

                    break


            print(str(last_date)+"の次の営業日は"+str(last_date2)+"です。")
            ws_watch_list.cell(row=1,column=last_col+1).value = last_date2

            wb_stock_data.close()
    last_col_last = ws_watch_list.max_column
    for i in range(last_col_last+1, 1, -1):

        if ws_watch_list.cell(row=1, column=i).value is None:
            ws_watch_list.delete_cols(i)
            print(str(i)+"列目は1行目に何もないため削除しました。")

    ws_watch_list.cell(row=1,column=last_col+1).value = last_date2
    last_col_last = ws_watch_list.max_column
    wb_watch_list.save(l)

    last_col_last = ws_watch_list.max_column


    print(l+'を保存しました。')

    wb_watch_list.close()
    shutil.move(l, dirstorage)
    
print(t)
t = datetime.datetime.now().time()
print(t)

#稼働終了アナウンス
import winsound
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾