#重複日付の削除
#2022/10/26
#銘柄データフォルダに適用
#銘柄データnを開いたら、最終行取得
#2行目から最終行まで繰り返し
#i行目の日付とi+j行目の日付が同じなら、i+j行目を削除

import os
from re import L
import openpyxl
import pandas
import requests
import bs4
import time
import datetime
import glob
import re
import sys
import winsound

#開始時間取得
t = datetime.datetime.now().time()
d = datetime.datetime.now()
d1 = d.strftime('%Y%m%d')
#開始時間取得

#対象フォルダ指定
dirmerge = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/"

stock_list = glob.glob(dirmerge + '*.xlsx')

for l in stock_list:
    wb_stock = openpyxl.load_workbook(l)
    print(l)
    sheetstock = wb_stock.worksheets[0]
    lastlaw = sheetstock.max_row+1

    for i in range(lastlaw, 2, -1):
        daycode = sheetstock.cell(i,1).value
        for j in range(i+1, lastlaw):
            daycode_same = sheetstock.cell(j,1).value
            if daycode == daycode_same:
                print(daycode_same)
                sheetstock.delete_rows(j-1)
            else:
                pass

    wb_stock.save(l)

#終了時間取得-経過時間
print(t)
t1 = datetime.datetime.now().time()
print(t1)
dt = t1-t
print(dt)
#終了時間取得-経過時間

#稼働終了アナウンス
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾