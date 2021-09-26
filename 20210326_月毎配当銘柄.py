import os
import openpyxl
import requests
import bs4
import time
import datetime

#どんな動きをさせるのか
#excelを開く
#参照excelの各シートに記載された証券コードを読み込む
#シート名は"株式", "マーケット", "為替", "投信"
#読み込んだ証券コードにより株探URLを作成し、銘柄ページへ移動
#必要な情報を取得する

#excelを開く

wb = openpyxl.load_workbook('stockcodelist.xlsx')
name = wb.get_sheet_names

#"マーケット"
sheet = wb.get_sheet_by_name('マーケット')
print(sheet.title)
print(name)

#読み込んだ証券コードにより株探URLを作成し、銘柄ページへ移動
#必要な情報を取得する

#日付を入れる
d_today = datetime.date.today()
sheet.cell(row=1, column=1).value = d_today

#参照excelの各シートに記載された証券コードを読み込む

sheet = wb.get_sheet_by_name('マーケット')
print(sheet)
for j in range(3, sheet.max_row):
    time.sleep(0.2)
    kabutan_URL_base = 'http://kabutan.jp/stock/?code='
    stock_code = sheet.cell(row=j, column=1).value
    kabutan_URL = kabutan_URL_base + str(stock_code)
    res = requests.get(kabutan_URL)
    print(kabutan_URL)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')

#読み込んだ情報をexcelファイルに書き込む
#書き出す初めのセルの列を指定
    write_column = 2

#           名称
    stock_name = soup.select('title')[0]
    print(stock_name)
    sheet.cell(row=j, column=write_column).value = str(soup.select('title')[0])
    write_column += 1

#           現在株価
    try:
        span_pick = str(soup.select('span')[14])
    except IndexError as e:
        print('現在株価存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = str(soup.select('span')[14])
    write_column += 1

#           前日比
    try:
        span_pick = str(soup.select('span')[15])
    except IndexError as e:
        print('前日比存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = str(soup.select('span')[15])
    write_column += 1
    
#           業種
    try:
        span_pick = str(soup.select('td')[104])
    except IndexError as e:
        print('業種存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = str(soup.select('td')[104])
    write_column += 1
    
#           PER
    try:
        span_pick = str(soup.select('td')[18])
    except IndexError as e:
        print('PER存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           PBR
    try:
        span_pick = str(soup.select('td')[19])
    except IndexError as e:
        print('PBR存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           利回り
    try:
        span_pick = str(soup.select('td')[20])
    except IndexError as e:
        print('利回り存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           出来高
    try:
        span_pick = str(soup.select('td')[35])
    except IndexError as e:
        print('出来高存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           約定回数
    try:
        span_pick = str(soup.select('td')[38])
    except IndexError as e:
        print('約定回数存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           単元株
    try:
        span_pick = str(soup.select('td')[40])
    except IndexError as e:
        print('単元株存在しない')
        print(span_pick)
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           時価総額
    try:
        span_pick = str(soup.select('td')[41])
    except IndexError as e:
        print('時価総額存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           発行済み株式数
    try:
        span_pick = str(soup.select('td')[42])
    except IndexError as e:
        print('発行済み株式数存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           最新信用売残
    try:
        span_pick = str(soup.select('td')[46])
    except IndexError as e:
        print('最新信用売残存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           最新信用買残
    try:
        span_pick = str(soup.select('td')[47])
    except IndexError as e:
        print('最新信用買残存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

wb.save('allkabu1.xlsx')

#"為替"
sheet = wb.get_sheet_by_name('為替')
print(sheet.title)
print(name)

#読み込んだ証券コードにより株探URLを作成し、銘柄ページへ移動
#必要な情報を取得する

#日付を入れる
d_today = datetime.date.today()
sheet.cell(row=1, column=1).value = d_today

#参照excelの各シートに記載された証券コードを読み込む

sheet = wb.get_sheet_by_name('為替')
print(sheet)
for j in range(3, sheet.max_row):
    time.sleep(0.2)
    kabutan_URL_base = 'http://kabutan.jp/stock/?code='
    stock_code = sheet.cell(row=j, column=1).value
    kabutan_URL = kabutan_URL_base + str(stock_code)
    res = requests.get(kabutan_URL)
    print(kabutan_URL)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')

#読み込んだ情報をexcelファイルに書き込む
#書き出す初めのセルの列を指定
    write_column = 2

#           名称
    stock_name = soup.select('title')[0]
    print(stock_name)
    sheet.cell(row=j, column=write_column).value = str(soup.select('title')[0])
    write_column += 1

#           現在株価
    try:
        span_pick = str(soup.select('span')[14])
    except IndexError as e:
        print('現在株価存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = str(soup.select('span')[14])
    write_column += 1

#           前日比
    try:
        span_pick = str(soup.select('span')[15])
    except IndexError as e:
        print('前日比存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = str(soup.select('span')[15])
    write_column += 1
    
#           業種
    try:
        span_pick = str(soup.select('td')[104])
    except IndexError as e:
        print('業種存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = str(soup.select('td')[104])
    write_column += 1
    
#           PER
    try:
        span_pick = str(soup.select('td')[18])
    except IndexError as e:
        print('PER存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           PBR
    try:
        span_pick = str(soup.select('td')[19])
    except IndexError as e:
        print('PBR存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           利回り
    try:
        span_pick = str(soup.select('td')[20])
    except IndexError as e:
        print('利回り存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           出来高
    try:
        span_pick = str(soup.select('td')[35])
    except IndexError as e:
        print('出来高存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           約定回数
    try:
        span_pick = str(soup.select('td')[38])
    except IndexError as e:
        print('約定回数存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           単元株
    try:
        span_pick = str(soup.select('td')[40])
    except IndexError as e:
        print('単元株存在しない')
        print(span_pick)
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           時価総額
    try:
        span_pick = str(soup.select('td')[41])
    except IndexError as e:
        print('時価総額存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           発行済み株式数
    try:
        span_pick = str(soup.select('td')[42])
    except IndexError as e:
        print('発行済み株式数存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           最新信用売残
    try:
        span_pick = str(soup.select('td')[46])
    except IndexError as e:
        print('最新信用売残存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           最新信用買残
    try:
        span_pick = str(soup.select('td')[47])
    except IndexError as e:
        print('最新信用買残存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           1株配当
    try:
        span_pick = str(soup.select('td')[94])
    except IndexError as e:
        print('1株配当存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick

wb.save('allkabu1.xlsx')

#"投信"
sheet = wb.get_sheet_by_name('投信')
print(sheet.title)
print(name)

#読み込んだ証券コードにより株探URLを作成し、銘柄ページへ移動
#必要な情報を取得する

#日付を入れる
d_today = datetime.date.today()
sheet.cell(row=1, column=1).value = d_today

#参照excelの各シートに記載された証券コードを読み込む

sheet = wb.get_sheet_by_name('投信')
print(sheet)
for j in range(3, sheet.max_row):
    time.sleep(0.2)
    kabutan_URL_base = 'http://kabutan.jp/stock/?code='
    stock_code = sheet.cell(row=j, column=1).value
    kabutan_URL = kabutan_URL_base + str(stock_code)
    res = requests.get(kabutan_URL)
    print(kabutan_URL)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')

#読み込んだ情報をexcelファイルに書き込む
#書き出す初めのセルの列を指定
    write_column = 2

#           名称
    stock_name = soup.select('title')[0]
    print(stock_name)
    sheet.cell(row=j, column=write_column).value = str(soup.select('title')[0])
    write_column += 1

#           現在株価
    try:
        span_pick = str(soup.select('span')[14])
    except IndexError as e:
        print('現在株価存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = str(soup.select('span')[14])
    write_column += 1

#           前日比
    try:
        span_pick = str(soup.select('span')[15])
    except IndexError as e:
        print('前日比存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = str(soup.select('span')[15])
    write_column += 1
    
#           業種
    try:
        span_pick = str(soup.select('td')[104])
    except IndexError as e:
        print('業種存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = str(soup.select('td')[104])
    write_column += 1
    
#           PER
    try:
        span_pick = str(soup.select('td')[18])
    except IndexError as e:
        print('PER存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           PBR
    try:
        span_pick = str(soup.select('td')[19])
    except IndexError as e:
        print('PBR存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           利回り
    try:
        span_pick = str(soup.select('td')[20])
    except IndexError as e:
        print('利回り存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           出来高
    try:
        span_pick = str(soup.select('td')[35])
    except IndexError as e:
        print('出来高存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           約定回数
    try:
        span_pick = str(soup.select('td')[38])
    except IndexError as e:
        print('約定回数存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           単元株
    try:
        span_pick = str(soup.select('td')[40])
    except IndexError as e:
        print('単元株存在しない')
        print(span_pick)
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           時価総額
    try:
        span_pick = str(soup.select('td')[41])
    except IndexError as e:
        print('時価総額存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           発行済み株式数
    try:
        span_pick = str(soup.select('td')[42])
    except IndexError as e:
        print('発行済み株式数存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           最新信用売残
    try:
        span_pick = str(soup.select('td')[46])
    except IndexError as e:
        print('最新信用売残存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           最新信用買残
    try:
        span_pick = str(soup.select('td')[47])
    except IndexError as e:
        print('最新信用買残存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           1株配当
    try:
        span_pick = str(soup.select('td')[94])
    except IndexError as e:
        print('1株配当存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick

wb.save('allkabu1.xlsx')

#"株式"
sheet = wb.get_sheet_by_name('株式')
print(sheet.title)
print(name)

#読み込んだ証券コードにより株探URLを作成し、銘柄ページへ移動
#必要な情報を取得する

#日付を入れる
d_today = datetime.date.today()
sheet.cell(row=1, column=1).value = d_today

#参照excelの各シートに記載された証券コードを読み込む

sheet = wb.get_sheet_by_name('株式')
print(sheet)
for j in range(3, sheet.max_row):
    time.sleep(0.2)
    kabutan_URL_base = 'http://kabutan.jp/stock/?code='
    stock_code = sheet.cell(row=j, column=1).value
    kabutan_URL = kabutan_URL_base + str(stock_code)
    res = requests.get(kabutan_URL)
    print(kabutan_URL)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')

#読み込んだ情報をexcelファイルに書き込む
#書き出す初めのセルの列を指定
    write_column = 2

#           名称
    stock_name = soup.select('title')[0]
    print(stock_name)
    sheet.cell(row=j, column=write_column).value = str(soup.select('title')[0])
    write_column += 1

#           現在株価
    try:
        span_pick = str(soup.select('span')[14])
    except IndexError as e:
        print('現在株価存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = str(soup.select('span')[14])
    write_column += 1

#           前日比
    try:
        span_pick = str(soup.select('span')[15])
    except IndexError as e:
        print('前日比存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = str(soup.select('span')[15])
    write_column += 1
    
#           業種
    try:
        span_pick = str(soup.select('td')[104])
    except IndexError as e:
        print('業種存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = str(soup.select('a')[104])
    write_column += 1
    
#           PER
    try:
        span_pick = str(soup.select('td')[18])
    except IndexError as e:
        print('PER存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           PBR
    try:
        span_pick = str(soup.select('td')[19])
    except IndexError as e:
        print('PBR存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           利回り
    try:
        span_pick = str(soup.select('td')[20])
    except IndexError as e:
        print('利回り存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           出来高
    try:
        span_pick = str(soup.select('td')[35])
    except IndexError as e:
        print('出来高存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           約定回数
    try:
        span_pick = str(soup.select('td')[38])
    except IndexError as e:
        print('約定回数存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           単元株
    try:
        span_pick = str(soup.select('td')[40])
    except IndexError as e:
        print('単元株存在しない')
        print(span_pick)
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           時価総額
    try:
        span_pick = str(soup.select('td')[41])
    except IndexError as e:
        print('時価総額存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           発行済み株式数
    try:
        span_pick = str(soup.select('td')[42])
    except IndexError as e:
        print('発行済み株式数存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           最新信用売残
    try:
        span_pick = str(soup.select('td')[46])
    except IndexError as e:
        print('最新信用売残存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           最新信用買残
    try:
        span_pick = str(soup.select('td')[47])
    except IndexError as e:
        print('最新信用買残存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick
    write_column += 1

#           1株配当
    try:
        span_pick = str(soup.select('td')[94])
    except IndexError as e:
        print('1株配当存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = span_pick

wb.save('allkabu1.xlsx')

