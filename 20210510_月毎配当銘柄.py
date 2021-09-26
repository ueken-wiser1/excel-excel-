import os
import openpyxl
import requests
import bs4
import time
import datetime

#どんな動きをさせるのか
#excelを開く
#参照excelの各シートに記載された証券コードを読み込む
#シート名は"株式", "マーケット", "為替", "投信"
#読み込んだ証券コードにより株探URLを作成し、銘柄ページへ移動
#必要な情報を取得する

#excelを開く

wb = openpyxl.load_workbook('stockcodelist.xlsx')
name = wb.get_sheet_names

#"マーケット"
sheet = wb.get_sheet_by_name('マーケット')
print(sheet.title)
print(name)

#読み込んだ証券コードにより株探URLを作成し、銘柄ページへ移動
#必要な情報を取得する

#日付を入れる
d_today = datetime.date.today()
sheet.cell(row=1, column=1).value = d_today

#参照excelの各シートに記載された証券コードを読み込む

sheet = wb.get_sheet_by_name('マーケット')
print(sheet)
for j in range(3, sheet.max_row + 1):
    time.sleep(0.2)
    kabutan_URL_base = 'http://kabutan.jp/stock/?code='
    stock_code = sheet.cell(row=j, column=1).value
    kabutan_URL = kabutan_URL_base + str(stock_code)
    res = requests.get(kabutan_URL)
    print(kabutan_URL)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')

#読み込んだ情報をexcelファイルに書き込む
#書き出す初めのセルの列を指定
    write_column = 2

#           名称
    stock_name = soup.select('h3')[0]
    print(stock_name)
    sheet.cell(row=j, column=write_column).value = str(soup.select('h3')[0])
    write_column += 1

#           前日終値
    try:
        stock_price = str(soup.select('dd')[7])
    except IndexError as e:
        print('前日終値存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
    write_column += 1

#           現在株価
    try:
        stock_price = str(soup.select('span')[14])
    except IndexError as e:
        print('現在株価存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
    write_column += 1

#           前日比
    try:
        DoD = str(soup.select('span')[15])
    except IndexError as e:
        print('前日比存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = DoD
    write_column += 1
    
#           出来高
    try:
        dekidaka_yield = str(soup.select('td')[35])
    except IndexError as e:
        print('出来高存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = dekidaka_yield
    write_column += 1

wb.save('allkabu1.xlsx')

'''
#"為替"
sheet = wb.get_sheet_by_name('為替')
print(sheet.title)
print(name)

#読み込んだ証券コードにより株探URLを作成し、銘柄ページへ移動
#必要な情報を取得する

#日付を入れる
d_today = datetime.date.today()
sheet.cell(row=1, column=1).value = d_today

#参照excelの各シートに記載された証券コードを読み込む

sheet = wb.get_sheet_by_name('為替')
print(sheet)
for j in range(3, sheet.max_row + 1):
    time.sleep(0.2)
    kabutan_URL_base = 'http://kabutan.jp/stock/?code='
    stock_code = sheet.cell(row=j, column=1).value
    kabutan_URL = kabutan_URL_base + str(stock_code)
    res = requests.get(kabutan_URL)
    print(kabutan_URL)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')

#読み込んだ情報をexcelファイルに書き込む
#書き出す初めのセルの列を指定
    write_column = 2

#           名称
    stock_name = soup.select('h3')[0]
    print(stock_name)
    sheet.cell(row=j, column=write_column).value = str(soup.select('h3')[0])
    write_column += 1

#           現在株価
    try:
        stock_price = str(soup.select('span')[14])
    except IndexError as e:
        print('現在株価存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
    write_column += 1

#           前日比
    try:
        DoD = str(soup.select('span')[15])
    except IndexError as e:
        print('前日比存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = DoD
    write_column += 1
    
wb.save('allkabu1.xlsx')

#"投信"
sheet = wb.get_sheet_by_name('投信')
print(sheet.title)
print(name)

#読み込んだ証券コードにより株探URLを作成し、銘柄ページへ移動
#必要な情報を取得する

#日付を入れる
d_today = datetime.date.today()
sheet.cell(row=1, column=1).value = d_today

#参照excelの各シートに記載された証券コードを読み込む

sheet = wb.get_sheet_by_name('投信')
print(sheet)
for j in range(3, sheet.max_row + 1):
    time.sleep(0.2)
    kabutan_URL_base = 'http://kabutan.jp/stock/?code='
    stock_code = sheet.cell(row=j, column=1).value
    kabutan_URL = kabutan_URL_base + str(stock_code)
    res = requests.get(kabutan_URL)
    print(kabutan_URL)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')

#読み込んだ情報をexcelファイルに書き込む
#書き出す初めのセルの列を指定
    write_column = 2

#           名称
    stock_name = soup.select('h3')[0]
    print(stock_name)
    sheet.cell(row=j, column=write_column).value = str(soup.select('h3')[0])
    write_column += 1

#           前日終値
    try:
        stock_price = str(soup.select('dd')[7])
    except IndexError as e:
        print('前日終値存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
    write_column += 1

#           始値
    try:
        stock_price = str(soup.select('td')[23])
    except IndexError as e:
        print('始値存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
    write_column += 1

#           始値時間
    try:
        stock_price = str(soup.select('td')[25])
    except IndexError as e:
        print('始値時間存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
    write_column += 1

#           高値
    try:
        stock_price = str(soup.select('td')[26])
    except IndexError as e:
        print('高値存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
    write_column += 1

#           高値時間
    try:
        stock_price = str(soup.select('td')[28])
    except IndexError as e:
        print('高値時間存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
    write_column += 1

#           安値
    try:
        stock_price = str(soup.select('td')[29])
    except IndexError as e:
        print('安値存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
    write_column += 1

#           安値時間
    try:
        stock_price = str(soup.select('td')[31])
    except IndexError as e:
        print('安値時間存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
    write_column += 1

#           終値
    try:
        stock_price = str(soup.select('td')[32])
    except IndexError as e:
        print('終値存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
    write_column += 1

#           終値時間
    try:
        stock_price = str(soup.select('td')[34])
    except IndexError as e:
        print('終値時間存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
    write_column += 1

#           現在株価
    try:
        stock_price = str(soup.select('span')[14])
    except IndexError as e:
        print('現在株価存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
    write_column += 1

#           前日比
    try:
        DoD = str(soup.select('span')[15])
    except IndexError as e:
        print('前日比存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = DoD
    write_column += 1
    
#           業種
    try:
        industry_type = str(soup.select('a')[30])
    except IndexError as e:
        print('業種存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = industry_type
    write_column += 1
    
#           出来高
    try:
        dekidaka_yield = str(soup.select('td')[35])
    except IndexError as e:
        print('出来高存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = dekidaka_yield
    write_column += 1

#           VWAP
    try:
        VWAP = str(soup.select('td')[37])
    except IndexError as e:
        print('VWAP存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = VWAP
    write_column += 1

#           約定回数
    try:
        number_of_contracts = str(soup.select('td')[38])
    except IndexError as e:
        print('約定回数存在しない')
        print(number_of_contracts)
        continue
    else:
        sheet.cell(row=j, column=write_column).value = number_of_contracts
    write_column += 1

#           単元株
    try:
        unit_share = str(soup.select('td')[40])
    except IndexError as e:
        print('単元株存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = unit_share
    write_column += 1

#           時価総額
    try:
        market_capitalization = str(soup.select('td')[41])
    except IndexError as e:
        print('時価総額存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = market_capitalization
    write_column += 1

#           発行済み株式数
    try:
        number_of_issued_shares = str(soup.select('td')[42])
    except IndexError as e:
        print('発行済み株式数存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = number_of_issued_shares
    write_column += 1

#           最新信用売残
    try:
        credit_unsold = str(soup.select('td')[43])
    except IndexError as e:
        print('最新信用売残存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = credit_unsold
    write_column += 1

#           最新信用買残
    try:
        credit_unpurchased = str(soup.select('td')[44])
    except IndexError as e:
        print('最新信用買残存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = credit_unpurchased

#           信用倍率
    try:
        credit_ratio = str(soup.select('td')[45])
    except IndexError as e:
        print('信用倍率存在しない')
        continue
    else:
        sheet.cell(row=j, column=write_column).value = credit_ratio
    write_column += 1

wb.save('allkabu1.xlsx')
'''

#"株式"
sheet = wb.get_sheet_by_name('株式')
print(sheet.title)
print(name)

#読み込んだ証券コードにより株探URLを作成し、銘柄ページへ移動
#必要な情報を取得する

#日付を入れる
d_today = datetime.date.today()
sheet.cell(row=1, column=1).value = d_today

#参照excelの各シートに記載された証券コードを読み込む

sheet = wb.get_sheet_by_name('株式')
print(sheet)
for j in range(3, sheet.max_row + 1):
    time.sleep(0.2)
    kabutan_URL_base = 'http://kabutan.jp/stock/?code='
    stock_code = sheet.cell(row=j, column=1).value
    kabutan_URL = kabutan_URL_base + str(stock_code)
    res = requests.get(kabutan_URL)
    print(kabutan_URL)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')

#読み込んだ情報をexcelファイルに書き込む
#書き出す初めのセルの列を指定
    write_column = 2

#           名称
    stock_name = soup.select('h3')[0]
    print(stock_name)
    sheet.cell(row=j, column=write_column).value = str(soup.select('h3')[0])
    write_column += 1

#           現在株価
    try:
        stock_price = str(soup.select('td')[32])
    except IndexError as e:
        print('現在株価存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
    write_column += 1

#           前日比
    try:
        DoD = str(soup.select('span')[15])
    except IndexError as e:
        print('前日比存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = DoD
    write_column += 1

#           前日比%
    try:
        DoD = str(soup.select('span')[16])
    except IndexError as e:
        print('前日比存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = DoD
    write_column += 1

#           PER
    try:
        kabu_PER = str(soup.select('td')[18])
    except IndexError as e:
        print('PER存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = kabu_PER
    write_column += 1

#           PBR
    try:
        kabu_PBR = str(soup.select('td')[19])
    except IndexError as e:
        print('PBR存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = kabu_PBR
    write_column += 1

#           出来高
    try:
        dekidaka_yield = str(soup.select('td')[35])
    except IndexError as e:
        print('出来高存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = dekidaka_yield
    write_column += 1

#           約定回数
    try:
        number_of_contracts = str(soup.select('td')[38])
    except IndexError as e:
        print('約定回数存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = number_of_contracts
    write_column += 15

#           始値
    try:
        stock_price = str(soup.select('td')[23])
    except IndexError as e:
        print('始値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
    write_column += 1

#           高値
    try:
        stock_price = str(soup.select('td')[26])
    except IndexError as e:
        print('高値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
    write_column += 1

#           安値
    try:
        stock_price = str(soup.select('td')[29])
    except IndexError as e:
        print('安値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
    write_column += 1

#           終値
    try:
        stock_price = str(soup.select('td')[32])
    except IndexError as e:
        print('終値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
    write_column += 1

#           上場市場
    try:
        stock_price = str(soup.select('span')[12])
    except IndexError as e:
        print('上場市場存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
    write_column += 1

#           前日終値
    try:
        stock_price = str(soup.select('dd')[8])
    except IndexError as e:
        print('前日終値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
    write_column += 1
    
#           利回り
    try:
        rimawari_yield = str(soup.select('td')[20])
    except IndexError as e:
        print('利回り存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = rimawari_yield
    write_column += 1

#           VWAP
    try:
        VWAP = str(soup.select('td')[37])
    except IndexError as e:
        print('VWAP存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = VWAP
    write_column += 1

#           単元株
    try:
        unit_share = str(soup.select('td')[40])
    except IndexError as e:
        print('単元株存在しない')
        print(unit_share)
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = unit_share
    write_column += 1

#           時価総額
    try:
        market_capitalization = str(soup.select('td')[41])
    except IndexError as e:
        print('時価総額存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = market_capitalization
    write_column += 1

#           発行済み株式数
    try:
        number_of_issued_shares = str(soup.select('td')[42])
    except IndexError as e:
        print('発行済み株式数存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = number_of_issued_shares
    write_column += 1

#           最新信用売残
    try:
        credit_unsold = str(soup.select('td')[46])
    except IndexError as e:
        print('最新信用売残存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = credit_unsold
    write_column += 1

#           最新信用買残
    try:
        credit_unpurchased = str(soup.select('td')[47])
    except IndexError as e:
        print('最新信用買残存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = credit_unpurchased
    write_column += 1

#           信用倍率
    try:
        credit_ratio = str(soup.select('td')[48])
    except IndexError as e:
        print('信用倍率存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = credit_ratio
    write_column += 1

#           1株配当
    try:
        dividend_share = str(soup.select('td')[94])
    except IndexError as e:
        print('1株配当存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = dividend_share

wb.save('allkabu1.xlsx')

