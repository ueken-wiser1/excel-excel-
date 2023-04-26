import imp
import os
from re import L
from turtle import color
import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd
import numpy as np
import bottleneck as bn
import requests
import bs4
import time
import datetime
import glob
import re
import sys
import winsound

dirmerge = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/"


stock_list = glob.glob(dirmerge + '*.xlsx')

#開始時間取得
t = datetime.datetime.now()
d = datetime.datetime.now()
d1 = d.strftime('%Y%m%d')
#開始時間取得

for l in stock_list:

        wb = openpyxl.load_workbook(l)
        sheetstock = wb.worksheets[0]
        lastrow_stockbook = sheetstock.max_row+1
        lastcolumn_stockbook = sheetstock.max_column
        if sheetstock.cell(2,5).value is None or sheetstock.cell(2,5).value == '－':
                sheetstock.cell(2,5).value = 0
                print('処理')
        else:
                pass

        for i in range(3,lastrow_stockbook):
                if sheetstock.cell(i,5).value is None or sheetstock.cell(i,5).value == '－':
                        sheetstock.cell(i,5).value = sheetstock.cell(i,4).value-sheetstock.cell(i-1,4).value
                        print('前日比処理')
                else:
                        pass
        


        wb.save(l)
        print(l)

#終了時間取得-経過時間
print(t)
t1 = datetime.datetime.now()
print(t1)
dt = t1-t
print(dt)
#終了時間取得-経過時間

#稼働終了アナウンス
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾