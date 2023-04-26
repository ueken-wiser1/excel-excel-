#売買sim
#2022/10/28
#銘柄データフォルダに適用
#銘柄データnを開いたら、最終行取得
#2行目から最終行まで繰り返し
#i行目の日付とi+j行目の日付が同じなら、i+j行目を削除

import os
from re import L
import openpyxl
from openpyxl.styles import PatternFill
import pandas
import requests
import bs4
import time
import datetime
import glob
import re
import sys
import winsound

#開始時間取得
t = datetime.datetime.now()
d = datetime.datetime.now()
d1 = d.strftime('%Y%m%d')
#開始時間取得


#対象フォルダ指定
dirdaily = "D:/株取引/20221031_日付データ/完了/"
dirmerge = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/"
dirsim = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/test/20221027/sim/"

file_list = glob.glob(dirdaily + '*.xlsx')
stock_list = glob.glob(dirmerge + '*.xlsx')
sim_list = glob.glob(dirsim + '*.xlsx')

#simデータのリストをglob関数で作成
for l in sim_list:
    #simデータを順番に開く
    #print(l)
    print(l)
    wb_sim = openpyxl.load_workbook(l)
    sheetsim = wb_sim.worksheets[0]
    lastrow_sim = sheetsim.max_row+1
    lastcolumn_sim = sheetsim.max_column
    basename = os.path.basename(l)
    daycode_format = basename[:8]
    daycode_format_int = int(basename[:8])
#    print(daycode_format)
    #print(l)
    for i in range(2,lastcolumn_sim+1):
        if type(sheetsim.cell(1,i).value)==str:
            #print(type(sheetsim.cell(i,i).value))
            sheetsim.cell(1,i).value = int(sheetsim.cell(1,i).value)
            #print(type(sheetsim.cell(1,i).value))
        else:
            pass
        #print(sheetsim.cell(1,i).value)
        k=5
        if sheetsim.cell(1,i).value == '5' or sheetsim.cell(1,i).value is None:
            print('パスしました')
            #売りフラグが立っているか、cellが空であればパスする
            pass
        elif sheetsim.cell(1,i).value == 1:
            #プラス超えフラグであれば、証券コードと最終行日付を取得
            codestock_sim = sheetsim.cell(2,i).value
            namestock_sim = sheetsim.cell(3,i).value
            print(codestock_sim+'_'+namestock_sim)
            daycode_sim = sheetsim.cell(lastrow_sim-1,1).value
#            print(codestock_sim + "_" + namestock_sim)
#            print(daycode_sim)
#            print(daycode_sim)
#            print(lastrow_sim)
            #日付は'yyyymmdd'フォーマットする-1
            #print(daycode_sim)
            print(lastrow_sim)
            daycode_sim_format = daycode_sim.strftime('%Y%m%d')
#            print(daycode_sim_format)
            #日付データフォルダ内のファイルを順番に開き、ファイル名の最初の8字を取得する-2
            stock_file = glob.glob(dirmerge + codestock_sim +'*.xlsx')
#            print(dirmerge)
#            print(codestock_sim)
#            print(stock_file)
            for m in stock_file:
                wb_stock = openpyxl.load_workbook(m)
                sheetstock = wb_stock.worksheets[0]
                lastrow_stock = sheetstock.max_row+1
#                print(lastrow_stock)
                for j in range(2,lastrow_stock):
#                    print(j)
#                    print(daycode_sim)
#                    print(sheetstock.cell(j,1).value)
                    if daycode_sim == sheetstock.cell(j,1).value:
#                        print(j)
                        daycode_stock = sheetstock.cell(j+1,1).value
#                        print(daycode_stock)
                        row_stock = j+1
#                        print(row_stock)
#                print(sheetsim.cell(1,i).value)
    
                while int(sheetsim.cell(1,i).value) < 5:

#                    print(i)
                    if int(sheetsim.cell(1,i).value) == 1:
                        sheetsim.cell(k,1).value = sheetstock.cell(row_stock, 1).value
                        sheetsim.cell(k,i).value = sheetstock.cell(row_stock, 12).value
                        if sheetstock.cell(row_stock,229).value == 4:
#                            print(row_stock)
                            sheetsim.cell(1,i).value = 4
                        else: 
                            print(sheetstock)
                            sheetsim.cell(1,i).value = 2
                        print(codestock_sim +'_' + namestock_sim +'は'+ str(sheetsim.cell(k,1).value) +'に買いました。')
                        row_stock += 1
                        k += 1
                    elif sheetsim.cell(1,i).value == 2:
                        sheetsim.cell(k,1).value = sheetstock.cell(row_stock, 1).value
                        sheetsim.cell(k,i).value = sheetstock.cell(row_stock, 15).value
                        if sheetstock.cell(row_stock,229).value == 4:
#                            print(row_stock)
                            sheetsim.cell(1,i).value = 4
                        else:
                            print(sheetstock)
                            sheetsim.cell(1,i).value = 3
                        print(codestock_sim +'_' + namestock_sim +'は'+str(sheetsim.cell(k,1).value)+'保有継続です。')
                        if sheetsim.cell(k,i).value <= sheetsim.cell(5,i).value:
                            fill = PatternFill(patternType="solid", fgColor="696969")
                            sheetsim.cell(row=k, column=i).fill = fill
                        row_stock += 1
#                        print(j)
                        k += 1
                    elif sheetsim.cell(1,i).value == 3:
                        sheetsim.cell(k,1).value = sheetstock.cell(row_stock, 1).value
                        sheetsim.cell(k,i).value = sheetstock.cell(row_stock, 15).value
#                        print(sheetstock.cell(row_stock,229).value)
                        if sheetstock.cell(row_stock,229).value == 4:
                            print(row_stock)
                            sheetsim.cell(1,i).value = 4
                        print(codestock_sim +'_' + namestock_sim +'は'+str(sheetsim.cell(k,1).value)+'現在保有中です。')
                        if sheetsim.cell(k,i).value <= sheetsim.cell(5,i).value:
                            fill = PatternFill(patternType="solid", fgColor="696969")
                            sheetsim.cell(row=k, column=i).fill = fill
                        row_stock += 1
#                        print(row_stock)
                        k += 1
                    elif sheetsim.cell(1,i).value == 4:
                        sheetsim.cell(k,1).value = sheetstock.cell(row_stock, 1).value
                        sheetsim.cell(k,i).value = sheetstock.cell(row_stock, 15).value
                        sheetsim.cell(k+1,1).value = sheetstock.cell(row_stock+1, 1).value
                        sheetsim.cell(k+1,i).value = sheetstock.cell(row_stock+1, 12).value
                        sheetsim.cell(1,i).value = 5
                        print(codestock_sim +'_' + namestock_sim +'は'+str(sheetsim.cell(k+1,1).value)+'に売りに出しました。')
                        if sheetsim.cell(k,i).value <= sheetsim.cell(5,i).value:
                            fill = PatternFill(patternType="solid", fgColor="696969")
                            sheetsim.cell(row=k, column=i).fill = fill
                        if sheetsim.cell(k+1,i).value <= sheetsim.cell(5,i).value:
                            fill = PatternFill(patternType="solid", fgColor="696969")
                            sheetsim.cell(row=k+1, column=i).fill = fill
                        row_stock += 1
                        k += 1
        lastrow_sim_2 = sheetsim.max_row
#        print(lastrow_sim_2)
        for h in reversed(range(2,lastrow_sim_2+1)):
            lastrow_sim_2 = sheetsim.max_row
            if sheetsim.cell(5,i).value is None:
                break
            elif sheetsim.cell(h,i).value is not None and sheetsim.cell(5,i).value is not None:
    #                print(h)
                print(sheetsim.cell(h,i).value)
                print(sheetsim.cell(5,i).value)
                sheetsim.cell(4,i).value = 100*(int(sheetsim.cell(h,i).value) - int(sheetsim.cell(5,i).value))
                print(codestock_sim + '_' + namestock_sim + 'の利益は' + sheetsim.cell(h,1).value.strftime("%Y/%m/%d") + 'に確定し、利益は' + str(sheetsim.cell(4,i).value) + '円です。')
                wb_sim.save(dirsim+daycode_format+'_'+'buyselsim.xlsx')
                break
            else:
                continue

            



    #print(l)        
    sourieki = 0
    for g in range(2, lastcolumn_sim+1):
        if sheetsim.cell(4,g).value is None:
            continue
        else:
            rieki = sheetsim.cell(4, g).value
            #print(sheetsim.cell(4,g).value)
            sourieki += rieki
   
    sheetsim.cell(2,1).value = sourieki
    if sourieki >= 0:
        fill = PatternFill(patternType="solid", fgColor="FF69B4")
        sheetsim.cell(row=2, column=1).fill = fill
    else:
        fill = PatternFill(patternType="solid", fgColor="00BFFF")
        sheetsim.cell(row=2, column=1).fill = fill

    print(str(daycode_format) +'から始めた売買の総利益は'+str(sourieki)+'円です。')

#    for k in range(2,lastcolumn_sim):
#        rieki = int(rieki) + int(sheetsim.cell(4,k).value)
#        sheetsim.cell(1,1).value = rieki
#    print(str(daycode_format) + 'からの利益は' + str(rieki) + '円です')
#    print('ここには来る')
    wb_sim.save(dirsim+daycode_format+'_'+'buyselsim.xlsx')










#終了時間取得-経過時間
print(t)
t1 = datetime.datetime.now()
print(t1)
dt = t1-t
print(dt)
#終了時間取得-経過時間

#稼働終了アナウンス
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾