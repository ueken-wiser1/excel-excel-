"""
・ver01のフロー
1. フォルダ1のファイルをリスト化。
2. リストのN番目のブックを開く。
    ブックの最終行を取得。
3. ブックNの日付列を下から(新しい順)取り込む。
    日付を取得　→　文字列化　→　スライス...①
4. 証券コードを取得。
    文字列化
5. 上場会社株探ニュースページのM番目のページを開く。
    スクレイピング
6. td class=news_timeのタグ内容と個数を取得。
7. L番目のnews_timeタグを取得。
    文字列化　→　スライス...②
8. ①と②が一致したら、親タグ内のカテゴリに応じてフラグチェック。
    ②からスライスした時間で15:00以降なら、翌日分にチェック。
・ver02のフロー
1. フォルダ1のファイルをリスト化。
2. リストのN番目のブックを開く。
    ブックの最終行を取得。
3. ブックNの日付列を下から(新しい順)取り込む。
    日付を取得　→　文字列化　→　スライス...①
4. 証券コードを取得。
    文字列化
5. 上場会社株探ニュースページのM番目のページを開く。
    スクレイピング
6. td class=news_timeのタグ内容と個数を取得。
7. L番目のnews_timeタグを取得。
    文字列化　→　スライス...②
8. ①と②が一致したら、親タグ内のカテゴリに応じてフラグチェック。
    ②からスライスした時間で15:00以降なら、翌日分にチェック。
    Q. ①の日付に対して、後ろから近づくのか、前から近づくのか。
    A. 後ろ-後の日付から近づく。
    ∴ ①の日付より前の日付を検知したら、ブック内の次の日付に移るようにしたい。
       検知したタグ番号を記録して、次の日付は記録したタグ番号からスタートさせたい。
9. プログラムを通したら、このプログラム自体のチェックを完了したフラグを立てたい。
    基準は日付。

廃棄→ver01にcolumn=999のフラグ付けとフラグのある日付はスキップするコードを実装で完了とする。
"""
import os
import openpyxl
import requests
import bs4
import time
import datetime
import re
import glob


#------------お約束開始---冒頭
#稼働時間計測開始
import datetime
t = datetime.datetime.now().time()
#------------お約束終了---冒頭
#------------プログラム本文---ここから
folder01 = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/完了/test/"
kabutan_newsURL_base = 'https://kabutan.jp/stock/news?code='
kabutan_newsURL_tale = "&nmode=0&page="

#銘柄データを順に開く
filelist01 = glob.glob(folder01+"*.xlsx")

#関数定義　→　使ってない
def is_include_listed_word(text, word_list):
    for listed_word in word_list:
        if listed_word in text:
            return True
    return False
    
#各銘柄データを開く
for l in filelist01:
    t1 = datetime.datetime.now().time()
    print("銘柄データの取り掛かり開始時間",t1)
    stockbook = openpyxl.load_workbook(l)
    sheet01 = stockbook.worksheets[0]
    print("銘柄データ",l)
    lastrow_stockbook =sheet01.max_row
    slicedate_site = datetime.date.today()
    # print(slicedate_site)

#各銘柄データの各日付列を取り込む
    for i in reversed(range(2,lastrow_stockbook+1)):
        date_stockbook = sheet01.cell(row=i,column=1).value
        date_stockbookstr = str(date_stockbook)
        slicedate_stockbookstr = date_stockbookstr[0:10]
        print("銘柄データから取得した日付の必要な個所切り抜き",slicedate_stockbookstr)
        print("銘柄データから取得した日付の生データ",date_stockbook)
        date_stockbook01 = date_stockbook.date()
        # print(type(date_stockbook), date_stockbook01)
        # print(type(slicedate_site), slicedate_site)
        # print(l)
        # stockcode = str(l[47:51])
        # print(stockcode)
        stockcode = str(l[52:56])
        # print(stockcode)
#対象の銘柄データの証券コードを開く
#対象銘柄の株探ニュースページの各ページ送り
        for j in range(1,4):
            if date_stockbook01 > slicedate_site:
                print("銘柄データの行番号i=",i,"株探データのニュースページ数j=",j,"ここで帰る")
                print(date_stockbook01, ">" ,slicedate_site)
                print("この先に求める日付はない")
                break
            else:
                # print(stockcode)
                time.sleep(0.1)
                res = requests.get(kabutan_newsURL_base+stockcode+kabutan_newsURL_tale+str(j))
                res.raise_for_status()
                soup = bs4.BeautifulSoup(res.text, 'html.parser')
                datetags = soup.find_all("td",class_="news_time")
                num_datetags = len(datetags)
#対象ニュースページの記載日付を取り込む
            for k in range(num_datetags):
                if date_stockbook01 > slicedate_site:
                    print("銘柄データの行番号i=",i,"株探データのニュースページ数j=",j,"news_timeタグ番号k=",k,"ここで帰る")
                    print(date_stockbook01, ">" ,slicedate_site)
                    break
                else:
                    # print(num_datetags)
                    slicedate_site01 = soup.select(".news_time")[k]
                    # print(slicedate_site01)
                    date_sitestr = str(slicedate_site01)
                    slicedate_sitestr = date_sitestr[38:48]
                    # print(slicedate_sitestr)
                    slicedate_sitetime = datetime.datetime.strptime(slicedate_sitestr, '%Y-%m-%d')
                    # print(slicedate_sitetime)
                    slicedate_site = datetime.date(slicedate_sitetime.year,slicedate_sitetime.month,slicedate_sitetime.day)
                    # print(type(slicedate_site),slicedate_site)
                    # print(type(date_stockbook01),date_stockbook01)
                    # slicedate_site = datetime.date(slicedate_sitestr,'%Y-%m-%d')
                    parent_sitedate = slicedate_site01.parent
#銘柄データで取り込んだ日付とニュースページで取り込んだ日付を突き合わせる
                    if date_stockbook01 > slicedate_site:
                        sheet01.cell(row=i,column=999).value = 1
                        break
                    elif date_stockbook == slicedate_site:
                        print(parent_sitedate)
#ニュースカテゴリーで分類分け
#ニュース掲載時間で分類分け－当日or翌日
                        if "ctg_kaiji" in str(parent_sitedate) \
                            or "ctg2" in str(parent_sitedate):
                            if int(date_sitestr[49:51]) >=15:
                                sheet01.cell(row=i+1,column=51).value = 1
                                print("翌日IRフラグをつけた")
                            else:
                                sheet01.cell(row=i,column=51).value = 1
                                print("当日IRフラグをつけた")
                        elif "ctg3_kk" in str(parent_sitedate):
                            if int(date_sitestr[49:51]) >=15:
                                sheet01.cell(row=i+1,column=52).value = 1
                                print("翌日決算IRフラグをつけた")
                            else:
                                sheet01.cell(row=i,column=52).value = 1
                                print("当日決算IRフラグをつけた")
                        elif "ctg3_ks" in str(parent_sitedate):
                            if int(date_sitestr[49:51]) >=15:
                                sheet01.cell(row=i+1,column=53).value = 1
                                print("翌日修正IRフラグをつけた")
                            else:
                                sheet01.cell(row=i,column=53).value = 1
                                print("当日修正IRフラグをつけた")
                        elif "ctg4" in str(parent_sitedate):
                            if int(date_sitestr[49:51]) >=15:
                                sheet01.cell(row=i+1,column=54).value = 1
                                print("翌日テクIRフラグをつけた")
                            else:
                                sheet01.cell(row=i,column=54).value = 1
                                print("当日テクIRフラグをつけた")
                        elif "ctg12" in str(parent_sitedate):
                            if int(date_sitestr[49:51]) >=15:
                                sheet01.cell(row=i+1,column=55).value = 1
                                print("翌日5%IRフラグをつけた")
                            else:
                                sheet01.cell(row=i,column=55).value = 1
                                print("当日5%IRフラグをつけた")
                        else:
                            pass
        sheet01.cell(row=i-1,column=999).value = 1
        print("フラグをつけた")
    stockbook.save(l)





#------------プログラム本文---ここまで

#------------お約束開始---末尾
#稼働時間表示
print(t)
t = datetime.datetime.now().time()
print(t)

#稼働終了アナウンス
import winsound
# winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
# winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
# winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
# winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
# winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾