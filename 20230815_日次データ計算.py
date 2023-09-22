import imp
import os
from re import L
from turtle import color
import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd
import numpy as np
import bottleneck as bn
import re
import datetime
import glob

import winsound


#どんな動きをさせるのか
#1. 銘柄データ格納フォルダを取得
#2. 銘柄データを順番に開く
#3. 指定の列に日次データを計算
#4.     
#5.     
#6.     
#7.     
#8.     
#9.     
#10.    
#11.    
#12.    

#要確認事項
#1. file_

#プログラム

#testのため、仮フォルダ設定
dirmerge = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/"


stock_list = glob.glob(dirmerge + '*.xlsx')

#開始時間取得
t = datetime.datetime.now()
d = datetime.datetime.now()
d1 = d.strftime('%Y%m%d')
#開始時間取得

#testのため、exceptリストは使わなくておｋ
#except_list = ['C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/0000', 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/0010', '0011', '0012', '0090', '0091', '0092', '0093', '0094', '0095', '0101', '0102', '0108', '0800', '0802']

for l in stock_list:
        if "None.xlsx" in l:
                continue #ファイル名に"None.xlsx"が含まれていた場合はスキップ

        wb = openpyxl.load_workbook(l)
        sheetstock = wb.worksheets[0]
        lastrow_stockbook = sheetstock.max_row+1
        lastcolumn_stockbook = sheetstock.max_column
        print("最終行は"+str(lastrow_stockbook))
        #列名命名
        '''
        sheetstock.cell(1,1).value ='日付'
        sheetstock.cell(1,2).value ='コード'
        sheetstock.cell(1,3).value ='会社名'
        sheetstock.cell(1,4).value ='株価'
        sheetstock.cell(1,5).value ='前日比'
        sheetstock.cell(1,6).value ='値幅'
        sheetstock.cell(1,7).value ='売買代金'
        sheetstock.cell(1,8).value ='約定回数'
        sheetstock.cell(1,9).value ='決算日'
        sheetstock.cell(1,10).value ='前日終値'
        sheetstock.cell(1,11).value ='出来高'
        sheetstock.cell(1,12).value ='始値'
        sheetstock.cell(1,13).value ='高値'
        sheetstock.cell(1,14).value ='安値'
        sheetstock.cell(1,15).value ='終値'
        sheetstock.cell(1,16).value ='前日比%'
        sheetstock.cell(1,17).value ='PER'
        sheetstock.cell(1,18).value ='PBR'
        sheetstock.cell(1,19).value ='上場市場'
        sheetstock.cell(1,20).value ='VWAP'
        sheetstock.cell(1,21).value ='発行済み株式数'
        sheetstock.cell(1,22).value ='最新信用売残'
        sheetstock.cell(1,23).value ='最新信用買残'
        sheetstock.cell(1,24).value ='信用倍率'
        sheetstock.cell(1,25).value ='売買代金2'
        sheetstock.cell(1,26).value ='信用売残前週比'
        sheetstock.cell(1,27).value ='信用買残前週比'
        sheetstock.cell(1,28).value ='出来高前日比'
        sheetstock.cell(1,29).value ='約定回数前日比'
        sheetstock.cell(1,30).value ='時価総額'
        sheetstock.cell(1,31).value ='浮動株総額'
        sheetstock.cell(1,32).value ='一株配当'
        sheetstock.cell(1,33).value ='平均約定金額'
        sheetstock.cell(1,34).value ='業界'
        sheetstock.cell(1,35).value ='時価総額2'
        sheetstock.cell(1,36).value ='ギャップアップフラグ'
        sheetstock.cell(1,37).value ='ギャップダウンフラグ'
        sheetstock.cell(1,38).value ='出来高急増フラグ'
        sheetstock.cell(1,51).value ='当日IR有無'
        sheetstock.cell(1,52).value ='決算IRフラグ有無'
        sheetstock.cell(1,53).value ='決算修正IRフラグ有無'
        sheetstock.cell(1,54).value ='当日テクニカルフラグ有無'
        sheetstock.cell(1,55).value ='当日5%フラグ有無'
        sheetstock.cell(1,56).value ='翌日用当日IR有無'
        sheetstock.cell(1,57).value ='翌日用決算IRフラグ有無'
        sheetstock.cell(1,58).value ='翌日用決算修正IRフラグ有無'
        sheetstock.cell(1,59).value ='翌日用当日テクニカルフラグ有無'
        sheetstock.cell(1,60).value ='翌日用当日5%フラグ有無'
        sheetstock.cell(1,61).value ='利回り'
        sheetstock.cell(1,62).value ='売上高'
        sheetstock.cell(1,63).value ='営業益'
        sheetstock.cell(1,64).value ='経常益'
        sheetstock.cell(1,65).value ='最終益'
        sheetstock.cell(1,66).value ='売上営業利益率'
        sheetstock.cell(1,67).value ='ROE'
        sheetstock.cell(1,68).value ='ROA'
        sheetstock.cell(1,69).value ='総資産回転率'
        sheetstock.cell(1,70).value ='フリーキャッシュフロー'
        sheetstock.cell(1,71).value ='営業キャッシュフロー'
        sheetstock.cell(1,72).value ='投資キャッシュフロー'
        sheetstock.cell(1,73).value ='財務キャッシュフロー'
        sheetstock.cell(1,74).value ='現金'
        sheetstock.cell(1,75).value ='一株益'
        sheetstock.cell(1,101).value ='信用取引規制中'
        sheetstock.cell(1,102).value ='貸借取引銘柄別増担保金徴収措置'
        sheetstock.cell(1,103).value ='貸借取引銘柄別増担保金徴収措置_措置内容'
        sheetstock.cell(1,106).value ='空売り規制対象'
        sheetstock.cell(1,111).value ='融資新規'
        sheetstock.cell(1,112).value ='融資返済'
        sheetstock.cell(1,113).value ='融資残高'
        sheetstock.cell(1,114).value ='貸株新規'
        sheetstock.cell(1,115).value ='貸株返済'
        sheetstock.cell(1,116).value ='貸株残高'
        sheetstock.cell(1,117).value ='差引残高'
        sheetstock.cell(1,118).value ='回転日数'
        sheetstock.cell(1,121).value ='貸株超過株数'
        sheetstock.cell(1,122).value ='最高料率'
        sheetstock.cell(1,123).value ='当日品貸料率'
        sheetstock.cell(1,124).value ='前日品貸料率'
        sheetstock.cell(1,151).value ='みんかぶ目標株価'
        sheetstock.cell(1,152).value ='現在株価との差'
        sheetstock.cell(1,199).value ='移動平均線数値5日'
        sheetstock.cell(1,200).value ='移動平均線数値25日'
        sheetstock.cell(1,201).value ='移動平均線数値75日'
        sheetstock.cell(1,202).value ='移動平均乖離率5日'
        sheetstock.cell(1,203).value ='移動平均乖離率25日'
        sheetstock.cell(1,204).value ='移動平均乖離率75日'
        sheetstock.cell(1,205).value ='移動平均線数値10日'
        sheetstock.cell(1,206).value ='移動平均線数値15日'
        sheetstock.cell(1,207).value ='移動平均線数値20日'
        sheetstock.cell(1,208).value ='移動平均線数値30日'
        sheetstock.cell(1,209).value ='移動平均線数値65日'
        sheetstock.cell(1,210).value ='移動平均線数値130日'
        sheetstock.cell(1,211).value ='移動平均線数値10日出来高'
        sheetstock.cell(1,212).value ='移動標準偏差+1σ25日'
        sheetstock.cell(1,213).value ='移動標準偏差-1σ25日'
        sheetstock.cell(1,214).value ='移動標準偏差+2σ25日'
        sheetstock.cell(1,215).value ='移動標準偏差-2σ25日'
        sheetstock.cell(1,216).value ='移動標準偏差+3σ25日'
        sheetstock.cell(1,217).value ='移動標準偏差-3σ25日'
        sheetstock.cell(1,218).value ='移動標準偏差+1σ65日'
        sheetstock.cell(1,219).value ='移動標準偏差-1σ65日'
        sheetstock.cell(1,220).value ='移動標準偏差+2σ65日'
        sheetstock.cell(1,221).value ='移動標準偏差-2σ65日'
        sheetstock.cell(1,222).value ='移動標準偏差+3σ65日'
        sheetstock.cell(1,223).value ='移動標準偏差-3σ65日'
        sheetstock.cell(1,224).value ='25日ボリンジャーバンドでの株価位置'
        sheetstock.cell(1,225).value ='13週ボリンジャーバンドでの株価位置'
        sheetstock.cell(1,226).value =''
        sheetstock.cell(1,227).value ='上影/下影/大/小'
        sheetstock.cell(1,228).value ='陽線/陰線/十字線'
        sheetstock.cell(1,229).value ='プラス超え/マイナス越え株価5日線超え'
        sheetstock.cell(1,230).value ='移動平均線数値5日出来高'
        sheetstock.cell(1,231).value ='プラス超え/マイナス越え出来高5日線超え'
        sheetstock.cell(1,251).value ='平均出来高'
        sheetstock.cell(1,252).value ='平均約定回数'
        sheetstock.cell(1,253).value ='平均回転日数'
        sheetstock.cell(1,254).value ='平均移動平均乖離率5日'
        sheetstock.cell(1,255).value ='平均移動平均乖離率25日'
        sheetstock.cell(1,256).value ='平均移動平均乖離率75日'
        sheetstock.cell(1,257).value ='平均信用買残'
        sheetstock.cell(1,258).value ='平均融資新規'
        sheetstock.cell(1,259).value ='平均融資返済'
        sheetstock.cell(1,260).value ='平均信用売残'
        sheetstock.cell(1,261).value ='平均貸株新規'
        sheetstock.cell(1,262).value ='平均貸株返済'
        sheetstock.cell(1,263).value ='平均貸株超過'
        sheetstock.cell(1,264).value ='平均出来高変化率'
        sheetstock.cell(1,265).value ='平均約定回数変化率'
        sheetstock.cell(1,266).value ='平均信用買残変化率'
        sheetstock.cell(1,267).value ='平均信用売残変化率'
        sheetstock.cell(1,268).value ='平均平均約定金額'
        sheetstock.cell(1,269).value ='標準偏差出来高'
        sheetstock.cell(1,270).value ='標準偏差約定回数'
        sheetstock.cell(1,271).value ='標準偏差回転日数'
        sheetstock.cell(1,272).value ='標準偏差移動平均乖離率5日'
        sheetstock.cell(1,273).value ='標準偏差移動平均乖離率25日'
        sheetstock.cell(1,274).value ='標準偏差移動平均乖離率75日'
        sheetstock.cell(1,275).value ='標準偏差信用買残'
        sheetstock.cell(1,276).value ='標準偏差融資新規'
        sheetstock.cell(1,277).value ='標準偏差融資返済'
        sheetstock.cell(1,278).value ='標準偏差信用売残'
        sheetstock.cell(1,279).value ='標準偏差貸株新規'
        sheetstock.cell(1,280).value ='標準偏差貸株返済'
        sheetstock.cell(1,281).value ='標準偏差貸株超過'
        sheetstock.cell(1,282).value ='標準偏差出来高変化率'
        sheetstock.cell(1,283).value ='標準偏差約定回数変化率'
        sheetstock.cell(1,284).value ='標準偏差信用買残変化率'
        sheetstock.cell(1,285).value ='標準偏差信用売残変化率'
        sheetstock.cell(1,286).value ='標準偏差平均約定金額'
        sheetstock.cell(1,301).value ='20日間の移動平均'
        sheetstock.cell(1,302).value ='終値と20日移動平均の差に二乗'
        sheetstock.cell(1,303).value ='20日間の分散'
        sheetstock.cell(1,304).value ='20日間の標準偏差'
        sheetstock.cell(1,305).value ='上限線'
        sheetstock.cell(1,306).value ='下限線'
        sheetstock.cell(1,307).value ='終値との比較-中間値に対してどちらよりか'
        sheetstock.cell(1,308).value =''
        sheetstock.cell(1,309).value ='前日比'
        sheetstock.cell(1,310).value ='上昇分'
        sheetstock.cell(1,311).value ='下落分'
        sheetstock.cell(1,312).value ='14日間の上昇分の平均'
        sheetstock.cell(1,313).value ='14日間の下落分の平均'
        sheetstock.cell(1,314).value ='相対力'
        sheetstock.cell(1,315).value ='RSI'
        sheetstock.cell(1,316).value ='RSI評価-買われすぎか売られすぎか'
        sheetstock.cell(1,317).value ='12日間の指数平滑移動平均'
        sheetstock.cell(1,318).value ='26日間の指数平滑移動平均'
        sheetstock.cell(1,319).value ='MACD'
        sheetstock.cell(1,320).value ='9日間の指数平滑移動平均'
        sheetstock.cell(1,321).value ='12日間の指数平滑移動平均と26日間の指数平滑移動平均の中間値'
        sheetstock.cell(1,322).value ='9日間の指数平滑移動平均と中間値の比較'
        sheetstock.cell(1,323).value ='MACDラインとシグナルラインの差分'
        sheetstock.cell(1,324).value ='RSIスコア'
        sheetstock.cell(1,325).value ='ボリンジャーバンドスコア'
        sheetstock.cell(1,326).value ='MACDスコア'
        sheetstock.cell(1,327).value ='合計点'
        sheetstock.cell(1,328).value ='仮スコア'
        sheetstock.cell(1,401).value ='業種'
        sheetstock.cell(1,402).value ='会社説明'
        sheetstock.cell(1,501).value ='標準化出来高'
        sheetstock.cell(1,502).value ='標準化約定回数'
        sheetstock.cell(1,503).value ='標準化回転日数'
        sheetstock.cell(1,504).value ='標準化移動平均乖離率5日'
        sheetstock.cell(1,505).value ='標準化移動平均乖離率25日'
        sheetstock.cell(1,506).value ='標準化移動平均乖離率75日'
        sheetstock.cell(1,507).value ='標準化信用買残'
        sheetstock.cell(1,508).value ='標準化融資新規'
        sheetstock.cell(1,509).value ='標準化融資返済'
        sheetstock.cell(1,510).value ='標準化信用売残'
        sheetstock.cell(1,511).value ='標準化貸株新規'
        sheetstock.cell(1,512).value ='標準化貸株返済'
        sheetstock.cell(1,513).value ='標準化貸株超過'
        sheetstock.cell(1,514).value ='標準化出来高変化率'
        sheetstock.cell(1,515).value ='標準化約定回数変化率'
        sheetstock.cell(1,516).value ='標準化信用買残変化率'
        sheetstock.cell(1,517).value ='標準化信用売残変化率'
        sheetstock.cell(1,518).value ='標準化平均約定金額'
        sheetstock.cell(1,601).value ='テーマ'
        sheetstock.cell(1,999).value ='getnewsinfoを実施したか'
        sheetstock.cell(1,1000).value ='売り時/買い時スコア'
        '''

        if sheetstock.cell(2,5).value is None or sheetstock.cell(2,5).value == '－':
                sheetstock.cell(2,5).value = 0
                print(sheetstock.cell(2,2).value + '_' + sheetstock.cell(2,3).value +'前日比の値がなかったので処理')
        else:
                pass
        if sheetstock.cell(lastrow_stockbook-1, 4).value == '－':
                sheetstock.cell(lastrow_stockbook-1,4).value = sheetstock.cell(lastrow_stockbook-2,4).value
                sheetstock.cell(lastrow_stockbook-1,11).value = 0
                sheetstock.cell(lastrow_stockbook-1,12).value = sheetstock.cell(lastrow_stockbook-2,4).value
                sheetstock.cell(lastrow_stockbook-1,13).value = sheetstock.cell(lastrow_stockbook-2,4).value
                sheetstock.cell(lastrow_stockbook-1,14).value = sheetstock.cell(lastrow_stockbook-2,4).value
                sheetstock.cell(lastrow_stockbook-1,15).value = sheetstock.cell(lastrow_stockbook-2,4).value
                print(sheetstock.cell(lastrow_stockbook-2,2).value + '_' + sheetstock.cell(lastrow_stockbook-2,3).value +'株価の値がなかったので処理')
        if sheetstock.cell(lastrow_stockbook-1,3).value == '株探からのお知らせ':
                sheetstock.delete_rows(lastrow_stockbook-1)
                print(str(sheetstock.cell(lastrow_stockbook-2,2).value) + '_' + str(sheetstock.cell(lastrow_stockbook-2,3).value) +'上場廃止対象のため削除')
                wb.save(l)
                continue
        for m in range(3,lastrow_stockbook):
                if sheetstock.cell(m,5).value is None or sheetstock.cell(m,5).value == '－':
                        sheetstock.cell(m,5).value = sheetstock.cell(m,4).value-sheetstock.cell(m-1,4).value
                        print(sheetstock.cell(m,2).value + '_' + sheetstock.cell(m,3).value +'：'+str(m)+'行目に前日比の値がなかったので処理')

                else:
                        pass


#配列宣言
        stockprice = [] #株価
        turnover = [] #出来高
        upprice = [] #上昇幅
        downprice = [] #下落幅
        stockprice_5dmovemean = [] #5日移動平均
        stockprice_25dmovemean = [] #25日移動平均
        stockprice_75dmovemean = [] #75日移動平均
        stockprice_10dmovemean = [] #10日移動平均
        stockprice_15dmovemean = [] #15日移動平均
        stockprice_20dmovemean = [] #20日移動平均
        stockprice_30dmovemean = [] #30日移動平均
        stockprice_65dmovemean = [] #13週移動平均
        stockprice_130dmovemean = [] #26週移動平均
        turnover_5movemean = [] #5日移動平均出来高
        turnover_10movemean = [] #10日移動平均出来高
        stockprice_25dmovestd = [] #25日移動標準偏差
        stockprice_65dmovestd = [] #13週移動標準偏差
#20230409追加
        stockprice_20dmovestd = [] #20日移動標準偏差ーボリンジャーバンド
        stockprice_20dmovedtl = [] #20日移動分散ーボリンジャーバンド
        stockprice_14dupmean = [] #14日上昇平均ーRSI
        stockprice_14ddownmean = [] #14日下落平均ーRSI
        stockprice_12dEmovemean = [] #12日指数平滑移動平均ーMACD
        stockprice_26dEmovemean = [] #26日指数平滑移動平均ーMACD
        stockprice_9dEmovemean = [] #9日指数平滑移動平均ーMACD
        macdscore=[] #macdスコア用配列
        diffarray=[] #macdスコア計算用
        diffarray_max=[]
        diffarray_min=[]
        macd_diff=[]



#シートの2行目~最終行をループ
        for i in range(2,lastrow_stockbook):

#A列を配列へ格納
#                print(i)
                stockprice = np.append(stockprice, sheetstock.cell(row=i, column=4).value) #株価の配列格納
                turnover = np.append(turnover, sheetstock.cell(row=i, column=11).value) #出来高の配列格納
                #前日比のセルから、upprice, downpriceの配列格納をしたい。
                #前日比マイナスの時はuppriceにゼロ、プラスの時はdownpriceにゼロを入れる形にする。

                         
                if int(sheetstock.cell(i, 5).value) >= 0:
                        upprice = np.append(upprice, sheetstock.cell(i, 5).value)
                        downprice = np.append(downprice, 0)
                        #print(sheetstock.cell(i,5).value)
                else:
                        upprice = np.append(upprice, 0)
                        downprice = np.append(downprice, abs(sheetstock.cell(i,5).value))
                
#売買代金2=VWAP(T)*出来高(K)
#                sheetstock.cell(row=i,column=25).value = sheetstock.cell(row=i,column=20).value * sheetstock.cell(row=i,column=11).value
                #ローソク足の分類分け
#陽線/陰線の場合分け
                if int(sheetstock.cell(row=i, column=12).value) > int(sheetstock.cell(row=i, column=15).value):
#                        print(str(i))
                        fill = PatternFill(patternType="solid", fgColor="00BFFF")
                        sheetstock.cell(row=i, column=228).value = '陰線'
                        sheetstock.cell(row=i, column=228).fill = fill
                elif int(sheetstock.cell(row=i, column=15).value) > int(sheetstock.cell(row=i, column=12).value):
                        fill = PatternFill(patternType="solid", fgColor="FF69B4")
                        sheetstock.cell(row=i,column=228).value = '陽線'
                        sheetstock.cell(row=i, column=228).fill = fill
                else:
                        sheetstock.cell(row=i, column=228).value = '十字線'
#                        print(str(i))


#実体の長さ
                candle = abs(int(sheetstock.cell(row=i,column=12).value) - int(sheetstock.cell(row=i, column=15).value))
#陰線の場合
                if sheetstock.cell(row=i, column=228).value == '陰線':
                        blue_beard_over = int(sheetstock.cell(row=i, column=13).value) - int(sheetstock.cell(row=i, column=12).value)
                        blue_beard_under = int(sheetstock.cell(row=i, column=15).value) - int(sheetstock.cell(row=i, column=14).value)
                        if candle > blue_beard_over and candle > blue_beard_under:
                                sheetstock.cell(row=i, column=227).value = '大'
                        elif candle >= blue_beard_over and candle <= blue_beard_under:
                                sheetstock.cell(row=i, column=227).value = '下影'
                        elif candle <= blue_beard_over and candle >= blue_beard_under:
                                sheetstock.cell(row=i, column=227).value = '上影'
                        elif candle < blue_beard_over and candle < blue_beard_under:
                                sheetstock.cell(row=i, column=227).value = '小'
                        else:
                                sheetstock.cell(row=i, column=227).value = '異常値'
#陽線の場合
                elif sheetstock.cell(row=i, column=228).value == '陽線':
                        red_beard_over = int(sheetstock.cell(row=i, column=13).value) - int(sheetstock.cell(row=i, column=15).value)
                        red_beard_under = int(sheetstock.cell(row=i, column=12).value) - int(sheetstock.cell(row=i, column=14).value)
                        if candle > red_beard_over and candle > red_beard_under:
                                sheetstock.cell(row=i, column=227).value = '大'
                        elif candle >= red_beard_over and candle <= red_beard_under:
                                sheetstock.cell(row=i, column=227).value = '下影'
                        elif candle <= red_beard_over and candle >= red_beard_under:
                                sheetstock.cell(row=i, column=227).value = '上影'
                        elif candle < red_beard_over and candle < red_beard_under:
                                sheetstock.cell(row=i, column=227).value = '小'
                        else:
                                sheetstock.cell(row=i, column=227).value = '異常値'
#        print(stockprice)
#        print(turnover)
#移動平均数値の取込
        if lastrow_stockbook > 7:
                stockprice_5dmovemean = bn.move_mean(stockprice, window=5)
                turnover_5dmovemean = bn.move_mean(turnover, window=5)

                if lastrow_stockbook > 12:
                        stockprice_10dmovemean = bn.move_mean(stockprice, window=10)
                        turnover_10dmovemean = bn.move_mean(turnover, window=10)
                        if lastrow_stockbook > 14:
                                stockprice_12dEmovemean = bn.move_mean(stockprice, window=12)
                                if lastrow_stockbook > 16:
                                        stockprice_14dupmean = bn.move_var(upprice, window=14)
                                        stockprice_14ddownmean = bn.move_var(downprice, window=14)
                                        if lastrow_stockbook > 17:
                                                stockprice_15dmovemean = bn.move_mean(stockprice, window=15)
                                                if lastrow_stockbook > 22:
                                                        stockprice_20dmovemean = bn.move_mean(stockprice, window=20)
                                                        stockprice_20dmovevar = bn.move_var(stockprice, window=20)
                                                        stockprice_20dmovestd = bn.move_std(stockprice, window=20)
                                                        if lastrow_stockbook > 27:
                                                                stockprice_25dmovemean = bn.move_mean(stockprice, window=25)
                                                                stockprice_25dmovestd = bn.move_std(stockprice, window=25)
                                                                if lastrow_stockbook > 28:
                                                                        stockprice_26dEmovemean = bn.move_mean(stockprice, window=26)
                                                                        if lastrow_stockbook > 32:
                                                                                stockprice_30dmovemean = bn.move_mean(stockprice, window=30)
                                                                                if lastrow_stockbook > 67:
                                                                                        stockprice_65dmovemean = bn.move_mean(stockprice, window=65)
                                                                                        stockprice_65dmovestd = bn.move_std(stockprice, window=65)
                                                                                        if lastrow_stockbook > 77:
                                                                                                stockprice_75dmovemean = bn.move_mean(stockprice, window=75)
                                                                                                if lastrow_stockbook > 132:
                                                                                                        stockprice_130dmovemean = bn.move_mean(stockprice, window=130)

                                                                        

#        print(stockprice1)
#        print('移動平均', bn.move_mean(stockprice, window = 5))
        for j in range(2, lastrow_stockbook):
                if lastrow_stockbook > 7:
                        sheetstock.cell(row=j,column=199).value = stockprice_5dmovemean[j-2]
                        sheetstock.cell(row=j,column=230).value = turnover_5dmovemean[j-2]
                        if lastrow_stockbook > 12:
                                sheetstock.cell(row=j,column=205).value = stockprice_10dmovemean[j-2]
                                sheetstock.cell(row=j,column=211).value = turnover_10dmovemean[j-2]
                                if lastrow_stockbook > 14:
                                        sheetstock.cell(row=j,column=317).value = stockprice_12dEmovemean[j-2] + (2/13)*(sheetstock.cell(j,4).value - stockprice_12dEmovemean[j-2])
                                        if lastrow_stockbook > 16:
                                                sheetstock.cell(row=j,column=312).value = stockprice_14dupmean[j-2]
                                                sheetstock.cell(row=j,column=313).value = stockprice_14ddownmean[j-2]
                                                sheetstock.cell(j,314).value = stockprice_14dupmean[j-2]/stockprice_14ddownmean[j-2]
                                                sheetstock.cell(j,315).value = 100-(100/(1+stockprice_14dupmean[j-2]/stockprice_14ddownmean[j-2]))
                                                if sheetstock.cell(j,315).value < 30:
                                                        sheetstock.cell(j,316).value = "売られすぎ"
                                                elif sheetstock.cell(j,315).value > 70:
                                                        sheetstock.cell(j,316).value = "買われすぎ"

                                                sheetstock.cell(j,324).value = 100-sheetstock.cell(j,315).value

                                                if lastrow_stockbook > 17:
                                                        sheetstock.cell(row=j,column=206).value = stockprice_15dmovemean[j-2]
                                                        if lastrow_stockbook > 22:
                                                                sheetstock.cell(row=j,column=207).value = stockprice_20dmovemean[j-2]
                                                                sheetstock.cell(row=j,column=301).value = stockprice_20dmovemean[j-2]
                                                                sheetstock.cell(row=j,column=302).value = (sheetstock.cell(row=j,column=4).value - stockprice_20dmovemean[j-2])**2
                                                                sheetstock.cell(row=j,column=303).value = stockprice_20dmovevar[j-2]
                                                                sheetstock.cell(row=j,column=304).value = stockprice_20dmovestd[j-2]
                                                                sheetstock.cell(row=j,column=305).value = stockprice_20dmovemean[j-2]+2*stockprice_20dmovestd[j-2]
                                                                sheetstock.cell(row=j,column=306).value = stockprice_20dmovemean[j-2]-2*stockprice_20dmovestd[j-2]
                                                                if sheetstock.cell(j,305).value-sheetstock.cell(j,4).value<0:
                                                                        sheetstock.cell(j,325).value = 0
                                                                elif sheetstock.cell(j,306).value-sheetstock.cell(j,4).value>0:
                                                                        sheetstock.cell(j,325).value = 100
                                                                else:
                                                                    sheetstock.cell(j,325).value = ((sheetstock.cell(j,305).value-sheetstock.cell(j,4).value)/(sheetstock.cell(j,305).value-sheetstock.cell(j,306).value))*100
                                                                if sheetstock.cell(j,4).value - stockprice_20dmovemean[j-2] > 0:
                                                                        sheetstock.cell(row=j,column=307).value = "上側"
                                                                else:
                                                                        sheetstock.cell(row=j,column=307).value = "下側"

                                                                if lastrow_stockbook > 27:
                                                                        sheetstock.cell(row=j,column=200).value = stockprice_25dmovemean[j-2]
                                                                        if lastrow_stockbook > 28:
                                                                                sheetstock.cell(row=j,column=318).value = stockprice_26dEmovemean[j-2] + (2/27)*(sheetstock.cell(j,4).value - stockprice_26dEmovemean[j-2])
                                                                                sheetstock.cell(row=j,column=319).value = sheetstock.cell(row=j,column=317).value - sheetstock.cell(row=j,column=318).value#macd
                                                                                if lastrow_stockbook > 32:
                                                                                        sheetstock.cell(row=j,column=208).value = stockprice_30dmovemean[j-2]
                                                                                        if lastrow_stockbook > 67:
                                                                                                sheetstock.cell(row=j,column=209).value = stockprice_65dmovemean[j-2]
                                                                                                if lastrow_stockbook > 77:
                                                                                                        sheetstock.cell(row=j,column=201).value = stockprice_75dmovemean[j-2]
                                                                                                        if lastrow_stockbook > 132:
                                                                                                                sheetstock.cell(row=j,column=210).value = stockprice_130dmovemean[j-2]
                                                                                        

                if lastrow_stockbook > 7:
                        if sheetstock.cell(j,15).value-sheetstock.cell(j,199).value>0 and sheetstock.cell(j-2,15).value-sheetstock.cell(j-2,199).value<=0:
                                sheetstock.cell(j,229).value=1
                        elif sheetstock.cell(j,15).value-sheetstock.cell(j,199).value<=0 and sheetstock.cell(j-2,15).value-sheetstock.cell(j-2,199).value>0:
                                sheetstock.cell(j,229).value=4
                        else:
                                pass

                if lastrow_stockbook > 7:
                        if sheetstock.cell(j,11).value-sheetstock.cell(j,230).value>0 and sheetstock.cell(j-2,11).value-sheetstock.cell(j-2,230).value<=0:
                                sheetstock.cell(j,231).value=1
                        elif sheetstock.cell(j,11).value-sheetstock.cell(j,230).value<=0 and sheetstock.cell(j-2,11).value-sheetstock.cell(j-2,230).value>0:
                                sheetstock.cell(j,231).value=4
                        else:
                                pass

                if lastrow_stockbook > 7:
                        if sheetstock.cell(j,15).value-sheetstock.cell(j-5,15).value>0:
                                sheetstock.cell(j,232).value=1
                        elif sheetstock.cell(j,15).value-sheetstock.cell(j-5,15).value<=0:
                                sheetstock.cell(j,232).value=4
                        else:
                                pass
                
                if lastrow_stockbook > 22:
                        if sheetstock.cell(j,15).value-sheetstock.cell(j-20,15).value>0:
                                sheetstock.cell(j,233).value=1
                        elif sheetstock.cell(j,15).value-sheetstock.cell(j-20,15).value<=0:
                                sheetstock.cell(j,233).value=4
                        else:
                                pass
                
                if lastrow_stockbook > 22:
                        if sheetstock.cell(j,232).value == 1 and sheetstock.cell(j,233).value == 4:
                                sheetstock.cell(j,234).value=1
                        elif sheetstock.cell(j,232).value == 4 and sheetstock.cell(j,233).value == 1:
                                sheetstock.cell(j,235).value=1
                        elif sheetstock.cell(j,232).value == 1 and sheetstock.cell(j,233).value == 1:
                                sheetstock.cell(j,236).value=1
                        elif sheetstock.cell(j,232).value == 4 and sheetstock.cell(j,233).value == 4:
                                sheetstock.cell(j,237).value=1
                        else:
                                pass

                if lastrow_stockbook > 27:
                        sheetstock.cell(row=j,column=212).value = stockprice_25dmovemean[j-2]+stockprice_25dmovestd[j-2] #25日+1σ
                        sheetstock.cell(row=j,column=213).value = stockprice_25dmovemean[j-2]-stockprice_25dmovestd[j-2] #25日-1σ
                        sheetstock.cell(row=j,column=214).value = stockprice_25dmovemean[j-2]+2*stockprice_25dmovestd[j-2] #25日+2σ
                        sheetstock.cell(row=j,column=215).value = stockprice_25dmovemean[j-2]-2*stockprice_25dmovestd[j-2] #25日-2σ
                        sheetstock.cell(row=j,column=216).value = stockprice_25dmovemean[j-2]+3*stockprice_25dmovestd[j-2] #25日+3σ
                        sheetstock.cell(row=j,column=217).value = stockprice_25dmovemean[j-2]-3*stockprice_25dmovestd[j-2] #25日-3σ

                if lastrow_stockbook > 67:
                        sheetstock.cell(row=j,column=218).value = stockprice_65dmovemean[j-2]+stockprice_65dmovestd[j-2] #13週+1σ
                        sheetstock.cell(row=j,column=219).value = stockprice_65dmovemean[j-2]-stockprice_65dmovestd[j-2] #13週-1σ
                        sheetstock.cell(row=j,column=220).value = stockprice_65dmovemean[j-2]+2*stockprice_65dmovestd[j-2] #13週+2σ
                        sheetstock.cell(row=j,column=221).value = stockprice_65dmovemean[j-2]-2*stockprice_65dmovestd[j-2] #13週-2σ
                        sheetstock.cell(row=j,column=222).value = stockprice_65dmovemean[j-2]+3*stockprice_65dmovestd[j-2] #13週+3σ
                        sheetstock.cell(row=j,column=223).value = stockprice_65dmovemean[j-2]-3*stockprice_65dmovestd[j-2] #13週-3σ

                if lastrow_stockbook > 27:
                        if sheetstock.cell(row=j, column=216).value<sheetstock.cell(row=j,column=4).value:
                                sheetstock.cell(row=j,column=224).value = 'プラスの異常値'
                        elif sheetstock.cell(row=j, column=216).value>=sheetstock.cell(row=j, column=4).value>sheetstock.cell(row=j,column=214).value:
                                sheetstock.cell(row=j, column=224).value = '+3σ範囲'
                        elif sheetstock.cell(row=j, column=214).value>=sheetstock.cell(row=j, column=4).value>sheetstock.cell(row=j,column=212).value:
                                sheetstock.cell(row=j, column=224).value = '+2σ範囲'
                        elif sheetstock.cell(row=j, column=212).value>=sheetstock.cell(row=j, column=4).value>sheetstock.cell(row=j,column=200).value:
                                sheetstock.cell(row=j, column=224).value = '+1σ範囲'
                        elif sheetstock.cell(row=j, column=200).value>=sheetstock.cell(row=j, column=4).value>sheetstock.cell(row=j,column=213).value:
                                sheetstock.cell(row=j, column=224).value = '-1σ範囲'
                        elif sheetstock.cell(row=j, column=213).value>=sheetstock.cell(row=j, column=4).value>sheetstock.cell(row=j,column=215).value:
                                sheetstock.cell(row=j, column=224).value = '-2σ範囲'
                        elif sheetstock.cell(row=j, column=215).value>=sheetstock.cell(row=j, column=4).value>sheetstock.cell(row=j,column=217).value:
                                sheetstock.cell(row=j, column=224).value = '-3σ範囲'
                        elif sheetstock.cell(row=j, column=217).value>sheetstock.cell(row=j, column=4).value:
                                sheetstock.cell(row=j, column=224).value = 'マイナスの異常値'
                        else:
                                sheetstock.cell(row=j, column=224).value = 'error'

                if lastrow_stockbook > 67:
                        if sheetstock.cell(row=j, column=222).value<sheetstock.cell(row=j,column=4).value:
                                sheetstock.cell(row=j,column=225).value = 'プラスの異常値'
                        elif sheetstock.cell(row=j, column=222).value>=sheetstock.cell(row=j, column=4).value>sheetstock.cell(row=j,column=220).value:
                                sheetstock.cell(row=j, column=225).value = '+3σ範囲'
                        elif sheetstock.cell(row=j, column=220).value>=sheetstock.cell(row=j, column=4).value>sheetstock.cell(row=j,column=218).value:
                                sheetstock.cell(row=j, column=225).value = '+2σ範囲'
                        elif sheetstock.cell(row=j, column=218).value>=sheetstock.cell(row=j, column=4).value>sheetstock.cell(row=j,column=209).value:
                                sheetstock.cell(row=j, column=225).value = '+1σ範囲'
                        elif sheetstock.cell(row=j, column=209).value>=sheetstock.cell(row=j, column=4).value>sheetstock.cell(row=j,column=219).value:
                                sheetstock.cell(row=j, column=225).value = '-1σ範囲'
                        elif sheetstock.cell(row=j, column=219).value>=sheetstock.cell(row=j, column=4).value>sheetstock.cell(row=j,column=221).value:
                                sheetstock.cell(row=j, column=225).value = '-2σ範囲'
                        elif sheetstock.cell(row=j, column=221).value>=sheetstock.cell(row=j, column=4).value>sheetstock.cell(row=j,column=223).value:
                                sheetstock.cell(row=j, column=225).value = '-3σ範囲'
                        elif sheetstock.cell(row=j, column=223).value>sheetstock.cell(row=j, column=4).value:
                                sheetstock.cell(row=j, column=225).value = 'マイナスの異常値'
                        else:
                                sheetstock.cell(row=j, column=225).value = 'error'

        for k in range(2, lastrow_stockbook):
                macdscore = np.append(macdscore, sheetstock.cell(row=k, column=319).value) #macdの配列格納
                if k > 11:
                        if sheetstock.cell(k,319).value is not None:
                                stockprice_9dEmovemean = bn.move_mean(macdscore, window=9)
                                sheetstock.cell(row=k,column=320).value = stockprice_9dEmovemean[k-2]#macd_signal
                                sheetstock.cell(k,323).value = sheetstock.cell(k,319).value-sheetstock.cell(k,320).value
                                diffarray = np.append(macdscore, sheetstock.cell(row=k, column=323).value) #macdの配列格納
                                diffarray_max = bn.move_max(macdscore, window=9)
                                diffarray_min = bn.move_min(macdscore, window=9)
                                macd = sheetstock.cell(k,319).value
                                macd_signal = sheetstock.cell(k,320).value
                                diff = macd - macd_signal
                                #print(diff)
                                macd_diff = np.append(macd_diff, diff)

                                macd_diff_abs_max = np.nanmax(macd_diff)
                                #print(macd_diff_abs_max)
                                sheetstock.cell(k,326).value =50+(diff/macd_diff_abs_max)*50

#MACDの計算2023/5/21段階
#macd = sheetstock.cell(k,323).valueの正負によって、スコアの計算を変えている。
#macd>=0で
#sheetstock.cell(k,326).value = (1-(sheetstock.cell(k,323).value-diffarray_min[k-2])/(diffarray_max[k-2]-diffarray_min[k-2]))*100
#macd<0で
#sheetstock.cell(k,326).value =((sheetstock.cell(k,323).value-diffarray_min[k-2])/(diffarray_max[k-2]-diffarray_min[k-2]))*100
#sheetstock.cell(k,323).valueとは
#sheetstock.cell(k,319).value-sheetstock.cell(k,320).value
#sheetstock.cell(k,319).valueとは
#sheetstock.cell(row=j,column=317).value - sheetstock.cell(row=j,column=318).value
#sheetstock.cell(row=j,column=317).valueとは
#stockprice_12dEmovemean[j-2] + (2/13)*(sheetstock.cell(j,4).value - stockprice_12dEmovemean[j-2])
#sheetstock.cell(row=j,column=318).valueとは
#stockprice_26dEmovemean[j-2] + (2/27)*(sheetstock.cell(j,4).value - stockprice_26dEmovemean[j-2])
#sheetstock.cell(k,320).valueとは
#stockprice_9dEmovemean[k-2]
        for m in range(2,lastrow_stockbook):
                if sheetstock.cell(m,324).value is not None and sheetstock.cell(m,325).value is not None and sheetstock.cell(m,326).value is not None:
                        sheetstock.cell(m, 327).value = -0.126*sheetstock.cell(m,324).value + 0.486*sheetstock.cell(m,325).value + 0.846*sheetstock.cell(m,326).value

        for n in range(2, lastrow_stockbook-10):
                sheetstock.cell(n,328).value = (sheetstock.cell(n+10,4).value-sheetstock.cell(n,4).value)/sheetstock.cell(n,4).value*1000
        #シグナルラインの計算
        #print(stockprice_25dmovemean[9])

        wb.save(l)
        print(l)

#終了時間取得-経過時間
print(t)
t1 = datetime.datetime.now()
print(t1)
dt = t1-t
print(dt)
#終了時間取得-経過時間

#稼働終了アナウンス
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾
#株価列はD=4列目
#13週(65日)移動平均     HA=209列目
#26週(130日)移動平均    HB=210列目
#移動平均5日線          GQ=199列目
#移動平均10日線         GW=205列目
#移動平均15日線         GX=206列目
#移動平均20日線         GY=207列目
#移動平均25日線         GR=200列目
#移動平均30日線         GZ=208列目
#ボリンジャーバンド 13週、25日
#標準偏差
#+1σ 25日               HD=212列目
#-1σ 25日               HE=213列目
#+2σ 25日               HF=214列目
#-2σ 25日               HG=215列目
#+3σ 25日               HH=216列目
#-3σ 25日               HI=217列目
#+1σ 13週               HJ=218列目
#-1σ 13週               HK=219列目
#+2σ 13週               HL=220列目
#-2σ 13週               HM=221列目
#+3σ 13週               HN=222列目
#-3σ 13週               HO=223列目
#25日ボリンジャーバンドでの株価位置   HP=224列目
#13週ボリンジャーバンドでの株価位置   HQ=225列目
#ローソク足分類         HQ=225列目




