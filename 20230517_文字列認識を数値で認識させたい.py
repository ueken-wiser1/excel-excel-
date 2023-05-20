#何を見るのか
#日付のフォーマットとそれらを照合して、一致と見なすかどうかの確認

import os
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import datetime
import jpholiday
from datetime import timedelta
import pandas as pd
from openpyxl.utils import column_index_from_string


t = datetime.datetime.now().time()
d = datetime.datetime.now()
d1 = d.strftime('%Y%m%d')

watch_list_file = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/'
completed_folder = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/完了/完/'

for watch_list_file in glob.glob(os.path.join(watch_list_file, '*.xlsx')):

    wb_watch_list = load_workbook(watch_list_file)
    ws_watch_list = wb_watch_list.active
    # Excelファイルを読み込みます
    print(watch_list_file)
    last_row = ws_watch_list.max_row
    #print('最終行は'+str(last_row))
    last_col = ws_watch_list.max_column
    #print('最終列は'+str(last_col))

    columns = ['D', 'E', 'L', 'M', 'N', 'O']
    
    for column in columns:


        column_index = column_index_from_string(column)


    # 最終行の数値データを数値として認識させる
        #print(column)
        cell_value = ws_watch_list.cell(row=last_row, column=column_index).value  # 最終行のセルの値を取得

# 数値データを数値型に変換
        if isinstance(cell_value, str):
            try:
                cell_value = int(cell_value)
                print("形式を変更しました。")
            except ValueError:
                pass

        # 変換後の数値をセルに設定
        ws_watch_list.cell(row=last_row, column=column_index).value = cell_value

    # Excelファイルを保存
    wb_watch_list.save(watch_list_file)


print(t)
t = datetime.datetime.now().time()
print(t)

#稼働終了アナウンス
import winsound
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾