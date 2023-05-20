'''#コメントここから
これは追跡開始初日データ処理を実施するコードで、過去のデータ用になります。
対象フォルダ
作業ファイルフォルダ：'C:/Users/touko/OneDrive/株価分析/excel/株式データ/追跡調査/過去/0日目/'
参照ファイルフォルダ：'C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/'
格納ファイルフォルダ：'C:/Users/touko/OneDrive/株価分析/excel/株式データ/追跡調査/過去/1日目/'

実施する処理
OSCIファイルについて、最終列の日付を取得する。
OSCIファイルの2列目の行の証券コードをキーにして、銘柄データフォルダを検索。
ヒットした銘柄データファイルを開く。
OSCIファイルの最終列の日付をキーとして、銘柄データをスキャンし、ヒットした行から、それぞれデータを取得し、OSCIファイルに転記。
始値、終値、終値-始値の値により、網掛けの色を変える、利益%が±2%を越えたら、TRUE/FALSE記載。
10行分、OSCIファイルの最終列が43に達したら、ファイルを格納ファイルフォルダに移す。
次のOSCIファイルに移る。
'''#コメントここまで

import os
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import datetime
import jpholiday

import shutil


t = datetime.datetime.now().time()
d = datetime.datetime.now()
d1 = d.strftime('%Y%m%d')

def find_date_row(ws, date):
    for row2 in range(2, ws.max_row + 1):
        #print(code)
        if ws.cell(row=row2, column=1).value == date:
            return row2
    return None

watch_list_folder = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/追跡調査/過去/2日目以降/'
stock_folder = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/'
dirstorage = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/test/20230518/期間満了/'
n_columns = 43
plus_fill = PatternFill(patternType='solid', fgColor='ee82ee') #前日より値上がり
minus_fill = PatternFill(patternType='solid', fgColor='00bfff') #前日より値下がり
attained_fill = PatternFill(patternType='solid', fgColor='adff2f') #+2%目標を達成したら
unattained_fill = PatternFill(patternType='solid', fgColor='696969') #+2%目標未達で、5日目を迎えたら


file_list = glob.glob(watch_list_folder+'*.xlsx')

for l in file_list:
    file_name = os.path.basename(l)
    print(str(file_name)+"に対し作業中。")
    wb_watch_list = load_workbook(l)
    ws_watch_list = wb_watch_list.active

    last_row_watch = ws_watch_list.max_row

    last_col_watch = ws_watch_list.max_column


    last_date = ws_watch_list.cell(row=1, column=last_col_watch).value
    basis_date = ws_watch_list.cell(row=1, column=last_col_watch).value
        

    
    for i in range(2, last_row_watch + 1):
        
        stock_code = ws_watch_list.cell(row=i, column=4).value
        company_name= ws_watch_list.cell(row=i, column=5).value
        stock_data_file = os.path.join(stock_folder, f'{stock_code}_{company_name}.xlsx')
        print(str(stock_data_file) +"を検索中")
        k=last_col_watch
        if os.path.exists(stock_data_file):
            wb_stock_data = load_workbook(stock_data_file)
            ws_stock_data = wb_stock_data.active
            last_row_stock=ws_stock_data.max_row
            today_date =ws_stock_data.cell(row=last_row_stock, column=1).value
            date_row = find_date_row(ws_stock_data, last_date)

            if date_row is None:
                continue
            
            for j in range(date_row, date_row+10):

                stock_date = ws_stock_data.cell(row=j,column=1).value
                ws_watch_list.cell(row=1, column=k).value = stock_date
                closing_price = ws_stock_data.cell(row=j, column=4).value
                ws_watch_list.cell(row=i, column=k).value = closing_price
                
                
                start_price = ws_watch_list.cell(row=i, column=29).value

                if ws_stock_data.cell(row=date_row, column=1).value is None:
                    print("証券コード："+str(stock_code)+"は上場廃止になったかも。")
                    continue
                diff=closing_price - ws_watch_list.cell(row=i, column=k-1).value
                if diff > 0:
                    ws_watch_list.cell(row=i, column=k).fill = plus_fill
                elif diff < 0:
                    ws_watch_list.cell(row=i,column=k).fill = minus_fill

                if closing_price is not None and start_price is not None:
                    ratio = (diff / start_price)*100
                    ws_watch_list.cell(row=i, column=31).value = ratio
                    ws_watch_list.cell(row=i, column=30).value = closing_price - start_price
                    if ratio > 2:
                        ws_watch_list.cell(row=i, column=1).value = True
                        for m in range(1, last_col_watch+1):
                            ws_watch_list.cell(row=i, column=m).fill =attained_fill

                        profit = closing_price - start_price
                        
                    elif ratio < -2:
                        ws_watch_list.cell(row=i, column=1).value = False
                        for m in range(1, last_col_watch+1):
                            ws_watch_list.cell(row=i, column=m).fill =unattained_fill
                        loss = closing_price - start_price
                        
                
                k += 1

                
                last_col_last = ws_watch_list.max_column


        

            wb_stock_data.close() 
            print(str(stock_data_file) +"を閉じました。")       


    last_col_last = ws_watch_list.max_column
    
    wb_watch_list.save(l)
    print(l+'を保存しました')
    last_col_last = ws_watch_list.max_column
    wb_watch_list.close()
    shutil.move(l, dirstorage)


    
print(t)
t = datetime.datetime.now().time()
print(t)

#稼働終了アナウンス
import winsound
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾