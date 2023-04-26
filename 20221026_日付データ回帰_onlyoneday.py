#重複日付の削除
#2022/10/26
#銘柄データフォルダに適用
#銘柄データnを開いたら、最終行取得
#2行目から最終行まで繰り返し
#i行目の日付とi+j行目の日付が同じなら、i+j行目を削除

import os
from re import L
import openpyxl
import pandas
import requests
import bs4
import time
import datetime
import glob
import re
import sys
import winsound

#開始時間取得
t = datetime.datetime.now()
d = datetime.datetime.now()
d1 = d.strftime('%Y%m%d')
#開始時間取得

#対象フォルダ指定
dirdaily = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/"
dirmerge = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/"

file_list = glob.glob(dirdaily + '*.xlsx')
stock_list = glob.glob(dirmerge + '*.xlsx')

#銘柄データのリストをglob関数で作成
for l in stock_list:
    wb_stock = openpyxl.load_workbook(l)
    print(l)
    sheetstock = wb_stock.worksheets[0]
    lastrow = sheetstock.max_row+1
    lastcolumn = 330

    for i in range(2, lastrow):
        daycode = sheetstock.cell(i,1).value
        stockcode = sheetstock.cell(i,2).value
        print(daycode)
        daycode_format = daycode.strftime('%Y%m%d')
#        print(daycode_format)
        book_search_list = glob.glob(dirdaily+str(daycode_format)+'*.xlsx')
        dailybook = book_search_list[0]
        wb_daily = openpyxl.load_workbook(dailybook)
        sheetdaily = wb_daily.worksheets[0]
        lastrow2=sheetdaily.max_row+1

        for j in range(2, lastrow2):
            daycode_same = sheetdaily.cell(j,1).value
            stockcode_same = sheetdaily.cell(j,2).value
            if daycode == daycode_same and stockcode == stockcode_same:
#                print(daycode_same)
                for k in range(1,lastcolumn):
                    row_copy = sheetstock.cell(i,k).value
                    sheetdaily.cell(j,k,value=row_copy)
                    k += 1
            else:
                pass
        wb_daily.save(dailybook)



#終了時間取得-経過時間
print(t)
t1 = datetime.datetime.now()
print(t1)
dt = t1-t
print(dt)
#終了時間取得-経過時間

#稼働終了アナウンス
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾