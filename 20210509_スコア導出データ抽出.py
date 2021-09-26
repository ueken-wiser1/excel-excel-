import os
import openpyxl
import requests
import bs4
import time
import datetime

#どんな動きをさせるのか
#excelを開く
#参照excelの各シートに記載された証券コードを読み込む
#シート名は"株式", "マーケット", "為替", "投信"
#読み込んだ証券コードにより株探URLを作成し、銘柄ページへ移動
#必要な情報を取得する

#excelを開く

wb = openpyxl.load_workbook('screeningcodelist.xlsx')
name = wb.get_sheet_names

#"株式"
sheet = wb.get_sheet_by_name('株式')
print(sheet.title)
print(name)

#読み込んだ証券コードにより株探URLを作成し、銘柄ページへ移動
#必要な情報を取得する

#日付を入れる
d_today = datetime.date.today()
sheet.cell(row=1, column=4).value = d_today

#参照excelの各シートに記載された証券コードを読み込む

print(sheet)
for j in range(1, 250):
    time.sleep(0.2)
    kabutan_URL_base = 'http://kabutan.jp/stock/?code='
    codearea=10*j-7
    stock_code = sheet.cell(row=j, column=2).value
    kabutan_URL = kabutan_URL_base + str(stock_code)
    res = requests.get(kabutan_URL)
    print(kabutan_URL)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')

#読み込んだ情報をexcelファイルに書き込む
#書き出す初めのセルの列を指定
    write_row = codearea - 1
    write_column = 4
    
    #           証券コード
    sheet.cell(row=write_row, column=write_column).value = stock_code
    print(write_row)
    write_row += 1

#           始値
    try:
        stock_price = str(soup.select('td')[23])
    except IndexError as e:
        print('始値存在しない')
        continue
    else:
        sheet.cell(row=write_row, column=write_column).value = stock_price
    print(write_row)
    write_row += 1

#           高値
    try:
        stock_price = str(soup.select('td')[26])
    except IndexError as e:
        print('高値存在しない')
        continue
    else:
        sheet.cell(row=write_row, column=write_column).value = stock_price
    write_row += 1

#           安値
    try:
        stock_price = str(soup.select('td')[29])
    except IndexError as e:
        print('安値存在しない')
        continue
    else:
        sheet.cell(row=write_row, column=write_column).value = stock_price
    write_row += 1

#           終値
    try:
        stock_price = str(soup.select('td')[32])
    except IndexError as e:
        print('終値存在しない')
        continue
    else:
        sheet.cell(row=write_row, column=write_column).value = stock_price
    write_row += 1

#           前日比
    try:
        span_pick = str(soup.select('span')[15])
    except IndexError as e:
        print('前日比存在しない')
        continue
    else:
        sheet.cell(row=write_row, column=write_column).value = span_pick
    write_row += 1
    
#           前日比%
    try:
        span_pick = str(soup.select('span')[16])
    except IndexError as e:
        print('前日比存在しない')
        continue
    else:
        sheet.cell(row=write_row, column=write_column).value = span_pick
    write_row += 1
    

#           出来高
    try:
        span_pick = str(soup.select('td')[35])
    except IndexError as e:  
        print('出来高存在しない')
        continue
    else:
        sheet.cell(row=write_row, column=write_column).value = span_pick
    write_row += 1

#           値幅
    try:
        stock_price01 = soup.select('td')[26]
        stock_price02 = soup.select('td')[29]
        stock_price03 = '-'
        stock_price04 = '壁'
        print(stock_price01)
        print(stock_price02)
        price_span = str(stock_price04)+str(soup.select('td')[26])+str(stock_price03)+str(soup.select('td')[29])
    except IndexError as e:
        print('出来高存在しない')
        continue
    else:
        sheet.cell(row=write_row, column=write_column).value = str(price_span)
    write_row += 1



wb.save('screeningkabu2.xlsx')

