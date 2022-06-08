import os
import openpyxl
import requests
import bs4
import time
import datetime
import sys
import sys
from selenium import webdriver
from selenium.webdriver.common.keys import Keys as keys
import winsound

#------------お約束開始---冒頭
#稼働時間計測開始
import datetime
t = datetime.datetime.now().time()
#------------お約束終了---冒頭

'''
仕様


'''

#------------プログラム本文---ここから
today = datetime.date.today()
d = today.strftime('%Y%m%d')
dir = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/スクレイピング生データ/"
wb = openpyxl.load_workbook(dir + 'stockcodelist.xlsx')
name = wb.get_sheet_names

#"株式"
sheet = wb.get_sheet_by_name('株式')
t = datetime.datetime.now().time()



for j in range(2, sheet.max_row + 1):
    time.sleep(0.1)
    kabutan_URL_base = 'http://kabutan.jp/stock/?code='
    stock_code = sheet.cell(row=j, column=2).value
    kabutan_URL = kabutan_URL_base + str(stock_code)
    res = requests.get(kabutan_URL)
    print(kabutan_URL)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')

#読み込んだ情報をexcelファイルに書き込む
#書き出す初めのセルの列を指定
    write_column = 1

#日付を入れる
    d_today = datetime.date.today()
    sheet.cell(row=j, column=1).value = d_today
    write_column += 2
#    t = soup.select('time')[4]
#    print(t)
#           名称
    stock_name = soup.select('h3')[0]
    print(stock_name)
    sheet.cell(row=j, column=3).value = str(soup.select('h3')[0].get_text())
    write_column += 1

#           現在株価
    try:
        stock_price = str(soup.select('td')[32])
    except IndexError as e:
        print('現在株価存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
        write_column += 1

#           前日比
    try:
        DoD = str(soup.select('dd')[0])
    except IndexError as e:
        print('前日比存在しない')
        write_column += 2
        pass
    else:
        sheet.cell(row=j, column=write_column).value = DoD
        write_column += 2

#           売買代金
    try:
        dekidaka_yield = str(soup.select('td')[36].get_text())
    except IndexError as e:
        print('売買代金存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = dekidaka_yield
        write_column += 1

#           約定回数
    try:
        number_of_contracts = str(soup.select('td')[38].get_text())
    except IndexError as e:
        print('約定回数存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = number_of_contracts
        write_column += 1

#           決算日
    try:
        stock_price = str(soup.select('dd')[7])
    except IndexError as e:
        print('前日終値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
        write_column += 1

#           前日終値1
    try:
        stock_price = str(soup.select('dd')[7])
    except IndexError as e:
        print('前日終値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
        write_column += 1

#           前日終値2
    try:
        stock_price = str(soup.select('dd')[8].get_text())
    except IndexError as e:
        print('前日終値存在しない')
        write_column += 2
        pass
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
        write_column += 2

#           出来高
    try:
        dekidaka_yield = str(soup.select('td')[35].get_text())
    except IndexError as e:
        print('出来高存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = dekidaka_yield
        write_column += 1

#           始値
    try:
        stock_price = str(soup.select('td')[23])
    except IndexError as e:
        print('始値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
        write_column += 1

#           高値
    try:
        stock_price = str(soup.select('td')[26])
    except IndexError as e:
        print('高値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
        write_column += 1

#           安値
    try:
        stock_price = str(soup.select('td')[29])
    except IndexError as e:
        print('安値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
        write_column += 1

#           終値
    try:
        stock_price = str(soup.select('td')[32])
    except IndexError as e:
        print('終値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
        write_column += 1

#           前日比%
    try:
        DoD = str(soup.select('span')[16].get_text())
    except IndexError as e:
        print('前日比存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = DoD
        write_column += 1

#           PER
    try:
        kabu_PER = str(soup.select('td')[18].get_text())
    except IndexError as e:
        print('PER存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = kabu_PER
        write_column += 1

#           PBR
    try:
        kabu_PBR = str(soup.select('td')[19].get_text())
    except IndexError as e:
        print('PBR存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = kabu_PBR
        write_column += 1

#           上場市場
    try:
        stock_price = str(soup.select('span')[13].get_text())
    except IndexError as e:
        print('上場市場存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
        write_column += 1

#           VWAP
    try:
        VWAP = str(soup.select('td')[37].get_text())
    except IndexError as e:
        print('VWAP存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = VWAP
        write_column += 1

#           発行済み株式数
    try:
        number_of_issued_shares = str(soup.select('td')[42].get_text())
    except IndexError as e:
        print('発行済み株式数存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = number_of_issued_shares
        write_column += 1

#           最新信用売残
    try:
        credit_unsold = str(soup.select('td')[46])
    except IndexError as e:
        print('最新信用売残存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = credit_unsold
        write_column += 1

#           最新信用買残
    try:
        credit_unpurchased = str(soup.select('td')[47])
    except IndexError as e:
        print('最新信用買残存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = credit_unpurchased
        write_column += 1

#           信用倍率
    try:
        credit_ratio = str(soup.select('td')[48])
    except IndexError as e:
        print('信用倍率存在しない')
        write_column += 1
        pass
    else:
        
        sheet.cell(row=j, column=write_column).value = credit_ratio
        write_column += 1
    


wb.save('C:/Users/touko/OneDrive/株価分析/excel/株式データ/allkabu1.xlsx')
#------------プログラム本文---ここまで

#------------お約束開始---末尾
#稼働時間表示
print(t)
t = datetime.datetime.now().time()
print(t)

#稼働終了アナウンス
import winsound
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾