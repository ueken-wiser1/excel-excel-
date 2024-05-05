#使ってない証券コードのデータを削除するプログラム
#株式会社の証券コードでないコードのリストはできた
#上記リストにあるコードを持つファイル/データがあれば、それを削除する
#結果：市場データ、銘柄データから、使われていない証券コードデータが消えている
#1. 不使用証券コードを取得する
#2. 市場データを読み込む
#3. 不使用証券コードと市場データの銘柄コードを比較して、同じであればその行を削除する
#4. 不使用証券コードと銘柄データのあるフォルダ内のファイル名を比較し、同じモノがあればそのファイルを削除する
import os
import openpyxl
import time
import sys
import winsound
import glob

#------------お約束開始---冒頭
#稼働時間計測開始
import datetime
t = datetime.datetime.now().time()
#------------お約束終了---冒頭

#------------プログラム本文---ここから

#使用フォルダ指定
folder01 = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/"

#銘柄データを順に開く
filelist01 = glob.glob(folder01+"*.xlsx")

for l in filelist01:
    t1 = datetime.datetime.now().time()
    # print(t1)
    stockbook = openpyxl.load_workbook(l)
    sheet01 = stockbook.worksheets[0]
    print(l)
    lastrow_stockbook =sheet01.max_row
    # cell01 = sheet01.cell(row=144,column=1).value
    # print(type(cell01),cell01)
    for i in reversed(range(2,lastrow_stockbook+1)):
        cell01 = sheet01.cell(row=i,column=1).value
        if cell01 is None:
            sheet01.delete_rows(i)
        elif cell01 == datetime.datetime(2023,7,10):
            sheet01.delete_rows(i)
        else:
            pass
    stockbook.save(l)



#------------プログラム本文---ここまで

#------------お約束開始---末尾
#稼働時間表示
print(t)
t = datetime.datetime.now().time()
print(t)

#稼働終了アナウンス
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾