#売買sim
#2022/10/28
#銘柄データフォルダに適用
#銘柄データnを開いたら、最終行取得
#2行目から最終行まで繰り返し
#i行目の日付とi+j行目の日付が同じなら、i+j行目を削除

import os
from re import L
import openpyxl
import pandas
import requests
import bs4
import time
import datetime
import glob
import re
import sys
import winsound

#開始時間取得
t = datetime.datetime.now()
d = datetime.datetime.now()
d1 = d.strftime('%Y%m%d')
#開始時間取得


#対象フォルダ指定
dirdaily = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/test/20221027/日付/"
dirmerge = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/test/20221027/銘柄/"
dirsim = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/test/20221027/sim/"

file_list = glob.glob(dirdaily + '*.xlsx')
stock_list = glob.glob(dirmerge + '*.xlsx')
sim_list = glob.glob(dirsim + '*.xlsx')

#simデータのリストをglob関数で作成
for l in sim_list:
    #simデータを順番に開く
    wb_sim = openpyxl.load_workbook(l)
    sheetsim = wb_sim.worksheets[0]
    k=5
    print(l)
    lastrow = sheetsim.max_row+1
    lastcolumn = sheetsim.max_column
    basename = os.path.basename(l)
    daycode_format = basename[:8]
#    print(daycode_format)
    for i in range(2,lastcolumn):
        if sheetsim.cell(1,i).value == '売り' or sheetsim.cell(1,i).value is None:
            pass
        else:
            stockcode = sheetsim.cell(2,i).value
            book_search_list = glob.glob(dirmerge+stockcode+'_'+'*.xlsx')
            stockbook = book_search_list[0]
            wb_stock = openpyxl.load_workbook(stockbook)
            sheetstock = wb_stock.worksheets[0]
            lastrow_stock = sheetstock.max_row+1
            for j in range(2,lastrow_stock):
                if sheetsim.cell(1,i).value == 'プラス越え':
                    daycode_stock = sheetstock.cell(j,1).value
                    daycode_stock_format = daycode_stock.strftime('%Y%m%d')
                    if daycode_stock_format>daycode_format and sheetsim.cell(1,i).value == 'プラス越え':
                        sheetsim.cell(k,1).value = daycode_stock
                        sheetsim.cell(k,i).value = sheetstock.cell(j,12).value
                        sheetsim.cell(1,i).value = '買い'
                        k += 1
                        continue
                    else:
                        pass
                elif sheetsim.cell(1,i).value == '買い':
                    daycode_stock = sheetstock.cell(j+1,1).value
                    daycode_stock_format = daycode_stock.strftime('%Y%m%d')
                    if daycode_stock_format>daycode_format and sheetsim.cell(1,i).value == '買い':
                        sheetsim.cell(k,1).value = daycode_stock
                        sheetsim.cell(k,i).value = sheetstock.cell(j+1,15).value
                        sheetsim.cell(1,i).value = '保有'
                        k += 1
                        continue
                    else:
                        pass
                elif sheetsim.cell(1,i).value == '保有':
                    daycode_stock = sheetstock.cell(j+1,1).value
                    daycode_stock_format = daycode_stock.strftime('%Y%m%d')
                    if daycode_stock_format>daycode_format and sheetsim.cell(1,i).value == '保有':
                        sheetsim.cell(k,1).value = daycode_stock
                        sheetsim.cell(k,i).value = sheetstock.cell(j+1,15).value
                        sheetsim.cell(1,i).value = '保有'
                        k += 1
                        continue
                    else:
                        pass







#終了時間取得-経過時間
print(t)
t1 = datetime.datetime.now()
print(t1)
dt = t1-t
print(dt)
#終了時間取得-経過時間

#稼働終了アナウンス
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾