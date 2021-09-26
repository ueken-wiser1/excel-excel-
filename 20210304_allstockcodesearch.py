import requests
import bs4
import os
import openpyxl
import time

#どんな動きをさせるのか
#excelを開く
#0000-9999まで証券コードを走査して、該当した銘柄のデータを抽出する
#エラーが出た場合はスキップ
#抽出するデータはこれまでと同じ

#excelを開く
wb = openpyxl.load_workbook('allkabu.xlsx')
#print(type(wb))
name = wb.get_sheet_names
#print(name)
#print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name('Sheet1')
print(sheet.title)
print(name)
#0000-9999までの証券コードを分解する

#0-9→'000'+0-9
for code_number in range(10):
    time.sleep(0.1)
#    sheet = wb.get_sheet_by_name('Sheet1')
    stock_code = '000' + str(code_number)
#    print(stock_code)
#    print(sheet.title)

    kabutan_URL_base = 'http://kabutan.jp/stock/?code='
    kabutan_URL = kabutan_URL_base + str(stock_code)
    print(kabutan_URL)

#銘柄ページから情報を読み込む

    res = requests.get(kabutan_URL)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')
#    wb = openpyxl.load_workbook('allkabu1.xlsx')
#    name = wb.get_sheet_names

#読み込んだ情報をexcelファイルに書き込む
#           コード
    row_number = code_number + 2
    write_column = 1
    sheet.cell(row=row_number, column=write_column).value = str(stock_code)
    print(row_number)
    print(write_column)
    print(sheet.cell(row=row_number, column=write_column).value)
    print(sheet.title)
    print(name)

    write_column += 1
#    wb.save('allkabu1.xlsx')
#           名称
    stock_name = soup.select('title')[0]
    print(stock_name)
    sheet.cell(row=row_number, column=write_column).value = str(soup.select('title')[0])
    write_column += 1
#           現在株価
    try:
        span_pick = str(soup.select('span')[14])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('span')[14])
    write_column += 1
#           PER
    try:
        span_pick = str(soup.select('td')[18])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[18])
    write_column += 1
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[18])
#    write_column = write_column + 1
#           PBR
    try:
        span_pick = str(soup.select('td')[19])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[19])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[19])
#    write_column = write_column + 1
#           利回り
    try:
        span_pick = str(soup.select('td')[20])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[20])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[20])
#    write_column = write_column + 1
#           出来高
    try:
        span_pick = str(soup.select('td')[35])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[35])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[35])
#    write_column = write_column + 1
#           時価総額
    try:
        span_pick = str(soup.select('td')[41])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[41])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[41])
#    write_column = write_column + 1
#           発行済み株式数
    try:
        span_pick = str(soup.select('td')[42])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[42])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[42])
#    write_column = write_column + 1
#           最新信用売残
    try:
        span_pick = str(soup.select('td')[46])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[46])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[46])
#    write_column = write_column + 1
#           最新信用買残
    try:
        span_pick = str(soup.select('td')[47])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[47])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[47])
#    write_column = write_column + 1
#           1株配当
    try:
        span_pick = str(soup.select('td')[82])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[82])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[82])
#    write_column = write_column + 1
#           決算タブに移動
#    kabutan_URL_base_finance = 'http://kabutan.jp/stock/finance?code='
#    kabutan_URL_finance = kabutan_URL_base_finance + str(stock_code)
#    res = requests.get(kabutan_URL_finance)
#    res.raise_for_status()
#    soup = bs4.BeautifulSoup(res.text, 'html.parser')
#    tag_title = soup.find_all('title')
#           総資産回転率
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[226])
#    pass
#    write_column = write_column + 1
#           1秒待機
wb.save('allkabu1.xlsx')
#time.sleep(1)
code_number += 1

#10-99→'00'+10-99
for code_number in range(10, 99):
    time.sleep(0.1)
    stock_code = '00' + str(code_number)
    row_number = code_number + 2
    write_column = 1
    sheet.cell(row=row_number, column=write_column).value = stock_code
    write_column += 1
    kabutan_URL_base = 'http://kabutan.jp/stock/?code='
    kabutan_URL = kabutan_URL_base + str(stock_code)

#銘柄ページから情報を読み込む

    res = requests.get(kabutan_URL)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')
#    wb = openpyxl.load_workbook('allkabu1.xlsx')
#    name = wb.get_sheet_names

#読み込んだ情報をexcelファイルに書き込む
#           名称
    stock_name = soup.select('title')[0]
    print(stock_name)
    sheet.cell(row=row_number, column=write_column).value = str(soup.select('title')[0])
    write_column += 1
#           現在株価
    try:
        span_pick = str(soup.select('span')[14])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('span')[14])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('span')[14])
        write_column += 1
#           PER
    try:
        span_pick = str(soup.select('td')[18])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[18])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[18])
        write_column += 1
#           PBR
    try:
        span_pick = str(soup.select('td')[19])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[19])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[19])
        write_column += 1
#           利回り
    try:
        span_pick = str(soup.select('td')[20])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[20])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[20])
        write_column += 1
#           出来高
    try:
        span_pick = str(soup.select('td')[35])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[35])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[35])
        write_column += 1
#           時価総額
    try:
        span_pick = str(soup.select('td')[41])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[41])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[41])
        write_column += 1
#           発行済み株式数
    try:
        span_pick = str(soup.select('td')[42])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[42])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[42])
        write_column += 1
#           最新信用売残
    try:
        span_pick = str(soup.select('td')[46])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[46])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[46])
        write_column += 1
#           最新信用買残
    try:
        span_pick = str(soup.select('td')[47])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[47])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[47])
        write_column += 1
#           1株配当
    try:
        span_pick = str(soup.select('td')[82])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[82])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[82])
        write_column += 1
#           決算タブに移動
#    kabutan_URL_base_finance = 'http://kabutan.jp/stock/finance?code='
#    kabutan_URL_finance = kabutan_URL_base_finance + str(stock_code)
#    res = requests.get(kabutan_URL_finance)
#    res.raise_for_status()
#    soup = bs4.BeautifulSoup(res.text, 'html.parser')
#    tag_title = soup.find_all('title')
#           総資産回転率
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[226])
#    write_column += 1
#           1秒待機
#    time.sleep(1)
    code_number += 1
    wb.save('allkabu1.xlsx')

#100-999→'00'+100-999
for code_number in range(100, 999):
    time.sleep(0.1)
    stock_code = '0' + str(code_number)
    row_number = code_number + 2
    write_column = 1
    sheet.cell(row=row_number, column=write_column).value = stock_code
    write_column += 1
    kabutan_URL_base = 'http://kabutan.jp/stock/?code='
    kabutan_URL = kabutan_URL_base + str(stock_code)

#銘柄ページから情報を読み込む

    res = requests.get(kabutan_URL)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')
#    wb = openpyxl.load_workbook('allkabu1.xlsx')
#    name = wb.get_sheet_names

#読み込んだ情報をexcelファイルに書き込む
#           名称
    stock_name = soup.select('title')[0]
    print(stock_name)
    sheet.cell(row=row_number, column=write_column).value = str(soup.select('title')[0])
    write_column += 1
#           現在株価
    try:
        span_pick = str(soup.select('span')[14])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('span')[14])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('span')[14])
        write_column += 1
#           PER
    try:
        span_pick = str(soup.select('td')[18])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[18])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[18])
        write_column += 1
#           PBR
    try:
        span_pick = str(soup.select('td')[19])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[19])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[19])
        write_column += 1
#           利回り
    try:
        span_pick = str(soup.select('td')[20])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[20])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[20])
        write_column += 1
#           出来高
    try:
        span_pick = str(soup.select('td')[35])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[35])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[35])
        write_column += 1
#           時価総額
    try:
        span_pick = str(soup.select('td')[41])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[41])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[41])
        write_column += 1
#           発行済み株式数
    try:
        span_pick = str(soup.select('td')[42])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[42])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[42])
        write_column += 1
#           最新信用売残
    try:
        span_pick = str(soup.select('td')[46])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[46])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[46])
        write_column += 1
#           最新信用買残
    try:
        span_pick = str(soup.select('td')[47])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[47])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[47])
        write_column += 1
#           1株配当
    try:
        span_pick = str(soup.select('td')[82])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[82])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[82])
        write_column += 1
#           決算タブに移動
#    kabutan_URL_base_finance = 'http://kabutan.jp/stock/finance?code='
#    kabutan_URL_finance = kabutan_URL_base_finance + str(stock_code)
#    res = requests.get(kabutan_URL_finance)
#    res.raise_for_status()
#    soup = bs4.BeautifulSoup(res.text, 'html.parser')
#    tag_title = soup.find_all('title')
#           総資産回転率
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[226])
#    write_column += 1
#           1秒待機
#    time.sleep(1)
    code_number += 1
    wb.save('allkabu1.xlsx')

#1000-9999→+1000-9999
for code_number in range(1000, 9999):
    time.sleep(0.2)
    stock_code = str(code_number)
    row_number = code_number + 2
    write_column = 1
    sheet.cell(row=row_number, column=write_column).value = stock_code
    write_column += 1
    kabutan_URL_base = 'http://kabutan.jp/stock/?code='
    kabutan_URL = kabutan_URL_base + str(stock_code)

#銘柄ページから情報を読み込む

    res = requests.get(kabutan_URL)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')
#    wb = openpyxl.load_workbook('allkabu1.xlsx')
#    name = wb.get_sheet_names

#読み込んだ情報をexcelファイルに書き込む
#           名称
    stock_name = soup.select('title')[0]
    print(stock_name)
    sheet.cell(row=row_number, column=write_column).value = str(soup.select('title')[0])
    write_column += 1
#           現在株価
    try:
        span_pick = str(soup.select('span')[14])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('span')[14])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('span')[14])
        write_column += 1
#           PER
    try:
        span_pick = str(soup.select('td')[18])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[18])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[18])
        write_column += 1
#           PBR
    try:
        span_pick = str(soup.select('td')[19])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[19])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[19])
        write_column += 1
#           利回り
    try:
        span_pick = str(soup.select('td')[20])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[20])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[20])
        write_column += 1
#           出来高
    try:
        span_pick = str(soup.select('td')[35])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[35])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[35])
        write_column += 1
#           時価総額
    try:
        span_pick = str(soup.select('td')[41])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[41])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[41])
        write_column += 1
#           発行済み株式数
    try:
        span_pick = str(soup.select('td')[42])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[42])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[42])
        write_column += 1
#           最新信用売残
    try:
        span_pick = str(soup.select('td')[46])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[46])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[46])
        write_column += 1
#           最新信用買残
    try:
        span_pick = str(soup.select('td')[47])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[47])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[47])
        write_column += 1
#           1株配当
    try:
        span_pick = str(soup.select('td')[82])
    except IndexError as e:
        print('存在しない')
        continue
    else:
        sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[82])
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[82])
        write_column += 1
#           決算タブに移動
#    kabutan_URL_base_finance = 'http://kabutan.jp/stock/finance?code='
#    kabutan_URL_finance = kabutan_URL_base_finance + str(stock_code)
#    res = requests.get(kabutan_URL_finance)
#    res.raise_for_status()
#    soup = bs4.BeautifulSoup(res.text, 'html.parser')
#    tag_title = soup.find_all('title')
#           総資産回転率
#    sheet.cell(row=row_number, column=write_column).value = str(soup.select('td')[226])
#    write_column += 1
#           1秒待機

#    time.sleep(1)
    code_number += 1
    wb.save('allkabu1.xlsx')
