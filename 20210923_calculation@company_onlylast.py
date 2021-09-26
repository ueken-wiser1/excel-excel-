import os
from re import L
import openpyxl
import requests
import bs4
import time
import datetime
import sys
import winsound
import glob

#どんな動きをさせるのか
#excelを開く
#参照excelの各シートに記載された証券コードを読み込む
#シート名は"株式", "マーケット", "為替", "投信"
#読み込んだ証券コードにより株探URLを作成し、銘柄ページへ移動
#必要な情報を取得する
t = datetime.datetime.now().time()
#excelを開く
#対象：信用規制
dir_stock = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/"
#ダウンロードしたexcel(以下databook)を開く
stock_list = glob.glob(dir_stock + '*.xlsx')

for l in stock_list:
    stockbook = openpyxl.load_workbook(l)
#    print(str(stockbook))
    sheet01 = stockbook.worksheets[0]
    lastrow = sheet01.max_row+1
#売買代金2=VWAP(T)*出来高(K)
    sheet01.cell(row=lastrow,column=25).value = str('=T')+str(lastrow)+str('*K')+str(lastrow)
#信用売残前週比：(V)
    sheet01.cell(row=lastrow,column=26).value = str('=V')+str(lastrow)+str('*V')+str(lastrow-1)
#信用買残前週比：(W)
    sheet01.cell(row=lastrow,column=27).value = str('=W')+str(lastrow)+str('*W')+str(lastrow-1)
#出来高前日比：(K)
    sheet01.cell(row=lastrow,column=28).value = str('=K')+str(lastrow)+str('*K')+str(lastrow-1)
#約定回数前日比：(H)
    sheet01.cell(row=lastrow,column=29).value = str('=H')+str(lastrow)+str('*H')+str(lastrow-1)
#平均約定金額平均約定金額=出来高(K)/約定回数(H)*VWAP(T)
    sheet01.cell(row=lastrow,column=33).value = str('=K')+str(lastrow)+str('/H')+str(lastrow)+str('*T')+str(lastrow)
#回転日数=((融資残(DI)+貸株残(DL))*2)/(融資新規(DG)+融資返済(DH)+貸株新規(Dlastrow)+貸株返済(DK)
    sheet01.cell(row=lastrow,column=118).value = str('=(DI')+str(lastrow)+str('+DL')+str(lastrow)+str(')*2/(DG')+str(lastrow)+str('+DH')+str(lastrow)+str('+Dlastrow')+str(lastrow)+str('+DK')+str(lastrow)+str(')')
#現在株価との差=株価(D)-みんかぶ目標株価(EU)
    sheet01.cell(row=lastrow,column=152).value = str('=D')+str(lastrow)+str('-EU')+str(lastrow)
#移動平均線数値　5日＝当日〜4日前の株価総和/5
    sheet01.cell(row=lastrow,column=201).value = str('=SUM(D')+str(lastrow-4)+str(':D')+str(lastrow)+str(')/5')
#移動平均線数値　25日＝当日〜24日前の株価総和/25
    sheet01.cell(row=lastrow,column=202).value = str('=SUM(D')+str(lastrow-24)+str(':D')+str(lastrow)+str(')/25')
#移動平均線数値　75日＝当日〜74日前の株価総和/75
    sheet01.cell(row=lastrow,column=203).value = str('=SUM(D')+str(lastrow-74)+str(':D')+str(lastrow)+str(')/75')
#移動平均乖離率    5日＝（株価ー移動平均5日）/移動平均5日
    sheet01.cell(row=lastrow,column=204).value = str('=(D')+str(lastrow)+str('-GS')+str(lastrow)+str(')/GS')+str(lastrow)
#移動平均乖離率    25日＝（株価ー移動平均25日）/移動平均25日
    sheet01.cell(row=lastrow,column=205).value = str('=(D')+str(lastrow)+str('-GT')+str(lastrow)+str(')/GT')+str(lastrow)
#移動平均乖離率    75日＝（株価ー移動平均75日）/移動平均75日
    sheet01.cell(row=lastrow,column=206).value = str('=(D')+str(lastrow)+str('-GU')+str(lastrow)+str(')/GU')+str(lastrow)
#標準化出来高＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
    sheet01.cell(row=lastrow,column=501).value = str('=STANDARDIZE(K')+str(lastrow)+str(',AVERAGE(K:K),STDEV.P(K:K))')+str(lastrow)
#標準化約定回数＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
    sheet01.cell(row=lastrow,column=502).value = str('=STANDARDIZE(H')+str(lastrow)+str(',AVERAGE(H:H),STDEV.P(H:H))')+str(lastrow)
#標準化回転日数＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
    sheet01.cell(row=lastrow,column=503).value = str('=STANDARDIZE(DN')+str(lastrow)+str(',AVERAGE(DN:DN),STDEV.P(DN:DN))')
#標準化移動平均乖離率5日＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
    sheet01.cell(row=lastrow,column=504).value = str('=STANDARDIZE(GV')+str(lastrow)+str(',AVERAGE(GV:GV),STDEV.P(GV:GV))')
#標準化移動平均乖離率25日＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
    sheet01.cell(row=lastrow,column=505).value = str('=STANDARDIZE(GW')+str(lastrow)+str(',AVERAGE(GW:GW),STDEV.P(GW:GW))')
#標準化移動平均乖離率75日＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
    sheet01.cell(row=lastrow,column=506).value = str('=STANDARDIZE(GX')+str(lastrow)+str(',AVERAGE(GX:GX),STDEV.P(GX:GX))')
#標準化信用買残＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
    sheet01.cell(row=lastrow,column=507).value = str('=STANDARDIZE(Y')+str(lastrow)+str(',AVERAGE(Y:Y),STDEV.P(Y:Y))')
#標準化融資新規＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
    sheet01.cell(row=lastrow,column=508).value = str('=STANDARDIZE(DG')+str(lastrow)+str(',AVERAGE(DG:DG),STDEV.P(DG:DG))')
#標準化融資返済＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
    sheet01.cell(row=lastrow,column=509).value = str('=STANDARDIZE(DH')+str(lastrow)+str(',AVERAGE(DH:DH),STDEV.P(DH:DH))')
#標準化信用売残＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
    sheet01.cell(row=lastrow,column=510).value = str('=STANDARDIZE(X')+str(lastrow)+str(',AVERAGE(X:X),STDEV.P(X:X))')
#標準化貸株新規＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
    sheet01.cell(row=lastrow,column=511).value = str('=STANDARDIZE(Dlastrow')+str(lastrow)+str(',AVERAGE(Dlastrow:Dlastrow),STDEV.P(Dlastrow:Dlastrow))')
#標準化貸株返済＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
    sheet01.cell(row=lastrow,column=512).value = str('=STANDARDIZE(DK')+str(lastrow)+str(',AVERAGE(DK:DK),STDEV.P(DK:DK))')
#標準化貸株超過＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
    sheet01.cell(row=lastrow,column=513).value = str('=STANDARDIZE(DQ')+str(lastrow)+str(',AVERAGE(DQ:DQ),STDEV.P(DQ:DQ))')
#標準化出来高変化率＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
    sheet01.cell(row=lastrow,column=514).value = str('=STANDARDIZE(AB')+str(lastrow)+str(',AVERAGE(AB:AB),STDEV.P(AB:AB))')
#標準化約定回数変化率＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
    sheet01.cell(row=lastrow,column=515).value = str('=STANDARDIZE(AC')+str(lastrow)+str(',AVERAGE(AC:AC),STDEV.P(AC:AC))')
#標準化信用買残変化率＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
    sheet01.cell(row=lastrow,column=516).value = str('=STANDARDIZE(AA')+str(lastrow)+str(',AVERAGE(AA:AA),STDEV.P(AA:AA))')
#標準化信用売残変化率＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
    sheet01.cell(row=lastrow,column=517).value = str('=STANDARDIZE(Z')+str(lastrow)+str(',AVERAGE(Z:Z),STDEV.P(Z:Z))')
#標準化平均約定金額＝standardize（当日の値，平均値（average(対象列)），標準偏差(stdevp(対象列)）
    sheet01.cell(row=lastrow,column=518).value = str('=STANDARDIZE(AG')+str(lastrow)+str(',AVERAGE(AG:AG),STDEV.P(AG:AG))')
#売り時/買い時スコア＝12*標準化出来高(SG)+15*標準化約定回数(SH)+6*標準化平均約定金額(SX)+3/標準化回転日数(SI)+9*標準化移動平均乖離率5日(Slastrow)+6*標準化信用買残(SM)+10*信用規制(CW)+8*増担規制(CX)+4*標準化融資新規(SN)+2/標準化融資返済(SO)+4/標準化信用売残(SP)+5*空売り規制(DB)+3/標準化貸株新規(SQ)+2*標準化貸株返済(SR)+1/標準化貸株超過(SS)+(20*IR)
    sheet01.cell(row=lastrow,column=1000).value = str('=12*SG')+str(lastrow)+str('+15*SH')+str(lastrow)+str('+6*SX')+str(lastrow)+str('+3/SI')+str(lastrow)+str('+9*Slastrow')+str(lastrow)+str('+6*SM')+str(lastrow)+str('+10*CW')+str(lastrow)+str('+8*CX')+str(lastrow)+str('+4*SN')+str(lastrow)+str('+2/SO')+str(lastrow)+str('+4/SP')+str(lastrow)+str('+5*DB')+str(lastrow)+str('+3/SQ')+str(lastrow)+str('+2*SR')+str(lastrow)+str('+1/SS')+str(lastrow)

    stockbook.save(l)
print(t)
t = datetime.datetime.now().time()
print(t)


winsound.Beep(500,50)  #ビープ音（500Hzの音を50msec流す）
