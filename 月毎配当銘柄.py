import os
import openpyxl
import requests
import bs4


wb = openpyxl.load_workbook('kabu.xlsx')
#print(type(wb))
#name = wb.get_sheet_names
#print(name)
#print(wb.get_sheet_names())

for i in range(1,13):
    print(i)
    a = '月配当'
    b = i
    c =str(b) + a
    sheet = wb.get_sheet_by_name(c)
    print(sheet)
        for j in range(1, sheet.max_row):
#            print(sheet.cell(row=j, column=1).value)
            d = 'http://kabutan.jp/stock/?code='
            e = sheet.cell(row=j, column=1).value
            f = d + str(e)
            res = requests.get(f)
            print(f)
            res.raise_for_status()
            soup = bs4.BeautifulSoup(res.text, 'html.parser')

            span_elem1 = soup.select('span')[8]
            span_elem1 = sheet.cell(row=j, column=2).value
            print(span_elem1)

#        span_elem2 = soup.select('span')[14]
#        span_elem2 = sheet.cell(row=j, column=2).value
#        print(span_elem2)
#for i in range(10):
#    print(i)
#    td_elem = soup.select('span')[i]
#    print(td_elem)
