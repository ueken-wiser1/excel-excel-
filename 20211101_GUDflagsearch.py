import os
import openpyxl
from openpyxl import cell
import xlrd
import time
import sys
import winsound
import glob

#------------お約束開始---冒頭
#稼働時間計測開始
import datetime
t = datetime.datetime.now().time()
#------------お約束終了---冒頭

#------------プログラム本文---ここから
#使用フォルダ指定
folder01 = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/完了/"
folder02 = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/集計結果/"

#記録ファイルを開く
record = openpyxl.Workbook()
recsheet = record.create_sheet(title='record')
j = 2
k = 2
#銘柄データを順に開く
filelist01 = glob.glob(folder01+"*.xlsx")
for l in filelist01:
    stockbook = openpyxl.load_workbook(l)
    sheet01 = stockbook.worksheets[0]
    lastrow_stockbook = sheet01.max_row
    lastcolumn_stockbook = sheet01.max_column
    for i in range(2,lastrow_stockbook+1):
        if sheet01.cell(row=i,column=36).value and sheet01.cell(row=i,column=38).value == 1:
            recsheet.cell(row=j,column=1).value = sheet01.cell(row=i,column=1).value
            recsheet.cell(row=j,column=2).value = sheet01.cell(row=i,column=2).value
            recsheet.cell(row=j,column=3).value = sheet01.cell(row=i,column=3).value
            recsheet.cell(row=j,column=4).value = sheet01.cell(row=i,column=51).value
            print('GU', recsheet.cell(row=j,column=1).value, recsheet.cell(row=j,column=2).value, recsheet.cell(row=j,column=3).value)
            j += 1

        else:
            pass
        if sheet01.cell(row=i,column=37).value and sheet01.cell(row=i,column=38).value == 1:
            recsheet.cell(row=k,column=5).value = sheet01.cell(row=i,column=1).value
            recsheet.cell(row=k,column=6).value = sheet01.cell(row=i,column=2).value
            recsheet.cell(row=k,column=7).value = sheet01.cell(row=i,column=3).value
            recsheet.cell(row=k,column=8).value = sheet01.cell(row=i,column=51).value
            print('GD', recsheet.cell(row=k,column=5).value, recsheet.cell(row=k,column=6).value, recsheet.cell(row=k,column=7).value)
            k += 1
        else:
            pass

record.save(folder02+'test.xlsx')

#------------プログラム本文---ここまで

#------------お約束開始---末尾
#稼働時間表示
print(t)
t = datetime.datetime.now().time()
print(t)

#稼働終了アナウンス
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾