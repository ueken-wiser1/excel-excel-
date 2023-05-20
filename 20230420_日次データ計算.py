import imp
import os
from re import L
from turtle import color
import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd
import numpy as np
import bottleneck as bn
import re
import datetime
import glob

import winsound


#どんな動きをさせるのか
#1. 銘柄データ格納フォルダを取得
#2. 銘柄データを順番に開く
#3. 指定の列に日次データを計算
#4.     
#5.     
#6.     
#7.     
#8.     
#9.     
#10.    
#11.    
#12.    

#要確認事項
#1. file_

#プログラム

#testのため、仮フォルダ設定
dirmerge = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/"


stock_list = glob.glob(dirmerge + '*.xlsx')

#開始時間取得
t = datetime.datetime.now()
d = datetime.datetime.now()
d1 = d.strftime('%Y%m%d')
#開始時間取得

#testのため、exceptリストは使わなくておｋ
#except_list = ['C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/0000', 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/株式/0010', '0011', '0012', '0090', '0091', '0092', '0093', '0094', '0095', '0101', '0102', '0108', '0800', '0802']

for l in stock_list:
        if "None.xlsx" in l:
                continue #ファイル名に"None.xlsx"が含まれていた場合はスキップ

        wb = openpyxl.load_workbook(l)
        sheetstock = wb.worksheets[0]
        lastrow_stockbook = sheetstock.max_row+1
        lastcolumn_stockbook = sheetstock.max_column
        print("最終行は"+str(lastrow_stockbook))
        if sheetstock.cell(2,5).value is None or sheetstock.cell(2,5).value == '－':
                sheetstock.cell(2,5).value = 0
                print(sheetstock.cell(2,2).value + '_' + sheetstock.cell(2,3).value +'前日比の値がなかったので処理')
        else:
                pass
        if sheetstock.cell(lastrow_stockbook-1, 4).value == '－':
                sheetstock.cell(lastrow_stockbook-1,4).value = sheetstock.cell(lastrow_stockbook-2,4).value
                sheetstock.cell(lastrow_stockbook-1,11).value = 0
                sheetstock.cell(lastrow_stockbook-1,12).value = sheetstock.cell(lastrow_stockbook-2,4).value
                sheetstock.cell(lastrow_stockbook-1,13).value = sheetstock.cell(lastrow_stockbook-2,4).value
                sheetstock.cell(lastrow_stockbook-1,14).value = sheetstock.cell(lastrow_stockbook-2,4).value
                sheetstock.cell(lastrow_stockbook-1,15).value = sheetstock.cell(lastrow_stockbook-2,4).value
                print(sheetstock.cell(lastrow_stockbook-2,2).value + '_' + sheetstock.cell(lastrow_stockbook-2,3).value +'株価の値がなかったので処理')
        if sheetstock.cell(lastrow_stockbook-1,3).value == '株探からのお知らせ':
                sheetstock.delete_rows(lastrow_stockbook-1)
                print(str(sheetstock.cell(lastrow_stockbook-2,2).value) + '_' + str(sheetstock.cell(lastrow_stockbook-2,3).value) +'上場廃止対象のため削除')
                wb.save(l)
                continue
        for m in range(3,lastrow_stockbook):
                if sheetstock.cell(m,5).value is None or sheetstock.cell(m,5).value == '－':
                        sheetstock.cell(m,5).value = sheetstock.cell(m,4).value-sheetstock.cell(m-1,4).value
                        print(sheetstock.cell(m,2).value + '_' + sheetstock.cell(m,3).value +'：'+str(m)+'行目に前日比の値がなかったので処理')

                else:
                        pass


#配列宣言
        stockprice = [] #株価
        turnover = [] #出来高
        upprice = [] #上昇幅
        downprice = [] #下落幅
        stockprice_5dmovemean = [] #5日移動平均
        stockprice_25dmovemean = [] #25日移動平均
        stockprice_75dmovemean = [] #75日移動平均
        stockprice_10dmovemean = [] #10日移動平均
        stockprice_15dmovemean = [] #15日移動平均
        stockprice_20dmovemean = [] #20日移動平均
        stockprice_30dmovemean = [] #30日移動平均
        stockprice_65dmovemean = [] #13週移動平均
        stockprice_130dmovemean = [] #26週移動平均
        turnover_5movemean = [] #5日移動平均出来高
        turnover_10movemean = [] #10日移動平均出来高
        stockprice_25dmovestd = [] #25日移動標準偏差
        stockprice_65dmovestd = [] #13週移動標準偏差
#20230409追加
        stockprice_20dmovestd = [] #20日移動標準偏差ーボリンジャーバンド
        stockprice_20dmovedtl = [] #20日移動分散ーボリンジャーバンド
        stockprice_14dupmean = [] #14日上昇平均ーRSI
        stockprice_14ddownmean = [] #14日下落平均ーRSI
        stockprice_12dEmovemean = [] #12日指数平滑移動平均ーMACD
        stockprice_26dEmovemean = [] #26日指数平滑移動平均ーMACD
        stockprice_9dEmovemean = [] #9日指数平滑移動平均ーMACD
        macdscore=[] #macdスコア用配列
        diffarray=[] #macdスコア計算用
        diffarray_max=[]
        diffarray_min=[]



#シートの2行目~最終行をループ
        for i in range(2,lastrow_stockbook):

#A列を配列へ格納
#                print(i)
                stockprice = np.append(stockprice, sheetstock.cell(row=i, column=4).value) #株価の配列格納
                turnover = np.append(turnover, sheetstock.cell(row=i, column=11).value) #出来高の配列格納
                #前日比のセルから、upprice, downpriceの配列格納をしたい。
                #前日比マイナスの時はuppriceにゼロ、プラスの時はdownpriceにゼロを入れる形にする。

                         
                if int(sheetstock.cell(i, 5).value) >= 0:
                        upprice = np.append(upprice, sheetstock.cell(i, 5).value)
                        downprice = np.append(downprice, 0)
                        #print(sheetstock.cell(i,5).value)
                else:
                        upprice = np.append(upprice, 0)
                        downprice = np.append(downprice, abs(sheetstock.cell(i,5).value))
                
#売買代金2=VWAP(T)*出来高(K)
#                sheetstock.cell(row=i,column=25).value = sheetstock.cell(row=i,column=20).value * sheetstock.cell(row=i,column=11).value
                #ローソク足の分類分け
#陽線/陰線の場合分け
                if int(sheetstock.cell(row=i, column=12).value) > int(sheetstock.cell(row=i, column=15).value):
#                        print(str(i))
                        fill = PatternFill(patternType="solid", fgColor="00BFFF")
                        sheetstock.cell(row=i, column=228).value = '陰線'
                        sheetstock.cell(row=i, column=228).fill = fill
                elif int(sheetstock.cell(row=i, column=15).value) > int(sheetstock.cell(row=i, column=12).value):
                        fill = PatternFill(patternType="solid", fgColor="FF69B4")
                        sheetstock.cell(row=i,column=228).value = '陽線'
                        sheetstock.cell(row=i, column=228).fill = fill
                else:
                        sheetstock.cell(row=i, column=228).value = '十字線'
#                        print(str(i))


#実体の長さ
                candle = abs(int(sheetstock.cell(row=i,column=12).value) - int(sheetstock.cell(row=i, column=15).value))
#陰線の場合
                if sheetstock.cell(row=i, column=228).value == '陰線':
                        blue_beard_over = int(sheetstock.cell(row=i, column=13).value) - int(sheetstock.cell(row=i, column=12).value)
                        blue_beard_under = int(sheetstock.cell(row=i, column=15).value) - int(sheetstock.cell(row=i, column=14).value)
                        if candle > blue_beard_over and candle > blue_beard_under:
                                sheetstock.cell(row=i, column=227).value = '大'
                        elif candle >= blue_beard_over and candle <= blue_beard_under:
                                sheetstock.cell(row=i, column=227).value = '下影'
                        elif candle <= blue_beard_over and candle >= blue_beard_under:
                                sheetstock.cell(row=i, column=227).value = '上影'
                        elif candle < blue_beard_over and candle < blue_beard_under:
                                sheetstock.cell(row=i, column=227).value = '小'
                        else:
                                sheetstock.cell(row=i, column=227).value = '異常値'
#陽線の場合
                elif sheetstock.cell(row=i, column=228).value == '陽線':
                        red_beard_over = int(sheetstock.cell(row=i, column=13).value) - int(sheetstock.cell(row=i, column=15).value)
                        red_beard_under = int(sheetstock.cell(row=i, column=12).value) - int(sheetstock.cell(row=i, column=14).value)
                        if candle > red_beard_over and candle > red_beard_under:
                                sheetstock.cell(row=i, column=227).value = '大'
                        elif candle >= red_beard_over and candle <= red_beard_under:
                                sheetstock.cell(row=i, column=227).value = '下影'
                        elif candle <= red_beard_over and candle >= red_beard_under:
                                sheetstock.cell(row=i, column=227).value = '上影'
                        elif candle < red_beard_over and candle < red_beard_under:
                                sheetstock.cell(row=i, column=227).value = '小'
                        else:
                                sheetstock.cell(row=i, column=227).value = '異常値'
#        print(stockprice)
#        print(turnover)
#移動平均数値の取込
        if lastrow_stockbook > 7:
                stockprice_5dmovemean = bn.move_mean(stockprice, window=5)
                turnover_5dmovemean = bn.move_mean(turnover, window=5)

                if lastrow_stockbook > 12:
                        stockprice_10dmovemean = bn.move_mean(stockprice, window=10)
                        turnover_10dmovemean = bn.move_mean(turnover, window=10)
                        if lastrow_stockbook > 14:
                                stockprice_12dEmovemean = bn.move_mean(stockprice, window=12)
                                if lastrow_stockbook > 16:
                                        stockprice_14dupmean = bn.move_var(upprice, window=14)
                                        stockprice_14ddownmean = bn.move_var(downprice, window=14)
                                        if lastrow_stockbook > 17:
                                                stockprice_15dmovemean = bn.move_mean(stockprice, window=15)
                                                if lastrow_stockbook > 22:
                                                        stockprice_20dmovemean = bn.move_mean(stockprice, window=20)
                                                        stockprice_20dmovevar = bn.move_var(stockprice, window=20)
                                                        stockprice_20dmovestd = bn.move_std(stockprice, window=20)
                                                        if lastrow_stockbook > 27:
                                                                stockprice_25dmovemean = bn.move_mean(stockprice, window=25)
                                                                stockprice_25dmovestd = bn.move_std(stockprice, window=25)
                                                                if lastrow_stockbook > 28:
                                                                        stockprice_26dEmovemean = bn.move_mean(stockprice, window=26)
                                                                        if lastrow_stockbook > 32:
                                                                                stockprice_30dmovemean = bn.move_mean(stockprice, window=30)
                                                                                if lastrow_stockbook > 67:
                                                                                        stockprice_65dmovemean = bn.move_mean(stockprice, window=65)
                                                                                        stockprice_65dmovestd = bn.move_std(stockprice, window=65)
                                                                                        if lastrow_stockbook > 77:
                                                                                                stockprice_75dmovemean = bn.move_mean(stockprice, window=75)
                                                                                                if lastrow_stockbook > 132:
                                                                                                        stockprice_130dmovemean = bn.move_mean(stockprice, window=130)

                                                                        

#        print(stockprice1)
#        print('移動平均', bn.move_mean(stockprice, window = 5))
        for j in range(2, lastrow_stockbook):
                if lastrow_stockbook > 7:
                        sheetstock.cell(row=j,column=199).value = stockprice_5dmovemean[j-2]
                        sheetstock.cell(row=j,column=230).value = turnover_5dmovemean[j-2]

                        if lastrow_stockbook > 12:
                                sheetstock.cell(row=j,column=205).value = stockprice_10dmovemean[j-2]
                                sheetstock.cell(row=j,column=211).value = turnover_10dmovemean[j-2]
                                if lastrow_stockbook > 14:
                                        sheetstock.cell(row=j,column=317).value = stockprice_12dEmovemean[j-2] + (2/13)*(sheetstock.cell(j,4).value - stockprice_12dEmovemean[j-2])
                                        if lastrow_stockbook > 16:
                                                sheetstock.cell(row=j,column=312).value = stockprice_14dupmean[j-2]
                                                sheetstock.cell(row=j,column=313).value = stockprice_14ddownmean[j-2]
                                                sheetstock.cell(j,314).value = stockprice_14dupmean[j-2]/stockprice_14ddownmean[j-2]
                                                sheetstock.cell(j,315).value = 100-(100/(1+stockprice_14dupmean[j-2]/stockprice_14ddownmean[j-2]))
                                                if sheetstock.cell(j,315).value < 30:
                                                        sheetstock.cell(j,316).value = "売られすぎ"
                                                elif sheetstock.cell(j,315).value > 70:
                                                        sheetstock.cell(j,316).value = "買われすぎ"

                                                sheetstock.cell(j,324).value = 100-sheetstock.cell(j,315).value

                                                if lastrow_stockbook > 17:
                                                        sheetstock.cell(row=j,column=206).value = stockprice_15dmovemean[j-2]
                                                        if lastrow_stockbook > 22:
                                                                sheetstock.cell(row=j,column=207).value = stockprice_20dmovemean[j-2]
                                                                sheetstock.cell(row=j,column=301).value = stockprice_20dmovemean[j-2]
                                                                sheetstock.cell(row=j,column=302).value = (sheetstock.cell(row=j,column=4).value - stockprice_20dmovemean[j-2])**2
                                                                sheetstock.cell(row=j,column=303).value = stockprice_20dmovevar[j-2]
                                                                sheetstock.cell(row=j,column=304).value = stockprice_20dmovestd[j-2]
                                                                sheetstock.cell(row=j,column=305).value = stockprice_20dmovemean[j-2]+2*stockprice_20dmovestd[j-2]
                                                                sheetstock.cell(row=j,column=306).value = stockprice_20dmovemean[j-2]-2*stockprice_20dmovestd[j-2]
                                                                sheetstock.cell(j,325).value = ((sheetstock.cell(j,305).value-sheetstock.cell(j,4).value)/(sheetstock.cell(j,305).value-sheetstock.cell(j,306).value))*100
                                                                if sheetstock.cell(j,4).value - stockprice_20dmovemean[j-2] > 0:
                                                                        sheetstock.cell(row=j,column=307).value = "上側"
                                                                else:
                                                                        sheetstock.cell(row=j,column=307).value = "下側"

                                                                if lastrow_stockbook > 27:
                                                                        sheetstock.cell(row=j,column=200).value = stockprice_25dmovemean[j-2]
                                                                        if lastrow_stockbook > 28:
                                                                                sheetstock.cell(row=j,column=318).value = stockprice_26dEmovemean[j-2] + (2/27)*(sheetstock.cell(j,4).value - stockprice_26dEmovemean[j-2])
                                                                                sheetstock.cell(row=j,column=319).value = sheetstock.cell(row=j,column=317).value - sheetstock.cell(row=j,column=318).value
                                                                                if lastrow_stockbook > 32:
                                                                                        sheetstock.cell(row=j,column=208).value = stockprice_30dmovemean[j-2]
                                                                                        if lastrow_stockbook > 67:
                                                                                                sheetstock.cell(row=j,column=209).value = stockprice_65dmovemean[j-2]
                                                                                                if lastrow_stockbook > 77:
                                                                                                        sheetstock.cell(row=j,column=201).value = stockprice_75dmovemean[j-2]
                                                                                                        if lastrow_stockbook > 132:
                                                                                                                sheetstock.cell(row=j,column=210).value = stockprice_130dmovemean[j-2]
                                                                                        

                if lastrow_stockbook > 7:
                        if sheetstock.cell(j,15).value-sheetstock.cell(j,199).value>0 and sheetstock.cell(j-1,15).value-sheetstock.cell(j-1,199).value<=0:
                                sheetstock.cell(j,229).value=1
                        elif sheetstock.cell(j,15).value-sheetstock.cell(j,199).value<=0 and sheetstock.cell(j-1,15).value-sheetstock.cell(j-1,199).value>0:
                                sheetstock.cell(j,229).value=4
                        else:
                                pass

                if lastrow_stockbook > 7:
                        if sheetstock.cell(j,11).value-sheetstock.cell(j,230).value>0 and sheetstock.cell(j-1,11).value-sheetstock.cell(j-1,230).value<=0:
                                sheetstock.cell(j,231).value=1
                        elif sheetstock.cell(j,11).value-sheetstock.cell(j,230).value<=0 and sheetstock.cell(j-1,11).value-sheetstock.cell(j-1,230).value>0:
                                sheetstock.cell(j,231).value=4
                        else:
                                pass

                if lastrow_stockbook > 27:
                        sheetstock.cell(row=j,column=212).value = stockprice_25dmovemean[j-2]+stockprice_25dmovestd[j-2] #25日+1σ
                        sheetstock.cell(row=j,column=213).value = stockprice_25dmovemean[j-2]-stockprice_25dmovestd[j-2] #25日-1σ
                        sheetstock.cell(row=j,column=214).value = stockprice_25dmovemean[j-2]+2*stockprice_25dmovestd[j-2] #25日+2σ
                        sheetstock.cell(row=j,column=215).value = stockprice_25dmovemean[j-2]-2*stockprice_25dmovestd[j-2] #25日-2σ
                        sheetstock.cell(row=j,column=216).value = stockprice_25dmovemean[j-2]+3*stockprice_25dmovestd[j-2] #25日+3σ
                        sheetstock.cell(row=j,column=217).value = stockprice_25dmovemean[j-2]-3*stockprice_25dmovestd[j-2] #25日-3σ

                if lastrow_stockbook > 67:
                        sheetstock.cell(row=j,column=218).value = stockprice_65dmovemean[j-2]+stockprice_65dmovestd[j-2] #13週+1σ
                        sheetstock.cell(row=j,column=219).value = stockprice_65dmovemean[j-2]-stockprice_65dmovestd[j-2] #13週-1σ
                        sheetstock.cell(row=j,column=220).value = stockprice_65dmovemean[j-2]+2*stockprice_65dmovestd[j-2] #13週+2σ
                        sheetstock.cell(row=j,column=221).value = stockprice_65dmovemean[j-2]-2*stockprice_65dmovestd[j-2] #13週-2σ
                        sheetstock.cell(row=j,column=222).value = stockprice_65dmovemean[j-2]+3*stockprice_65dmovestd[j-2] #13週+3σ
                        sheetstock.cell(row=j,column=223).value = stockprice_65dmovemean[j-2]-3*stockprice_65dmovestd[j-2] #13週-3σ

                if lastrow_stockbook > 27:
                        if sheetstock.cell(row=j, column=216).value<sheetstock.cell(row=j,column=4).value:
                                sheetstock.cell(row=j,column=224).value = 'プラスの異常値'
                        elif sheetstock.cell(row=j, column=216).value>=sheetstock.cell(row=j, column=4).value>sheetstock.cell(row=j,column=214).value:
                                sheetstock.cell(row=j, column=224).value = '+3σ範囲'
                        elif sheetstock.cell(row=j, column=214).value>=sheetstock.cell(row=j, column=4).value>sheetstock.cell(row=j,column=212).value:
                                sheetstock.cell(row=j, column=224).value = '+2σ範囲'
                        elif sheetstock.cell(row=j, column=212).value>=sheetstock.cell(row=j, column=4).value>sheetstock.cell(row=j,column=200).value:
                                sheetstock.cell(row=j, column=224).value = '+1σ範囲'
                        elif sheetstock.cell(row=j, column=200).value>=sheetstock.cell(row=j, column=4).value>sheetstock.cell(row=j,column=213).value:
                                sheetstock.cell(row=j, column=224).value = '-1σ範囲'
                        elif sheetstock.cell(row=j, column=213).value>=sheetstock.cell(row=j, column=4).value>sheetstock.cell(row=j,column=215).value:
                                sheetstock.cell(row=j, column=224).value = '-2σ範囲'
                        elif sheetstock.cell(row=j, column=215).value>=sheetstock.cell(row=j, column=4).value>sheetstock.cell(row=j,column=217).value:
                                sheetstock.cell(row=j, column=224).value = '-3σ範囲'
                        elif sheetstock.cell(row=j, column=217).value>sheetstock.cell(row=j, column=4).value:
                                sheetstock.cell(row=j, column=224).value = 'マイナスの異常値'
                        else:
                                sheetstock.cell(row=j, column=224).value = 'error'

                if lastrow_stockbook > 67:
                        if sheetstock.cell(row=j, column=222).value<sheetstock.cell(row=j,column=4).value:
                                sheetstock.cell(row=j,column=225).value = 'プラスの異常値'
                        elif sheetstock.cell(row=j, column=222).value>=sheetstock.cell(row=j, column=4).value>sheetstock.cell(row=j,column=220).value:
                                sheetstock.cell(row=j, column=225).value = '+3σ範囲'
                        elif sheetstock.cell(row=j, column=220).value>=sheetstock.cell(row=j, column=4).value>sheetstock.cell(row=j,column=218).value:
                                sheetstock.cell(row=j, column=225).value = '+2σ範囲'
                        elif sheetstock.cell(row=j, column=218).value>=sheetstock.cell(row=j, column=4).value>sheetstock.cell(row=j,column=209).value:
                                sheetstock.cell(row=j, column=225).value = '+1σ範囲'
                        elif sheetstock.cell(row=j, column=209).value>=sheetstock.cell(row=j, column=4).value>sheetstock.cell(row=j,column=219).value:
                                sheetstock.cell(row=j, column=225).value = '-1σ範囲'
                        elif sheetstock.cell(row=j, column=219).value>=sheetstock.cell(row=j, column=4).value>sheetstock.cell(row=j,column=221).value:
                                sheetstock.cell(row=j, column=225).value = '-2σ範囲'
                        elif sheetstock.cell(row=j, column=221).value>=sheetstock.cell(row=j, column=4).value>sheetstock.cell(row=j,column=223).value:
                                sheetstock.cell(row=j, column=225).value = '-3σ範囲'
                        elif sheetstock.cell(row=j, column=223).value>sheetstock.cell(row=j, column=4).value:
                                sheetstock.cell(row=j, column=225).value = 'マイナスの異常値'
                        else:
                                sheetstock.cell(row=j, column=225).value = 'error'

        for k in range(2, lastrow_stockbook):
                macdscore = np.append(macdscore, sheetstock.cell(row=k, column=319).value) #macdの配列格納
                if k > 11:
                        if sheetstock.cell(k,319).value is not None:
                                stockprice_9dEmovemean = bn.move_mean(macdscore, window=9)
                                sheetstock.cell(row=k,column=320).value = stockprice_9dEmovemean[k-2]
                                sheetstock.cell(k,323).value = sheetstock.cell(k,319).value-sheetstock.cell(k,320).value
                                diffarray = np.append(macdscore, sheetstock.cell(row=k, column=323).value) #macdの配列格納
                                diffarray_max = bn.move_max(macdscore, window=9)
                                diffarray_min = bn.move_min(macdscore, window=9)
                                macd = sheetstock.cell(k,323).value
                                if macd >= 0:
                                        sheetstock.cell(k,326).value = (1-(sheetstock.cell(k,323).value-diffarray_min[k-2])/(diffarray_max[k-2]-diffarray_min[k-2]))*100
                                else:
                                        sheetstock.cell(k,326).value =((sheetstock.cell(k,323).value-diffarray_min[k-2])/(diffarray_max[k-2]-diffarray_min[k-2]))*100
                
        for m in range(2,lastrow_stockbook):
                if sheetstock.cell(m,324).value is not None and sheetstock.cell(m,325).value is not None and sheetstock.cell(m,326).value is not None:
                        sheetstock.cell(m, 327).value = 0.5*sheetstock.cell(m,324).value + 0.2*sheetstock.cell(m,325).value + 0.3*sheetstock.cell(m,326).value

        #シグナルラインの計算
        #print(stockprice_25dmovemean[9])

        wb.save(l)
        print(l)

#終了時間取得-経過時間
print(t)
t1 = datetime.datetime.now()
print(t1)
dt = t1-t
print(dt)
#終了時間取得-経過時間

#稼働終了アナウンス
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾
#株価列はD=4列目
#13週(65日)移動平均     HA=209列目
#26週(130日)移動平均    HB=210列目
#移動平均5日線          GQ=199列目
#移動平均10日線         GW=205列目
#移動平均15日線         GX=206列目
#移動平均20日線         GY=207列目
#移動平均25日線         GR=200列目
#移動平均30日線         GZ=208列目
#ボリンジャーバンド 13週、25日
#標準偏差
#+1σ 25日               HD=212列目
#-1σ 25日               HE=213列目
#+2σ 25日               HF=214列目
#-2σ 25日               HG=215列目
#+3σ 25日               HH=216列目
#-3σ 25日               HI=217列目
#+1σ 13週               HJ=218列目
#-1σ 13週               HK=219列目
#+2σ 13週               HL=220列目
#-2σ 13週               HM=221列目
#+3σ 13週               HN=222列目
#-3σ 13週               HO=223列目
#25日ボリンジャーバンドでの株価位置   HP=224列目
#13週ボリンジャーバンドでの株価位置   HQ=225列目
#ローソク足分類         HQ=225列目




