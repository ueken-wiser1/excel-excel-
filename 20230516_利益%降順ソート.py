import os
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import datetime
import jpholiday
from datetime import timedelta
import pandas as pd

t = datetime.datetime.now().time()

# Excelファイルを読み込みます
watch_list_folder = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/追跡調査/調査期間満了/'
file_list = glob.glob(watch_list_folder+'*.xlsx')
sorted_data = []
for l in file_list:
    df = pd.read_excel(l)

# 並べ替えたい列の列番号（0から始まるインデックス）を指定します
    column_index = 30  # 例えば、3列目を指定する場合

    # 列番号を使用してデータフレームを並べ替えます
    sorted_df = df.sort_values(by=df.columns[column_index], ascending=False)
    sorted_df.to_excel(l, index=False)
    sorted_data.append(sorted_df)
    # 並べ替えた結果を表示します
    print(sorted_df)
    wb_watch_list = load_workbook(l)
    ws_watch_list = wb_watch_list.active
    last_row = ws_watch_list.max_row
    for i in range(2,last_row+1):
        ws_watch_list.cell(i,26).value = (last_row-i)/last_row*100
    wb_watch_list.save(l)



print(t)
t = datetime.datetime.now().time()
print(t)

#稼働終了アナウンス
import winsound
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾