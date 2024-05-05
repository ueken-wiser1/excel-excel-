'''#コメントここから
これは追跡開始初日データ処理を実施するコードで、現在のデータ用になります。
対象フォルダ
作業ファイルフォルダ：'C:/Users/touko/OneDrive/株価分析/excel/株式データ/追跡調査/0日目/'
参照ファイルフォルダ：'C:/Users/touko/OneDrive/株価分析/excel/株式データ/'
格納ファイルフォルダ：'C:/Users/touko/OneDrive/株価分析/excel/株式データ/追跡調査/1日目/'

実施する処理
OSCIファイルについて、最終列の日付を取得し、フォルダ(株式データ)内のファイルでその日付と同じ文字列を持つファイル(日次データ)を開く。
OSCIファイルの2列目の行の証券コードをキーにして、全行繰り返し。
OSCIファイルの証券コードを取得して、日次データをスキャンし、ヒットした行から、それぞれデータを取得し、OSCIファイルに転記。
始値、終値、終値-始値の値により、網掛けの色を変える、利益%が±2%を越えたら、TRUE/FALSE記載。
全行完了したら、最終列の日付の次の日付で、平日である日付を次の列に記載する。
次のOSCIファイルに移る。
'''#コメントここまで

import os
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import datetime
import jpholiday
import shutil

t = datetime.datetime.now().time()
d = datetime.datetime.now()
d1 = d.strftime('%Y%m%d')

def find_security_code_row(ws, code):
    for row2 in range(2, ws.max_row + 1):
        if int(ws.cell(row=row2, column=2).value) == int(code):
            #print(code)
            return row2
    raise ValueError(f"security code {code} not found")

watch_list_folder = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/追跡調査/0日目/'
completed_folder = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/'
dirstorage = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/追跡調査/1日目/'

file_list = glob.glob(watch_list_folder)

plus_fill = PatternFill(patternType='solid', fgColor='ee82ee') #前日より値上がり
minus_fill = PatternFill(patternType='solid', fgColor='00bfff') #前日より値下がり
attained_fill = PatternFill(patternType='solid', fgColor='adff2f') #+2%目標を達成したら
unattained_fill = PatternFill(patternType='solid', fgColor='696969') #-2%達成したら

for l in file_list:
    wb_watch_list = load_workbook(l)
    ws_watch_list = wb_watch_list.active

    last_row = ws_watch_list.max_row
    #print('最終行は'+str(last_row))
    last_col = ws_watch_list.max_column
    #print('最終列は'+str(last_col))
    if ws_watch_list.cell(row=2, column=2).value is None:
        print("この日は対象無し")
        continue
    basis_date = ws_watch_list.cell(row=2, column=2).value + datetime.timedelta(days=1)
    #print(basis_date)

    #print('対象日付は'+str(file_date))

    # 指定日が平日になるまでループ
    while True:
        # 指定日が土曜日または日曜日の場合
        if basis_date.weekday() == 5 or basis_date.weekday() == 6:
            #print(f"{basis_date}は土曜日または日曜日です。")
            # 指定日を翌日に上書き
            basis_date = basis_date + datetime.timedelta(days=1)
        # 指定日が祝日の場合
        elif jpholiday.is_holiday(basis_date):
            #print(f"{basis_date}は{str(jpholiday.is_holiday_name(basis_date))}です。")
            # 指定日を翌日に上書き
            basis_date = basis_date + datetime.timedelta(days=1)
        # 指定日が平日の場合
        else:
            #print(f"{basis_date}は平日です。")
            # ループを終了
            break
    file_date = basis_date.strftime('%Y%m%d')
    ws_watch_list.cell(row=1,column=last_col+1).value = basis_date

    wb_watch_list.save(l)
    print(l+'を保存しました')
    wb_watch_list.close()
    shutil.move(l, dirstorage)

print(t)
t = datetime.datetime.now().time()
print(t)

#稼働終了アナウンス
import winsound
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾