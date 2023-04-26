#5日線をプラス越えした銘柄リスト
#2022/10/28
#日付データフォルダに適用
#日付データnを開いたら、最終行取得
#2行目から最終行まで繰り返し
#プラス越えフラグあるデータを取得して、新規作成ファイルに記録

import os
from re import L
import openpyxl
import pandas
import requests
import bs4
import time
import datetime
import glob
import re
import sys
import shutil
from openpyxl.styles import PatternFill
import winsound

#開始時間取得
t = datetime.datetime.now()
d = datetime.datetime.now()
d1 = d.strftime('%Y%m%d')
#開始時間取得

#対象フォルダ指定
dirdaily = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/"
dirsim = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/test/20230115/検討/"
dirstorage = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/完了/"

file_list = glob.glob(dirdaily + '*.xlsx')

#日付データのリストをglob関数で作成
for l in file_list:
    #日付データを順番に開く
    wb_daily = openpyxl.load_workbook(l)
    wb_sim = openpyxl.Workbook()
    sheetsim = wb_sim.worksheets[0]
    print(l)
    sheetdaily = wb_daily.worksheets[0]
    lastrow = sheetdaily.max_row+1
    lastcolumn = sheetdaily.max_column
    target=0
    daycode = sheetdaily.cell(2,1).value
    daycode_format = daycode.strftime('%Y%m%d')
    sheetsim.cell(1,1).value = daycode
    sheetsim.cell(1,2).value = "証券コード"
    sheetsim.cell(1,3).value = "会社名"
    sheetsim.cell(1,4).value = "株価"
    sheetsim.cell(1,5).value = "前日比"
    sheetsim.cell(1,6).value = "前日終値"
    sheetsim.cell(1,7).value = "VWAP"
    sheetsim.cell(1,8).value = "VWAP/株価"
    sheetsim.cell(1,9).value = "出来高"
    sheetsim.cell(1,10).value = "約定回数"
    sheetsim.cell(1,11).value = "買閾値"
    sheetsim.cell(1,12).value = "静観閾値"
    sheetsim.cell(1,13).value = "出来高/出来高移動平均"
    sheetsim.cell(1,14).value = "ローソク足の長さ"
    sheetsim.cell(1,15).value = "値幅"
    sheetsim.cell(1,16).value = "ヒゲの長さ/ローソク足の長さ"
    sheetsim.cell(1,17).value = "決算日"
    sheetsim.cell(1,18).value = "IR情報"
    sheetsim.cell(1,19).value = "業界"
    sheetsim.cell(1,20).value = "1株配当"
    sheetsim.cell(1,21).value = "株価/1株配当"
    sheetsim.cell(1,22).value = "信用規制"
    sheetsim.cell(1,23).value = "増担措置"
    sheetsim.cell(1,24).value = "増担措置内容"
    sheetsim.cell(1,25).value = "空売規制"
    k=2
    for i in range(2, lastrow):
        if sheetdaily.cell(i,229).value==4 and sheetdaily.cell(i,4).value <= 1000:
            if sheetdaily.cell(i,231).value==4:
                fill = PatternFill(patternType='solid', fgColor='008080')
            else:
                fill = PatternFill(patternType='solid', fgColor='ffffff')

            stockcode = sheetdaily.cell(i,2).value
            stockname = sheetdaily.cell(i,3).value
            print(stockcode + '＿' + stockname)
#項番
            sheetsim.cell(k,1).value = k-1
#証券コード
            sheetsim.cell(k,2).value = sheetdaily.cell(i,2).value
            sheetsim.cell(k,2).fill = fill
#会社名
            sheetsim.cell(k,3).value = sheetdaily.cell(i,3).value
            sheetsim.cell(k,3).fill = fill
#株価
            sheetsim.cell(k,4).value = sheetdaily.cell(i,4).value
            sheetsim.cell(k,4).fill = fill
#前日比
            sheetsim.cell(k,5).value = sheetdaily.cell(i,5).value
            sheetsim.cell(k,5).fill = fill
#前日終値
            sheetsim.cell(k,6).value = sheetdaily.cell(i,10).value
            sheetsim.cell(k,6).fill = fill
#VWAP
            sheetsim.cell(k,7).value = sheetdaily.cell(i,20).value
            sheetsim.cell(k,7).fill = fill
#買い参考VWAP/株価ー1以下なら安値で買われている様子、1以上なら高値で買われている様子
            if sheetsim.cell(k,7).value == "－":
                sheetsim.cell(k,7).value = 0
            else:
                pass
            sheetsim.cell(k,8).value = sheetsim.cell(k,7).value / sheetsim.cell(k,4).value
            sheetsim.cell(k,8).fill = fill
#出来高
            sheetsim.cell(k,9).value =sheetdaily.cell(i,11).value
            sheetsim.cell(k,9).fill = fill
#約定回数
            sheetsim.cell(k,10).value =sheetdaily.cell(i,8).value
            sheetsim.cell(k,10).fill = fill
#買い閾値 株価の+1%
            sheetsim.cell(k,11).value =sheetdaily.cell(i,4).value*1.01
            sheetsim.cell(k,11).fill = fill
#静観閾値 株価の-3%
            sheetsim.cell(k,12).value =sheetdaily.cell(i,4).value*0.97
            sheetsim.cell(k,12).fill = fill
#出来高/出来高移動平均
            sheetsim.cell(k,13).value = sheetdaily.cell(i,11).value/sheetdaily.cell(i,230).value
            sheetsim.cell(k,13).fill = fill
#ローソク足の長さ
            sheetsim.cell(k,14).value = abs(sheetdaily.cell(i,12).value-sheetdaily.cell(i,15).value)
            sheetsim.cell(k,14).fill = fill
#値幅
            sheetsim.cell(k,15).value =sheetdaily.cell(i,13).value-sheetdaily.cell(i,14).value
            sheetsim.cell(k,15).fill = fill
#ヒゲの長さ/ローソク足の長さ
            if sheetsim.cell(k,14).value ==0:
                pass
            else:
                sheetsim.cell(k,16).value =(sheetsim.cell(k,15).value-sheetsim.cell(k,14).value)/sheetsim.cell(k,14).value
                sheetsim.cell(k,16).fill = fill
#決算日
            sheetsim.cell(k,17).value = sheetdaily.cell(i,9).value
            sheetsim.cell(k,17).fill = fill
#IR情報
            sheetsim.cell(k,18).value =sheetdaily.cell(i,51).value
            sheetsim.cell(k,18).fill = fill
#業界
            sheetsim.cell(k,19).value =sheetdaily.cell(i,34).value
            sheetsim.cell(k,19).fill = fill
#1株配当
            sheetsim.cell(k,20).value =sheetdaily.cell(i,32).value
            sheetsim.cell(k,20).fill = fill
#1株配当/株価
            if sheetsim.cell(k,20).value is None or sheetsim.cell(k,20).value =="－":
                sheetsim.cell(k,20).value =0
            else:
                pass
            #print(sheetsim.cell(k,20).value)
            #print(sheetsim.cell(k,4).value)
            sheetsim.cell(k,21).value =sheetsim.cell(k,20).value/sheetsim.cell(k,4).value
            sheetsim.cell(k,21).fill = fill
#信用規制
            sheetsim.cell(k,22).value =sheetdaily.cell(i,101).value
            sheetsim.cell(k,22).fill = fill
#増担措置
            sheetsim.cell(k,23).value =sheetdaily.cell(i,102).value
            sheetsim.cell(k,23).fill = fill
#増担措置内容
            sheetsim.cell(k,24).value =sheetdaily.cell(i,103).value
            sheetsim.cell(k,24).fill = fill
#空売規制
            sheetsim.cell(k,25).value =sheetdaily.cell(i,106).value
            sheetsim.cell(k,25).fill = fill
            k+=1
        else:
            pass

    
        wb_sim.save(dirsim+daycode_format+'_'+'sel.xlsx')
        
#os.rename(dirdaily+'allkabu1.xlsx', dirdaily+d1+'_allkabu1.xlsx')
#new_path = shutil.move(dirdaily+d1+'_allkabu1.xlsx', dirstorage)


#終了時間取得-経過時間
print(t)
t1 = datetime.datetime.now()
print(t1)
dt = t1-t
print(dt)
#終了時間取得-経過時間

#稼働終了アナウンス
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾

#    for j in range(2, lastrow):
#        if sheetdaily.cell(j,229).value=='プラス超え':
#            target+=1
#        else:
#            pass
#    print(target)


#        print(target)