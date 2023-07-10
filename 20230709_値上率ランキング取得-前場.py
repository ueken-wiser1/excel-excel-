import openpyxl
from openpyxl import Workbook
import os
import requests
import bs4
import time
import datetime
import winsound

#------------お約束開始---冒頭
#稼働時間計測開始
import datetime
t = datetime.datetime.now().time()
#------------お約束終了---冒頭

'''
目的
前場開始直後の値上がり率ランキングの推移から、利益が得られそうかの感触を見る。

仕様
株探ページから値上がり率ランキングを取得。
9:20-10:20に実施←これはタスクスケジューラで指定。
値上がり率ランキングページの50件を取得し、コード、銘柄名、株価、前日比、前日比％、出来高を合わせて記録。
日経も併せて記録。株価と前日比、前日比％のみ。
5分後、値上がり率ランキングページの50件を取得し、同様の記録。
前のランキングリストのコードを、今のランキングリストでスキャン。
ヒットすれば、パス。ヒットしなければ、コードから株探銘柄ページを取得し、同様の数字を記録。
前のランキングリストのコードを全てスキャンした今のランキングリストを新たなランキングリストとして、次のランキング取得時の参照とする。
重要な数字は値上がり率だから、順位自体はそこまで重要でない。
全てのリストが出来たら、新しいシートに、9:00段階のランキングを値上がり率と併せて記録する。次の行には出来高を記録する。
そのランキングの上からコードを参照して、次のランキングリストをスキャン。
ヒットすれば、その時の値上がり率、出来高を記録し、次のランキングリストに移る。
9:00段階のランキングリストを全て終わったら、次のランキングリストに移り、コードを上から参照し、前のランキングリストでスキャン。
ヒットすればパス。ヒットしなければ、値上がり率、出来高を記録し、次のランキングリストをスキャン。
これを繰り返し、10:00までのランキングリストのソート完了。

1行目は日付と時間の記載。
使うフォルダは新規に設定。
ファイルは毎日新規に作成。
次のランキングの記録は右に続けていく。
'''

#------------プログラム本文---ここから

#ファイル準備-指定のフォルダにワークブックを作成。シートとして、データを入力する"data"とデータから結果を導く"summary"のシートを作る。
dir = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/値上がりランキング調査/前場/"
wb = Workbook()
sheet01 = wb.create_sheet(title="data")
sheet02 = wb.create_sheet(title="summary")
today = datetime.date.today()
d = today.strftime('%Y%m%d')
wb.save(dir+d+"_値上がりランキング-前場.xlsx")
#記録開始行、列を定義。
i = 2
j = 1
k = 2

#株探ページから、ランキングデータをスクレイピングする。
#ランキングデータの必要なデータ位置を定義。
row_indices = list(range(5,20))
cell_indices = [0, 1, 2, 5, 7, 8, 9]

#繰り返し　9:00ー10:00の13回
for n in range(1,13):

    kabutan_ranking_URL = 'https://kabutan.jp/warning/?mode=2_1'
    res = requests.get(kabutan_ranking_URL)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')
    rows = soup.find_all('tr')
    time.sleep(1)

    n_now       = 8*(n-1)+2
    #print(n_now)

    t = datetime.datetime.now().time()
    sheet01.cell(1,n_now-1).value = t
    sheet01.cell(1,n_now+0).value = "コード"
    sheet01.cell(1,n_now+1).value = "会社名"
    sheet01.cell(1,n_now+2).value = "市場"
    sheet01.cell(1,n_now+3).value = "株価"
    sheet01.cell(1,n_now+4).value = "値上幅"
    sheet01.cell(1,n_now+5).value = "値上率"
    sheet01.cell(1,n_now+6).value = "出来高"
#ランキングデータの必要なデータ位置を定義。
    if n == 1: 
        i = 2

        #ランキングデータの取得
        for row_index in row_indices:
            if len(rows) > row_index:
                row = rows[row_index]
                cells = row.find_all(['td', 'th'])
                k = 0
                for cell_index in cell_indices:
                    if len(cells) > cell_index:

                        cell = cells[cell_index]
                        sheet01.cell(i, n_now+k).value = cell.text
                        j += 1
                        k += 1
                i += 1
    #nが2以上なら、その前のランキングデータがあるから、前のデータを参照して、今のデータになり銘柄についてデータを個別処理する。
        #for k in range(2,15):
            #print(sheet01.cell(k, n_now).value)
    #    print(n)
        #print(sheet01.cell(2,10).value)
        #code_nowを上から順番に取り出す。
        #取り出したcode_nowをcode_beforeと比較する。
        #もしcode_beforeがnone、またはcode_now=code_beforeであれば、そのループはスキップ。：code_beforeが終わった or code_nowと同じコードがcode_beforeにあったから。
        #もし、code_beforeがnoneでなく、かつ全てのcode_beforeとcode_nowが一致しなければ、それはそのcode_nowは新規に出現したコードであるから、それは記録する。　
    if n >= 2:
        i = 2
        for row_index in row_indices:
            if len(rows) > row_index:
                row = rows[row_index]
                cells = row.find_all(['td', 'th'])
                k = 0
                for cell_index in cell_indices:
                    if len(cells) > cell_index:

                        cell = cells[cell_index]
                        sheet01.cell(i, n_now+k).value = cell.text
                        print(n_now+k)
                        j += 1
                        k += 1
                i += 1
        
        n_before       = 8*(n-2)+2
        lastrow = sheet01.max_row
        #print(lastrow)
        for l in range(2, lastrow+1):
            code_before    = sheet01.cell(l,n_before).value
            if code_before is None:
                continue
            flag = False
            a = 1
            for m in range(2, lastrow+1):
                code_now = sheet01.cell(m,n_now).value
                if flag == True:
                    continue
                else:

                    if code_before != code_now:
                        continue
                    else:
                        flag = True
                        print(code_now+'は直前に取得した銘柄と一致しました。')
                
                if flag == False and code_before is not None:
                    #print(flag)
                    kabutan_URL_base = 'http://kabutan.jp/stock/?code='
                    kabutan_URL = kabutan_URL_base + str(code_before)
                    res = requests.get(kabutan_URL)
                    res.raise_for_status()
                    soup = bs4.BeautifulSoup(res.text, 'html.parser')
                    name_elem = soup.select('span')[8].text
                    market_elem = soup.select('span')[11].text
                    price_elem = soup.select('td')[32].text
                    up_elem = soup.select('span')[14].text
                    upper_elem = soup.select('span')[15].text
                    amount = soup.select('td')[35].text
                    #取得したデータをexcelに書き込む。
                    sheet01.cell(lastrow+a, n_now).value = code_before
                    sheet01.cell(lastrow+a, n_now+1).value = name_elem
                    sheet01.cell(lastrow+a, n_now+2).value =market_elem
                    sheet01.cell(lastrow+a, n_now+3).value =price_elem
                    sheet01.cell(lastrow+a, n_now+4).value =up_elem
                    sheet01.cell(lastrow+a, n_now+5).value =upper_elem
                    sheet01.cell(lastrow+a, n_now+6).value =amount
                    a += 1
                    time.sleep(1)
    print(n)
    time.sleep(300)
wb.save(dir+d+"_値上がりランキング-前場.xlsx")


#------------プログラム本文---ここまで

#------------お約束開始---末尾
#稼働時間表示
print(t)
t = datetime.datetime.now().time()
print(t)

#稼働終了アナウンス
import winsound
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾