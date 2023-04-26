import requests
import bs4
import os
import openpyxl
import time
import datetime
import urllib


#どんな動きをさせるのか
#株探URLに読み込んだ銘柄コードを組み合わせて銘柄ページに移動
#行【銘柄コード】に数値があれば、値をURLに組み込む
#       print(code_row)
kabutan_URL = 'https://kabutan.jp/stock/?code=1301'
kabutankessan_URL = 'https://kabutan.jp/stock/finance/?code=1301'

#銘柄ページから情報を読み込む
res2 = requests.get(kabutankessan_URL)
print(kabutan_URL)
res2.raise_for_status()
soup2 = bs4.BeautifulSoup(res2.text, 'html.parser')

#td_elem = soup.select('span')[1]
#print(td_elem)

#d_today = datetime.date.today()
#print(d_today)

#a = str(d_today) in str(td_elem)
#print(a)
#欲しい情報：<span class="market"><a href="/themes/?industry=25&market=2">
#tdtags=soup.find_all('td')       #全aタグ取得
#print('tdタグ数：', len(tdtags))  #aタグ数取得
tags = soup2.find_all('td')       #全aタグ取得
#print('tdタグ数：', len(tdtags))  #aタグ数取得a53 td25
num_tags = int(len(tags))
print(num_tags)

for i in range(num_tags):
    print(i)
#    print(d_today)
    td_elem = soup2.select('td')[i]
#    td_elem2 = soup.select('td')[i+2].get_text()
    print(td_elem)

#    if str(d_today) in str(td_elem):
#        print(td_elem2)





'''
#情報
#読み込んだ情報をexcelファイルに書き込む
            write_column = 2
#           名称
            stock_name = soup.select('title')[0]
            print(stock_name)
#           print(code_row)
#           print(write_column)
            sheet.cell(row=code_row, column=write_column).value = str(soup.select('span')[8])
#            wb.save('kabu.xlsx')
#           print(sheet.cell(row=code_row, column=write_column).value)
            write_column = write_column + 1
#           現在株価
#           stock_price = soup.select('span')[13]
            sheet.cell(row=code_row, column=write_column).value = str(soup.select('span')[14])
#           stock_price = sheet.cell(row=code_row, column=write_column).value
            write_column = write_column + 1
#           PER
#           stock_PER = soup.select('td')[17]
            sheet.cell(row=code_row, column=write_column).value = str(soup.select('td')[18])
#           stock_PER = sheet.cell(row=code_row, column=write_column).value
            write_column = write_column + 1
#           PBR
#           stock_PBR = soup.select('td')[18]
            sheet.cell(row=code_row, column=write_column).value = str(soup.select('td')[19])
#           stock_PBR = sheet.cell(row=code_row, column=write_column).value
            write_column = write_column + 1
#           利回り
#           stock_yield = soup.select('td')[19]
            sheet.cell(row=code_row, column=write_column).value = str(soup.select('td')[20])
#           stock_yield = sheet.cell(row=code_row, column=write_column).value
            write_column = write_column + 1
#           出来高
#           stock_volume = soup.select('td')[34]
            sheet.cell(row=code_row, column=write_column).value = str(soup.select('td')[35])
#           stock_volume = sheet.cell(row=code_row, column=write_column).value
            write_column = write_column + 1
#           時価総額
#           stock_capital = soup.select('td')[40]
            sheet.cell(row=code_row, column=write_column).value = str(soup.select('td')[41])
#           stock_capital = sheet.cell(row=code_row, column=write_column).value
            write_column = write_column + 1
#           発行済み株式数
#           stock_issued_number = soup.select('td')[41]
            sheet.cell(row=code_row, column=write_column).value = str(soup.select('td')[42])
#           stock_issued_number = sheet.cell(row=code_row, column=write_column).value
#           write_column = write_column + 1
#           最新信用売残
#           stock_unsold_credit = soup.select('td')[45]
            sheet.cell(row=code_row, column=write_column).value = str(soup.select('td')[46])
#           stock_unsold_credit = sheet.cell(row=code_row, column=write_column).value
            write_column = write_column + 1
#           最新信用買残
#           stock_unpurchace_credit = soup.select('td')[46]
            sheet.cell(row=code_row, column=write_column).value = str(soup.select('td')[47])
#           stock_unpurchace_credit = sheet.cell(row=code_row, column=write_column).value
            write_column = write_column + 1
#           1株配当
#           stock_dividend = soup.select('td')[82]
            sheet.cell(row=code_row, column=write_column).value = str(soup.select('td')[82])
#           stock_dividend = sheet.cell(row=code_row, column=write_column).value
            write_column = write_column + 1
#           決算タブに移動
            kabutan_URL_base_finance = 'http://kabutan.jp/stock/finance?code='
            kabutan_URL_finance = kabutan_URL_base_finance + str(stock_code)
            res = requests.get(kabutan_URL_finance)
#           print(stock_code)
#           print(kabutan_URL)
            res.raise_for_status()
            soup = bs4.BeautifulSoup(res.text, 'html.parser')
            tag_title = soup.find_all('title')
#           print(tag_title)

#           総資産回転率
#           stock_issued_number = soup.select('td')[225]
            sheet.cell(row=code_row, column=write_column).value = str(soup.select('td')[226])
#           stock_issued_number = sheet.cell(row=code_row, column=write_column).value
            write_column = write_column + 1

#           1秒待機
            time.sleep(1)
    code_row += 1
#
#           print(sheet.cell(row=code_row, column=2).value)
#sheet.cell(row=4, column=2).value = 'hello world'
#span_elem2 = sheet.cell(row=4, column=2).value
#span_elem = soup.select('span')[14]
#print(span_elem)

#書き込んだexcelを保存する

    wb.save('kabu1.xlsx')

    month_number = month_number + 1



#soup = bs4.BeautifulSoup(open('http://kabutan.jp/stock/?code=2292.html'))
#title_elem =soup.select('id')[0]
#print(title_elem.get('content'))




#for i in range(10):
#    print(i)
#    td_elem = soup.select('span')[i]
#    print(td_elem)


#name_text = soup.find('title').get_text()
#print(name_text)
#type(kabu.txt)

#print(res.status_code == requests.codes.ok)
#print(len(res.text))
#print(res.text[:250])

#soup = bs4.BeautifulSoup(open(c))
#title_elem =soup.select('id')[0]
#print(title_elem.get('content'))
'''