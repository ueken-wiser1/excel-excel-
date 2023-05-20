
import os
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import datetime
import jpholiday

import shutil


t = datetime.datetime.now().time()
d = datetime.datetime.now()
d1 = d.strftime('%Y%m%d')

def find_date_row(ws, code):
    for row2 in range(2, ws.max_row + 1):

        if ws.cell(row=row2, column=2).value == code:
            return row2
    return None

watch_list_folder = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/追跡調査/現在/1日目/'
daily_folder = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/'
dirstorage = 'C:/Users/touko/OneDrive/株価分析/excel/株式データ/追跡調査/現在/2日目以降/'

plus_fill = PatternFill(patternType='solid', fgColor='ee82ee') #前日より値上がり
minus_fill = PatternFill(patternType='solid', fgColor='00bfff') #前日より値下がり
attained_fill = PatternFill(patternType='solid', fgColor='adff2f') #+2%目標を達成したら
unattained_fill = PatternFill(patternType='solid', fgColor='696969') #+2%目標未達で、5日目を迎えたら

file_list = glob.glob(watch_list_folder + '*.xlsx')

for l in file_list:
    file_name = os.path.basename(l)
    print(str(file_name)+"に対し作業中。")
    wb_watch_list = load_workbook(l)
    ws_watch_list = wb_watch_list.active

    last_row_watch = ws_watch_list.max_row
    last_col_watch = ws_watch_list.max_column

    last_date = ws_watch_list.cell(row=1, column=last_col_watch).value
    last_date_str = last_date.strftime('%Y%m%d')
    daily_data_files = glob.glob(daily_folder+'*allkabu1.xlsx')
    daily_data_file = daily_data_files[0]
    if os.path.exists(daily_data_file):
        wb_daily_data = load_workbook(daily_data_file)
        ws_daily_data = wb_daily_data.active
        last_row_daily=ws_daily_data.max_row

    for i in range(2, last_row_watch + 1):

        stock_code = ws_watch_list.cell(row=i, column=4).value
        company_name= ws_watch_list.cell(row=i, column=5).value

        code_row = find_date_row(ws_daily_data, stock_code)
        print(str(daily_data_file) +"から"+str(stock_code)+"_"+str(company_name)+"を検索中")

        if code_row is None:
            print("見当たらず")
            continue  
        daily_date0 = ws_daily_data.cell(row=2,column=1).value
        if ws_watch_list.cell(row=1, column=last_col_watch).value == daily_date0:
            closing_price = ws_daily_data.cell(row=code_row, column=4).value
            start_price = ws_daily_data.cell(row=code_row, column=12).value
            
            ws_watch_list.cell(row=i, column=last_col_watch).value = closing_price
            ws_watch_list.cell(row=i, column=29).value = start_price
            print("証券コード"+str(stock_code) +"は、終値"+str(ws_watch_list.cell(row=i, column=last_col_watch).value)+"を示しました")

            diff=closing_price - start_price
            if diff > 0:
                ws_watch_list.cell(row=i, column=last_col_watch).fill = plus_fill
            elif diff < 0:
                ws_watch_list.cell(row=i,column=last_col_watch).fill = minus_fill

            if closing_price is not None and start_price is not None:
                ratio = (diff / start_price)*100
                ws_watch_list.cell(row=i, column=31).value = ratio
                ws_watch_list.cell(row=i, column=30).value = diff
                if ratio > 2:
                    ws_watch_list.cell(row=i, column=1).value = True
                    for j in range(1, last_col_watch):
                        ws_watch_list.cell(row=i, column=j).fill =attained_fill

                    profit = ws_watch_list.cell(row=i,column=last_col_watch).value - ws_watch_list.cell(row=i,column=29).value
                    print(str(ws_watch_list.cell(row=i,column=4).value)+"_"+str(ws_watch_list.cell(row=i,column=5).value) + "は2%目標を達成しました。利益は"+str(profit)+"円です。")

                elif ratio < -2:
                    ws_watch_list.cell(row=i, column=1).value = False
                    for j in range(1, last_col_watch):
                        ws_watch_list.cell(row=i, column=j).fill =unattained_fill
                    loss = ws_watch_list.cell(row=i,column=last_col_watch).value - ws_watch_list.cell(row=i,column=29).value
                    print(str(ws_watch_list.cell(row=i,column=4).value)+"_"+str(ws_watch_list.cell(row=i,column=5).value) + "は-2%目標に達してしまいました。損失は"+str(loss)+"円です。")

            last_date2 = last_date + datetime.timedelta(days=1)

            while True:

                if last_date2.weekday() == 5 or last_date2.weekday() == 6:

                    last_date2 = last_date2 + datetime.timedelta(days=1)

                elif jpholiday.is_holiday(last_date2):

                    last_date2 = last_date2 + datetime.timedelta(days=1)

                else:

                    break

            ws_watch_list.cell(row=1,column=last_col_watch+1).value = last_date2

            wb_daily_data.close()
    last_col_last = ws_watch_list.max_column
    for i in range(last_col_last+1, 1, -1):

        if ws_watch_list.cell(row=1, column=i).value is None:
            ws_watch_list.delete_cols(i)
            print(str(i)+"列目は1行目に何もないため削除しました。")
    wb_watch_list.save(l)

    print(l+'を保存しました')

    wb_watch_list.close()
    shutil.move(l, dirstorage)
    
print(t)
t = datetime.datetime.now().time()
print(t)

#稼働終了アナウンス
import winsound
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(500,100)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
winsound.Beep(750,50)  #ビープ音（500Hzの音を50msec流す）
#------------お約束終了---末尾