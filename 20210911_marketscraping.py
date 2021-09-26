import os
import openpyxl
import requests
import bs4
import time
import datetime
import sys
import sys
from selenium import webdriver
from selenium.webdriver.common.keys import Keys as keys
import winsound

#どんな動きをさせるのか
#excelを開く
#参照excelの各シートに記載された証券コードを読み込む
#シート名は"株式", "マーケット", "為替", "投信"
#読み込んだ証券コードにより株探URLを作成し、銘柄ページへ移動
#必要な情報を取得する

#excelを開く

dir = "C:/Users/touko/OneDrive/株価分析/excel/株式データ/スクレイピング生データ/"
wb = openpyxl.load_workbook(dir + 'stockcodelist.xlsx')
name = wb.get_sheet_names

#"株式"
sheet = wb.get_sheet_by_name('株式')
print(sheet.title)
print(name)

#読み込んだ証券コードにより株探URLを作成し、銘柄ページへ移動
#必要な情報を取得する



#参照excelの各シートに記載された証券コードを読み込む
t = datetime.datetime.now().time()
#print(t)
'''
def AutoLogin():
  # 起動するブラウザを宣言します 
  browser = webdriver.Chrome('C:/Users/touko/program/chromedriver.exe') 
  # ログイン対象のWebページURLを宣言します 
  url = "https://account.kabutan.jp/login" 
  # 対象URLをブラウザで表示します。 
  browser.get(url)
  # ログインIdとパスワードの入力領域を取得します。 
  login_id = browser.find_element_by_xpath("//input[@id='session_email']") 
  login_pw = browser.find_element_by_xpath("//input[@id='session_password']")
  # ログインIDとパスワードを入力します。
  userid = "toukouikitai@hotmail.com" 
  userpw = "s4b4egqekabutan"
  login_id.send_keys(userid) 
  login_pw.send_keys(userpw)
  # ログインボタンをクリックします。 

  login_btn = browser.find_element_by_xpath(".//input[@type='submit']")
  login_btn.click()

# AutoLogin関数を実行します。
#
ret = AutoLogin()
'''

for j in range(2, sheet.max_row + 1):
    time.sleep(0.1)
    kabutan_URL_base = 'http://kabutan.jp/stock/?code='
    stock_code = sheet.cell(row=j, column=2).value
    kabutan_URL = kabutan_URL_base + str(stock_code)
    res = requests.get(kabutan_URL)
    print(kabutan_URL)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')

#読み込んだ情報をexcelファイルに書き込む
#書き出す初めのセルの列を指定
    write_column = 1

#日付を入れる
    d_today = datetime.date.today()
    sheet.cell(row=j, column=1).value = d_today
    write_column += 2
#    t = soup.select('time')[4]
#    print(t)
#           名称
    stock_name = soup.select('h3')[0]
    print(stock_name)
    sheet.cell(row=j, column=3).value = str(soup.select('h3')[0].get_text())
    write_column += 1

#           現在株価
    try:
        stock_price = str(soup.select('td')[32])
    except IndexError as e:
        print('現在株価存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
        write_column += 1

#           前日比
    try:
        DoD = str(soup.select('dd')[0])
    except IndexError as e:
        print('前日比存在しない')
        write_column += 2
        pass
    else:
        sheet.cell(row=j, column=write_column).value = DoD
        write_column += 2

#           売買代金
    try:
        dekidaka_yield = str(soup.select('td')[36].get_text())
    except IndexError as e:
        print('売買代金存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = dekidaka_yield
        write_column += 1

#           約定回数
    try:
        number_of_contracts = str(soup.select('td')[38].get_text())
    except IndexError as e:
        print('約定回数存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = number_of_contracts
        write_column += 1

#           決算日
    try:
        stock_price = str(soup.select('dd')[7].get_text())
    except IndexError as e:
        print('前日終値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
        write_column += 1

#           前日終値1
    try:
        stock_price = str(soup.select('dd')[7].get_text())
    except IndexError as e:
        print('前日終値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
        write_column += 1

#           前日終値2
    try:
        stock_price = str(soup.select('dd')[8].get_text())
    except IndexError as e:
        print('前日終値存在しない')
        write_column += 2
        pass
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
        write_column += 2

#           出来高
    try:
        dekidaka_yield = str(soup.select('td')[35].get_text())
    except IndexError as e:
        print('出来高存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = dekidaka_yield
        write_column += 1

#           始値
    try:
        stock_price = str(soup.select('td')[23])
    except IndexError as e:
        print('始値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
        write_column += 1

#           高値
    try:
        stock_price = str(soup.select('td')[26])
    except IndexError as e:
        print('高値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
        write_column += 1

#           安値
    try:
        stock_price = str(soup.select('td')[29])
    except IndexError as e:
        print('安値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
        write_column += 1

#           終値
    try:
        stock_price = str(soup.select('td')[32])
    except IndexError as e:
        print('終値存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
        write_column += 1

#           前日比%
    try:
        DoD = str(soup.select('span')[16].get_text())
    except IndexError as e:
        print('前日比存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = DoD
        write_column += 1

#           PER
    try:
        kabu_PER = str(soup.select('td')[18].get_text())
    except IndexError as e:
        print('PER存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = kabu_PER
        write_column += 1

#           PBR
    try:
        kabu_PBR = str(soup.select('td')[19].get_text())
    except IndexError as e:
        print('PBR存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = kabu_PBR
        write_column += 1

#           上場市場
    try:
        stock_price = str(soup.select('span')[12].get_text())
    except IndexError as e:
        print('上場市場存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = stock_price
        write_column += 1

#           VWAP
    try:
        VWAP = str(soup.select('td')[37].get_text())
    except IndexError as e:
        print('VWAP存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = VWAP
        write_column += 1

#           発行済み株式数
    try:
        number_of_issued_shares = str(soup.select('td')[42].get_text())
    except IndexError as e:
        print('発行済み株式数存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = number_of_issued_shares
        write_column += 1

#           最新信用売残
    try:
        credit_unsold = str(soup.select('td')[46])
    except IndexError as e:
        print('最新信用売残存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = credit_unsold
        write_column += 1

#           最新信用買残
    try:
        credit_unpurchased = str(soup.select('td')[47])
    except IndexError as e:
        print('最新信用買残存在しない')
        write_column += 1
        pass
    else:
        sheet.cell(row=j, column=write_column).value = credit_unpurchased
        write_column += 1

#           信用倍率
    try:
        credit_ratio = str(soup.select('td')[48])
    except IndexError as e:
        print('信用倍率存在しない')
        write_column += 1
        pass
    else:
        
        sheet.cell(row=j, column=write_column).value = credit_ratio
        write_column += 1
    
print(t)
t = datetime.datetime.now().time()
print(t)

wb.save('allkabu1.xlsx')
winsound.Beep(500,50)  #ビープ音（500Hzの音を50msec流す）
